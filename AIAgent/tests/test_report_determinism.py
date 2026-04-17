"""
Report-layer determinism tests for the Full Tenant Assessment pipeline.

Validates that every report generator produces identical output when called
twice with the same evaluation results.  Timestamps and report IDs embedded
in display HTML/MD are stripped before comparison (Category B — volatile by
design).  Binary formats (PDF, Excel .xlsx) are validated structurally rather
than byte-for-byte.

Covers:
  - SARIF export          (Category A — data file, must be byte-identical)
  - OSCAL export          (Category A — data file, must be byte-identical)
  - Data exports CSV/JSON (already fixed — regression guard)
  - Compliance HTML       (Category B — strip timestamps)
  - Compliance MD         (Category B — strip timestamps)
  - Executive Summary HTML/MD (Category B)
  - Gaps Report HTML/MD   (Category B)
  - Master Report HTML/MD (Category B)
  - Executive Dashboard   (Category B)
  - Methodology Report    (Category B)
  - JSON Report           (Category B — filename + generated field)
  - Markdown Report       (Category B — filename + generated line)
  - HTML Report           (Category B — filename + generated line)
  - Excel Report          (Category C — structural equivalence)
"""

from __future__ import annotations

import copy
import json
import os
import pathlib
import re
import shutil
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.evaluators.engine import evaluate_all
from app.config import ThresholdConfig

# ── Report generators ────────────────────────────────────────────────
from app.reports.sarif_export import export_sarif
from app.reports.oscal_export import export_oscal
from app.reports.data_exports import export_data_files, save_raw_evidence
from app.reports.compliance_report_html import generate_compliance_report_html
from app.reports.compliance_report_md import generate_compliance_report_md
from app.reports.executive_summary import generate_executive_summary_html, generate_executive_summary_md
from app.reports.gaps_report import generate_gaps_report_html, generate_gaps_report_md
from app.reports.master_report import generate_master_report
from app.reports.executive_dashboard import generate_executive_dashboard
from app.reports.methodology_report import generate_methodology_html
from app.reports.json_report import generate_json_report
from app.reports.markdown_report import generate_markdown_report
from app.reports.html_report import generate_html_report
from app.reports.excel_export import generate_excel_report

# ── Volatile-field stripping ─────────────────────────────────────────

# Regex patterns for timestamps, report-IDs, and generated dates in HTML/MD
_TS_PATTERNS = [
    # ISO 8601 timestamps  e.g. 2026-04-08T12:34:56.123456+00:00  or  ...Z
    re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}[\.\d]*(?:\+\d{2}:\d{2}|Z)"),
    # Report IDs  e.g. CIQ-20260408-123456
    re.compile(r"CIQ-\d{8}-\d{6}"),
    # Date-time strings  e.g. 2026-04-08 12:34:56 UTC  or  April 08, 2026 12:34 UTC
    re.compile(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} UTC"),
    re.compile(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2} UTC"),
    re.compile(r"[A-Z][a-z]+ \d{2}, \d{4} \d{2}:\d{2} UTC"),
    re.compile(r"[A-Z][a-z]+ \d{2}, \d{4}"),
    # Bare YYYYMMDD-HHMMSS  e.g. in filenames
    re.compile(r"\d{8}-\d{6}"),
    # SHA-256 hashes (master report embeds one)
    re.compile(r"[a-f0-9]{64}"),
]


def _strip_timestamps(text: str) -> str:
    """Remove volatile timestamp/ID fragments from report text."""
    for pat in _TS_PATTERNS:
        text = pat.sub("__VOLATILE__", text)
    return text


# ── Shared frozen test fixtures ──────────────────────────────────────

TENANT_INFO = {
    "tenant_id": "00000000-0000-0000-0000-000000000001",
    "display_name": "Determinism-Test-Tenant",
    "TenantId": "00000000-0000-0000-0000-000000000001",
    "DisplayName": "Determinism-Test-Tenant",
}


def _build_frozen_evidence() -> list[dict]:
    """Minimal but realistic evidence set for report generation."""
    evidence = []
    # RBAC
    for i in range(5):
        role = "Owner" if i < 2 else "Reader"
        evidence.append({
            "EvidenceType": "azure-role-assignment",
            "Source": "Azure",
            "Collector": "collect_rbac",
            "ResourceId": f"/subscriptions/sub1/providers/Microsoft.Authorization/roleAssignments/ra-{i:03d}",
            "Data": {
                "RoleDefinitionName": role,
                "PrincipalDisplayName": f"user{i}@test.com",
                "PrincipalId": f"pid-{i:04d}",
                "PrincipalType": "User",
                "ScopeLevel": "Subscription",
                "IsPrivileged": i < 2,
                "ResourceId": f"/subscriptions/sub1/providers/Microsoft.Authorization/roleAssignments/ra-{i:03d}",
            },
        })
    # Identity — MFA
    evidence.append({
        "EvidenceType": "entra-mfa-summary",
        "Source": "Entra",
        "Collector": "collect_entra_users",
        "ResourceId": "tenant-mfa-summary",
        "Data": {"TotalUsers": 50, "MfaRegistered": 40, "MfaPercentage": 80.0, "NotRegistered": 10},
    })
    # CA policies
    for i in range(2):
        evidence.append({
            "EvidenceType": "entra-conditional-access-policy",
            "Source": "Entra",
            "Collector": "collect_entra_conditional_access",
            "ResourceId": f"ca-policy-{i}",
            "Data": {"DisplayName": f"CA Policy {i}", "State": "enabled", "Id": f"ca-{i}"},
        })
    # Storage
    evidence.append({
        "EvidenceType": "azure-storage-security",
        "Source": "Azure",
        "Collector": "collect_storage",
        "ResourceId": "/subscriptions/sub1/rg1/Microsoft.Storage/storageAccounts/store1",
        "Data": {
            "Name": "store1", "StorageAccountName": "store1",
            "AllowBlobPublicAccess": True, "EnableHttpsTrafficOnly": False, "HttpsOnly": False,
            "NetworkDefaultAction": "Allow", "MinimumTlsVersion": "TLS1_0",
            "BlobSoftDeleteEnabled": False,
            "id": "/subscriptions/sub1/rg1/Microsoft.Storage/storageAccounts/store1",
            "type": "microsoft.storage/storageaccounts",
        },
    })
    # SQL
    evidence.append({
        "EvidenceType": "azure-sql-server",
        "Source": "Azure",
        "Collector": "collect_databases",
        "ResourceId": "/subscriptions/sub1/rg1/Microsoft.Sql/servers/sql1",
        "Data": {
            "Name": "sql1", "TransparentDataEncryption": False, "TdeEnabled": False,
            "AuditingEnabled": False, "AdvancedThreatProtection": False, "MinimalTlsVersion": "1.0",
            "ResourceId": "/subscriptions/sub1/rg1/Microsoft.Sql/servers/sql1",
            "ResourceType": "Microsoft.Sql/servers",
        },
    })
    # VM
    evidence.append({
        "EvidenceType": "azure-vm-config",
        "Source": "Azure",
        "Collector": "collect_compute",
        "ResourceId": "/subscriptions/sub1/rg1/Microsoft.Compute/virtualMachines/vm1",
        "Data": {
            "Name": "vm1", "OsDiskEncrypted": False, "DataDiskCount": 1, "DataDisksEncrypted": False,
            "ResourceId": "/subscriptions/sub1/rg1/Microsoft.Compute/virtualMachines/vm1",
            "ResourceType": "Microsoft.Compute/virtualMachines",
        },
    })
    # Defender plans
    for plan in ["VirtualMachines", "SqlServers", "StorageAccounts"]:
        evidence.append({
            "EvidenceType": "azure-defender-plan",
            "Source": "Azure",
            "Collector": "collect_defender_plans",
            "ResourceId": f"/subscriptions/sub1/providers/Microsoft.Security/pricings/{plan}",
            "Data": {"Name": plan, "PricingTier": "Standard", "Enabled": True},
        })
    # Policy
    for i in range(3):
        evidence.append({
            "EvidenceType": "azure-policy-assignment",
            "Source": "Azure",
            "Collector": "collect_policy",
            "ResourceId": f"/subscriptions/sub1/providers/Microsoft.Authorization/policyAssignments/pa-{i}",
            "Data": {
                "DisplayName": f"Policy {i}",
                "ComplianceState": "Compliant" if i < 2 else "NonCompliant",
                "PolicyDefinitionId": f"/providers/Microsoft.Authorization/policyDefinitions/pd-{i}",
            },
        })
    # KeyVault
    evidence.append({
        "EvidenceType": "azure-keyvault",
        "Source": "Azure",
        "Collector": "collect_keyvault",
        "ResourceId": "/subscriptions/sub1/rg1/Microsoft.KeyVault/vaults/kv1",
        "Data": {
            "name": "kv1", "PurgeProtectionEnabled": False, "SoftDeleteEnabled": False,
            "EnableRbacAuthorization": False, "NetworkDefaultAction": "Allow",
            "id": "/subscriptions/sub1/rg1/Microsoft.KeyVault/vaults/kv1",
            "type": "microsoft.keyvault/vaults",
        },
    })
    # Diagnostics
    for i in range(3):
        evidence.append({
            "EvidenceType": "azure-diagnostic-setting",
            "Source": "Azure",
            "Collector": "collect_diagnostics",
            "ResourceId": f"/subscriptions/sub1/rg1/providers/res-{i}/diagnosticSettings/ds-{i}",
            "Data": {
                "Name": f"diag-{i}", "ResourceId": f"/subscriptions/sub1/rg1/providers/res-{i}",
                "Enabled": True, "LogCategories": ["AuditEvent"], "DestinationType": "LogAnalytics",
            },
        })
    # NSG
    evidence.append({
        "EvidenceType": "azure-nsg-rule",
        "Source": "Azure",
        "Collector": "collect_network",
        "ResourceId": "/subscriptions/sub1/rg1/Microsoft.Network/nsg/nsg1/rules/AllowAll",
        "Data": {
            "RuleName": "AllowAll", "Direction": "Inbound", "Access": "Allow",
            "SourceAddressPrefix": "*", "DestinationPortRange": "*", "Priority": 100,
            "NsgName": "nsg1",
        },
    })
    return evidence


@pytest.fixture(scope="module")
def frozen_evaluation():
    """Evaluate once and cache for the whole module."""
    evidence = _build_frozen_evidence()
    results = evaluate_all(
        evidence=evidence,
        frameworks=["FedRAMP"],
        thresholds=ThresholdConfig(),
    )
    return results, evidence


@pytest.fixture()
def tmp_output(tmp_path):
    """Provide two clean temporary output dirs for A/B comparison."""
    dir_a = tmp_path / "run_a"
    dir_b = tmp_path / "run_b"
    dir_a.mkdir()
    dir_b.mkdir()
    return str(dir_a), str(dir_b)


# ═══════════════════════════════════════════════════════════════════════
# Category A: Data files — must be byte-identical
# ═══════════════════════════════════════════════════════════════════════

class TestCategoryA_DataFiles:
    """Data-integrity files that must be byte-identical across runs."""

    def test_sarif_deterministic(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = export_sarif(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = export_sarif(results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        content_a = pathlib.Path(p1).read_text(encoding="utf-8")
        content_b = pathlib.Path(p2).read_text(encoding="utf-8")
        assert content_a == content_b, "SARIF output must be byte-identical across runs"

    def test_oscal_deterministic(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = export_oscal(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = export_oscal(results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        content_a = pathlib.Path(p1).read_text(encoding="utf-8")
        content_b = pathlib.Path(p2).read_text(encoding="utf-8")

        # Strip the volatile timestamp fields (last-modified, collected, start)
        content_a = _strip_timestamps(content_a)
        content_b = _strip_timestamps(content_b)
        assert content_a == content_b, "OSCAL output (modulo timestamps) must be identical"

    def test_oscal_uses_deterministic_uuids(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = export_oscal(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = export_oscal(results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        doc1 = json.loads(pathlib.Path(p1).read_text(encoding="utf-8"))
        doc2 = json.loads(pathlib.Path(p2).read_text(encoding="utf-8"))

        # Document UUID
        assert doc1["assessment-results"]["uuid"] == doc2["assessment-results"]["uuid"]
        # Result UUID
        r1 = doc1["assessment-results"]["results"][0]
        r2 = doc2["assessment-results"]["results"][0]
        assert r1["uuid"] == r2["uuid"]
        # Finding UUIDs
        for f1, f2 in zip(r1["findings"], r2["findings"]):
            assert f1["uuid"] == f2["uuid"]
        # Observation UUIDs
        for o1, o2 in zip(r1["observations"], r2["observations"]):
            assert o1["uuid"] == o2["uuid"]

    def test_oscal_filename_is_fixed(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, _ = tmp_output

        p = export_oscal(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        assert pathlib.Path(p).name == "oscal-assessment-results.json"

    def test_data_exports_deterministic(self, frozen_evaluation, tmp_output):
        """Regression guard: CSV/JSON data exports remain deterministic."""
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        paths_a = export_data_files(results, evidence, dir_a)
        paths_b = export_data_files(results, evidence, dir_b)

        assert len(paths_a) == len(paths_b)
        for pa, pb in zip(sorted(paths_a), sorted(paths_b)):
            ca = pathlib.Path(pa).read_text(encoding="utf-8")
            cb = pathlib.Path(pb).read_text(encoding="utf-8")
            assert ca == cb, f"Data export mismatch: {pathlib.Path(pa).name}"

    def test_raw_evidence_deterministic(self, frozen_evaluation, tmp_output):
        """Regression guard: raw evidence JSON files remain deterministic."""
        _, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        paths_a = save_raw_evidence(evidence, dir_a)
        paths_b = save_raw_evidence(evidence, dir_b)

        # save_raw_evidence returns a directory path string
        dir_a_raw = pathlib.Path(paths_a)
        dir_b_raw = pathlib.Path(paths_b)
        files_a = sorted(dir_a_raw.glob("*.json"))
        files_b = sorted(dir_b_raw.glob("*.json"))
        assert len(files_a) == len(files_b)
        for fa, fb in zip(files_a, files_b):
            assert fa.name == fb.name
            assert fa.read_text(encoding="utf-8") == fb.read_text(encoding="utf-8"), \
                f"Raw evidence mismatch: {fa.name}"


# ═══════════════════════════════════════════════════════════════════════
# Category B: Display reports — identical after stripping timestamps
# ═══════════════════════════════════════════════════════════════════════

class TestCategoryB_DisplayReports:
    """HTML/MD reports that embed timestamps. Content must match after stripping."""

    def test_compliance_html_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_compliance_report_html(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_compliance_report_html(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Compliance HTML differs after stripping timestamps"

    def test_compliance_md_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_compliance_report_md(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_compliance_report_md(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Compliance MD differs after stripping timestamps"

    def test_executive_summary_html_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_executive_summary_html(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_executive_summary_html(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Executive summary HTML differs after stripping timestamps"

    def test_executive_summary_md_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_executive_summary_md(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_executive_summary_md(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Executive summary MD differs after stripping timestamps"

    def test_gaps_report_html_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_gaps_report_html(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_gaps_report_html(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Gaps report HTML differs after stripping timestamps"

    def test_gaps_report_md_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_gaps_report_md(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_gaps_report_md(
            results=results, evidence=evidence, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Gaps report MD differs after stripping timestamps"

    def test_master_report_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        paths_a = generate_master_report(
            results=results, output_dir=dir_a, framework="FedRAMP",
            evidence=evidence, tenant_info=TENANT_INFO)
        paths_b = generate_master_report(
            results=results, output_dir=dir_b, framework="FedRAMP",
            evidence=evidence, tenant_info=TENANT_INFO)

        # HTML
        if "html" in paths_a and "html" in paths_b:
            a = _strip_timestamps(pathlib.Path(paths_a["html"]).read_text(encoding="utf-8"))
            b = _strip_timestamps(pathlib.Path(paths_b["html"]).read_text(encoding="utf-8"))
            assert a == b, "Master report HTML differs after stripping timestamps"

        # MD
        if "md" in paths_a and "md" in paths_b:
            a = _strip_timestamps(pathlib.Path(paths_a["md"]).read_text(encoding="utf-8"))
            b = _strip_timestamps(pathlib.Path(paths_b["md"]).read_text(encoding="utf-8"))
            assert a == b, "Master report MD differs after stripping timestamps"

    def test_executive_dashboard_deterministic(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_executive_dashboard(
            results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_executive_dashboard(
            results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Executive dashboard differs after stripping timestamps"

    def test_methodology_report_deterministic(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        kwargs = dict(
            results=results, evidence=evidence, tenant_info=TENANT_INFO,
            access_denied=[], collector_stats=[], elapsed_seconds=10.0,
        )
        p1 = generate_methodology_html(output_dir=dir_a, **kwargs)
        p2 = generate_methodology_html(output_dir=dir_b, **kwargs)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Methodology report differs after stripping timestamps"

    def test_json_report_deterministic(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_json_report(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_json_report(results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        # JSON report has timestamp in filename and "generated" field — strip both
        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "JSON report content differs after stripping timestamps"

    def test_markdown_report_deterministic(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_markdown_report(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_markdown_report(results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "Markdown report content differs after stripping timestamps"

    def test_html_report_deterministic(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_html_report(results=results, tenant_info=TENANT_INFO, output_dir=dir_a)
        p2 = generate_html_report(results=results, tenant_info=TENANT_INFO, output_dir=dir_b)

        a = _strip_timestamps(pathlib.Path(p1).read_text(encoding="utf-8"))
        b = _strip_timestamps(pathlib.Path(p2).read_text(encoding="utf-8"))
        assert a == b, "HTML report content differs after stripping timestamps"


# ═══════════════════════════════════════════════════════════════════════
# Category C: Binary/structured — validate content equivalence
# ═══════════════════════════════════════════════════════════════════════

class TestCategoryC_StructuralEquivalence:
    """Excel output validated structurally (sheet names, row counts, cell values)."""

    def test_excel_structural_equivalence(self, frozen_evaluation, tmp_output):
        results, _ = frozen_evaluation
        dir_a, dir_b = tmp_output

        p1 = generate_excel_report(results=results, output_dir=dir_a, framework="FedRAMP")
        p2 = generate_excel_report(results=results, output_dir=dir_b, framework="FedRAMP")

        from openpyxl import load_workbook

        wb1 = load_workbook(p1, data_only=True)
        wb2 = load_workbook(p2, data_only=True)

        assert wb1.sheetnames == wb2.sheetnames, "Excel sheet names differ"

        for sheet_name in wb1.sheetnames:
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]
            assert ws1.max_row == ws2.max_row, f"Row count differs in sheet '{sheet_name}'"
            assert ws1.max_column == ws2.max_column, f"Column count differs in sheet '{sheet_name}'"

            # Compare all cell values (ignoring formatting)
            for row in range(1, ws1.max_row + 1):
                for col in range(1, ws1.max_column + 1):
                    v1 = ws1.cell(row=row, column=col).value
                    v2 = ws2.cell(row=row, column=col).value
                    assert v1 == v2, \
                        f"Cell ({row},{col}) in '{sheet_name}': {v1!r} != {v2!r}"

    def test_master_excel_structural_equivalence(self, frozen_evaluation, tmp_output):
        results, evidence = frozen_evaluation
        dir_a, dir_b = tmp_output

        paths_a = generate_master_report(
            results=results, output_dir=dir_a, framework="FedRAMP",
            evidence=evidence, tenant_info=TENANT_INFO)
        paths_b = generate_master_report(
            results=results, output_dir=dir_b, framework="FedRAMP",
            evidence=evidence, tenant_info=TENANT_INFO)

        if "xlsx" not in paths_a or "xlsx" not in paths_b:
            pytest.skip("Master report did not produce Excel output")

        from openpyxl import load_workbook

        wb1 = load_workbook(paths_a["xlsx"], data_only=True)
        wb2 = load_workbook(paths_b["xlsx"], data_only=True)

        assert wb1.sheetnames == wb2.sheetnames

        for sheet_name in wb1.sheetnames:
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]
            assert ws1.max_row == ws2.max_row
            assert ws1.max_column == ws2.max_column
            for row in range(1, ws1.max_row + 1):
                for col in range(1, ws1.max_column + 1):
                    v1 = ws1.cell(row=row, column=col).value
                    v2 = ws2.cell(row=row, column=col).value
                    # Strip timestamps from string cells
                    if isinstance(v1, str) and isinstance(v2, str):
                        v1 = _strip_timestamps(v1)
                        v2 = _strip_timestamps(v2)
                    assert v1 == v2, \
                        f"Master Excel cell ({row},{col}) in '{sheet_name}': {v1!r} != {v2!r}"
