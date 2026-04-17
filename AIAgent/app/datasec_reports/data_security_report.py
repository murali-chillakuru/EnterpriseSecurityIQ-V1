"""
Data Security Assessment Report — Interactive HTML
Full-width professional report showing data security posture analysis
with executive summary, category breakdown, severity distribution,
top findings, detailed findings with remediation, and methodology.
"""

from __future__ import annotations

import hashlib
import json
import pathlib
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.reports.shared_theme import (
    get_css, get_js, esc, format_date_short, VERSION,
    SEVERITY_COLORS,
)
from app.logger import log


# ── Category metadata ────────────────────────────────────────────────────

_CATEGORY_META: dict[str, dict] = {
    "storage": {
        "icon": "&#128451;",  # 🗃
        "name": "Storage Exposure",
        "color": "#0078D4",
        "description": "Public blob access, HTTPS enforcement, network access rules, soft-delete, min TLS, shared-key access, infrastructure encryption, anonymous containers.",
    },
    "database": {
        "icon": "&#128450;",  # 🗂
        "name": "Database Security",
        "color": "#00B7C3",
        "description": "SQL Transparent Data Encryption, auditing, Advanced Threat Protection, firewall rules (deep ARM sub-resource checks).",
    },
    "cosmosdb": {
        "icon": "&#127760;",  # 🌐
        "name": "Cosmos DB Security",
        "color": "#744DA9",
        "description": "Public network access, IP/VNet firewall, key-based auth, backup policy (periodic vs continuous).",
    },
    "pgmysql": {
        "icon": "&#128024;",  # 🐘
        "name": "PostgreSQL / MySQL",
        "color": "#336791",
        "description": "SSL enforcement, public network access, open firewall rules, geo-redundant backup.",
    },
    "keyvault": {
        "icon": "&#128273;",  # 🔑
        "name": "Key / Secret Hygiene",
        "color": "#F7630C",
        "description": "Key Vault access policies, purge protection, expired and expiring secrets/keys/certificates.",
    },
    "encryption": {
        "icon": "&#128274;",  # 🔒
        "name": "Encryption Posture",
        "color": "#FFB900",
        "description": "VM disk encryption (ADE, DES, encryption-at-host), storage encryption key source, customer-managed key adoption.",
    },
    "data_access": {
        "icon": "&#128101;",  # 👥
        "name": "Data Access Controls",
        "color": "#E74856",
        "description": "Overly-broad data-plane RBAC, Defender for Storage/SQL/Key Vault, diagnostic settings, sensitive-tag classification.",
    },
    "private_endpoints": {
        "icon": "&#128279;",  # 🔗
        "name": "Private Endpoints",
        "color": "#16C60C",
        "description": "Checks whether data services (storage, SQL, Cosmos, Key Vault, PG, MySQL) are reachable only over private link.",
    },
    "purview": {
        "icon": "&#128269;",  # 🔍
        "name": "Purview / Info Protection",
        "color": "#8764B8",
        "description": "Microsoft Purview account presence, public network access, managed identity, private endpoints for data governance and classification.",
    },
    "file_sync": {
        "icon": "&#128259;",  # 🔃
        "name": "Azure File Sync",
        "color": "#0063B1",
        "description": "Storage Sync Services discovery, incoming traffic policy, private endpoints for hybrid file sync.",
    },
    "m365_dlp": {
        "icon": "&#128737;",  # 🛡
        "name": "M365 DLP Policies",
        "color": "#C239B3",
        "description": "Microsoft 365 Data Loss Prevention policy existence, enablement status, and workload coverage gaps.",
    },
    "data_classification": {
        "icon": "&#127991;",  # 🎷 → 🏷️
        "name": "Data Classification & Labeling",
        "color": "#008575",
        "description": "SQL sensitivity labels & data discovery, Defender sensitive-data discovery & alerts, Purview scan coverage & scan run results, and detection of unclassified data stores.",
    },
    "backup_dr": {
        "icon": "&#128190;",  # 💾
        "name": "Backup & Disaster Recovery",
        "color": "#0078D4",
        "description": "Recovery Services vault redundancy, VM backup protection, SQL long-term retention, Cosmos DB continuous backup.",
    },
    "container_security": {
        "icon": "&#128230;",  # 📦
        "name": "Container Security",
        "color": "#326CE5",
        "description": "ACR admin access, container vulnerability scanning, AKS RBAC & Azure AD integration, Kubernetes network policies.",
    },
    "network_segmentation": {
        "icon": "&#128737;",  # 🛡
        "name": "Network Segmentation",
        "color": "#FF4500",
        "description": "NSG data-port rules, DDoS Protection, VNet service endpoints for data services.",
    },
    "data_residency": {
        "icon": "&#127758;",  # 🌎
        "name": "Data Residency & Sovereignty",
        "color": "#2E8B57",
        "description": "Data service region compliance, cross-boundary geo-replication checks.",
    },
    "threat_detection": {
        "icon": "&#128680;",  # 🚨
        "name": "Threat Detection",
        "color": "#C72C41",
        "description": "Defender for Cloud coverage gaps for data services, security alert action group configuration.",
    },
    "redis": {
        "icon": "&#9889;",  # ⚡
        "name": "Redis Cache Security",
        "color": "#D82C20",
        "description": "TLS enforcement, non-SSL port disabled, firewall rules for Azure Cache for Redis instances.",
    },
    "messaging": {
        "icon": "&#128233;",  # 📩
        "name": "Messaging Security",
        "color": "#FF6F00",
        "description": "Network access rules, local auth disabling, capture, and TLS enforcement for Event Hub and Service Bus namespaces.",
    },
    "ai_services": {
        "icon": "&#129302;",  # 🤖
        "name": "AI Services Security",
        "color": "#6B3FA0",
        "description": "Key authentication, managed identity, CMK encryption, and network isolation for Azure AI / Cognitive Services.",
    },
    "data_pipeline": {
        "icon": "&#128640;",  # 🚀
        "name": "Data Pipeline Security",
        "color": "#00897B",
        "description": "Data Factory and Synapse workspace security: public access, managed identity, Git integration, AAD-only auth.",
    },
    "identity": {
        "icon": "&#128100;",  # 👤
        "name": "Identity & Managed Identity",
        "color": "#5C6BC0",
        "description": "Managed identity adoption across data services, credential-free authentication, and identity lifecycle management.",
    },
    "sharepoint_governance": {
        "icon": "&#128196;",  # 📄
        "name": "SharePoint Governance",
        "color": "#0B6A46",
        "description": "SharePoint site permissions, sharing links, external sharing configuration, stale sites, and sensitivity labeling.",
    },
    "data_lifecycle": {
        "icon": "&#9851;",  # ♻
        "name": "M365 Data Lifecycle",
        "color": "#795548",
        "description": "Retention labels, retention policies, eDiscovery cases for Microsoft 365 data compliance.",
    },
    "dlp_alert": {
        "icon": "&#128276;",  # 🔔
        "name": "DLP Alert Effectiveness",
        "color": "#E91E63",
        "description": "DLP alert metrics, false positive rates, and policy enforcement effectiveness.",
    },
    "app_config": {
        "icon": "&#9881;",  # ⚙
        "name": "App Configuration Security",
        "color": "#4DB6AC",
        "description": "Azure App Configuration public access, private endpoints, soft-delete protection.",
    },
    "databricks": {
        "icon": "&#128293;",  # 🔥
        "name": "Databricks Security",
        "color": "#FF3621",
        "description": "Azure Databricks VNet injection, customer-managed keys, public workspace access.",
    },
    "apim": {
        "icon": "&#128256;",  # 🔀
        "name": "API Management Security",
        "color": "#68217A",
        "description": "Azure API Management VNet integration, managed identity, subscription key policies.",
    },
    "frontdoor": {
        "icon": "&#128737;",  # 🛡
        "name": "Front Door / WAF",
        "color": "#0078D4",
        "description": "Azure Front Door WAF policy association, minimum TLS version enforcement.",
    },
    "secret_sprawl": {
        "icon": "&#128270;",  # 🔎
        "name": "Secret Sprawl",
        "color": "#FF6D00",
        "description": "Plain-text secrets in app settings, Key Vault reference adoption for web apps.",
    },
    "firewall": {
        "icon": "&#128293;",  # 🔥
        "name": "Firewall & App Gateway",
        "color": "#DD2C00",
        "description": "Azure Firewall threat intelligence, IDPS, Application Gateway WAF SKU.",
    },
    "bastion": {
        "icon": "&#128272;",  # 🔐
        "name": "Bastion & Remote Access",
        "color": "#1565C0",
        "description": "Azure Bastion coverage for RDP/SSH, shareable-link security.",
    },
    "policy_compliance": {
        "icon": "&#128220;",  # 📜
        "name": "Policy Compliance",
        "color": "#2E7D32",
        "description": "Azure Policy non-compliance for data-related policies, governance enforcement.",
    },
    "defender_score": {
        "icon": "&#128737;",  # 🛡
        "name": "Defender Score",
        "color": "#D32F2F",
        "description": "Microsoft Defender for Cloud unhealthy data-related security recommendations.",
    },
    "stale_permissions": {
        "icon": "&#128274;",  # 🔒
        "name": "Stale Permissions",
        "color": "#F57C00",
        "description": "RBAC role assignments with no sign-in activity over 90 days on data resources.",
    },
    "data_exfiltration": {
        "icon": "&#128680;",  # 🚨
        "name": "Data Exfiltration Risk",
        "color": "#B71C1C",
        "description": "Storage bypass rules, cross-subscription private endpoints, unrestricted NSG outbound.",
    },
    "conditional_access": {
        "icon": "&#128100;",  # 👤
        "name": "Conditional Access & PIM",
        "color": "#6A1B9A",
        "description": "Conditional Access MFA policies, PIM permanent role assignments on data resources.",
    },
    "config_drift": {
        "icon": "&#128260;",  # 🔄
        "name": "Configuration Drift",
        "color": "#E65100",
        "description": "Security-sensitive configuration changes detected between assessment snapshots.",
    },
    "supply_chain": {
        "icon": "&#128230;",  # 📦
        "name": "Supply Chain Risk",
        "color": "#4E342E",
        "description": "Container registry, external package, and third-party dependency risks.",
    },
}

_SEV_COLORS: dict[str, str] = {
    "critical": "#D13438",
    "high": "#F7630C",
    "medium": "#FFB900",
    "low": "#107C10",
    "informational": "#A8A6A3",
}

_SCORE_LEVEL_META: dict[str, dict] = {
    "critical": {"color": "#D13438", "label": "Critical Risk", "icon": "&#128680;"},
    "high":     {"color": "#F7630C", "label": "High Risk",     "icon": "&#9888;&#65039;"},
    "medium":   {"color": "#FFB900", "label": "Medium Risk",   "icon": "&#9432;"},
    "low":      {"color": "#107C10", "label": "Low Risk",      "icon": "&#10003;"},
}

# ── Data Security Relevance Justifications ───────────────────────────────────────
# Loaded from config/data-security-relevance.json

_RELEVANCE_PATH = pathlib.Path(__file__).resolve().parents[2] / "config" / "data-security-relevance.json"
if _RELEVANCE_PATH.exists():
    with open(_RELEVANCE_PATH, encoding="utf-8") as _fh:
        _DATA_SECURITY_RELEVANCE: dict[str, str] = json.load(_fh)
else:
    _DATA_SECURITY_RELEVANCE: dict[str, str] = {}


# ── SVG helpers ──────────────────────────────────────────────────────────

def _donut_svg(slices: list[tuple[str, float, str]], size: int = 160, hole: float = 0.6, center_text: str | None = None) -> str:
    total = sum(v for _, v, _ in slices) or 1
    r = size // 2 - 4
    circ = 2 * 3.14159 * r
    cx = cy = size // 2
    parts = [f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" role="img" aria-label="Donut chart">']
    offset = 0
    for label, val, color in slices:
        if val <= 0:
            continue
        pct = val / total
        dash = circ * pct
        gap = circ - dash
        parts.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" '
            f'stroke-width="{r * (1 - hole)}" stroke-dasharray="{dash:.2f} {gap:.2f}" '
            f'stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})">'
            f'<title>{esc(label)}: {int(val)} ({pct*100:.0f}%)</title></circle>'
        )
        offset += dash
    ct = center_text if center_text is not None else str(int(total))
    parts.append(
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" '
        f'font-size="22" font-weight="700" fill="var(--text)" '
        f'font-family="var(--font-mono)">{esc(ct)}</text>'
    )
    parts.append("</svg>")
    return "".join(parts)


def _ring_score_svg(score: float, size: int = 140) -> str:
    r = size // 2 - 8
    circ = 2 * 3.14159 * r
    pct = min(score, 100) / 100
    dash = circ * pct
    gap = circ - dash
    cx = cy = size // 2
    color = "#D13438" if score >= 75 else "#F7630C" if score >= 50 else "#FFB900" if score >= 25 else "#107C10"
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" class="ring" role="img" aria-label="Security score {score:.0f} out of 100">'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="var(--ring-track)" stroke-width="10"/>'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="10" '
        f'stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-dashoffset="{circ * 0.25:.2f}" '
        f'stroke-linecap="round" style="transition:stroke-dasharray 1s ease"/>'
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" font-size="28" font-weight="700" '
        f'fill="{color}" font-family="var(--font-mono)">{score:.0f}</text>'
        f'<text x="{cx}" y="{cy + 18}" text-anchor="middle" font-size="10" fill="var(--text-muted)">/100</text>'
        f'</svg>'
    )


def _bar_chart_svg(items: list[tuple[str, int, str]], width: int = 480, bar_h: int = 24) -> str:
    if not items:
        return '<p class="empty">No data</p>'
    max_val = max(v for _, v, _ in items) or 1
    # Compute label column width from longest label (~8px per char at 12px font)
    max_label_len = max((len(label) for label, _, _ in items), default=10)
    label_col = max(220, max_label_len * 8 + 20)
    gap = 8
    svg_w = label_col + 480  # bar area + value text space
    h = (bar_h + gap) * len(items) + 10
    bar_area = svg_w - label_col - 50  # reserve space for value text
    parts = [f'<svg width="{svg_w}" height="{h}" role="img" aria-label="Bar chart">']
    for i, (label, val, color) in enumerate(items):
        y = i * (bar_h + gap) + 4
        bw = int((val / max_val) * bar_area) if val > 0 else 0
        parts.append(
            f'<text x="{label_col - 12}" y="{y + bar_h - 6}" font-size="12" fill="var(--text-secondary)" '
            f'font-family="var(--font-primary)" text-anchor="end">{esc(label)}</text>'
        )
        parts.append(
            f'<rect x="{label_col}" y="{y}" width="{max(bw, 3)}" height="{bar_h}" rx="4" fill="{color}" opacity="0.85">'
            f'<title>{esc(label)}: {val}</title></rect>'
        )
        parts.append(
            f'<text x="{label_col + max(bw, 3) + 8}" y="{y + bar_h - 6}" font-size="12" fill="var(--text)" '
            f'font-family="var(--font-mono)" font-weight="600">{val}</text>'
        )
    parts.append("</svg>")
    return "".join(parts)


def _heatmap_svg(category_scores: dict[str, dict], width: int = 720) -> str:
    """Generate a severity heatmap SVG for category findings distribution."""
    cats = sorted(
        [(k, v) for k, v in category_scores.items() if v.get("findings", 0) > 0],
        key=lambda x: x[0],
    )
    if not cats:
        return '<p class="empty">No findings to display in heatmap</p>'
    cell_w = 70
    cell_h = 36
    label_w = 240
    severities = ["critical", "high", "medium", "low"]
    header_h = 30
    svg_h = header_h + len(cats) * (cell_h + 4) + 10
    svg_w = label_w + len(severities) * (cell_w + 4) + 20
    parts = [f'<svg width="{svg_w}" height="{svg_h}" role="img" aria-label="Category severity heatmap">']
    # Header
    for i, s in enumerate(severities):
        x = label_w + i * (cell_w + 4)
        parts.append(f'<text x="{x + cell_w // 2}" y="{header_h - 8}" text-anchor="middle" '
                     f'font-size="11" fill="var(--text-secondary)" font-family="var(--font-primary)" '
                     f'text-transform="uppercase">{s.title()}</text>')
    # Rows
    for row_i, (cat_key, cat_data) in enumerate(cats):
        y = header_h + row_i * (cell_h + 4)
        meta = _CATEGORY_META.get(cat_key, {})
        label = meta.get("name", cat_key.replace("_", " ").title())
        parts.append(f'<text x="0" y="{y + cell_h // 2 + 4}" font-size="12" fill="var(--text-secondary)" '
                     f'font-family="var(--font-primary)">{esc(label)}</text>')
        sev_dist = cat_data.get("severity_distribution", {})
        for col_i, s in enumerate(severities):
            x = label_w + col_i * (cell_w + 4)
            count = sev_dist.get(s, 0)
            opacity = min(0.15 + (count / max(cat_data.get("findings", 1), 1)) * 0.85, 1.0) if count > 0 else 0.08
            parts.append(
                f'<rect x="{x}" y="{y}" width="{cell_w}" height="{cell_h}" rx="4" '
                f'fill="{_SEV_COLORS.get(s, "#A8A6A3")}" opacity="{opacity:.2f}">'
                f'<title>{esc(label)} — {s}: {count}</title></rect>')
            txt_fill = 'var(--text-muted)' if count == 0 else 'var(--text)'
            parts.append(
                f'<text x="{x + cell_w // 2}" y="{y + cell_h // 2 + 5}" text-anchor="middle" '
                f'font-size="13" font-weight="700" fill="{txt_fill}" '
                f'font-family="var(--font-mono)">{count}</text>')
    parts.append("</svg>")
    return "".join(parts)


# ── Report-specific CSS ─────────────────────────────────────────────────



def _compliance_gap_matrix_svg(findings: list[dict], width: int = 780) -> str:
    """Generate a compliance gap matrix SVG showing framework control coverage."""
    fw_controls: dict[str, dict[str, int]] = {}
    for f in findings:
        mapping = f.get("ComplianceMapping", {})
        sev = f.get("Severity", "medium").lower()
        for fw, controls in mapping.items():
            if fw not in fw_controls:
                fw_controls[fw] = {"total": 0, "critical": 0, "high": 0, "medium": 0, "low": 0}
            if isinstance(controls, list):
                fw_controls[fw]["total"] += len(controls)
            elif isinstance(controls, str):
                fw_controls[fw]["total"] += 1
            fw_controls[fw][sev] = fw_controls[fw].get(sev, 0) + 1
    if not fw_controls:
        return '<p class="empty">No compliance mappings found</p>'
    cell_w = 80
    cell_h = 36
    label_w = 180
    severities = ["critical", "high", "medium", "low"]
    header_h = 34
    sorted_fws = sorted(fw_controls.keys())
    svg_h = header_h + len(sorted_fws) * (cell_h + 4) + 10
    svg_w = label_w + len(severities) * (cell_w + 4) + (cell_w + 4) + 20  # extra col for total
    parts = [f'<svg width="{svg_w}" height="{svg_h}" role="img" aria-label="Compliance gap matrix">']
    # Header
    for i, s in enumerate(severities):
        x = label_w + i * (cell_w + 4)
        parts.append(f'<text x="{x + cell_w // 2}" y="{header_h - 10}" text-anchor="middle" '
                     f'font-size="11" fill="var(--text-secondary)" font-family="var(--font-primary)">{s.title()}</text>')
    tx = label_w + len(severities) * (cell_w + 4)
    parts.append(f'<text x="{tx + cell_w // 2}" y="{header_h - 10}" text-anchor="middle" '
                 f'font-size="11" fill="var(--text-secondary)" font-family="var(--font-primary)" font-weight="700">Controls</text>')
    for row_i, fw in enumerate(sorted_fws):
        y = header_h + row_i * (cell_h + 4)
        data = fw_controls[fw]
        parts.append(f'<text x="0" y="{y + cell_h // 2 + 4}" font-size="12" fill="var(--text-secondary)" '
                     f'font-family="var(--font-primary)" font-weight="600">{esc(fw)}</text>')
        for col_i, s in enumerate(severities):
            x = label_w + col_i * (cell_w + 4)
            count = data.get(s, 0)
            opacity = min(0.15 + (count / max(data.get("total", 1), 1)) * 0.85, 1.0) if count > 0 else 0.08
            parts.append(f'<rect x="{x}" y="{y}" width="{cell_w}" height="{cell_h}" rx="4" '
                         f'fill="{_SEV_COLORS.get(s, "#888")}" opacity="{opacity:.2f}">'
                         f'<title>{esc(fw)} {s}: {count} finding(s)</title></rect>')
            txt_fill = 'var(--text-muted)' if count == 0 else 'var(--text)'
            parts.append(f'<text x="{x + cell_w // 2}" y="{y + cell_h // 2 + 5}" text-anchor="middle" '
                         f'font-size="13" font-weight="700" fill="{txt_fill}" '
                         f'font-family="var(--font-mono)">{count}</text>')
        # Total controls col
        tx = label_w + len(severities) * (cell_w + 4)
        parts.append(f'<rect x="{tx}" y="{y}" width="{cell_w}" height="{cell_h}" rx="4" '
                     f'fill="var(--primary)" opacity="0.15"><title>{esc(fw)}: {data["total"]} controls mapped</title></rect>')
        parts.append(f'<text x="{tx + cell_w // 2}" y="{y + cell_h // 2 + 5}" text-anchor="middle" '
                     f'font-size="13" font-weight="700" fill="var(--primary)" font-family="var(--font-mono)">{data["total"]}</text>')
    parts.append("</svg>")
    return "".join(parts)


def _sparkline_svg(values: list[float], width: int = 80, height: int = 24, color: str = "#0078D4") -> str:
    """Generate a tiny sparkline SVG from a list of values."""
    if not values or len(values) < 2:
        return ""
    mn = min(values)
    mx = max(values)
    rng = mx - mn if mx != mn else 1
    step = width / (len(values) - 1)
    points = []
    for i, v in enumerate(values):
        x = i * step
        y = height - ((v - mn) / rng) * (height - 4) - 2
        points.append(f"{x:.1f},{y:.1f}")
    polyline = " ".join(points)
    last_y = height - ((values[-1] - mn) / rng) * (height - 4) - 2
    return (
        f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
        f'style="vertical-align:middle;margin-left:6px" role="img" aria-label="Trend sparkline">'
        f'<polyline points="{polyline}" fill="none" stroke="{color}" stroke-width="1.5" stroke-linejoin="round"/>'
        f'<circle cx="{width:.1f}" cy="{last_y:.1f}" r="2" fill="{color}"/>'
        f'</svg>'
    )


def _priority_quadrant_svg(findings: list[dict], width: int = 900, height: int = 540) -> str:
    """Generate a professional priority quadrant chart SVG with external labels."""
    if not findings:
        return '<p class="empty">No findings to plot</p>'
    sev_impact = {"critical": 95, "high": 75, "medium": 50, "low": 25, "informational": 10}
    sev_effort = {"critical": 30, "high": 45, "medium": 60, "low": 75, "informational": 85}

    # Layout: extra top/bottom margins for external labels
    label_zone = 22  # space above/below chart for quadrant names
    margin_left = 56
    margin_right = 10
    margin_top = 44 + label_zone
    margin_bottom = 44 + label_zone
    legend_w = 150
    plot_w = width - margin_left - margin_right - legend_w
    plot_h = height - margin_top - margin_bottom

    parts: list[str] = [
        f'<svg width="100%" viewBox="0 0 {width} {height}" '
        f'preserveAspectRatio="xMidYMid meet" role="img" '
        f'aria-label="Remediation Priority Quadrant">'
    ]

    # Defs — arrowhead markers
    parts.append(
        '<defs>'
        '<marker id="ax" viewBox="0 0 10 10" refX="10" refY="5" '
        'markerWidth="7" markerHeight="7" orient="auto">'
        '<path d="M0 0 L10 5 L0 10z" fill="var(--text)"/></marker>'
        '<marker id="ay" viewBox="0 0 10 10" refX="10" refY="5" '
        'markerWidth="7" markerHeight="7" orient="auto">'
        '<path d="M0 0 L10 5 L0 10z" fill="var(--text)"/></marker>'
        '</defs>'
    )

    hw = plot_w // 2
    hh = plot_h // 2

    # Background quadrants with subtle rounded rects
    parts.append(f'<rect x="{margin_left}" y="{margin_top}" width="{hw}" height="{hh}" fill="rgba(16,124,16,0.18)" rx="6"/>')
    parts.append(f'<rect x="{margin_left + hw}" y="{margin_top}" width="{hw}" height="{hh}" fill="rgba(160,130,0,0.14)" rx="6"/>')
    parts.append(f'<rect x="{margin_left}" y="{margin_top + hh}" width="{hw}" height="{hh}" fill="rgba(100,120,140,0.12)" rx="6"/>')
    parts.append(f'<rect x="{margin_left + hw}" y="{margin_top + hh}" width="{hw}" height="{hh}" fill="rgba(140,40,50,0.14)" rx="6"/>')

    # Midlines (dashed, subtle)
    parts.append(f'<line x1="{margin_left + hw}" y1="{margin_top}" x2="{margin_left + hw}" y2="{margin_top + plot_h}" stroke="var(--border)" stroke-width="0.8" stroke-dasharray="6,4" opacity="0.6"/>')
    parts.append(f'<line x1="{margin_left}" y1="{margin_top + hh}" x2="{margin_left + plot_w}" y2="{margin_top + hh}" stroke="var(--border)" stroke-width="0.8" stroke-dasharray="6,4" opacity="0.6"/>')

    # ── External quadrant labels (above and below the chart) ──
    q1x = margin_left + hw // 2
    q2x = margin_left + hw + hw // 2
    label_top_y = margin_top - 8    # above the chart
    label_bot_y = margin_top + plot_h + 16  # below the chart

    parts.append(f'<text x="{q1x}" y="{label_top_y}" text-anchor="middle" font-size="12" fill="var(--text-muted)" font-weight="700" letter-spacing="0.5">Quick Wins</text>')
    parts.append(f'<text x="{q2x}" y="{label_top_y}" text-anchor="middle" font-size="12" fill="var(--text-muted)" font-weight="700" letter-spacing="0.5">Major Projects</text>')
    parts.append(f'<text x="{q1x}" y="{label_bot_y}" text-anchor="middle" font-size="12" fill="var(--text-muted)" font-weight="700" letter-spacing="0.5">Low Priority</text>')
    parts.append(f'<text x="{q2x}" y="{label_bot_y}" text-anchor="middle" font-size="12" fill="var(--text-muted)" font-weight="700" letter-spacing="0.5">Consider</text>')

    # ── Axes ──
    # X-axis
    parts.append(f'<line x1="{margin_left}" y1="{margin_top + plot_h}" x2="{margin_left + plot_w}" y2="{margin_top + plot_h}" stroke="var(--text)" stroke-width="1.5" marker-end="url(#ax)"/>')
    # Y-axis
    parts.append(f'<line x1="{margin_left}" y1="{margin_top + plot_h}" x2="{margin_left}" y2="{margin_top}" stroke="var(--text)" stroke-width="1.5" marker-end="url(#ay)"/>')

    # Axis labels
    x_label_y = height - 8
    parts.append(f'<text x="{margin_left + plot_w // 2}" y="{x_label_y}" text-anchor="middle" font-size="12" fill="var(--text)" font-weight="600">Remediation Effort &#8594;</text>')
    y_label_x = 16
    y_label_y = margin_top + plot_h // 2
    parts.append(f'<text x="{y_label_x}" y="{y_label_y}" text-anchor="middle" font-size="12" fill="var(--text)" font-weight="600" transform="rotate(-90 {y_label_x} {y_label_y})">Security Impact &#8594;</text>')

    # ── Plot findings (aggregate by severity) ──
    sev_counts: dict[str, int] = {}
    for f in findings:
        s = f.get("Severity", "medium").lower()
        sev_counts[s] = sev_counts.get(s, 0) + 1

    _all_bubbles: list[tuple[float, float, float]] = []
    for sev, count in sev_counts.items():
        impact = sev_impact.get(sev, 50)
        effort = sev_effort.get(sev, 50)
        cx = margin_left + (effort / 100) * plot_w
        cy = margin_top + plot_h - (impact / 100) * plot_h
        r = min(8 + count * 0.9, 24)
        _all_bubbles.append((cx, cy, r))
        color = _SEV_COLORS.get(sev, "#888")
        # Drop shadow for depth
        parts.append(f'<circle cx="{cx + 1:.0f}" cy="{cy + 1:.0f}" r="{r:.0f}" fill="#000" opacity="0.15"/>')
        # Main bubble
        parts.append(
            f'<circle cx="{cx:.0f}" cy="{cy:.0f}" r="{r:.0f}" fill="{color}" opacity="0.82" '
            f'stroke="{color}" stroke-width="1.5">'
            f'<title>{sev.title()}: {count} finding{"s" if count != 1 else ""} '
            f'(Impact: {impact}, Effort: {effort})</title></circle>'
        )
        # Count label inside bubble
        parts.append(f'<text x="{cx:.0f}" y="{cy + 4:.0f}" text-anchor="middle" font-size="11" font-weight="700" fill="#fff">{count}</text>')
        # Severity label beside bubble — place left if bubble is near right edge
        if cx + r + 60 > margin_left + plot_w:
            lbl_x = cx - r - 5
            anchor = "end"
        else:
            lbl_x = cx + r + 5
            anchor = "start"
        parts.append(f'<text x="{lbl_x:.0f}" y="{cy + 4:.0f}" text-anchor="{anchor}" font-size="10" fill="var(--text-secondary)" font-weight="600">{sev.title()}</text>')

    # ── Legend ──
    lx = margin_left + plot_w + 20
    ly = margin_top + 10
    parts.append(f'<text x="{lx}" y="{ly}" font-size="11" fill="var(--text-secondary)" font-weight="700" letter-spacing="0.5">LEGEND (RISK)</text>')
    for i, (sev_name, sev_col) in enumerate(_SEV_COLORS.items()):
        yy = ly + 22 + i * 24
        cnt = sev_counts.get(sev_name, 0)
        parts.append(f'<circle cx="{lx + 7}" cy="{yy - 4}" r="6" fill="{sev_col}" opacity="0.85"/>')
        parts.append(f'<text x="{lx + 20}" y="{yy}" font-size="10" fill="var(--text-secondary)">{sev_name.title()} ({cnt})</text>')

    parts.append("</svg>")
    return "".join(parts)
def _priority_quadrant_detail(findings: list[dict]) -> str:
    """Generate supplementary detail panel for the priority quadrant.

    Part 1: Quadrant Breakdown by Severity table (summary).
    Part 2: Remediation Action Plan — grouped by quadrant with all findings,
            affected resources, and remediation guidance.
    All descriptions are data-driven from actual findings.
    """
    sev_order = ["critical", "high", "medium", "low", "informational"]
    quadrant_map = {"critical": "Quick Wins", "high": "Quick Wins", "medium": "Major Projects", "low": "Low Priority", "informational": "Consider"}
    effort_map = {"critical": "Low (config change)", "high": "Low-Medium", "medium": "Medium", "low": "Medium-High", "informational": "Varies"}
    sev_counts: dict[str, list[dict]] = {}
    for f in findings:
        s = f.get("Severity", "medium").lower()
        sev_counts.setdefault(s, []).append(f)

    # ── Part 1: Quadrant Breakdown by Severity (summary table) ──
    rows = ""
    for s in sev_order:
        items = sev_counts.get(s, [])
        if not items:
            continue
        color = _SEV_COLORS.get(s, "#888")
        quad = quadrant_map.get(s, "Consider")
        effort = effort_map.get(s, "Varies")
        rows += (
            f'<tr style="border-bottom:1px solid var(--border)">'
            f'<td style="padding:8px 12px;white-space:nowrap">'
            f'<span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{color};margin-right:6px;vertical-align:middle"></span>'
            f'<strong style="color:{color}">{s.title()} Risk</strong></td>'
            f'<td style="padding:8px 12px;text-align:center;font-family:var(--font-mono);font-weight:700">{len(items)}</td>'
            f'<td style="padding:8px 12px">{quad}</td>'
            f'<td style="padding:8px 12px">{effort}</td>'
            f'</tr>'
        )

    # ── Part 2: Remediation Action Plan — grouped by quadrant ──
    quadrant_order = ["Quick Wins", "Major Projects", "Low Priority", "Consider"]
    quadrant_icons = {"Quick Wins": "&#9889;", "Major Projects": "&#128736;", "Low Priority": "&#128337;", "Consider": "&#128269;"}
    quadrant_colors = {"Quick Wins": "#107C10", "Major Projects": "#A08200", "Low Priority": "#6A7A8C", "Consider": "#8C2838"}
    impact_effort = {"Quick Wins": "High impact, low effort", "Major Projects": "High impact, high effort", "Low Priority": "Low impact, low effort", "Consider": "Low impact, high effort"}

    # Group findings by quadrant
    by_quadrant: dict[str, list[dict]] = {}
    for f in findings:
        s = f.get("Severity", "medium").lower()
        quad = quadrant_map.get(s, "Consider")
        by_quadrant.setdefault(quad, []).append(f)

    quadrant_sections = ""
    for quad in quadrant_order:
        q_findings = by_quadrant.get(quad, [])
        if not q_findings:
            continue
        q_color = quadrant_colors.get(quad, "#888")
        q_icon = quadrant_icons.get(quad, "")

        # Dynamic description from actual findings
        cats = sorted({f.get("Category", "unknown").replace("_", " ").title() for f in q_findings})
        total_res = sum(f.get("AffectedCount", 0) for f in q_findings)
        ie = impact_effort.get(quad, "")
        q_desc = f"{ie}. {len(q_findings)} finding{'s' if len(q_findings) != 1 else ''} across {', '.join(cats)} affecting {total_res} resource{'s' if total_res != 1 else ''}."

        # Build findings table for this quadrant
        finding_rows = ""
        _res_overflow_id = getattr(_priority_quadrant_detail, '_rid', 0)
        for f in q_findings:
            f_sev = f.get("Severity", "medium").lower()
            f_color = _SEV_COLORS.get(f_sev, "#888")
            f_title = esc(f.get("Title", "Untitled"))
            f_cat = esc(f.get("Category", "").replace("_", " ").title())
            f_affected = f.get("AffectedCount", 0)
            f_resources = f.get("AffectedResources", [])
            f_remediation = f.get("Remediation", {})
            f_rem_desc = esc(f_remediation.get("Description", "")) if f_remediation else ""

            # Resource names – show first 10, rest hidden behind expandable toggle
            _MAX_VISIBLE = 10
            all_res_names: list[str] = []
            for ar in (f_resources or []):
                if isinstance(ar, dict):
                    rname = ar.get("Name", ar.get("name", ""))
                    rtype = ar.get("Type", ar.get("type", ""))
                    if rname:
                        short_type = rtype.split("/")[-1] if rtype else ""
                        all_res_names.append(f"{esc(str(rname))}" + (f" <span style='color:var(--text-muted);font-size:10px'>({esc(short_type)})</span>" if short_type else ""))
                elif ar:
                    all_res_names.append(esc(str(ar)))

            if not all_res_names:
                res_html = f'<span style="color:var(--text-muted)">{f_affected} resource{"s" if f_affected != 1 else ""}</span>'
            elif len(all_res_names) <= _MAX_VISIBLE:
                res_html = ", ".join(all_res_names)
            else:
                visible = ", ".join(all_res_names[:_MAX_VISIBLE])
                hidden = ", ".join(all_res_names[_MAX_VISIBLE:])
                _res_overflow_id += 1
                eid = f"res-overflow-{_res_overflow_id}"
                extra = len(all_res_names) - _MAX_VISIBLE
                res_html = (
                    f'{visible}'
                    f'<span id="{eid}-hidden" style="display:none">, {hidden}</span> '
                    f'<a id="{eid}-btn" href="javascript:void(0)" onclick="'
                    f"document.getElementById('{eid}-hidden').style.display='inline';"
                    f"this.style.display='none';"
                    f'" style="color:#0078D4;font-size:10px;cursor:pointer;white-space:nowrap">+{extra} more</a>'
                )

            # Truncate remediation for table view
            rem_short = f_rem_desc[:120] + ("…" if len(f_rem_desc) > 120 else "") if f_rem_desc else '<span style="color:var(--text-muted)">—</span>'

            finding_rows += (
                f'<tr style="border-bottom:1px solid var(--border)">'
                f'<td style="padding:6px 10px;white-space:nowrap">'
                f'<span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{f_color};margin-right:4px;vertical-align:middle"></span>'
                f'<span style="color:{f_color};font-weight:600;font-size:11px">{f_sev.title()}</span></td>'
                f'<td style="padding:6px 10px;font-size:12px">{f_title}</td>'
                f'<td style="padding:6px 10px;color:var(--text-muted);font-size:11px">{f_cat}</td>'
                f'<td style="padding:6px 10px;text-align:center;font-family:var(--font-mono);font-size:12px">{f_affected}</td>'
                f'<td style="padding:6px 10px;font-size:11px">{res_html}</td>'
                f'<td style="padding:6px 10px;font-size:11px;color:var(--text-secondary)">{rem_short}</td>'
                f'</tr>'
            )

        _priority_quadrant_detail._rid = _res_overflow_id

        quadrant_sections += (
            f'<div style="margin-top:20px">'
            f'<h5 style="margin:0 0 4px 0;font-size:14px;color:{q_color}">'
            f'{q_icon} {quad} <span style="font-weight:400;font-size:12px;color:var(--text-muted)">({len(q_findings)} finding{"s" if len(q_findings) != 1 else ""})</span></h5>'
            f'<p style="margin:0 0 10px 0;font-size:12px;color:var(--text-secondary)">{esc(q_desc)}</p>'
            f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px">'
            f'<thead><tr style="border-bottom:2px solid var(--border)">'
            f'<th style="text-align:left;padding:6px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Sev</th>'
            f'<th style="text-align:left;padding:6px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Finding</th>'
            f'<th style="text-align:left;padding:6px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Category</th>'
            f'<th style="text-align:center;padding:6px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Affected</th>'
            f'<th style="text-align:left;padding:6px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Resources</th>'
            f'<th style="text-align:left;padding:6px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Remediation</th>'
            f'</tr></thead><tbody>{finding_rows}</tbody></table></div>'
            f'</div>'
        )

    return (
        f'<div style="margin-top:20px">'
        f'<h4 style="margin-bottom:10px">Quadrant Breakdown by Severity</h4>'
        f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:13px">'
        f'<thead><tr style="border-bottom:2px solid var(--border)">'
        f'<th style="text-align:left;padding:8px 12px;font-size:11px;text-transform:uppercase;color:var(--text-secondary)">Severity</th>'
        f'<th style="text-align:center;padding:8px 12px;font-size:11px;text-transform:uppercase;color:var(--text-secondary)">Count</th>'
        f'<th style="text-align:left;padding:8px 12px;font-size:11px;text-transform:uppercase;color:var(--text-secondary)">Quadrant</th>'
        f'<th style="text-align:left;padding:8px 12px;font-size:11px;text-transform:uppercase;color:var(--text-secondary)">Est. Effort</th>'
        f'</tr></thead><tbody>{rows}</tbody></table></div>'
        f'<details style="margin-top:20px"><summary style="cursor:pointer;font-size:14px;font-weight:700;color:var(--text)">'
        f'Remediation Action Plan by Quadrant</summary>'
        f'{quadrant_sections}'
        f'</details>'
        f'</div>'
    )


def _quadrant_guide_bullets(findings: list[dict]) -> str:
    """Generate data-driven quadrant guide bullets from actual findings."""
    quadrant_map = {"critical": "Quick Wins", "high": "Quick Wins", "medium": "Major Projects", "low": "Low Priority", "informational": "Consider"}
    impact_effort = {"Quick Wins": "High impact, low effort", "Major Projects": "High impact, high effort", "Low Priority": "Low impact, low effort", "Consider": "Low impact, high effort"}
    positions = {"Quick Wins": "top-left", "Major Projects": "top-right", "Low Priority": "bottom-left", "Consider": "bottom-right"}

    by_quadrant: dict[str, list[dict]] = {}
    for f in findings:
        s = f.get("Severity", "medium").lower()
        quad = quadrant_map.get(s, "Consider")
        by_quadrant.setdefault(quad, []).append(f)

    items = []
    for quad in ["Quick Wins", "Major Projects", "Low Priority", "Consider"]:
        q_findings = by_quadrant.get(quad, [])
        ie = impact_effort[quad]
        pos = positions[quad]
        if q_findings:
            cats = sorted({f.get("Category", "unknown").replace("_", " ").title() for f in q_findings})
            total_res = sum(f.get("AffectedCount", 0) for f in q_findings)
            desc = f"{ie}. {len(q_findings)} finding{'s' if len(q_findings) != 1 else ''} across {', '.join(cats)} affecting {total_res} resource{'s' if total_res != 1 else ''}."
        else:
            desc = f"{ie}. No findings in this quadrant."
        items.append(
            f'<li><strong style="color:#0078D4">{quad} ({pos})</strong> &mdash; {esc(desc)}</li>'
        )
    return "\n      ".join(items)


def _estimate_noncompliance_cost(findings: list[dict]) -> dict:
    """Estimate the financial impact of non-compliance based on findings."""
    # Rough cost multipliers per severity (USD per finding per year)
    cost_per_sev = {
        "critical": 250_000,  # Potential breach, regulatory fine
        "high": 75_000,       # Significant risk exposure
        "medium": 15_000,     # Moderate compliance gap
        "low": 2_500,         # Minor best-practice deviation
        "informational": 500,
    }
    total = 0
    by_sev: dict[str, dict] = {}
    for f in findings:
        sev = f.get("Severity", "medium").lower()
        cost = cost_per_sev.get(sev, 5_000)
        total += cost
        if sev not in by_sev:
            by_sev[sev] = {"count": 0, "cost": 0}
        by_sev[sev]["count"] += 1
        by_sev[sev]["cost"] += cost
    return {"total": total, "by_severity": by_sev}


def _ds_css() -> str:
    return """
a{color:#6CB4EE;text-decoration:underline;text-decoration-thickness:1px;text-underline-offset:2px}
a:visited{color:#B39DDB}
a:hover,a:focus{color:#90CAF9;outline:2px solid #6CB4EE;outline-offset:2px;border-radius:2px}
a:active{color:#E1BEE7}
.top-nav a{text-decoration:none;outline:none}
.top-nav a:hover{text-decoration:none}
.top-nav{position:sticky;top:0;z-index:500;display:flex;align-items:center;gap:4px;padding:8px 24px;
  background:var(--bg-elevated);border-bottom:1px solid var(--border);font-size:13px;flex-wrap:wrap}
.top-nav .brand{font-weight:700;color:var(--primary);font-size:14px;margin-right:12px}
.top-nav a{color:var(--text-secondary);text-decoration:none;padding:6px 10px;border-radius:6px;transition:all .2s;min-height:36px;display:inline-flex;align-items:center}
.top-nav a:hover{color:var(--text);background:var(--bg-card)}
.nav-dropdown{position:relative}
.nav-dropdown>.nav-toggle{cursor:pointer;user-select:none;padding:6px 10px;border-radius:6px;color:var(--text-secondary);font-size:13px;display:inline-flex;align-items:center;gap:4px;transition:all .2s;min-height:36px;border:none;background:none;font-family:inherit}
.nav-dropdown>.nav-toggle:hover,.nav-dropdown:focus-within>.nav-toggle{color:var(--text);background:var(--bg-card)}
.nav-dropdown>.nav-toggle::after{content:'\\25BE';font-size:10px;margin-left:2px}
.nav-menu{display:none;position:absolute;top:100%;left:0;min-width:220px;background:var(--bg-elevated);border:1px solid var(--border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.3);padding:6px 0;z-index:600;margin-top:4px}
.nav-dropdown:hover>.nav-menu,.nav-dropdown:focus-within>.nav-menu{display:block}
.nav-menu a{display:flex;padding:8px 16px;color:var(--text-secondary);font-size:12px;border-radius:0;min-height:auto;white-space:nowrap}
.nav-menu a:hover{color:var(--text);background:var(--bg-card)}
.nav-menu .nav-sep{height:1px;background:var(--border);margin:4px 12px}
.full-width-content{padding:32px 40px;max-width:1200px;margin:0 auto}
.exec-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:20px;margin:24px 0}
.exec-panel{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px}
.exec-panel h3{font-size:14px;color:var(--text-secondary);margin-bottom:12px;border:none;padding:0}
.legend{display:flex;gap:16px;flex-wrap:wrap}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--text-secondary)}
.legend-dot{width:10px;height:10px;border-radius:50%;display:inline-block}
.category-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin:16px 0}
.category-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px;text-align:center;transition:all .3s;cursor:default}
.category-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md)}
.category-icon{font-size:32px;margin-bottom:8px}
.category-name{font-size:13px;color:var(--text-secondary);margin-bottom:4px}
.category-score{font-size:28px;font-weight:700;font-family:var(--font-mono)}
.category-level{font-size:11px;text-transform:uppercase;font-weight:600;letter-spacing:.5px;margin-top:2px}
.category-findings{font-size:11px;color:var(--text-muted);margin-top:4px}
.score-display{display:flex;align-items:center;gap:40px;flex-wrap:wrap;margin:20px 0}
.score-info{display:flex;flex-direction:column;gap:6px}
.level-badge{display:inline-block;padding:4px 12px;border-radius:6px;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.sev-bars{display:flex;flex-direction:column;gap:8px;margin:12px 0}
.sev-row{display:flex;align-items:center;gap:12px}
.sev-label{width:70px;font-size:12px;text-transform:uppercase;color:var(--text-secondary);font-weight:600}
.sev-track{flex:1;height:10px;background:var(--bar-bg);border-radius:5px;overflow:hidden}
.sev-fill{height:100%;border-radius:5px;transition:width .6s cubic-bezier(.16,1,.3,1)}
.sev-count{width:30px;text-align:right;font-family:var(--font-mono);font-size:13px}
.finding-card{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:20px;margin-bottom:12px;transition:all .2s}
.finding-card:hover{background:var(--bg-card-hover)}
.finding-card:focus{outline:2px solid #0078D4;outline-offset:2px}
.finding-card.expanded .finding-desc,.finding-card.expanded .ds-relevance,.finding-card.expanded .affected-resources,.finding-card.expanded .remediation-block{max-height:none}
.finding-card.critical{border-left:4px solid #D13438}
.finding-card.high{border-left:4px solid #F7630C}
.finding-card.medium{border-left:4px solid #FFB900}
.finding-card.low{border-left:4px solid #107C10}
.finding-card.informational{border-left:4px solid #A8A6A3}
.finding-title{font-size:15px;font-weight:600;margin-bottom:6px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.finding-desc{font-size:13px;color:var(--text-secondary);line-height:1.6;margin-bottom:8px}
.finding-meta{display:flex;gap:20px;flex-wrap:wrap;font-size:12px;color:var(--text-muted);margin-bottom:10px;padding:6px 0}
.finding-meta span{display:inline-flex;align-items:center;gap:5px;white-space:nowrap}
.remediation-box{margin-top:10px;padding:14px;background:var(--remediation-bg);border-left:3px solid var(--remediation-border);border-radius:6px}
.remediation-box h4{font-size:12px;color:var(--remediation-border);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}
.remediation-box .rem-desc{font-size:13px;color:#A5D6A7;margin-bottom:8px;line-height:1.5}
.remediation-box pre{font-family:var(--font-mono);font-size:12px;background:var(--code-bg);border:1px solid var(--code-border);border-radius:4px;padding:10px;overflow-x:auto;color:var(--text);margin:6px 0;white-space:pre-wrap;word-break:break-all}
.remediation-box .portal-steps{margin:6px 0 0;padding-left:20px;font-size:12px;color:var(--text-secondary)}
.remediation-box .portal-steps li{margin-bottom:3px}
.affected-details{margin-top:8px}
.affected-details summary{cursor:pointer;color:var(--primary);font-weight:500;font-size:12px;padding:6px 0}
.affected-details summary:hover{text-decoration:underline}
.resource-table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;margin:8px 0;border:1px solid var(--border);border-radius:8px;overflow:hidden}
.res-table-wrap{overflow-x:auto;margin:8px 0;-webkit-overflow-scrolling:touch}
.resource-table thead{background:var(--bg-elevated)}
.resource-table th{padding:8px 12px;text-align:left;font-weight:600;color:var(--text-secondary);text-transform:uppercase;font-size:11px;letter-spacing:.3px;border-bottom:2px solid var(--border);white-space:nowrap}
.resource-table td{padding:8px 12px;border-bottom:1px solid var(--border-light,var(--border));color:var(--text);vertical-align:top}
.resource-table tbody tr:last-child td{border-bottom:none}
.resource-table tbody tr:hover{background:var(--bg-card-hover,rgba(255,255,255,.03))}
.resource-table tbody tr.res-alt{background:rgba(255,255,255,.015)}
.resource-table .res-sev{white-space:nowrap;width:70px}
.resource-table .res-name{min-width:140px}
.resource-table .res-name-primary{font-weight:600;color:var(--text);white-space:nowrap}
.resource-table .res-type-sub{font-size:10px;color:var(--text-muted);margin-top:2px}
.resource-table .res-id{font-family:var(--font-mono);font-size:11px;color:var(--text-secondary);word-break:break-all;max-width:280px}
.resource-table .res-detail{font-size:11px;color:var(--text-secondary)}
.resource-table .res-risk{font-size:11px;color:var(--text-secondary);line-height:1.5;max-width:300px;min-width:180px}
.resource-table .res-rem{font-size:11px;max-width:320px;min-width:160px}
.resource-table .res-rem code{font-family:var(--font-mono);font-size:10px;background:var(--code-bg);padding:4px 6px;border-radius:4px;display:block;word-break:break-all;white-space:pre-wrap}
.resource-table .risk-tag{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600;letter-spacing:.3px}
.resource-table .risk-tag.critical{background:rgba(209,52,56,.15);color:#D13438}
.resource-table .risk-tag.high{background:rgba(247,99,12,.15);color:#F7630C}
.resource-table .risk-tag.medium{background:rgba(255,185,0,.15);color:#FFB900}
.resource-table .risk-tag.low{background:rgba(16,124,16,.15);color:#107C10}
.resource-table .risk-tag.info{background:rgba(168,166,163,.15);color:#A8A6A3}
.more-row td{font-style:italic;color:var(--text-muted)!important}
.filter-bar{display:flex;align-items:center;gap:8px;margin-bottom:16px;flex-wrap:wrap;font-size:13px}
.filter-bar input[type="search"]{min-width:240px;padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar select{padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar button{padding:6px 14px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:12px;min-height:36px;transition:all .2s}
.filter-bar button:hover{border-color:var(--primary);color:var(--primary)}
.method-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin:16px 0}
.method-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:20px}
.method-card .method-icon{font-size:28px;margin-bottom:8px}
.method-card h4{font-size:14px;margin-bottom:8px;color:var(--text)}
.method-card p{font-size:12px;color:var(--text-secondary);line-height:1.6}
.top-findings-table{border:1px solid var(--border);border-radius:8px;overflow:hidden;margin:12px 0}
.top-finding-header{display:grid;grid-template-columns:40px 90px 1fr 180px 80px;gap:0;padding:10px 16px;background:var(--bg-elevated);font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:var(--text-secondary);border-bottom:2px solid var(--border)}
.top-finding-row{display:grid;grid-template-columns:40px 90px 1fr 180px 80px;gap:0;padding:10px 16px;align-items:center;border-bottom:1px solid var(--border-light,var(--border));transition:background .15s}
.top-finding-row:last-child{border-bottom:none}
.top-finding-row:hover{background:var(--bg-card-hover,rgba(255,255,255,.03))}
.top-finding-row.tf-alt{background:rgba(255,255,255,.02)}
.tf-rank{font-size:16px;font-weight:700;font-family:var(--font-mono);color:var(--text-muted);text-align:center}
.tf-sev{display:flex;align-items:center}
.tf-title{font-size:13px;font-weight:600;padding-right:12px}
.tf-cat{font-size:12px;color:var(--text-muted)}
.tf-count{font-size:13px;font-family:var(--font-mono);text-align:right;font-weight:600}
.how-to-read{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;margin-bottom:24px}
.how-to-read h4{margin-bottom:8px}
.how-to-read p{font-size:13px;color:var(--text-secondary);margin-bottom:6px}
.zoom-controls{display:flex;align-items:center;gap:4px;margin-left:auto}
.zoom-controls button{padding:4px 10px;border:1px solid var(--border);border-radius:4px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:14px;min-height:32px;transition:all .2s}
.zoom-controls button:hover{border-color:var(--primary);color:var(--primary)}
#zoom-label{font-size:12px;font-family:var(--font-mono);width:40px;text-align:center}
.export-bar{display:flex;gap:4px}
/* ── Hover tooltips ── */
[data-tip]{cursor:help}
th [data-tip],h2 [data-tip],h3 [data-tip],.legend [data-tip]{border-bottom:1px dotted var(--text-muted)}
.stat-card[data-tip],.badge[data-tip],.level-badge[data-tip]{border-bottom:none;cursor:help}
.exec-panel-title{cursor:help;display:inline-block}
.score-ring-wrap{cursor:help;display:inline-block}
.nav-tip{cursor:help}
#ciq-tooltip{position:fixed;z-index:99999;pointer-events:none;opacity:0;transition:opacity .18s ease;max-width:380px;min-width:200px;padding:14px 18px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  color:var(--text);
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  border-radius:10px;
  font-size:12.5px;line-height:1.6;font-weight:400;text-transform:none;letter-spacing:normal;white-space:normal;
  box-shadow:0 2px 6px rgba(0,0,0,.18),0 8px 24px rgba(0,0,0,.32),0 0 0 1px rgba(255,255,255,.06) inset;
}
#ciq-tooltip.visible{opacity:1}
#ciq-tooltip::before{content:'';position:absolute;width:12px;height:12px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  transform:rotate(45deg);z-index:-1}
#ciq-tooltip.arrow-bottom::before{bottom:-7px;left:var(--arrow-x,24px);border-top:none;border-left:none}
#ciq-tooltip.arrow-top::before{top:-7px;left:var(--arrow-x,24px);border-bottom:none;border-right:none}
#ciq-tooltip .t-sep{display:block;border-top:1px solid rgba(255,255,255,.15);margin:8px 0 4px;padding-top:6px;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--primary)}
@media(max-width:768px){.full-width-content{padding:16px}.exec-grid{grid-template-columns:1fr}.stat-grid{grid-template-columns:repeat(2,1fr)}.top-nav{padding:8px 12px}}

/* ── Data Security Relevance Shield ─────────────────────────── */
.ds-relevance{position:relative;margin:16px 0;padding:14px 18px 14px 18px;
  background:linear-gradient(135deg,rgba(0,120,212,.08),rgba(107,63,160,.08));
  border:1.5px solid rgba(0,120,212,.25);border-radius:10px;
  font-size:12.5px;line-height:1.7;color:var(--text-secondary);
  overflow:hidden;clear:both}
.ds-relevance::before{content:none}
.ds-relevance::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,#0078D4,#6B3FA0,#00B7C3,#F7630C);
  border-radius:10px 10px 0 0}
.ds-relevance .ds-tag{display:inline-block;padding:2px 8px;border-radius:4px;
  font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;
  background:linear-gradient(135deg,#0078D4,#6B3FA0);color:#fff;
  margin-bottom:6px;animation:dsPulse 3s ease-in-out infinite}
@keyframes dsPulse{0%,100%{opacity:1;box-shadow:0 0 4px rgba(0,120,212,.3)}
  50%{opacity:.85;box-shadow:0 0 12px rgba(0,120,212,.6)}}
@media print{.ds-relevance{background:#f0f6ff;border-color:#0078D4;-webkit-print-color-adjust:exact}
  .ds-relevance .ds-tag{animation:none}}
@media print{.top-nav,.back-to-top,.filter-bar,.zoom-controls,.export-bar{display:none!important}.full-width-content{padding:16px;max-width:100%}body{background:#fff;color:#000;font-size:12px}.finding-card,.category-card,.exec-panel,.stat-card{border:1px solid #ccc;background:#fff}.badge{border:1px solid #333;print-color-adjust:exact;-webkit-print-color-adjust:exact}.remediation-box{background:#f1faf1;border-left-color:#107C10}}
/* ── Severity badge WCAG overrides ──────────────────────────── */
.badge-medium{color:#1A1600!important}
.badge-informational{color:#1A1A1A!important}
.badge-high{color:#1A0A00!important}
/* ── Compliance Framework Popup ─────────────────────────────── */
.compliance-fw-wrap{position:relative;margin:8px 0}
.fw-link{display:inline-flex;align-items:center;gap:4px;padding:4px 12px;
  border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;
  color:var(--primary);background:color-mix(in srgb,var(--primary) 8%,transparent);
  border:1px solid color-mix(in srgb,var(--primary) 20%,transparent);
  text-decoration:none!important;transition:all .2s}
.fw-link:hover{background:color-mix(in srgb,var(--primary) 15%,transparent)}
/* Backdrop overlay */
.fw-backdrop{display:none;position:fixed;inset:0;z-index:998;
  background:rgba(0,0,0,.55);backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px)}
.fw-backdrop.fw-show{display:block}
@keyframes fwPopIn{0%{opacity:0;transform:translate(-50%,-50%) scale(.88)}100%{opacity:1;transform:translate(-50%,-50%) scale(1)}}
.fw-popup{display:none;flex-direction:column;position:fixed;
  left:50%;top:50%;transform:translate(-50%,-50%);
  z-index:999;min-width:560px;max-width:720px;max-height:70vh;
  background:var(--bg-elevated,#1e1e1e);
  border:5px ridge color-mix(in srgb,var(--primary) 60%,var(--border,#555));
  border-radius:14px;
  box-shadow:0 20px 60px rgba(0,0,0,.6),
    inset 0 1px 0 rgba(255,255,255,.1),
    0 0 0 2px rgba(0,0,0,.2);
  overflow:hidden}
.fw-popup.fw-open{animation:fwPopIn .28s cubic-bezier(.34,1.56,.64,1) both}
.fw-popup-hdr{display:flex;justify-content:space-between;align-items:center;
  padding:12px 16px;border-bottom:1px solid var(--border,#333);font-weight:700;font-size:13px}
.fw-popup-close{background:none;border:none;color:var(--text-secondary);
  font-size:16px;cursor:pointer;padding:4px 8px;border-radius:4px}
.fw-popup-close:hover{background:color-mix(in srgb,var(--text-secondary) 20%,transparent)}
.fw-popup-body{overflow-y:auto;padding:12px 16px;display:flex;flex-direction:column;gap:12px}
@keyframes fwSlideIn{0%{opacity:0;transform:translateY(8px)}100%{opacity:1;transform:translateY(0)}}
.fw-open .fw-section-hdr,.fw-open .fw-ctrl{animation:fwSlideIn .3s ease both}
.fw-open .fw-section-hdr:nth-child(1){animation-delay:.05s}
.fw-open .fw-ctrl:nth-child(2){animation-delay:.08s}
.fw-open .fw-ctrl:nth-child(3){animation-delay:.11s}
.fw-open .fw-ctrl:nth-child(4){animation-delay:.14s}
.fw-open .fw-ctrl:nth-child(5){animation-delay:.17s}
.fw-section-hdr{font-weight:700;font-size:12px;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px;
  color:var(--primary);border-left:3px solid var(--primary);padding-left:8px}
.fw-ctrl{padding:6px 8px;border-radius:6px;font-size:12px;
  background:color-mix(in srgb,var(--text-secondary) 5%,transparent);margin-bottom:4px}
.fw-ctrl:hover{background:color-mix(in srgb,var(--text-secondary) 10%,transparent)}
.fw-ctrl-id{font-weight:700;margin-right:8px;font-family:var(--font-mono,monospace);font-size:11px;color:var(--primary)}
.fw-ctrl-title{color:var(--text-primary)}
.fw-rec{margin-top:3px;font-size:11px;color:var(--text-secondary);line-height:1.4}

"""


# ── Report-specific JS ──────────────────────────────────────────────────

def _ds_js() -> str:
    return """
// Backdrop overlay for compliance popup
(function(){var d=document.createElement('div');d.id='fw-backdrop';d.className='fw-backdrop';d.addEventListener('click',function(){this.classList.remove('fw-show');document.querySelectorAll('.fw-popup').forEach(function(p){p.style.display='none';p.classList.remove('fw-open')})});document.body.prepend(d)})()
// Zoom
var zoomLevel=100;
function zoomIn(){zoomLevel=Math.min(zoomLevel+10,150);applyZoom()}
function zoomOut(){zoomLevel=Math.max(zoomLevel-10,70);applyZoom()}
function zoomReset(){zoomLevel=100;applyZoom()}
function applyZoom(){document.querySelector('.full-width-content').style.zoom=(zoomLevel/100);document.getElementById('zoom-label').textContent=zoomLevel+'%'}

// Filter findings
var _PAGE_SIZE=20;
var _visibleCount=_PAGE_SIZE;
function _applyPagination(){
  var cards=document.querySelectorAll('.finding-card[data-category]');
  var shown=0;
  cards.forEach(function(card){
    if(card.style.display==='none')return;// already hidden by filter
    shown++;
  });
  // if filtered, reset page count
  var btn=document.getElementById('show-more-btn');
  if(!btn)return;
  var idx=0;
  cards.forEach(function(card){
    if(card._filterHidden)return;
    idx++;
    card.style.display=idx<=_visibleCount?'':'none';
  });
  var remaining=shown-_visibleCount;
  if(remaining>0){btn.style.display='';btn.textContent='Show '+Math.min(remaining,_PAGE_SIZE)+' more findings ('+remaining+' remaining)';}
  else{btn.style.display='none';}
}
function showMoreFindings(){_visibleCount+=_PAGE_SIZE;_applyPagination();}
function filterFindings(){
  _visibleCount=_PAGE_SIZE;
  var q=(document.getElementById('finding-filter').value||'').toLowerCase();
  var sev=(document.getElementById('filter-severity').value||'').toLowerCase();
  var cat=(document.getElementById('filter-category').value||'').toLowerCase();
  var cards=document.querySelectorAll('.finding-card[data-category]');
  var shown=0;
  cards.forEach(function(card){
    var text=card.textContent.toLowerCase();
    var cSev=card.getAttribute('data-severity')||'';
    var cCat=card.getAttribute('data-category')||'';
    var match=true;
    if(q&&text.indexOf(q)<0)match=false;
    if(sev&&cSev!==sev)match=false;
    if(cat&&cCat!==cat)match=false;
    card._filterHidden=!match;
    card.style.display=match?'':'none';
    if(match)shown++;
  });
  var live=document.getElementById('findings-live');
  if(live)live.textContent=shown+' findings shown';
  _applyPagination();
}
document.addEventListener('DOMContentLoaded',function(){
  _applyPagination();
  // Keyboard navigation for finding cards
  document.addEventListener('keydown',function(e){
    if(e.target.classList.contains('finding-card')){
      var cards=Array.from(document.querySelectorAll('.finding-card[tabindex]')).filter(function(c){return c.style.display!=='none';});
      var idx=cards.indexOf(e.target);
      if(e.key==='ArrowDown'&&idx<cards.length-1){e.preventDefault();cards[idx+1].focus();}
      else if(e.key==='ArrowUp'&&idx>0){e.preventDefault();cards[idx-1].focus();}
      else if(e.key==='Enter'||e.key===' '){e.preventDefault();e.target.classList.toggle('expanded');}
    }
  });
});

// ── Tooltip engine (viewport-aware positioning) ──
(function(){
  var tip=document.getElementById('ciq-tooltip');
  if(!tip)return;
  var GAP=10,MARGIN=12;
  function show(ev){
    var tgt=ev.target.closest('[data-tip]');
    if(!tgt)return;
    var text=tgt.getAttribute('data-tip');
    if(!text)return;
    var d=document.createElement('span');d.textContent=text;var safe=d.innerHTML;
    safe=safe.replace(/\\n+YOUR TENANT:/g,'<span class="t-sep">&#x1F4CA; Your Tenant</span>');
    safe=safe.replace(/\\n/g,'<br>');
    tip.innerHTML=safe;
    tip.classList.add('visible');
    tip.setAttribute('aria-hidden','false');
    requestAnimationFrame(function(){
      var r=tgt.getBoundingClientRect();
      var tw=tip.offsetWidth,th=tip.offsetHeight;
      var vw=window.innerWidth,vh=window.innerHeight;
      var above=r.top-GAP-th;
      var below=r.bottom+GAP;
      var top,arrow;
      if(above>=MARGIN){top=above;arrow='arrow-bottom';}
      else if(below+th<=vh-MARGIN){top=below;arrow='arrow-top';}
      else{top=Math.max(MARGIN,vh-th-MARGIN);arrow='';}
      var left=r.left+r.width/2-tw/2;
      left=Math.max(MARGIN,Math.min(left,vw-tw-MARGIN));
      var arrowX=r.left+r.width/2-left;
      arrowX=Math.max(16,Math.min(arrowX,tw-16));
      tip.style.top=top+'px';
      tip.style.left=left+'px';
      tip.style.setProperty('--arrow-x',arrowX+'px');
      tip.className='visible'+(arrow?' '+arrow:'');
    });
  }
  function hide(){
    tip.classList.remove('visible');
    tip.setAttribute('aria-hidden','true');
    tip.className='';
  }
  document.addEventListener('mouseenter',show,true);
  document.addEventListener('mouseleave',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
  document.addEventListener('focusin',show,true);
  document.addEventListener('focusout',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
})();
"""


# ── Rendering helpers ────────────────────────────────────────────────────

def _severity_badge(sev: str) -> str:
    color = SEVERITY_COLORS.get(sev.lower(), "#A8A6A3")
    return f'<span class="badge badge-{esc(sev.lower())}" style="background:{color}">{esc(sev.upper())} RISK</span>'


def _render_finding(f: dict) -> str:
    sev = f.get("Severity", "medium").lower()
    cat = f.get("Category", "unknown")
    subcat = f.get("Subcategory", "")
    title = f.get("Title", "Untitled")
    desc = f.get("Description", "")
    affected = f.get("AffectedCount", 0)
    affected_resources = f.get("AffectedResources", [])
    remediation = f.get("Remediation", {})
    detected = f.get("DetectedAt", "")
    finding_id = f.get("DataSecurityFindingId", f"{cat}_{subcat}")

    # Remediation block
    rem_html = ""
    if remediation:
        rem_parts = []
        rem_desc = remediation.get("Description", "")
        if rem_desc:
            rem_parts.append(f'<div class="rem-desc">{esc(rem_desc)}</div>')
        cli = remediation.get("AzureCLI", "")
        if cli:
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">Azure CLI:</strong><pre>{esc(cli)}</pre></div>')
        ps = remediation.get("PowerShell", "")
        if ps:
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">PowerShell:</strong><pre>{esc(ps)}</pre></div>')
        steps = remediation.get("PortalSteps", [])
        if steps:
            step_items = "".join(f"<li>{esc(s)}</li>" for s in steps)
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">Portal Steps:</strong><ol class="portal-steps">{step_items}</ol></div>')
        if rem_parts:
            rem_html = f'<div class="remediation-box"><h4>&#128736; Remediation</h4>{"".join(rem_parts)}</div>'

    # Affected resources — rich per-resource table with severity, risk, remediation
    affected_html = ""
    if affected_resources and len(affected_resources) > 0:
        # Detect enriched fields and extra detail keys
        _STANDARD_KEYS = {"Name", "name", "Type", "type", "ResourceId", "resource_id", "id",
                          "Severity", "Risk", "ResourceRemediation"}
        detail_keys: list[str] = []
        for ar in affected_resources[:1]:
            if isinstance(ar, dict):
                detail_keys = [k for k in ar if k not in _STANDARD_KEYS]

        # Build header
        header_cells = '<th data-tip="Severity level for this specific resource.">Sev</th>'
        header_cells += '<th data-tip="Name of the affected Azure resource.">Resource</th>'
        header_cells += '<th data-tip="Full Azure Resource Manager (ARM) resource ID.">Resource ID</th>'
        for dk in detail_keys:
            header_cells += f'<th data-tip="Additional detail collected for this resource.">{esc(dk.replace("_", " ").title())}</th>'

        # Build rows
        rows_html = ""
        for ri, ar in enumerate(affected_resources[:30]):
            if isinstance(ar, dict):
                res_sev = ar.get("Severity", sev).lower()
                res_name = ar.get("Name", ar.get("name", "—"))
                res_type = ar.get("Type", ar.get("type", "—"))
                res_id = ar.get("ResourceId", ar.get("resource_id", ar.get("id", "—")))
                sev_badge = _severity_badge(res_sev)
                res_id_str = str(res_id)
                if res_id_str.startswith("/subscriptions/"):
                    portal_url = f"https://portal.azure.com/#@/resource{res_id_str}"
                    res_id_cell = f'<td class="res-id"><a href="{esc(portal_url)}" target="_blank" rel="noopener" title="Open in Azure Portal">{esc(res_id_str)}</a></td>'
                else:
                    res_id_cell = f'<td class="res-id">{esc(res_id_str)}</td>'

                zebra = ' class="res-alt"' if ri % 2 == 1 else ''
                row = f'<td class="res-sev">{sev_badge}</td>'
                row += (f'<td class="res-name">'
                        f'<div class="res-name-primary">{esc(str(res_name))}</div>'
                        f'<div class="res-type-sub">{esc(str(res_type))}</div></td>')
                row += res_id_cell
                for dk in detail_keys:
                    val = ar.get(dk, "—")
                    row += f'<td class="res-detail">{esc(str(val))}</td>'
                rows_html += f'<tr{zebra}>{row}</tr>'
            else:
                total_cols = 3 + len(detail_keys)
                rows_html += f'<tr><td colspan="{total_cols}">{esc(str(ar))}</td></tr>'
        if len(affected_resources) > 30:
            total_cols = 3 + len(detail_keys)
            rows_html += (f'<tr class="more-row"><td colspan="{total_cols}">'
                          f'… and {len(affected_resources) - 30} more resources</td></tr>')

        sev_class = sev if sev in ("critical", "high", "medium", "low") else "info"
        affected_html = (
            f'<details class="affected-details" open>'
            f'<summary><span class="risk-tag {sev_class}">{sev.upper()}</span> '
            f'{affected} affected resource(s)</summary>'
            f'<div class="res-table-wrap"><table class="resource-table"><thead><tr>{header_cells}</tr></thead>'
            f'<tbody>{rows_html}</tbody></table></div></details>'
        )

    cat_meta = _CATEGORY_META.get(cat, {})
    cat_name = cat_meta.get("name", cat.replace("_", " ").title())

    # Compliance framework mapping — "N Framework Mappings" link + popup
    compliance_mapping = f.get("ComplianceMapping", {})
    compliance_details = f.get("ComplianceDetails", {})
    compliance_html = ""
    if compliance_mapping:
        # Colors handled by CSS (--primary) — no per-framework colors
        total_ctrls = sum(len(ctrls) for ctrls in compliance_mapping.values())
        popup_id = f"fw-popup-{esc(finding_id)}"

        # Build popup content grouped by framework
        popup_sections = []
        for fw, controls in sorted(compliance_mapping.items()):
            # color handled by CSS
            rows = []
            for ctrl in controls:
                key = f"{fw}:{ctrl}"
                det = compliance_details.get(key, {})
                ctrl_title = esc(det.get("title", ""))
                ctrl_rationale = esc(det.get("rationale", ""))
                ctrl_rec = esc(det.get("recommendation", ""))
                tooltip = f' title="{ctrl_rationale}"' if ctrl_rationale else ""
                rec_html = f'<div class="fw-rec">{ctrl_rec}</div>' if ctrl_rec else ""
                rows.append(
                    f'<div class="fw-ctrl"{tooltip}>'
                    f'<span class="fw-ctrl-id">{esc(fw)} {esc(ctrl)}</span>'
                    f'<span class="fw-ctrl-title">{ctrl_title}</span>'
                    f'{rec_html}'
                    f'</div>'
                )
            popup_sections.append(
                f'<div class="fw-section">'
                f'<div class="fw-section-hdr">'
                f'{esc(fw)} ({len(controls)})</div>'
                f'{"".join(rows)}</div>'
            )
        popup_body = "".join(popup_sections)

        compliance_html = (
            f'<div class="compliance-fw-wrap">'
            f'<a href="javascript:void(0)" class="fw-link" '
            f'onclick="(function(e){{var p=document.getElementById(\'{popup_id}\');'
            f'var bk=document.getElementById(\'fw-backdrop\');'
            f'if(p.style.display===\'flex\'){{p.style.display=\'none\';p.classList.remove(\'fw-open\');bk.classList.remove(\'fw-show\')}}'
            f'else{{document.querySelectorAll(\'.fw-popup\').forEach(function(x){{x.style.display=\'none\';x.classList.remove(\'fw-open\')}});'
            f'p.style.display=\'flex\';p.classList.remove(\'fw-open\');void p.offsetWidth;p.classList.add(\'fw-open\');bk.classList.add(\'fw-show\')}}'
            f'e.stopPropagation()}})(event)">'
            f'\U0001f4cb {total_ctrls} Framework Mapping{"s" if total_ctrls != 1 else ""} \u25b8</a>'
            f'<div class="fw-popup" id="{popup_id}" style="display:none" '
            f'onclick="event.stopPropagation()">'
            f'<div class="fw-popup-hdr">'
            f'<span>Compliance Framework Mappings</span>'
            f'<button class="fw-popup-close" '
            f'onclick="this.closest(\'.fw-popup\').style.display=\'none\';this.closest(\'.fw-popup\').classList.remove(\'fw-open\');document.getElementById(\'fw-backdrop\').classList.remove(\'fw-show\')">\u2715</button>'
            f'</div>'
            f'<div class="fw-popup-body">{popup_body}</div>'
            f'</div></div>'
        )

    # Merged-from badge for consolidated findings
    merged_from = f.get("MergedFrom", [])
    if merged_from:
        merged_badge = (
            f'<span title="Consolidated from {len(merged_from)} checks">'
            f'&#128279; {len(merged_from)} checks merged</span>'
        )
        merged_table_rows = "".join(
            f'<tr>'
            f'<td style="padding:4px 10px;white-space:nowrap">{_severity_badge(m.get("Severity", "medium").lower())}</td>'
            f'<td style="padding:4px 10px;font-weight:600;white-space:nowrap">{esc(m.get("Subcategory", "").replace("_", " ").title())}</td>'
            f'<td style="padding:4px 10px;color:var(--text-secondary)">{esc(m.get("Title", ""))}</td>'
            f'</tr>'
            for m in merged_from
        )
        merged_html = (
            f'<details class="merged-details" style="margin:6px 0">'
            f'<summary style="cursor:pointer;font-size:12px;color:var(--text-muted)">'
            f'View {len(merged_from)} individual checks</summary>'
            f'<table style="width:100%;border-collapse:collapse;font-size:12px;margin:6px 0">'
            f'<thead><tr style="border-bottom:1px solid var(--border)">'
            f'<th style="text-align:left;padding:4px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Sev</th>'
            f'<th style="text-align:left;padding:4px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Check</th>'
            f'<th style="text-align:left;padding:4px 10px;font-size:10px;text-transform:uppercase;color:var(--text-muted)">Description</th>'
            f'</tr></thead>'
            f'<tbody>{merged_table_rows}</tbody></table></details>'
        )
    else:
        merged_badge = f'<span>&#128196; {esc(subcat.replace("_", " ").title())}</span>'
        merged_html = ""

    # Data Security Relevance callout
    ds_rel_text = _DATA_SECURITY_RELEVANCE.get(cat, "")
    ds_relevance_html = ""
    if ds_rel_text:
        ds_relevance_html = (
            f'<div class="ds-relevance">'
            f'<div class="ds-tag">&#128737; Why This is Data Security</div>'
            f'<span>{esc(ds_rel_text)}</span></div>'
        )

    return (
        f'<div class="finding-card {sev}" id="finding-{esc(finding_id)}" data-severity="{sev}" data-category="{esc(cat)}" '
        f'data-subcategory="{esc(subcat)}" data-affected="{affected}" tabindex="0" role="article">'
        f'<div class="finding-title">{_severity_badge(sev)} {esc(title)}</div>'
        f'<div class="finding-meta">'
        f'<span>&#128193; {esc(cat_name)}</span>'
        f'{merged_badge}'
        f'<span>&#128202; {affected} affected</span>'
        + (f'<span>&#128337; {esc(detected[:19])}</span>' if detected else "")
        + f'</div>'
        f'{compliance_html}'
        f'{merged_html}'
        f'<div class="finding-desc">{esc(desc)}</div>'
        f'{ds_relevance_html}'
        f'{affected_html}'
        f'{rem_html}'
        f'</div>'
    )


def _render_top_findings(top_findings: list[dict]) -> str:
    if not top_findings:
        return '<p class="empty">No findings — data layer is secure. Well done!</p>'
    header = (
        '<div class="top-findings-table">'
        '<div class="top-finding-header">'
        '<div class="tf-rank">#</div>'
        '<div class="tf-sev">Severity</div>'
        '<div class="tf-title">Finding</div>'
        '<div class="tf-cat">Category</div>'
        '<div class="tf-count">Affected</div>'
        '</div>'
    )
    rows = []
    for i, r in enumerate(top_findings[:10], 1):
        sev = r.get("Severity", "medium").lower()
        cat = r.get("Category", "")
        cat_meta = _CATEGORY_META.get(cat, {})
        cat_name = cat_meta.get("name", cat.replace("_", " ").title())
        zebra = ' tf-alt' if i % 2 == 0 else ''
        rows.append(
            f'<div class="top-finding-row{zebra}">'
            f'<div class="tf-rank">{i}</div>'
            f'<div class="tf-sev">{_severity_badge(sev)}</div>'
            f'<div class="tf-title">{esc(r.get("Title", ""))}</div>'
            f'<div class="tf-cat">{esc(cat_name)}</div>'
            f'<div class="tf-count">{r.get("AffectedCount", 0)}</div>'
            f'</div>'
        )
    return header + "".join(rows) + "</div>"


def _generate_exec_summary(scores: dict, finding_count: int, sub_count: int) -> str:
    """Generate a natural-language executive summary paragraph."""
    overall = scores.get("OverallScore", 0)
    level = scores.get("OverallLevel", "low")
    sev = scores.get("SeverityDistribution", {})
    n_crit = sev.get("critical", 0)
    n_high = sev.get("high", 0)
    cat_scores = scores.get("CategoryScores", {})

    # Identify weakest categories
    weak = sorted(
        [(k, v.get("Score", 0)) for k, v in cat_scores.items() if v.get("FindingCount", 0) > 0],
        key=lambda x: -x[1],
    )[:3]
    weak_names = [_CATEGORY_META.get(k, {"name": k})["name"] for k, _ in weak]

    if level == "critical":
        posture = "critically exposed"
    elif level == "high":
        posture = "at elevated risk"
    elif level == "medium":
        posture = "moderately secure with notable gaps"
    else:
        posture = "well-secured"

    para = (
        f"Across {sub_count} subscription{'s' if sub_count != 1 else ''}, "
        f"the data security posture is <strong>{posture}</strong> "
        f"with an overall risk score of <strong>{overall:.0f}/100</strong>. "
    )
    if n_crit or n_high:
        para += (
            f"There {'are' if n_crit + n_high != 1 else 'is'} "
            f"<strong>{n_crit} critical</strong> and <strong>{n_high} high</strong> "
            f"severity finding{'s' if n_crit + n_high != 1 else ''} requiring immediate attention. "
        )
    if finding_count == 0:
        para += "No findings were detected — excellent data-layer hygiene."
    elif weak_names:
        para += f"Priority areas: <strong>{', '.join(weak_names)}</strong>."
    return para


def _render_trend_section(trend: dict) -> str:
    """Render a trend comparison HTML section."""
    delta = trend.get("ScoreDelta", 0)
    arrow = "↑" if delta > 0 else "↓" if delta < 0 else "→"
    delta_color = "#D13438" if delta > 0 else "#107C10" if delta < 0 else "#FFB900"
    new_count = trend.get("NewCount", 0)
    resolved_count = trend.get("ResolvedCount", 0)
    prev_at = trend.get("PreviousAssessedAt", "")[:19]
    prev_total = trend.get("PreviousFindingCount", 0)
    curr_total = trend.get("CurrentFindingCount", 0)

    new_html = ""
    for f in trend.get("NewFindings", [])[:5]:
        sev = f.get("Severity", "medium").lower()
        sev_color = SEVERITY_COLORS.get(sev, "#FFB900")
        new_html += (
            f'<div style="display:flex;align-items:center;gap:8px;margin:4px 0">'
            f'<span class="severity-badge" style="background:{sev_color}">{esc(sev.upper())}</span>'
            f'<span>{esc(f.get("Title", ""))}</span></div>'
        )

    resolved_html = ""
    for f in trend.get("ResolvedFindings", [])[:5]:
        resolved_html += (
            f'<div style="margin:4px 0;color:var(--text-secondary)">'
            f'<span style="color:#107C10">✓</span> {esc(f.get("Title", ""))}</div>'
        )

    return f"""
<section id="trend" class="section" aria-labelledby="trend-heading">
  <h2 id="trend-heading" data-tip="Compare current assessment with the previous run to track score movement and finding changes.">&#128200; Trend Comparison</h2>
  <div class="exec-grid">
    <div class="exec-panel" data-tip="Score delta between the current and previous assessments. Green arrow = improvement, red = regression.">
      <h3>Score Trend</h3>
      <div style="text-align:center;font-size:32px;font-weight:700;color:{delta_color}">
        {trend.get("PreviousScore", 0):.0f} {arrow} {trend.get("CurrentScore", 0):.0f}
      </div>
      <div style="text-align:center;color:var(--text-secondary);font-size:13px;margin-top:4px">
        Delta: <strong style="color:{delta_color}">{'+' if delta > 0 else ''}{delta:.1f}</strong>
      </div>
      <div style="text-align:center;color:var(--text-secondary);font-size:12px;margin-top:8px">
        Previous: {esc(prev_at)} ({prev_total} findings) → Current: {curr_total} findings
      </div>
    </div>
    <div class="exec-panel" data-tip="Findings detected in this assessment that were not present in the previous run.">
      <h3>New Findings ({new_count})</h3>
      {new_html if new_html else '<p style="color:var(--text-secondary)">No new findings</p>'}
    </div>
    <div class="exec-panel" data-tip="Findings from the previous assessment that are no longer detected — remediation confirmed.">
      <h3>Resolved Findings ({resolved_count})</h3>
      {resolved_html if resolved_html else '<p style="color:var(--text-secondary)">No resolved findings</p>'}
    </div>
  </div>
</section>
"""


# ── Main generator ───────────────────────────────────────────────────────

def generate_data_security_report(results: dict, output_dir: str | pathlib.Path) -> pathlib.Path:
    """Generate the Data Security Assessment HTML report.

    Args:
        results: Result dict from run_data_security_assessment (keys:
                 DataSecurityScores, Findings, FindingCount, SubscriptionCount,
                 EvidenceSource, AssessedAt, etc.).
        output_dir: Directory to write the report into.

    Returns:
        Path to the generated HTML file.
    """
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "data-security-assessment.html"

    ts = format_date_short()
    scores = results.get("DataSecurityScores", {})
    findings = results.get("Findings", [])
    finding_count = results.get("FindingCount", len(findings))
    sub_count = results.get("SubscriptionCount", 0)
    evidence_src = results.get("EvidenceSource", "unknown")
    assessed_at = results.get("AssessedAt", "")
    tenant_id = results.get("TenantId", "")
    tenant_display = results.get("TenantDisplayName", "")
    evidence_record_count = results.get("EvidenceRecordCount", 0)
    report_id = f"CIQ-DS-{ts.replace('-', '').replace(':', '').replace(' ', '-')}"

    overall_score = scores.get("OverallScore", 0)
    overall_level = scores.get("OverallLevel", "low")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})
    top_findings = scores.get("TopFindings", [])

    n_critical = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_medium = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)

    # Level badge
    level_meta = _SCORE_LEVEL_META.get(overall_level, _SCORE_LEVEL_META["medium"])
    level_color = level_meta["color"]
    level_label = level_meta["label"]

    # Score ring
    score_ring = _ring_score_svg(overall_score, size=160)

    # Severity donut
    sev_donut = _donut_svg([
        ("Critical", n_critical, "#D13438"),
        ("High", n_high, "#F7630C"),
        ("Medium", n_medium, "#FFB900"),
        ("Low", n_low, "#107C10"),
        ("Info", n_info, "#A8A6A3"),
    ], size=140)

    # Severity bars
    max_sev = max(n_critical, n_high, n_medium, n_low, n_info, 1)
    sev_bars_html = ""
    for sev_name, sev_count, sev_color in [
        ("Critical", n_critical, _SEV_COLORS["critical"]),
        ("High", n_high, _SEV_COLORS["high"]),
        ("Medium", n_medium, _SEV_COLORS["medium"]),
        ("Low", n_low, _SEV_COLORS["low"]),
        ("Info", n_info, _SEV_COLORS["informational"]),
    ]:
        pct = (sev_count / max_sev) * 100 if max_sev > 0 else 0
        sev_bars_html += (
            f'<div class="sev-row">'
            f'<span class="sev-label">{sev_name}</span>'
            f'<div class="sev-track"><div class="sev-fill" style="width:{pct:.0f}%;background:{sev_color}"></div></div>'
            f'<span class="sev-count">{sev_count}</span>'
            f'</div>'
        )

    # Category sparklines (from trend data if available)
    cat_sparklines: dict[str, str] = {}
    _trend_s = results.get("Trend")
    if _trend_s and isinstance(_trend_s, dict):
        _cat_hist = _trend_s.get("CategoryHistory", {})
        for _ck, _hv in sorted(_cat_hist.items()):
            if isinstance(_hv, list) and len(_hv) >= 2:
                _m = _CATEGORY_META.get(_ck, {"color": "#888"})
                cat_sparklines[_ck] = _sparkline_svg(_hv, color=_m.get("color", "#0078D4"))

    # Category cards
    _ALL_CATS = ["storage", "database", "cosmosdb", "pgmysql", "keyvault", "encryption", "data_access", "private_endpoints", "purview", "file_sync", "m365_dlp", "data_classification", "backup_dr", "container_security", "network_segmentation", "data_residency", "threat_detection", "redis", "messaging", "ai_services", "data_pipeline", "identity", "sharepoint_governance", "data_lifecycle", "dlp_alert", "app_config", "databricks", "apim", "frontdoor", "secret_sprawl", "firewall", "bastion", "policy_compliance", "defender_score", "stale_permissions", "data_exfiltration", "conditional_access", "config_drift", "supply_chain"]
    # Ensure any category found in actual findings/scores is included
    for _extra_cat in sorted(cat_scores):
        if _extra_cat not in _ALL_CATS:
            _ALL_CATS.append(_extra_cat)
    cat_cards_html = ""
    for cat_key in _ALL_CATS:
        meta = _CATEGORY_META.get(cat_key, {"icon": "&#128196;", "name": cat_key.replace("_", " ").title(), "color": "#888", "description": cat_key})
        cs = cat_scores.get(cat_key, {"Score": 0, "Level": "low", "FindingCount": 0})
        c_score = cs.get("Score", 0)
        c_level = cs.get("Level", "low")
        c_count = cs.get("FindingCount", 0)
        c_color = _SCORE_LEVEL_META.get(c_level, _SCORE_LEVEL_META["low"])["color"]
        cat_cards_html += (
            f'<div class="category-card" data-tip="{esc(meta["description"])} Score: {c_score:.0f}/100 ({c_level.upper()}).">'
            f'<div class="category-icon">{meta["icon"]}</div>'
            f'<div class="category-name">{esc(meta["name"])}</div>'
            f'<div class="category-score" style="color:{c_color}">{c_score:.0f}</div>'
            f'<div class="category-level" style="color:{c_color}">{esc(c_level.upper())}</div>'
            f'<div class="category-findings">{c_count} finding{"s" if c_count != 1 else ""}{cat_sparklines.get(cat_key, "")}</div>'
            f'</div>'
        )



    # All findings HTML, grouped by category
    findings_by_cat: dict[str, list[dict]] = {}
    for f in findings:
        findings_by_cat.setdefault(f.get("Category", "unknown"), []).append(f)

    # Remediation impact panel
    impact = results.get("RemediationImpact", {})
    impact_panel_html = ""
    if impact:
        cur = impact.get("CurrentScore", 0)
        if_crit = impact.get("IfCriticalFixed", 0)
        if_high = impact.get("IfHighFixed", 0)
        crit_delta = cur - if_crit
        high_delta = cur - if_high
        impact_panel_html = (
            f'<div class="exec-panel" style="grid-column:span 2;margin-top:20px" '
            f'data-tip="Projected score improvement if findings are remediated. Shows the score impact of fixing critical, high, and all findings.">'
            f'<h3>&#128200; Remediation Impact</h3>'
            f'<div style="display:flex;gap:24px;flex-wrap:wrap;margin-top:12px">'
            f'<div style="text-align:center;flex:1;min-width:140px">'
            f'<div style="font-size:28px;font-weight:700;color:var(--text)">{cur:.0f}</div>'
            f'<div style="font-size:12px;color:var(--text-secondary)">Current Score</div></div>'
            f'<div style="text-align:center;flex:1;min-width:140px">'
            f'<div style="font-size:28px;font-weight:700;color:#D13438">{if_crit:.0f}</div>'
            f'<div style="font-size:12px;color:var(--text-secondary)">If Critical Fixed</div>'
            f'<div style="font-size:11px;color:#107C10">&#9660; {crit_delta:.0f} pts</div></div>'
            f'<div style="text-align:center;flex:1;min-width:140px">'
            f'<div style="font-size:28px;font-weight:700;color:#F7630C">{if_high:.0f}</div>'
            f'<div style="font-size:12px;color:var(--text-secondary)">If Critical+High Fixed</div>'
            f'<div style="font-size:11px;color:#107C10">&#9660; {high_delta:.0f} pts</div></div>'
            f'<div style="text-align:center;flex:1;min-width:140px">'
            f'<div style="font-size:28px;font-weight:700;color:#107C10">0</div>'
            f'<div style="font-size:12px;color:var(--text-secondary)">If All Fixed</div>'
            f'<div style="font-size:11px;color:#107C10">&#9660; {cur:.0f} pts</div></div>'
            f'</div></div>'
        )

    all_findings_html = ""
    cat_options = ""
    findings_nav_items = ""
    for cat_key, cat_findings in sorted(findings_by_cat.items()):
        meta = _CATEGORY_META.get(cat_key, {"name": cat_key.title(), "icon": "&#128196;"})
        cat_options += f'<option value="{esc(cat_key)}">{esc(meta["name"])} ({len(cat_findings)})</option>'
        findings_nav_items += f'      <a href="#cat-{esc(cat_key)}">{meta["icon"]} {esc(meta["name"])} ({len(cat_findings)})</a>\n'

        sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
        sorted_findings = sorted(cat_findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))

        all_findings_html += f'<h3 id="cat-{esc(cat_key)}" style="margin-top:24px">{meta["icon"]} {esc(meta["name"])} ({len(cat_findings)} findings)</h3>'
        for f in sorted_findings:
            all_findings_html += _render_finding(f)

    # Executive summary paragraph
    exec_summary_text = _generate_exec_summary(scores, finding_count, sub_count)

    # Trend comparison (optional)
    trend = results.get("Trend")
    trend_section_html = _render_trend_section(trend) if trend else ""

    # Suppressed findings info
    suppressed_count = results.get("SuppressedCount", 0)
    suppressed_html = ""
    if suppressed_count:
        suppressed_html = (
            f'<div style="background:var(--bg-card);border:1px solid var(--border);'
            f'border-radius:8px;padding:12px 16px;margin-top:12px;font-size:13px">'
            f'&#128683; <strong>{suppressed_count}</strong> finding(s) suppressed '
            f'(accepted risk — excluded from scoring).</div>'
        )


    # Collection errors banner
    collection_errors = results.get("CollectionErrors", [])
    enrichment_errors = results.get("EnrichmentErrors", [])
    collection_errors_html = ""
    if collection_errors:
        err_items = "".join(
            f'<li><code>{esc(e.get("query", ""))}</code> &mdash; {esc(e.get("error", "unknown"))}</li>'
            for e in collection_errors
        )
        collection_errors_html = (
            f'<div style="background:rgba(209,52,56,0.08);border:1px solid rgba(209,52,56,0.3);'
            f'border-radius:8px;padding:14px 18px;margin:16px 0;font-size:13px;'
            f'border-left:4px solid #D13438">'
            f'<strong style="color:#D13438">&#9888; {len(collection_errors)} data source(s) could not be queried</strong>'
            f'<p style="margin:8px 0 4px;color:var(--text-secondary)">Findings for these categories '
            f'may be incomplete. A score of 0 below does not guarantee compliance.</p>'
            f'<ul style="margin:4px 0 0;padding-left:20px;color:var(--text-secondary)">{err_items}</ul>'
            f'</div>'
        )
    if enrichment_errors:
        enrich_items = "".join(
            f'<li><strong>{esc(e.get("enrichment_step", ""))}</strong> &mdash; {esc(e.get("error", "unknown"))}</li>'
            for e in enrichment_errors
        )
        collection_errors_html += (
            f'<div style="background:rgba(255,140,0,0.08);border:1px solid rgba(255,140,0,0.3);'
            f'border-radius:8px;padding:14px 18px;margin:16px 0;font-size:13px;'
            f'border-left:4px solid #FF8C00">'
            f'<strong style="color:#FF8C00">&#9888; {len(enrichment_errors)} ARM enrichment step(s) '
            f'returned partial data</strong>'
            f'<p style="margin:8px 0 4px;color:var(--text-secondary)">Some deeper security checks '
            f'may have incomplete evidence. Categories below may under-report findings.</p>'
            f'<ul style="margin:4px 0 0;padding-left:20px;color:var(--text-secondary)">{enrich_items}</ul>'
            f'</div>'
        )

    # Compliance Gap Matrix
    compliance_gap_html = _compliance_gap_matrix_svg(findings)

    # Priority Quadrant
    priority_quadrant_html = _priority_quadrant_svg(findings)
    priority_quadrant_detail_html = _priority_quadrant_detail(findings)
    quadrant_guide_html = _quadrant_guide_bullets(findings)

    # Cost methodology computed for standalone report only (not shown in main report)

    # Assemble HTML
    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Data Security Assessment — EnterpriseSecurityIQ</title>
<style>{get_css()}{_ds_css()}</style>
</head>
<body>
<a href="#main" class="skip-nav">Skip to content</a>

<!-- Top Navigation -->
<nav class="top-nav" aria-label="Report sections">
  <span class="brand" aria-hidden="true">&#128274; EnterpriseSecurityIQ Data Security</span>
  <div class="nav-dropdown">
    <button class="nav-toggle">Document Control</button>
    <div class="nav-menu">
      <a href="#doc-control">Report Metadata</a>
      <a href="#doc-control" onclick="setTimeout(function(){{document.querySelector('#doc-control .conf-notice').scrollIntoView({{behavior:'smooth'}})}},50)">Confidentiality Notice</a>
      <a href="#doc-control" onclick="setTimeout(function(){{document.querySelector('#doc-control h3').scrollIntoView({{behavior:'smooth'}})}},50)">Audit Attestation</a>
    </div>
  </div>
  <div class="nav-dropdown">
    <button class="nav-toggle">Executive Summary</button>
    <div class="nav-menu">
      <a href="#summary">Overview &amp; KPIs</a>
      <a href="#summary" onclick="setTimeout(function(){{document.querySelector('.score-display').scrollIntoView({{behavior:'smooth'}})}},50)">Security Score</a>
      <a href="#summary" onclick="setTimeout(function(){{document.querySelector('.sev-bars').scrollIntoView({{behavior:'smooth'}})}},50)">Severity Breakdown</a>
    </div>
  </div>
  <div class="nav-dropdown">
    <button class="nav-toggle">Categories</button>
    <div class="nav-menu">
      <a href="#categories">Category Cards</a>
    </div>
  </div>
  <div class="nav-dropdown">
    <button class="nav-toggle">Analytics</button>
    <div class="nav-menu">
      <a href="#compliance-gap">Compliance Gap Matrix</a>
      <a href="#priority-quadrant">Priority Quadrant</a>
      <a href="#trend">Trend Comparison</a>
    </div>
  </div>
  <div class="nav-dropdown">
    <button class="nav-toggle">All Findings</button>
    <div class="nav-menu">
      <a href="#findings">All Findings</a>
      <div class="nav-sep"></div>
{findings_nav_items}    </div>
  </div>
  <div class="zoom-controls" aria-label="Page zoom">
    <button onclick="zoomOut()" aria-label="Zoom out" data-tip="Decrease page zoom level by 10%.">&minus;</button>
    <span id="zoom-label">100%</span>
    <button onclick="zoomIn()" aria-label="Zoom in" data-tip="Increase page zoom level by 10%.">&plus;</button>
    <button onclick="zoomReset()" aria-label="Reset zoom" data-tip="Reset page zoom to 100%." style="font-size:11px">Reset</button>
  </div>
  <button class="theme-btn" onclick="toggleTheme()" style="margin:0;padding:6px 14px"
          aria-label="Toggle dark and light theme" data-tip="Switch between dark and light colour themes for readability.">
    Switch to Light
  </button>
</nav>

<main id="main" class="full-width-content">

<!-- ── Document Control ── -->
<section id="doc-control" class="section">
  <h1 class="page-title">EnterpriseSecurityIQ &mdash; Data Security Assessment Report</h1>
  <table class="doc-control-table">
    <tr><th data-tip="Unique identifier for this report instance. Use this ID when referencing findings in audit trails or remediation tickets.">Report Identifier</th><td>{esc(report_id)}</td></tr>
    <tr><th data-tip="The name of the compliance assessment that generated this report.">Assessment Name</th><td>EnterpriseSecurityIQ Data Security Assessment</td></tr>
    <tr><th data-tip="Date and time when this report was generated.">Date Generated</th><td>{esc(ts)}</td></tr>
    <tr><th data-tip="Entra ID (Azure AD) tenant identifier for the assessed environment.">Tenant ID</th><td><code>{esc(tenant_id) if tenant_id else 'N/A'}</code></td></tr>
    <tr><th data-tip="Display name of the Entra ID tenant.">Tenant Name</th><td>{esc(tenant_display) if tenant_display else 'Unknown'}</td></tr>
    <tr><th data-tip="The 17 data-security domains evaluated in this assessment.">Assessment Domains</th><td>Storage, Database, Cosmos DB, PostgreSQL/MySQL, Key Vault, Encryption, Access Controls, Private Endpoints, Purview, File Sync, DLP, Classification, Backup/DR, Container, Network, Residency, Threat Detection</td></tr>
    <tr><th data-tip="Data classification level of this report. Handle according to your organization\u2019s information protection policy.">Classification</th><td>CONFIDENTIAL &mdash; Authorized Recipients Only</td></tr>
    <tr><th data-tip="The EnterpriseSecurityIQ tool version and agent that produced this report.">Tool</th><td>EnterpriseSecurityIQ AI Agent v{VERSION}</td></tr>
    <tr><th data-tip="How evidence was collected. Read-only API calls ensure no changes were made to your tenant.">Collection Method</th><td>Azure Resource Manager API + Microsoft Graph API (Read-Only)</td></tr>
  </table>
  <div class="conf-notice">
    <strong>CONFIDENTIALITY NOTICE:</strong> This document contains sensitive security and compliance
    information about the assessed environment. Distribution is restricted to authorized personnel only.
  </div>
  <h3>Audit Attestation</h3>
  <table class="doc-control-table">
    <tr><th data-tip="Scope of the assessment \u2014 how many subscriptions and which security domains were evaluated.">Assessment Scope</th><td>Data security posture analysis across {sub_count} subscription(s) covering 17 security domains</td></tr>
    <tr><th data-tip="Confirms data collection was non-invasive. All API calls are read-only with no modifications to your environment.">Data Integrity</th><td>All evidence collected via read-only API calls; no tenant modifications were made</td></tr>
    <tr><th data-tip="Total number of evidence records gathered from Azure APIs and evaluated against security controls.">Evidence Records</th><td>{evidence_record_count:,} records collected and evaluated</td></tr>
    <tr><th data-tip="SHA-256 hash of the complete report HTML. Use this to verify the report has not been tampered with.">Report Hash (SHA-256)</th><td><code id="report-hash">Computed at render</code></td></tr>
    <tr><th data-tip="When the data was collected. This is a point-in-time snapshot \u2014 your environment may have changed since.">Assessment Period</th><td>{esc(assessed_at[:19]) if assessed_at else esc(ts)} (point-in-time snapshot)</td></tr>
  </table>
</section>

<!-- ── Executive Summary ── -->
<section id="summary" class="section" aria-labelledby="summary-heading">
  <h2 id="summary-heading" class="page-title">Executive Summary</h2>
  <p style="color:var(--text-secondary);font-size:14px;line-height:1.6;max-width:960px;margin:8px 0 20px">
    This report presents a comprehensive data-layer security assessment across your Azure tenant,
    analyzing seventeen key domains: storage exposure, database security, Cosmos DB security,
    PostgreSQL/MySQL security, key &amp; secret hygiene, encryption posture, data access controls,
    private endpoint coverage, Purview / Information Protection, Azure File Sync,
    Microsoft 365 DLP policies, data classification &amp; labeling,
    backup &amp; disaster recovery, container security, network segmentation,
    data residency &amp; sovereignty, and threat detection. Each finding includes severity,
    affected resources with per-resource detail tables, and actionable remediation steps
    with Azure CLI, PowerShell, and portal instructions.
  </p>
  <div class="meta-bar">
    <span>Generated: {esc(ts)}</span>
    <span>Assessed: {esc(assessed_at[:19]) if assessed_at else "N/A"}</span>
    {f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_display)} ({esc(tenant_id[:8])}\u2026)</span>' if tenant_id and tenant_display else f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_id)}</span>' if tenant_id else ''}
    <span>Evidence: {esc(evidence_src.replace("_", " ").title())}</span>
    <span>EnterpriseSecurityIQ v{VERSION}</span>
  </div>

  <!-- Executive summary paragraph -->
  <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin-top:16px;line-height:1.65;font-size:14px">
    {exec_summary_text}
  </div>
  {suppressed_html}

  <!-- KPI cards -->
  <div class="stat-grid" style="grid-template-columns:repeat(auto-fit,minmax(140px,1fr))">
    <div class="stat-card" data-tip="Composite security score (0\u2013100). Higher = more risk. Calculated from severity-weighted findings across all 17 data-security categories.\nThresholds: 0\u201324 = Low (green), 25\u201349 = Medium (yellow), 50\u201374 = High (orange), 75\u2013100 = Critical (red).\nYOUR TENANT: Score {overall_score:.0f}/100 ({level_label}), {finding_count} total findings."><div class="stat-value" style="color:{level_color}">{overall_score:.0f}</div><div class="stat-label">Security Score /100</div></div>
    <div class="stat-card" data-tip="Total data-security findings detected across all categories.\nEach finding maps to a specific Azure resource or configuration gap with severity, affected resources, and remediation.\nYOUR TENANT: {finding_count} findings \u2014 {n_critical} critical, {n_high} high, {n_medium} medium, {n_low} low."><div class="stat-value">{finding_count}</div><div class="stat-label">Total Findings</div></div>
    <div class="stat-card" data-tip="Critical findings require immediate remediation \u2014 active data exposure, missing encryption, or public access.\nThese carry the highest severity weight in the composite score.\nYOUR TENANT: {n_critical} critical finding(s) detected."><div class="stat-value" style="color:#D13438">{n_critical}</div><div class="stat-label">Critical</div></div>
    <div class="stat-card" data-tip="High-severity findings indicating significant gaps in data protection controls.\nAddress within the current sprint after critical items.\nYOUR TENANT: {n_high} high-severity finding(s) detected."><div class="stat-value" style="color:#F7630C">{n_high}</div><div class="stat-label">High</div></div>
    <div class="stat-card" data-tip="Medium findings represent configuration gaps to address in the next review cycle.\nModerate risk \u2014 not immediately exploitable but weakens data security posture.\nYOUR TENANT: {n_medium} medium finding(s) detected."><div class="stat-value" style="color:#FFB900">{n_medium}</div><div class="stat-label">Medium</div></div>
    <div class="stat-card" data-tip="Low-severity findings are informational or best-practice recommendations.\nLimited direct impact but improve overall security hygiene.\nYOUR TENANT: {n_low} low-severity finding(s) detected."><div class="stat-value" style="color:#107C10">{n_low}</div><div class="stat-label">Low</div></div>
    <div class="stat-card" data-tip="Number of Azure subscriptions included in this data-security assessment.\nEach subscription is evaluated across all 17 data-security domains.\nYOUR TENANT: {sub_count} subscription(s) assessed."><div class="stat-value">{sub_count}</div><div class="stat-label">Subscriptions</div></div>
    <div class="stat-card" data-tip="Number of data-security categories evaluated (17 domains max).\nDomains: Storage, Database, Cosmos DB, PostgreSQL/MySQL, Key Vault, Encryption, Access Controls, Private Endpoints, Purview, File Sync, DLP, Classification, Backup/DR, Container, Network, Residency, Threat Detection.\nYOUR TENANT: {len(cat_scores)} categories assessed."><div class="stat-value">{len(cat_scores)}</div><div class="stat-label">Categories</div></div>
  </div>

  <!-- Score ring + severity -->
  <div class="exec-grid">
    <div class="exec-panel" data-tip="Overall security score ring (0\u2013100). 0 = fully secure, 100 = maximum risk.\nThresholds: Critical \u226575, High \u226550, Medium \u226525, Low <25. Target: below 25.\nYOUR TENANT: Score {overall_score:.0f}/100 ({level_label}), {finding_count} findings across {len(cat_scores)} categories.">
      <h3>Overall Security Score</h3>
      <div class="score-display">
        <div style="text-align:center">{score_ring}</div>
        <div class="score-info">
          <div><span class="level-badge" style="background:{level_color};color:#fff" data-tip="Risk level based on composite score.\nCritical (\u226575): Immediate action. High (50\u201374): Address this sprint. Medium (25\u201349): Plan remediation. Low (<25): Healthy posture.\nYOUR TENANT: {level_label} ({overall_score:.0f}/100).">{level_label}</span></div>
          <div style="font-size:13px;color:var(--text-secondary);margin-top:8px">
            Higher score = more risk.<br>
            Target: below 25 (Low).
          </div>
        </div>
      </div>
    </div>
    <div class="exec-panel" data-tip="Donut chart showing the proportion of findings by severity level across all data-security categories.">
      <h3>Severity Distribution</h3>
      <div style="text-align:center">{sev_donut}</div>
      <div class="legend" style="justify-content:center;margin-top:16px">
        <span class="legend-item" data-tip="Critical: Immediate action required — active data exposure, missing encryption, or public access."><span class="legend-dot" style="background:#D13438"></span> Critical ({n_critical})</span>
        <span class="legend-item" data-tip="High: Address this sprint — significant gaps in data protection controls."><span class="legend-dot" style="background:#F7630C"></span> High ({n_high})</span>
        <span class="legend-item" data-tip="Medium: Plan remediation in the next review cycle to harden data security posture."><span class="legend-dot" style="background:#FFB900"></span> Medium ({n_medium})</span>
        <span class="legend-item" data-tip="Low: Informational or best-practice enhancements with limited immediate risk."><span class="legend-dot" style="background:#107C10"></span> Low ({n_low})</span>
        <span class="legend-item" data-tip="Info: Advisory items for awareness — no direct remediation required."><span class="legend-dot" style="background:#A8A6A3"></span> Info ({n_info})</span>
      </div>
    </div>
    <div class="exec-panel" style="grid-column:span 2" data-tip="Horizontal bars showing the count of findings at each severity level for quick visual comparison.">
      <h3>Severity Breakdown</h3>
      <div class="sev-bars">
        {sev_bars_html}
      </div>
    </div>
  </div>
  {impact_panel_html}
</section>

{collection_errors_html}

<!-- ── Category Breakdown ── -->
<section id="categories" class="section" aria-labelledby="categories-heading">
  <h2 id="categories-heading" data-tip="Per-category security scores across all seventeen data-security domains. Higher score = more risk.">&#128202; Category Breakdown</h2>
  <div class="how-to-read">
    <h4>How to read this section</h4>
    <p>Each category scores 0&ndash;100 based on severity-weighted findings. Higher = more risk.
    Categories with score 0 and no findings have a clean bill of health.</p>
  </div>
  <div class="category-grid">
    {cat_cards_html}
  </div>
</section>

<!-- ── Compliance Gap Matrix ── -->
<section id="compliance-gap" class="section" aria-labelledby="gap-heading">
  <h2 id="gap-heading" data-tip="Shows which compliance frameworks have findings at each severity level. Wider coverage = more regulatory exposure.">&#128202; Compliance Gap Matrix</h2>
  <div class="how-to-read">
    <h4>How to read this chart</h4>
    <p>Each row is a compliance framework. Columns are color-coded by risk level:
    <span style="display:inline-block;width:14px;height:14px;border-radius:3px;background:#D13438;opacity:0.5;vertical-align:middle;margin-right:3px"></span><strong>Critical</strong>,
    <span style="display:inline-block;width:14px;height:14px;border-radius:3px;background:#F7630C;opacity:0.5;vertical-align:middle;margin-right:3px"></span><strong>High</strong>,
    <span style="display:inline-block;width:14px;height:14px;border-radius:3px;background:#FFB900;opacity:0.5;vertical-align:middle;margin-right:3px"></span><strong>Medium</strong>,
    <span style="display:inline-block;width:14px;height:14px;border-radius:3px;background:#107C10;opacity:0.5;vertical-align:middle;margin-right:3px"></span><strong>Low</strong>.
    Cell numbers show how many findings at that severity map to framework controls.
    More opaque cells indicate higher concentrations. The Controls column shows total mapped control IDs.</p>
  </div>
  <div class="exec-panel" style="overflow-x:auto">
    {compliance_gap_html}
  </div>
</section>

<!-- ── Priority Quadrant ── -->
<section id="priority-quadrant" class="section" aria-labelledby="quadrant-heading">
  <h2 id="quadrant-heading" data-tip="Remediation priority quadrant plotting findings by estimated impact vs effort. Quick wins (high impact, low effort) are top-left.">&#127919; Remediation Priority Quadrant</h2>
  <div class="how-to-read">
    <h4>How to read this chart</h4>
    <p>This quadrant maps every finding along two axes: <strong>Security Impact</strong> (vertical &mdash;
    how much risk the finding poses if left unresolved) and <strong>Remediation Effort</strong> (horizontal &mdash;
    the estimated work required to fix it). Findings are aggregated by severity and plotted as circles whose
    size reflects the number of findings at that level.</p>
    <p><strong>Quadrant guide:</strong></p>
    <ul style="margin:6px 0 0 18px;font-size:13px;color:var(--text-secondary);line-height:1.8">
      {quadrant_guide_html}
    </ul>
  </div>
  <div class="exec-panel" style="text-align:center">
    {priority_quadrant_html}
  </div>
  <div class="exec-panel">
    {priority_quadrant_detail_html}
  </div>
</section>


{trend_section_html}

<!-- ── All Findings ── -->
<section id="findings" class="section" aria-labelledby="findings-heading">
  <h2 id="findings-heading" data-tip="Complete table of all data-security findings — use the filters below to narrow by severity, keyword, or category.">&#128270; All Findings ({finding_count})</h2>
  <div class="filter-bar" role="search" aria-label="Filter findings">
    <label for="finding-filter">Search:</label>
    <input id="finding-filter" type="search" placeholder="Search findings…"
           oninput="filterFindings()" aria-label="Filter findings by keyword"
           aria-describedby="findings-live" autocomplete="off"
           data-tip="Type a keyword to instantly filter findings by title, resource, or description.">
    <label for="filter-severity">Severity:</label>
    <select id="filter-severity" onchange="filterFindings()" aria-label="Filter by severity" data-tip="Show only findings at the selected severity level.">
      <option value="">All</option>
      <option value="critical">Critical</option>
      <option value="high">High</option>
      <option value="medium">Medium</option>
      <option value="low">Low</option>
      <option value="informational">Info</option>
    </select>
    <label for="filter-category">Category:</label>
    <select id="filter-category" onchange="filterFindings()" aria-label="Filter by category" data-tip="Show only findings belonging to the selected data-security category.">
      <option value="">All</option>
      {cat_options}
    </select>
  </div>
  <div id="findings-live" class="sr-only" aria-live="polite" aria-atomic="true"></div>
  {all_findings_html if all_findings_html else '<p class="empty">No findings — your data layer is secure. Well done!</p>'}
  <div style="text-align:center;margin:18px 0 0 0"><button id="show-more-btn" onclick="showMoreFindings()" style="display:none;padding:10px 28px;font-size:15px;border-radius:8px;border:1px solid #bbb;background:#f5f5f5;cursor:pointer">Show more findings</button></div>
</section>

</main>

<button class="back-to-top" aria-label="Back to top">&#8593;</button>

<div id="ciq-tooltip" role="tooltip" aria-hidden="true"></div>
<script>{get_js()}</script>
<script>{_ds_js()}</script>

<script>
document.addEventListener('click',function(){{
  document.querySelectorAll('.fw-popup').forEach(function(p){{p.style.display='none'}});
}});
</script>
</body>
</html>"""

    report_hash = hashlib.sha256(html.encode("utf-8")).hexdigest()
    html = html.replace("Computed at render", report_hash)
    out_path.write_text(html, encoding="utf-8")
    log.info("[DataSecurityReport] Written to %s (%d KB)", out_path, len(html) // 1024)
    return out_path


# ── Excel styling ────────────────────────────────────────────────────────
_XL_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_XL_HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
_XL_BODY_FONT = Font(name="Segoe UI", size=10)
_XL_BODY_FONT_BOLD = Font(name="Segoe UI", size=10, bold=True)
_XL_SECTION_FONT = Font(name="Segoe UI", bold=True, size=12, color="0078D4")
_XL_TOTAL_FONT = Font(name="Segoe UI", bold=True, size=11, color="0078D4")
_XL_THIN_BORDER = Border(
    left=Side(style="thin", color="EDEBE9"),
    right=Side(style="thin", color="EDEBE9"),
    top=Side(style="thin", color="EDEBE9"),
    bottom=Side(style="thin", color="EDEBE9"),
)
_XL_BOTTOM_THICK = Border(
    left=Side(style="thin", color="EDEBE9"),
    right=Side(style="thin", color="EDEBE9"),
    top=Side(style="thin", color="EDEBE9"),
    bottom=Side(style="medium", color="0078D4"),
)
_XL_SEV_FILLS = {
    "critical": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "high": PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),
    "medium": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "low": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
    "informational": PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid"),
}
_XL_ZEBRA_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
_XL_SEV_FONTS = {
    "critical": Font(name="Segoe UI", size=10, bold=True, color="D13438"),
    "high": Font(name="Segoe UI", size=10, bold=True, color="F7630C"),
    "medium": Font(name="Segoe UI", size=10, bold=True, color="986F0B"),
    "low": Font(name="Segoe UI", size=10, bold=True, color="107C10"),
    "informational": Font(name="Segoe UI", size=10, color="605E5C"),
}


def _xl_apply_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _XL_HEADER_FONT
        cell.fill = _XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _XL_THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _xl_write_row(ws, row_num: int, values: list,
                  col_fills: dict[int, PatternFill] | None = None,
                  col_fonts: dict[int, Font] | None = None,
                  row_fill: PatternFill | None = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = (col_fonts or {}).get(col, _XL_BODY_FONT)
        cell.border = _XL_THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_fills and col in col_fills:
            cell.fill = col_fills[col]
        elif row_fill:
            cell.fill = row_fill


def _xl_auto_width(ws, headers: list[str], max_width: int = 55):
    min_widths = {"Severity": 12, "Priority": 18, "Category": 22, "Finding": 40,
                  "Title": 40, "Description": 50, "Resources": 40,
                  "Azure CLI": 45, "PowerShell": 45, "Portal Steps": 40,
                  "Resource Name": 28, "Resource ID": 40, "Remediation Description": 45,
                  "Control Title": 35, "Rationale": 40, "Recommendation": 40,
                  "Recommended Action": 50, "Example Findings": 55,
                  "Finding Remediation (CLI)": 45, "Finding Remediation (PS)": 45,
                  "Finding Remediation (Portal)": 40}
    for col, h in enumerate(headers, 1):
        min_w = min_widths.get(h, len(h) + 4)
        best = max(min_w, len(h) + 4)
        for row in range(2, min(ws.max_row + 1, 52)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                line_len = max(len(ln) for ln in str(cell.value).split("\n")) if "\n" in str(cell.value) else len(str(cell.value))
                best = max(best, min(line_len + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = min(best, max_width)


# ── Executive Brief ──────────────────────────────────────────────────────

def generate_executive_brief(results: dict, output_dir: str) -> pathlib.Path:
    """Generate a rich, data-dense executive brief HTML for data security posture."""
    out_dir = pathlib.Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    scores = results.get("DataSecurityScores", {})
    findings = results.get("Findings", [])
    impact = results.get("RemediationImpact", {})
    overall = scores.get("OverallScore", 0)
    level = scores.get("OverallLevel", "low")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})
    tenant = results.get("TenantDisplayName", results.get("TenantId", ""))
    tenant_id = results.get("TenantId", "")
    assessed_at = results.get("AssessedAt", "")
    sub_count = results.get("SubscriptionCount", 0)
    finding_count = results.get("FindingCount", len(findings))
    suppressed = results.get("SuppressedCount", 0)
    evidence_count = results.get("EvidenceRecordCount", 0)

    n_critical = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_medium = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)

    level_meta = _SCORE_LEVEL_META.get(level, _SCORE_LEVEL_META["medium"])
    level_color = level_meta["color"]
    level_label = level_meta["label"]
    score_ring = _ring_score_svg(overall, size=120)
    sev_donut = _donut_svg([
        ("Critical", n_critical, "#D13438"),
        ("High", n_high, "#F7630C"),
        ("Medium", n_medium, "#FFB900"),
        ("Low", n_low, "#107C10"),
        ("Info", n_info, "#A8A6A3"),
    ], size=110)

    # Compliance coverage
    frameworks_hit: dict[str, set[str]] = {}
    for f in findings:
        mapping = f.get("ComplianceMapping", {})
        for fw, controls in mapping.items():
            frameworks_hit.setdefault(fw, set()).update(controls)
    compliance_rows = ""
    for fw in sorted(frameworks_hit.keys()):
        controls = sorted(frameworks_hit[fw])
        sev_counts_fw: dict[str, int] = {}
        for f in findings:
            if fw in f.get("ComplianceMapping", {}):
                s = f.get("Severity", "medium").lower()
                sev_counts_fw[s] = sev_counts_fw.get(s, 0) + 1
        sev_pills = ""
        for s, color in [("critical", "#D13438"), ("high", "#F7630C"), ("medium", "#FFB900"), ("low", "#107C10")]:
            cnt = sev_counts_fw.get(s, 0)
            if cnt:
                sev_pills += f'<span class="pill" style="background:{color}">{cnt} {s}</span> '
        compliance_rows += (
            f'<tr><td class="fw-name">{esc(fw)}</td>'
            f'<td class="fw-count">{len(controls)}</td>'
            f'<td>{sev_pills}</td>'
            f'<td class="fw-ids">{esc(", ".join(controls[:8]))}'
            f'{"&hellip;" if len(controls) > 8 else ""}</td></tr>'
        )

    # Category breakdown
    cat_items = []
    for cat_key, cs in sorted(cat_scores.items(), key=lambda x: -x[1].get("Score", 0) if isinstance(x[1], dict) else 0):
        if not isinstance(cs, dict):
            continue
        meta = _CATEGORY_META.get(cat_key, {"name": cat_key.replace("_", " ").title(), "icon": "&#128196;"})
        c_score = cs.get("Score", 0)
        c_level = cs.get("Level", "low")
        c_count = cs.get("FindingCount", 0)
        if c_count == 0 and c_score == 0:
            continue
        c_color = _SCORE_LEVEL_META.get(c_level, _SCORE_LEVEL_META["low"])["color"]
        cat_items.append((meta, c_score, c_level, c_count, c_color))

    cat_cards_html = ""
    for meta, c_score, c_level, c_count, c_color in cat_items[:12]:
        cat_cards_html += (
            f'<div class="cat-card">'
            f'<div class="cat-icon">{meta["icon"]}</div>'
            f'<div class="cat-info">'
            f'<div class="cat-name">{esc(meta["name"])}</div>'
            f'<div class="cat-score" style="color:{c_color}">{c_score:.0f}/100</div>'
            f'</div>'
            f'<div class="cat-badge" style="background:{c_color}">{c_count}</div>'
            f'</div>'
        )

    # Impact projection
    impact_html = ""
    if impact:
        cur = impact.get("CurrentScore", 0)
        if_crit = impact.get("IfCriticalFixed", 0)
        if_high = impact.get("IfHighFixed", 0)
        impact_html = (
            '<div class="panel">'
            '<h3>&#128200; Remediation Impact Projection</h3>'
            '<div class="impact-grid">'
            f'<div class="impact-item"><div class="impact-val">{cur:.0f}</div><div class="impact-lbl">Current</div></div>'
            '<div class="impact-arrow">&#8594;</div>'
            f'<div class="impact-item"><div class="impact-val" style="color:#D13438">{if_crit:.0f}</div><div class="impact-lbl">If Critical Fixed</div></div>'
            '<div class="impact-arrow">&#8594;</div>'
            f'<div class="impact-item"><div class="impact-val" style="color:#F7630C">{if_high:.0f}</div><div class="impact-lbl">If Crit+High Fixed</div></div>'
            '<div class="impact-arrow">&#8594;</div>'
            '<div class="impact-item"><div class="impact-val" style="color:#107C10">0</div><div class="impact-lbl">If All Fixed</div></div>'
            '</div></div>'
        )

    # Cost of non-compliance
    cost_est = _estimate_noncompliance_cost(findings)
    cost_html = ""
    if cost_est["total"] > 0:
        cost_html = (
            '<div class="panel">'
            '<h3>&#128176; Estimated Annual Exposure</h3>'
            f'<div class="cost-total">${cost_est["total"]:,.0f}</div>'
            '<div class="cost-note">Based on IBM/Ponemon breach-cost benchmarks (directional estimate)</div>'
            '</div>'
        )

    # Top findings
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    sorted_top = sorted(findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))[:10]
    top_rows = ""
    for i, t in enumerate(sorted_top, 1):
        sev = t.get("Severity", "").lower()
        sev_color = _SEV_COLORS.get(sev, "#A8A6A3")
        cat_key = t.get("Category", "")
        cat_name = _CATEGORY_META.get(cat_key, {}).get("name", cat_key.replace("_", " ").title())
        cls = " class=\"row-alt\"" if i % 2 == 0 else ""
        top_rows += (
            f'<tr{cls}>'
            f'<td class="rank">{i}</td>'
            f'<td><span class="sev-badge" style="background:{sev_color}">{esc(sev.upper())}</span></td>'
            f'<td class="finding-title">{esc(t.get("Title", ""))}</td>'
            f'<td class="cat">{esc(cat_name)}</td>'
            f'<td class="count">{t.get("AffectedCount", 0)}</td></tr>'
        )

    # Priority quadrant summary
    quad_map = {"critical": "Quick Wins", "high": "Quick Wins", "medium": "Major Projects", "low": "Low Priority", "informational": "Consider"}
    quad_counts: dict[str, int] = {}
    for f in findings:
        q = quad_map.get(f.get("Severity", "medium").lower(), "Consider")
        quad_counts[q] = quad_counts.get(q, 0) + 1
    quad_html = ""
    quad_colors = {"Quick Wins": "#0078D4", "Major Projects": "#0078D4", "Low Priority": "#0078D4", "Consider": "#0078D4"}
    for q_name in ["Quick Wins", "Major Projects", "Low Priority", "Consider"]:
        cnt = quad_counts.get(q_name, 0)
        if cnt:
            qc = quad_colors.get(q_name, "#888")
            quad_html += f'<div class="quad-item" style="border-left:4px solid {qc}"><span class="quad-count" style="color:{qc}">{cnt}</span><span class="quad-name">{q_name}</span></div>'

    # Trend
    trend = results.get("Trend")
    trend_html = ""
    if trend and isinstance(trend, dict):
        delta = trend.get("ScoreDelta", 0)
        arrow = "&#8593;" if delta > 0 else "&#8595;" if delta < 0 else "&#8594;"
        delta_color = "#D13438" if delta > 0 else "#107C10" if delta < 0 else "#FFB900"
        trend_html = (
            '<div class="panel trend-panel">'
            '<h3>&#128200; Trend vs Previous</h3>'
            '<div class="trend-grid">'
            f'<div class="trend-item"><span class="trend-val">{trend.get("PreviousScore", "?")}</span><span class="trend-lbl">Previous</span></div>'
            f'<div class="trend-item"><span class="trend-val" style="color:{delta_color}">{arrow} {abs(delta):.0f}</span><span class="trend-lbl">Delta</span></div>'
            f'<div class="trend-item"><span class="trend-val" style="color:#107C10">{trend.get("ResolvedCount", 0)}</span><span class="trend-lbl">Resolved</span></div>'
            f'<div class="trend-item"><span class="trend-val" style="color:#D13438">{trend.get("NewCount", 0)}</span><span class="trend-lbl">New</span></div>'
            '</div></div>'
        )

    # Severity bars data
    max_sev = max(n_critical, n_high, n_medium, n_low, n_info, 1)
    sev_bars_html = ""
    for name, cnt, color in [("Critical", n_critical, "#D13438"), ("High", n_high, "#F7630C"),
                              ("Medium", n_medium, "#FFB900"), ("Low", n_low, "#107C10"), ("Info", n_info, "#A8A6A3")]:
        pct = cnt / max_sev * 100
        sev_bars_html += (
            f'<div class="sev-row">'
            f'<span class="sev-lbl" style="color:{color}">{name}</span>'
            f'<div class="sev-track"><div class="sev-fill" style="width:{pct:.0f}%;background:{color}"></div></div>'
            f'<span class="sev-cnt">{cnt}</span></div>'
        )

    # Narrative
    posture = (
        "critically exposed" if level == "critical" else
        "at elevated risk" if level == "high" else
        "moderately secure with notable gaps" if level == "medium" else
        "well-secured"
    )
    weak = sorted(
        [(k, v.get("Score", 0)) for k, v in cat_scores.items() if isinstance(v, dict) and v.get("FindingCount", 0) > 0],
        key=lambda x: -x[1],
    )[:3]
    weak_names = [_CATEGORY_META.get(k, {"name": k})["name"] for k, _ in weak]
    narrative = (
        f"Across <strong>{sub_count}</strong> subscription(s), the data security posture is "
        f"<strong>{posture}</strong> with an overall risk score of <strong>{overall:.0f}/100</strong>. "
    )
    if n_critical or n_high:
        narrative += (
            f"There {'are' if n_critical + n_high != 1 else 'is'} "
            f"<strong>{n_critical} critical</strong> and <strong>{n_high} high</strong> "
            f"severity finding(s) requiring immediate attention. "
        )
    if weak_names:
        narrative += f"Priority areas: <strong>{', '.join(weak_names)}</strong>."

    suppressed_kpi = f'<div class="kpi"><div class="kpi-val">{suppressed}</div><div class="kpi-lbl">Suppressed</div></div>' if suppressed else ""

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Data Security Executive Brief — EnterpriseSecurityIQ</title>
<style>
:root{{--bg:#1b1b1f;--bg-card:#2d2d30;--bg-elevated:#252528;--text:#e0e0e0;--text-secondary:#a0a0a0;
--text-muted:#888;--border:#3e3e42;--primary:#4fc3f7;--font:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
--font-mono:'Cascadia Code','Fira Code',Consolas,monospace}}
[data-theme=light]{{--bg:#f5f5f5;--bg-card:#fff;--bg-elevated:#fff;--text:#1a1a1a;--text-secondary:#555;
--text-muted:#888;--border:#e0e0e0;--primary:#0078D4}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:var(--font);background:var(--bg);color:var(--text);font-size:14px;line-height:1.6;max-width:1100px;margin:0 auto;padding:24px 32px}}
a{{color:var(--primary)}}
h1{{font-size:24px;font-weight:700;border-bottom:3px solid var(--primary);padding-bottom:10px;margin-bottom:16px}}
h2{{font-size:17px;color:var(--primary);margin:24px 0 10px;font-weight:600}}
h3{{font-size:14px;margin:0 0 10px;color:var(--text)}}
.meta-bar{{display:flex;flex-wrap:wrap;gap:16px;font-size:12px;color:var(--text-secondary);padding:8px 0;border-bottom:1px solid var(--border);margin-bottom:20px}}
.meta-bar span{{white-space:nowrap}}
.theme-btn{{position:fixed;top:12px;right:16px;z-index:100;padding:6px 14px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:12px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:12px;margin:16px 0}}
.kpi{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:14px 12px;text-align:center}}
.kpi-val{{font-size:26px;font-weight:700;font-family:var(--font-mono)}}
.kpi-lbl{{font-size:11px;color:var(--text-secondary);margin-top:4px}}
.panel{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin:16px 0}}
.narrative{{line-height:1.7;font-size:14px}}
.score-row{{display:flex;gap:24px;align-items:center;flex-wrap:wrap;justify-content:center}}
.score-block{{text-align:center}}
.score-label{{font-size:12px;color:var(--text-secondary);margin-top:6px}}
.level-badge{{display:inline-block;padding:4px 14px;border-radius:20px;font-size:13px;font-weight:700;color:#fff}}
.sev-row{{display:flex;align-items:center;gap:10px;margin:4px 0}}
.sev-lbl{{width:65px;font-size:12px;font-weight:600;text-align:right}}
.sev-track{{flex:1;height:16px;background:var(--bg-elevated);border-radius:8px;overflow:hidden}}
.sev-fill{{height:100%;border-radius:8px;transition:width .3s}}
.sev-cnt{{width:30px;font-size:13px;font-weight:700;font-family:var(--font-mono)}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin:8px 0}}
th{{background:var(--bg-elevated);padding:8px 10px;text-align:left;font-size:11px;text-transform:uppercase;color:var(--text-secondary);border-bottom:2px solid var(--border);font-weight:600}}
td{{padding:7px 10px;border-bottom:1px solid var(--border)}}
.row-alt td{{background:rgba(128,128,128,0.04)}}
.sev-badge{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:700;color:#fff;letter-spacing:.5px}}
.rank{{font-weight:700;color:var(--text-muted);text-align:center;width:32px}}
.count{{text-align:center;font-family:var(--font-mono);font-weight:700}}
.cat{{color:var(--text-secondary);font-size:12px}}
.finding-title{{font-weight:500}}
.fw-name{{font-weight:600}}
.fw-count{{text-align:center;font-family:var(--font-mono);font-weight:700}}
.fw-ids{{font-size:11px;color:var(--text-muted);font-family:var(--font-mono)}}
.pill{{display:inline-block;padding:1px 7px;border-radius:10px;font-size:10px;font-weight:700;color:#fff;margin:1px 2px}}
.cat-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px;margin:10px 0}}
.cat-card{{display:flex;align-items:center;gap:10px;background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:10px 12px}}
.cat-icon{{font-size:20px}}
.cat-info{{flex:1}}
.cat-name{{font-size:12px;font-weight:600}}
.cat-score{{font-size:16px;font-weight:700;font-family:var(--font-mono)}}
.cat-badge{{min-width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:#fff}}
.impact-grid{{display:flex;align-items:center;gap:12px;flex-wrap:wrap;justify-content:center}}
.impact-item{{text-align:center;min-width:90px}}
.impact-val{{font-size:28px;font-weight:700;font-family:var(--font-mono)}}
.impact-lbl{{font-size:11px;color:var(--text-secondary)}}
.impact-arrow{{font-size:20px;color:var(--text-muted)}}
.cost-total{{font-size:36px;font-weight:700;color:#D13438;font-family:var(--font-mono);text-align:center;margin:8px 0}}
.cost-note{{font-size:11px;color:var(--text-muted);text-align:center}}
.quad-grid{{display:flex;gap:12px;flex-wrap:wrap;margin:10px 0}}
.quad-item{{display:flex;align-items:center;gap:8px;padding:8px 14px;background:var(--bg-card);border:1px solid var(--border);border-radius:8px}}
.quad-count{{font-size:20px;font-weight:700;font-family:var(--font-mono)}}
.quad-name{{font-size:12px;color:var(--text-secondary)}}
.trend-grid{{display:flex;gap:20px;flex-wrap:wrap;justify-content:center}}
.trend-item{{text-align:center;min-width:80px}}
.trend-val{{font-size:22px;font-weight:700;font-family:var(--font-mono);display:block}}
.trend-lbl{{font-size:11px;color:var(--text-secondary)}}
.footer{{margin-top:32px;padding:14px 18px;background:var(--bg-card);border:1px solid var(--border);border-radius:8px;font-size:12px;color:var(--text-muted);text-align:center}}
@media print{{body{{background:#fff;color:#000;max-width:none;padding:12px}}
.theme-btn,.no-print{{display:none}}
.panel,.kpi,.cat-card,.quad-item{{border-color:#ccc;background:#f9f9f9}}
h1{{border-color:#0078D4}}}}
</style>
</head>
<body>
<button class="theme-btn" onclick="var t=document.documentElement;t.dataset.theme=t.dataset.theme==='dark'?'light':'dark';this.textContent=t.dataset.theme==='dark'?'Switch to Light':'Switch to Dark'">Switch to Light</button>

<h1>&#128274; EnterpriseSecurityIQ — Data Security Executive Brief</h1>
<div class="meta-bar">
  <span>Tenant: <strong>{esc(tenant)}</strong></span>
  {f'<span>ID: <code>{esc(tenant_id[:8])}&hellip;</code></span>' if tenant_id else ''}
  <span>Assessed: {esc(assessed_at[:19]) if assessed_at else "N/A"} UTC</span>
  <span>{sub_count} subscription(s)</span>
  <span>{evidence_count:,} evidence records</span>
  <span>EnterpriseSecurityIQ v{VERSION}</span>
</div>

<div class="panel narrative">{narrative}</div>

<div class="panel">
  <div class="score-row">
    <div class="score-block">
      {score_ring}
      <div class="score-label">Risk Score</div>
      <div><span class="level-badge" style="background:{level_color}">{level_label}</span></div>
    </div>
    <div class="score-block">
      {sev_donut}
      <div class="score-label">Severity Distribution</div>
    </div>
    <div style="flex:1;min-width:200px">
      {sev_bars_html}
    </div>
  </div>
</div>

<div class="kpi-grid">
  <div class="kpi"><div class="kpi-val" style="color:{level_color}">{overall:.0f}</div><div class="kpi-lbl">Score /100</div></div>
  <div class="kpi"><div class="kpi-val">{finding_count}</div><div class="kpi-lbl">Findings</div></div>
  <div class="kpi"><div class="kpi-val" style="color:#D13438">{n_critical}</div><div class="kpi-lbl">Critical</div></div>
  <div class="kpi"><div class="kpi-val" style="color:#F7630C">{n_high}</div><div class="kpi-lbl">High</div></div>
  <div class="kpi"><div class="kpi-val" style="color:#FFB900">{n_medium}</div><div class="kpi-lbl">Medium</div></div>
  <div class="kpi"><div class="kpi-val">{sub_count}</div><div class="kpi-lbl">Subscriptions</div></div>
  <div class="kpi"><div class="kpi-val">{len(cat_scores)}</div><div class="kpi-lbl">Categories</div></div>
  {suppressed_kpi}
</div>

{f'<h2>&#127919; Remediation Priority</h2><div class="quad-grid">{quad_html}</div>' if quad_html else ''}

{impact_html}
{cost_html}

<h2>&#128202; Category Breakdown (Top {len(cat_items[:12])})</h2>
<div class="cat-grid">{cat_cards_html}</div>

<h2>&#128270; Top {min(10, len(sorted_top))} Priority Findings</h2>
<div class="panel" style="overflow-x:auto">
  <table>
    <thead><tr><th>#</th><th>Sev</th><th>Finding</th><th>Category</th><th>Affected</th></tr></thead>
    <tbody>{top_rows}</tbody>
  </table>
</div>

{"<h2>&#128203; Compliance Framework Coverage</h2><div class='panel' style='overflow-x:auto'><table><thead><tr><th>Framework</th><th>Controls</th><th>Severities</th><th>Mapped IDs</th></tr></thead><tbody>" + compliance_rows + "</tbody></table></div>" if compliance_rows else ""}

{trend_html}

<div class="footer">
  <strong>CONFIDENTIAL</strong> — Generated by EnterpriseSecurityIQ v{VERSION}<br>
  For detailed per-resource findings, remediation commands, and compliance mappings, refer to the full Data Security Assessment Report.
</div>

<script>
if(window.matchMedia&&window.matchMedia('(prefers-color-scheme:light)').matches){{
  document.documentElement.dataset.theme='light';
  document.querySelector('.theme-btn').textContent='Switch to Dark';
}}
</script>
</body></html>"""

    path = out_dir / "executive-brief.html"
    path.write_text(html, encoding="utf-8")
    log.info("Executive brief -> %s", path)
    return path


# ── Excel export ─────────────────────────────────────────────────────────

def generate_data_security_excel(results: dict, output_dir: str) -> pathlib.Path:
    """Generate a comprehensive multi-tab Data Security Excel workbook.

    Sheets: Executive Summary, All Findings, Affected Resources, Category Scores,
    Remediation, Compliance Mappings, Priority Quadrant, Cost of Non-Compliance,
    Trend Comparison, Collection Errors.
    """

    out_dir = pathlib.Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "data-security.xlsx"

    scores = results.get("DataSecurityScores", {})
    findings = results.get("Findings", [])
    overall = scores.get("OverallScore", 0)
    overall_level = scores.get("OverallLevel", "low")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})
    top_findings = scores.get("TopFindings", [])
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}

    level = (
        "Critical" if overall >= 75 else
        "High" if overall >= 50 else
        "Medium" if overall >= 25 else
        "Low"
    )

    # Severity counts
    n_critical = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_medium = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)

    wb = Workbook()

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    label_font = Font(name="Segoe UI", bold=True, size=11)
    value_font = Font(name="Segoe UI", size=11)
    section_font = Font(name="Segoe UI", bold=True, size=12, color="0078D4")

    ts_now = results.get("AssessedAt", datetime.now().strftime("%Y-%m-%d %H:%M"))
    if "T" in ts_now:
        # Trim ISO format to "YYYY-MM-DD HH:MM" for display
        ts_now = ts_now[:16].replace("T", " ")
    tenant_id = results.get("TenantId", "N/A")
    tenant_display = results.get("TenantDisplayName", "Unknown")
    sub_count = results.get("SubscriptionCount", 0)
    evidence_src = results.get("EvidenceSource", "N/A")
    assessed_at = results.get("AssessedAt", "")
    evidence_record_count = results.get("EvidenceRecordCount", 0)
    suppressed_count = results.get("SuppressedCount", 0)
    finding_count = results.get("FindingCount", len(findings))

    # Executive summary narrative
    exec_text = ""
    if findings:
        weak = sorted(
            [(k, v.get("Score", 0)) for k, v in cat_scores.items() if v.get("FindingCount", 0) > 0],
            key=lambda x: -x[1],
        )[:3]
        weak_names = [_CATEGORY_META.get(k, {"name": k})["name"] for k, _ in weak]
        posture = (
            "critically exposed" if overall_level == "critical" else
            "at elevated risk" if overall_level == "high" else
            "moderately secure with notable gaps" if overall_level == "medium" else
            "well-secured"
        )
        exec_text = (
            f"Across {sub_count} subscription(s), the data security posture is {posture} "
            f"with an overall risk score of {overall:.0f}/100. "
            f"{n_critical} critical and {n_high} high severity findings require immediate attention. "
        )
        if weak_names:
            exec_text += f"Priority areas: {', '.join(weak_names)}."

    summary_rows: list[tuple[str, str]] = [
        ("DOCUMENT CONTROL", ""),
        ("Report", "Data Security Assessment"),
        ("Date Generated", ts_now),
        ("EnterpriseSecurityIQ Version", VERSION),
        ("Tenant ID", tenant_id),
        ("Tenant Name", tenant_display),
        ("Assessment Period", assessed_at[:19] if assessed_at else ts_now),
        ("Evidence Source", evidence_src),
        ("Subscription Count", str(sub_count)),
        ("Evidence Records", f"{evidence_record_count:,}"),
        ("Collection Method", "Azure Resource Manager API + Microsoft Graph API (Read-Only)"),
        ("Classification", "CONFIDENTIAL — Authorized Recipients Only"),
        ("", ""),
        ("EXECUTIVE SUMMARY", ""),
        ("Overall Security Score", f"{overall:.1f} / 100"),
        ("Risk Level", level),
        ("Total Findings", str(finding_count)),
        ("Suppressed Findings", str(suppressed_count)),
        ("", ""),
        ("SEVERITY DISTRIBUTION", ""),
        ("  Critical", str(n_critical)),
        ("  High", str(n_high)),
        ("  Medium", str(n_medium)),
        ("  Low", str(n_low)),
        ("  Informational", str(n_info)),
    ]

    if exec_text:
        summary_rows.append(("", ""))
        summary_rows.append(("NARRATIVE", ""))
        summary_rows.append(("Summary", exec_text))

    # Remediation Impact
    impact = results.get("RemediationImpact", {})
    if impact:
        cur = impact.get("CurrentScore", 0)
        if_crit = impact.get("IfCriticalFixed", 0)
        if_high = impact.get("IfHighFixed", 0)
        summary_rows.append(("", ""))
        summary_rows.append(("REMEDIATION IMPACT", ""))
        summary_rows.append(("Current Score", f"{cur:.0f}"))
        summary_rows.append(("If Critical Fixed", f"{if_crit:.0f}"))
        summary_rows.append(("If Critical+High Fixed", f"{if_high:.0f}"))
        summary_rows.append(("If All Fixed", "0"))

    # Category scores
    all_cats = ["storage", "database", "cosmosdb", "pgmysql", "keyvault", "encryption",
                "data_access", "private_endpoints", "purview", "file_sync", "m365_dlp",
                "data_classification", "backup_dr", "container_security",
                "network_segmentation", "data_residency", "threat_detection",
                "redis", "messaging", "ai_services", "data_pipeline", "identity",
                "sharepoint_governance", "data_lifecycle", "dlp_alert",
                "app_config", "databricks", "apim", "frontdoor", "secret_sprawl",
                "firewall", "bastion", "policy_compliance", "defender_score",
                "stale_permissions", "data_exfiltration", "conditional_access", "config_drift", "supply_chain"]

    summary_rows.append(("", ""))
    summary_rows.append(("CATEGORY SCORES", ""))
    for cat_key in all_cats:
        meta = _CATEGORY_META.get(cat_key)
        if not meta:
            continue
        cs = cat_scores.get(cat_key, {})
        if isinstance(cs, dict):
            summary_rows.append((
                f"  {meta['name']}",
                f"{cs.get('Score', 0):.1f} / 100 ({cs.get('Level', 'N/A')})"
            ))

    # Top 10 findings
    sorted_top = sorted(findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))[:10]
    if sorted_top:
        summary_rows.append(("", ""))
        summary_rows.append(("TOP 10 PRIORITY FINDINGS", ""))
        for i, f in enumerate(sorted_top, 1):
            summary_rows.append((
                f"  {i}. [{f.get('Severity', '').upper()}]",
                f.get("Title", "")
            ))

    for r, (label, value) in enumerate(summary_rows, 1):
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_v = ws.cell(row=r, column=2, value=value)
        # Section headers
        if label and label.isupper() and label.strip() and not label.startswith(" "):
            cell_l.font = section_font
            cell_v.font = section_font
        else:
            cell_l.font = label_font
            cell_v.font = value_font
        cell_l.border = _XL_THIN_BORDER
        cell_v.border = _XL_THIN_BORDER
        cell_v.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 75

    # ── Sheet 2: All Findings ──
    ws2 = wb.create_sheet("All Findings")
    ws2.sheet_properties.tabColor = "D13438"
    headers2 = [
        "Finding ID", "Severity", "Category", "Subcategory", "Title",
        "Description", "Affected Count", "Subscription",
        "Detected At", "Data Security Relevance",
        "Compliance Frameworks",
        "Remediation Description",
        "Remediation (Azure CLI)", "Remediation (PowerShell)",
        "Remediation (Portal Steps)", "Resources",
    ]
    _xl_apply_header(ws2, headers2)

    sorted_findings = sorted(findings, key=lambda f: (sev_order.get(f.get("Severity", "medium").lower(), 5), f.get("Category", "")))
    for i, f in enumerate(sorted_findings, 2):
        sev = f.get("Severity", "medium").lower()
        cat_key = f.get("Category", "")
        cat_name = _CATEGORY_META.get(cat_key, {}).get("name", cat_key)
        subcat = f.get("Subcategory", "")

        resources = f.get("AffectedResources", [])
        resource_names = ", ".join(
            r.get("Name", r) if isinstance(r, dict) else str(r)
            for r in (resources if isinstance(resources, list) else [])
        )

        remediation = f.get("Remediation", "")
        rem_desc = ""
        rem_cli = ""
        rem_ps = ""
        rem_portal = ""
        if isinstance(remediation, dict):
            rem_desc = remediation.get("Description", "")
            rem_cli = remediation.get("AzureCLI", remediation.get("CLI", ""))
            rem_ps = remediation.get("PowerShell", "")
            portal_steps = remediation.get("PortalSteps", [])
            if isinstance(portal_steps, list) and portal_steps:
                rem_portal = "\n".join(f"{j}. {s}" for j, s in enumerate(portal_steps, 1))
            else:
                rem_portal = remediation.get("Portal", "")
        elif isinstance(remediation, str):
            rem_cli = remediation

        subs = set()
        for r in (resources if isinstance(resources, list) else []):
            if isinstance(r, dict):
                rid = r.get("ResourceId", "")
                parts = rid.split("/")
                if len(parts) > 2 and parts[1].lower() == "subscriptions":
                    subs.add(parts[2])

        # Compliance mapping summary
        comp_mapping = f.get("ComplianceMapping", {})
        comp_summary = "; ".join(
            f"{fw}: {', '.join(ctrls)}"
            for fw, ctrls in sorted(comp_mapping.items())
        ) if comp_mapping else ""

        # Data security relevance
        ds_rel = _DATA_SECURITY_RELEVANCE.get(cat_key, "")

        finding_id = f.get("DataSecurityFindingId", f"{cat_key}_{subcat}")

        fills: dict[int, PatternFill] = {}
        if sev in _XL_SEV_FILLS:
            fills[2] = _XL_SEV_FILLS[sev]
        fonts2: dict[int, Font] = {}
        if sev in _XL_SEV_FONTS:
            fonts2[2] = _XL_SEV_FONTS[sev]
        _row2_fill = _XL_ZEBRA_FILL if i % 2 == 0 else None

        _xl_write_row(ws2, i, [
            finding_id,
            sev.capitalize(),
            cat_name,
            subcat.replace("_", " ").title(),
            f.get("Title", ""),
            f.get("Description", ""),
            f.get("AffectedCount", len(resources) if isinstance(resources, list) else 0),
            ", ".join(subs) if subs else "",
            f.get("DetectedAt", "")[:19] if f.get("DetectedAt") else "",
            ds_rel,
            comp_summary,
            rem_desc,
            rem_cli,
            rem_ps,
            rem_portal,
            resource_names,
        ], fills, fonts2, _row2_fill)

    _xl_auto_width(ws2, headers2)

    # ── Sheet 3: Affected Resources ──
    ws3 = wb.create_sheet("Affected Resources")
    ws3.sheet_properties.tabColor = "F7630C"

    # Detect all enriched detail keys across all findings
    _STANDARD_KEYS = {"Name", "name", "Type", "type", "ResourceId", "resource_id", "id",
                      "Severity", "Risk", "ResourceRemediation", "Detail", "Reason"}
    all_detail_keys: list[str] = []
    seen_keys: set[str] = set()
    for f in sorted_findings:
        for ar in (f.get("AffectedResources", []) if isinstance(f.get("AffectedResources"), list) else []):
            if isinstance(ar, dict):
                for k in ar:
                    if k not in _STANDARD_KEYS and k not in seen_keys:
                        all_detail_keys.append(k)
                        seen_keys.add(k)

    headers3 = [
        "Resource Name", "Resource Type", "Resource Group", "Subscription",
        "Resource ID", "Finding", "Finding ID", "Severity", "Category",
        "Detail", "Risk", "Resource Remediation",
        "Finding Remediation (CLI)", "Finding Remediation (PS)", "Finding Remediation (Portal)",
    ] + [k.replace("_", " ").title() for k in all_detail_keys]
    _xl_apply_header(ws3, headers3)

    row3 = 2
    for f in sorted_findings:
        sev = f.get("Severity", "medium").lower()
        cat_key = f.get("Category", "")
        cat_name = _CATEGORY_META.get(cat_key, {}).get("name", cat_key)
        finding_id = f.get("DataSecurityFindingId", "")
        resources = f.get("AffectedResources", [])
        if not isinstance(resources, list):
            continue

        # Extract finding-level remediation
        _f_rem = f.get("Remediation", "")
        _f_rem_cli = ""
        _f_rem_ps = ""
        _f_rem_portal = ""
        if isinstance(_f_rem, dict):
            _f_rem_cli = _f_rem.get("AzureCLI", _f_rem.get("CLI", ""))
            _f_rem_ps = _f_rem.get("PowerShell", "")
            _f_rem_portal_steps = _f_rem.get("PortalSteps", [])
            if isinstance(_f_rem_portal_steps, list) and _f_rem_portal_steps:
                _f_rem_portal = "\n".join(f"{j}. {s}" for j, s in enumerate(_f_rem_portal_steps, 1))
            else:
                _f_rem_portal = _f_rem.get("Portal", "")
        elif isinstance(_f_rem, str):
            _f_rem_cli = _f_rem

        for r in resources:
            if isinstance(r, dict):
                rid = r.get("ResourceId", r.get("resource_id", r.get("id", "")))
                parts = str(rid).split("/")
                rg = ""
                sub = ""
                rtype = r.get("Type", r.get("type", ""))
                if len(parts) > 4:
                    sub = parts[2] if parts[1].lower() == "subscriptions" else ""
                    rg = parts[4] if parts[3].lower() == "resourcegroups" else ""
                res_sev = r.get("Severity", sev).lower()
                fills3: dict[int, PatternFill] = {}
                fonts3: dict[int, Font] = {}
                if res_sev in _XL_SEV_FILLS:
                    fills3[8] = _XL_SEV_FILLS[res_sev]
                if res_sev in _XL_SEV_FONTS:
                    fonts3[8] = _XL_SEV_FONTS[res_sev]
                _row3_fill = _XL_ZEBRA_FILL if row3 % 2 == 0 else None

                base_vals = [
                    r.get("Name", r.get("name", "")),
                    rtype,
                    rg,
                    sub,
                    str(rid),
                    f.get("Title", ""),
                    finding_id,
                    res_sev.capitalize(),
                    cat_name,
                    r.get("Detail", r.get("Reason", "")),
                    r.get("Risk", ""),
                    r.get("ResourceRemediation", ""),
                    _f_rem_cli,
                    _f_rem_ps,
                    _f_rem_portal,
                ]
                extra_vals = [str(r.get(k, "")) for k in all_detail_keys]
                _xl_write_row(ws3, row3, base_vals + extra_vals, fills3, fonts3, _row3_fill)
                row3 += 1
            else:
                _xl_write_row(ws3, row3, [
                    str(r), "", "", "", "",
                    f.get("Title", ""),
                    finding_id,
                    sev.capitalize(),
                    cat_name, "", "", "",
                    _f_rem_cli, _f_rem_ps, _f_rem_portal,
                ] + [""] * len(all_detail_keys))
                row3 += 1

    _xl_auto_width(ws3, headers3)

    # ── Sheet 4: Category Scores ──
    ws4 = wb.create_sheet("Category Scores")
    ws4.sheet_properties.tabColor = "FFB900"
    headers4 = ["Category", "Score", "Level", "Finding Count", "Description"]
    _xl_apply_header(ws4, headers4)

    row4 = 2
    for cat_key in all_cats:
        meta = _CATEGORY_META.get(cat_key)
        if not meta:
            continue
        cs = cat_scores.get(cat_key, {})
        if not isinstance(cs, dict):
            continue
        cat_findings = [f for f in findings if f.get("Category") == cat_key]
        _cat_level = cs.get("Level", "low")
        _cat_fills: dict[int, PatternFill] = {}
        _cat_fonts: dict[int, Font] = {}
        if _cat_level in _XL_SEV_FILLS:
            _cat_fills[3] = _XL_SEV_FILLS[_cat_level]
        if _cat_level in _XL_SEV_FONTS:
            _cat_fonts[3] = _XL_SEV_FONTS[_cat_level]
        _xl_write_row(ws4, row4, [
            meta["name"],
            round(cs.get("Score", 0), 1),
            cs.get("Level", "N/A"),
            len(cat_findings),
            meta["description"],
        ], _cat_fills, _cat_fonts, _XL_ZEBRA_FILL if row4 % 2 == 0 else None)
        row4 += 1

    _xl_auto_width(ws4, headers4)

    # ── Sheet 5: Remediation ──
    ws5 = wb.create_sheet("Remediation")
    ws5.sheet_properties.tabColor = "107C10"
    headers5 = [
        "Priority", "Severity", "Category", "Finding", "Finding ID",
        "Affected Count", "Remediation Description",
        "Azure CLI", "PowerShell", "Portal Steps",
    ]
    _xl_apply_header(ws5, headers5)

    priority_map = {
        "critical": "P0 — Immediate",
        "high": "P1 — Short-term",
        "medium": "P2 — Medium-term",
        "low": "P3 — Long-term",
        "informational": "P4 — Advisory",
    }

    row5 = 2
    for f in sorted_findings:
        sev = f.get("Severity", "medium").lower()
        cat_key = f.get("Category", "")
        cat_name = _CATEGORY_META.get(cat_key, {}).get("name", cat_key)
        resources = f.get("AffectedResources", [])
        finding_id = f.get("DataSecurityFindingId", "")

        remediation = f.get("Remediation", "")
        rem_desc = ""
        rem_cli = ""
        rem_ps = ""
        rem_portal = ""
        if isinstance(remediation, dict):
            rem_desc = remediation.get("Description", "")
            rem_cli = remediation.get("AzureCLI", remediation.get("CLI", ""))
            rem_ps = remediation.get("PowerShell", "")
            portal_steps = remediation.get("PortalSteps", [])
            if isinstance(portal_steps, list) and portal_steps:
                rem_portal = "\n".join(f"{j}. {s}" for j, s in enumerate(portal_steps, 1))
            else:
                rem_portal = remediation.get("Portal", "")
        elif isinstance(remediation, str):
            rem_cli = remediation

        fills5: dict[int, PatternFill] = {}
        fonts5: dict[int, Font] = {}
        if sev in _XL_SEV_FILLS:
            fills5[2] = _XL_SEV_FILLS[sev]
        if sev in _XL_SEV_FONTS:
            fonts5[2] = _XL_SEV_FONTS[sev]
        _row5_fill = _XL_ZEBRA_FILL if row5 % 2 == 0 else None

        _xl_write_row(ws5, row5, [
            priority_map.get(sev, "P3 — Long-term"),
            sev.capitalize(),
            cat_name,
            f.get("Title", ""),
            finding_id,
            f.get("AffectedCount", len(resources) if isinstance(resources, list) else 0),
            rem_desc,
            rem_cli,
            rem_ps,
            rem_portal,
        ], fills5, fonts5, _row5_fill)
        row5 += 1

    _xl_auto_width(ws5, headers5)

    # ── Sheet 6: Compliance Mappings ──
    ws6 = wb.create_sheet("Compliance Mappings")
    headers6 = [
        "Framework", "Control ID", "Control Title", "Rationale",
        "Recommendation", "Finding Title", "Finding ID", "Severity", "Category",
    ]
    _xl_apply_header(ws6, headers6)

    row6 = 2
    for f in sorted_findings:
        sev = f.get("Severity", "medium").lower()
        cat_key = f.get("Category", "")
        cat_name = _CATEGORY_META.get(cat_key, {}).get("name", cat_key)
        finding_id = f.get("DataSecurityFindingId", "")
        comp_mapping = f.get("ComplianceMapping", {})
        comp_details = f.get("ComplianceDetails", {})

        for fw, controls in sorted(comp_mapping.items()):
            for ctrl in controls:
                key = f"{fw}:{ctrl}"
                det = comp_details.get(key, {})
                fills6: dict[int, PatternFill] = {}
                if sev in _XL_SEV_FILLS:
                    fills6[8] = _XL_SEV_FILLS[sev]
                _xl_write_row(ws6, row6, [
                    fw,
                    ctrl,
                    det.get("title", ""),
                    det.get("rationale", ""),
                    det.get("recommendation", ""),
                    f.get("Title", ""),
                    finding_id,
                    sev.capitalize(),
                    cat_name,
                ], fills6)
                row6 += 1

    _xl_auto_width(ws6, headers6)

    # ── Sheet 7: Priority Quadrant ──
    ws7 = wb.create_sheet("Priority Quadrant")
    headers7 = [
        "Quadrant", "Severity", "Finding Count", "Estimated Effort",
        "Recommended Action", "Example Findings",
    ]
    _xl_apply_header(ws7, headers7)

    quadrant_map = {"critical": "Quick Wins", "high": "Quick Wins",
                    "medium": "Major Projects", "low": "Low Priority",
                    "informational": "Consider"}
    effort_map = {"critical": "Low (config change)", "high": "Low-Medium",
                  "medium": "Medium", "low": "Medium-High",
                  "informational": "Varies"}
    action_map = {
        "critical": "Remediate immediately — active breach risk or regulatory violation.",
        "high": "Remediate within 1-2 sprints — significant risk exposure.",
        "medium": "Plan for next quarter — moderate compliance gaps for defense-in-depth.",
        "low": "Address during regular maintenance — minor best-practice deviations.",
        "informational": "Review periodically — awareness items with minimal direct risk.",
    }

    sev_groups: dict[str, list[dict]] = {}
    for f in findings:
        s = f.get("Severity", "medium").lower()
        sev_groups.setdefault(s, []).append(f)

    row7 = 2
    for s in ["critical", "high", "medium", "low", "informational"]:
        items = sev_groups.get(s, [])
        if not items:
            continue
        examples = ", ".join(f.get("Title", "")[:60] for f in items[:5])
        if len(items) > 5:
            examples += f", +{len(items) - 5} more"
        fills7: dict[int, PatternFill] = {}
        if s in _XL_SEV_FILLS:
            fills7[2] = _XL_SEV_FILLS[s]
        _xl_write_row(ws7, row7, [
            quadrant_map.get(s, "Consider"),
            s.capitalize(),
            len(items),
            effort_map.get(s, "Varies"),
            action_map.get(s, "Review as needed."),
            examples,
        ], fills7)
        row7 += 1

    _xl_auto_width(ws7, headers7)

    # ── Sheet 8: Cost of Non-Compliance ──
    ws8 = wb.create_sheet("Cost of Non-Compliance")
    headers8 = [
        "Severity", "Finding Count", "Cost per Finding (USD)",
        "Calculation", "Total Exposure (USD)", "Methodology",
    ]
    _xl_apply_header(ws8, headers8)

    cost_est = _estimate_noncompliance_cost(findings)
    cost_per_sev_display = {"critical": 250_000, "high": 75_000, "medium": 15_000, "low": 2_500, "informational": 500}

    row8 = 2
    for s in ["critical", "high", "medium", "low", "informational"]:
        d = cost_est["by_severity"].get(s)
        if not d:
            continue
        unit = cost_per_sev_display.get(s, 5_000)
        fills8: dict[int, PatternFill] = {}
        if s in _XL_SEV_FILLS:
            fills8[1] = _XL_SEV_FILLS[s]
        _xl_write_row(ws8, row8, [
            f"{s.title()} Risk",
            d["count"],
            f"${unit:,}",
            f"{d['count']} x ${unit:,}",
            f"${d['cost']:,}",
            "",
        ], fills8)
        row8 += 1

    # Total row
    total_font = Font(name="Segoe UI", bold=True, size=11)
    for col, val in enumerate(["TOTAL", "", "", "", f"${cost_est['total']:,}", ""], 1):
        cell = ws8.cell(row=row8, column=col, value=val)
        cell.font = total_font
        cell.border = _XL_THIN_BORDER
    row8 += 2

    # Methodology note
    meth_note = (
        "Cost-per-finding multipliers derived from: IBM Cost of a Data Breach Report 2024 "
        "(avg. breach cost $4.88M), Ponemon Institute per-record cost studies, NIST SP 800-184 "
        "cyber-event recovery guidance. Critical ($250K) = high breach likelihood with regulatory "
        "fine exposure; High ($75K) = significant risk; Medium ($15K) = moderate compliance gap; "
        "Low ($2.5K) = minor deviation; Informational ($500) = awareness item. "
        "Disclaimer: These are directional estimates."
    )
    cell_m = ws8.cell(row=row8, column=1, value="Methodology & Sources")
    cell_m.font = Font(name="Segoe UI", bold=True, size=10)
    cell_mn = ws8.cell(row=row8 + 1, column=1, value=meth_note)
    cell_mn.font = Font(name="Segoe UI", size=9)
    cell_mn.alignment = Alignment(wrap_text=True)
    ws8.merge_cells(start_row=row8 + 1, start_column=1, end_row=row8 + 1, end_column=6)

    _xl_auto_width(ws8, headers8)

    # ── Sheet 9: Trend Comparison (if available) ──
    trend = results.get("Trend")
    if trend and isinstance(trend, dict):
        ws9 = wb.create_sheet("Trend Comparison")
        headers9 = ["Metric", "Value"]
        _xl_apply_header(ws9, headers9)

        prev_at = trend.get("PreviousAssessedAt", "")[:19]
        trend_rows: list[tuple[str, str]] = [
            ("Previous Assessment", prev_at),
            ("Previous Score", str(trend.get("PreviousScore", "N/A"))),
            ("Current Score", str(trend.get("CurrentScore", "N/A"))),
            ("Score Delta", str(trend.get("ScoreDelta", 0))),
            ("Previous Finding Count", str(trend.get("PreviousFindingCount", 0))),
            ("Current Finding Count", str(trend.get("CurrentFindingCount", 0))),
            ("New Findings", str(trend.get("NewCount", 0))),
            ("Resolved Findings", str(trend.get("ResolvedCount", 0))),
        ]

        for r, (label, value) in enumerate(trend_rows, 2):
            _xl_write_row(ws9, r, [label, value])

        # New findings detail
        new_findings = trend.get("NewFindings", [])
        if new_findings:
            gap_row = len(trend_rows) + 3
            cell_h = ws9.cell(row=gap_row, column=1, value="NEW FINDINGS")
            cell_h.font = section_font
            nr = gap_row + 1
            for nf in new_findings:
                _xl_write_row(ws9, nr, [
                    f"[{nf.get('Severity', '').upper()}] {nf.get('Title', '')}",
                    nf.get("Category", ""),
                ])
                nr += 1

        # Resolved findings detail
        resolved_findings = trend.get("ResolvedFindings", [])
        if resolved_findings:
            gap_row2 = (len(trend_rows) + 3 + len(new_findings) + 2) if new_findings else len(trend_rows) + 3
            cell_h2 = ws9.cell(row=gap_row2, column=1, value="RESOLVED FINDINGS")
            cell_h2.font = section_font
            rr = gap_row2 + 1
            for rf in resolved_findings:
                _xl_write_row(ws9, rr, [
                    f"[{rf.get('Severity', '').upper()}] {rf.get('Title', '')}",
                    rf.get("Category", ""),
                ])
                rr += 1

        ws9.column_dimensions["A"].width = 55
        ws9.column_dimensions["B"].width = 40

    # ── Sheet 10: Collection Errors (if any) ──
    collection_errors = results.get("CollectionErrors", [])
    if collection_errors:
        ws10 = wb.create_sheet("Collection Errors")
        headers10 = ["Data Source / Query", "Error", "Impact"]
        _xl_apply_header(ws10, headers10)

        for r, e in enumerate(collection_errors, 2):
            _xl_write_row(ws10, r, [
                e.get("query", ""),
                e.get("error", "unknown"),
                "Findings for this category may be incomplete. Score of 0 does not guarantee compliance.",
            ])

        _xl_auto_width(ws10, headers10)

    # ── Sheet 10b: Enrichment Errors (if any) ──
    enrichment_errors = results.get("EnrichmentErrors", [])
    if enrichment_errors:
        ws10b = wb.create_sheet("Enrichment Errors")
        headers10b = ["Enrichment Step", "Error", "Impact"]
        _xl_apply_header(ws10b, headers10b)

        for r, e in enumerate(enrichment_errors, 2):
            _xl_write_row(ws10b, r, [
                e.get("enrichment_step", ""),
                e.get("error", "unknown"),
                e.get("impact", "Findings for affected categories may be incomplete."),
            ])

        _xl_auto_width(ws10b, headers10b)

    # ── Sheet 11: Compliance Gap Matrix ──
    # Reproduce the framework × severity matrix from the HTML
    fw_data: dict[str, dict[str, int]] = {}
    fw_controls: dict[str, set[str]] = {}
    for f in findings:
        sev_f = f.get("Severity", "medium").lower()
        comp_map = f.get("ComplianceMapping", {})
        for fw, ctrls in comp_map.items():
            if fw not in fw_data:
                fw_data[fw] = {}
                fw_controls[fw] = set()
            fw_data[fw][sev_f] = fw_data[fw].get(sev_f, 0) + 1
            fw_controls[fw].update(ctrls)

    if fw_data:
        ws11 = wb.create_sheet("Compliance Gap Matrix")
        headers11 = ["Framework", "Critical", "High", "Medium", "Low", "Informational", "Total Controls"]
        _xl_apply_header(ws11, headers11)

        sev_fill_map = {
            2: _XL_SEV_FILLS.get("critical"),
            3: _XL_SEV_FILLS.get("high"),
            4: _XL_SEV_FILLS.get("medium"),
            5: _XL_SEV_FILLS.get("low"),
        }

        row11 = 2
        for fw in sorted(fw_data.keys()):
            d = fw_data[fw]
            fills11: dict[int, PatternFill] = {}
            for col_idx, sev_key in [(2, "critical"), (3, "high"), (4, "medium"), (5, "low")]:
                if d.get(sev_key, 0) > 0 and sev_key in _XL_SEV_FILLS:
                    fills11[col_idx] = _XL_SEV_FILLS[sev_key]
            _xl_write_row(ws11, row11, [
                fw,
                d.get("critical", 0),
                d.get("high", 0),
                d.get("medium", 0),
                d.get("low", 0),
                d.get("informational", 0),
                len(fw_controls.get(fw, set())),
            ], fills11)
            row11 += 1

        _xl_auto_width(ws11, headers11)

    wb.save(out_path)
    log.info("[DataSecurityExcel] Written to %s", out_path)
    return out_path


# ── Cost Methodology Report ──────────────────────────────────────────

def generate_cost_methodology_report(results: dict, output_dir: str) -> pathlib.Path:
    """Generate a standalone Cost of Non-Compliance Methodology HTML report.

    Produces a professional HTML report explaining how estimated costs are
    calculated, with authoritative references from compliance standards and
    Microsoft documentation.
    """
    out_dir = pathlib.Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    findings = results.get("Findings", [])
    scores = results.get("DataSecurityScores", results.get("Scores", {}))
    overall = scores.get("OverallScore", 0)
    level = scores.get("OverallLevel", "low")
    tenant = results.get("TenantDisplayName", results.get("TenantId", ""))
    assessed_at = results.get("AssessedAt", "")

    cost_est = _estimate_noncompliance_cost(findings)
    cost_total = cost_est["total"]
    cost_by_sev = cost_est["by_severity"]

    # Severity tier definitions
    _tiers = [
        {
            "key": "critical", "label": "Critical", "color": "#D13438",
            "cost": 250_000,
            "probability": "60 – 90%",
            "rationale": (
                "Critical findings represent active breach risk or imminent regulatory violation. "
                "Examples include unencrypted data at rest, publicly exposed storage accounts, "
                "or missing encryption on virtual machines. The $250,000 multiplier reflects the "
                "combination of high breach probability (60–90%) and significant per-incident cost "
                "including regulatory fines, breach notification expenses, forensic investigation, "
                "and reputational damage."
            ),
            "references": [
                "IBM Cost of a Data Breach Report 2024 — average breach cost $4.88M globally",
                "GDPR Article 83(5) — fines up to €20M or 4% of annual turnover",
                "HIPAA Breach Notification Rule (45 CFR §164.404) — penalties $100–$50,000 per violation",
                "PCI DSS v4.0 §12.10 — incident response requirements; non-compliance fines $5,000–$100,000/month",
                "NIST SP 800-53 Rev. 5 SC-28 (Protection of Information at Rest)",
            ],
        },
        {
            "key": "high", "label": "High", "color": "#F7630C",
            "cost": 75_000,
            "probability": "30 – 60%",
            "rationale": (
                "High severity findings indicate significant risk exposure that increases attack "
                "surface substantially. Examples include Key Vaults without purge protection, "
                "services without private endpoints, or AI services with public access. The $75,000 "
                "multiplier represents moderate-to-high breach probability with meaningful regulatory "
                "and operational impact."
            ),
            "references": [
                "IBM Cost of a Data Breach Report 2024 — $4.88M avg; high-severity vectors ~$1.2M incremental",
                "Ponemon Institute 2024 — per-record cost $165 (regulated industries up to $239)",
                "NIST SP 800-53 Rev. 5 SC-7 (Boundary Protection), SC-8 (Transmission Confidentiality)",
                "Microsoft Security Benchmark v3 — NS-2 (Secure cloud services with network controls)",
                "PCI DSS v4.0 §1.3.1 — restrict inbound/outbound traffic; penalty exposure for segmentation gaps",
            ],
        },
        {
            "key": "medium", "label": "Medium", "color": "#FFB900",
            "cost": 15_000,
            "probability": "10 – 30%",
            "rationale": (
                "Medium findings represent moderate compliance gaps in defense-in-depth posture. "
                "Examples include missing diagnostic settings, absent managed identities, "
                "DDoS protection gaps, or missing resource locks. The $15,000 multiplier reflects "
                "lower direct breach probability but meaningful regulatory audit exposure and "
                "operational risk from reduced visibility or recovery capacity."
            ),
            "references": [
                "IBM Cost of a Data Breach Report 2024 — organizations with security AI/automation saved $2.22M on avg",
                "NIST SP 800-184 (Guide for Cybersecurity Event Recovery) — recovery cost modeling",
                "NIST SP 800-53 Rev. 5 AU-6 (Audit Record Review), SI-4 (System Monitoring)",
                "Microsoft Security Benchmark v3 — LT-4 (Enable logging for security investigation)",
                "HIPAA Security Rule §164.312(b) — audit controls requirement",
            ],
        },
        {
            "key": "low", "label": "Low", "color": "#107C10",
            "cost": 2_500,
            "probability": "< 10%",
            "rationale": (
                "Low severity findings are minor best-practice deviations with limited immediate risk. "
                "Examples include managed disks using platform-managed keys instead of customer-managed "
                "keys. The $2,500 multiplier accounts for incremental risk during compliance audits, "
                "potential audit findings, and the operational overhead of remediation during incidents."
            ),
            "references": [
                "NIST SP 800-53 Rev. 5 SC-12 (Cryptographic Key Establishment and Management)",
                "Microsoft Security Benchmark v3 — DP-5 (Use customer-managed key option in data at rest encryption when required)",
                "CIS Azure Foundations Benchmark v2.0 — Section 7 (Key Management)",
                "Ponemon Institute 2024 — encryption reduces breach cost by avg $252K",
            ],
        },
        {
            "key": "informational", "label": "Informational", "color": "#8A8886",
            "cost": 500,
            "probability": "< 5%",
            "rationale": (
                "Informational items represent awareness opportunities with minimal direct risk. "
                "The $500 multiplier accounts for the marginal cost of unaddressed improvement "
                "opportunities that could compound over time or surface during regulatory reviews."
            ),
            "references": [
                "NIST Cybersecurity Framework v2.0 — ID.RA (Risk Assessment) continuous improvement",
                "ISO 27001:2022 Annex A.8.10 — Information deletion and lifecycle",
                "Microsoft Well-Architected Framework — Security pillar (continuous posture improvement)",
            ],
        },
    ]

    # Build cost breakdown rows
    cost_rows_html = ""
    for tier in _tiers:
        k = tier["key"]
        if k not in cost_by_sev:
            continue
        d = cost_by_sev[k]
        cost_rows_html += (
            f'<tr>'
            f'<td style="color:{tier["color"]};font-weight:700;padding:10px 16px">{tier["label"]}</td>'
            f'<td style="text-align:center;padding:10px 16px">{d["count"]}</td>'
            f'<td style="text-align:right;padding:10px 16px;font-family:monospace">${tier["cost"]:,.0f}</td>'
            f'<td style="text-align:center;padding:10px 16px;color:var(--text-muted);font-size:12px">'
            f'{d["count"]} &times; ${tier["cost"]:,.0f}</td>'
            f'<td style="text-align:right;padding:10px 16px;font-family:monospace;font-weight:700">${d["cost"]:,.0f}</td>'
            f'</tr>'
        )

    # Build tier rationale sections
    tier_sections = ""
    for tier in _tiers:
        k = tier["key"]
        d = cost_by_sev.get(k, {"count": 0, "cost": 0})
        refs_html = "".join(
            f'<li style="margin:4px 0">{esc(r)}</li>' for r in tier["references"]
        )
        tier_sections += f"""
    <div style="margin:24px 0;padding:20px 24px;background:var(--bg-elevated);border-radius:10px;border-left:4px solid {tier['color']}">
      <h3 style="margin:0 0 8px 0;color:{tier['color']}">{tier['label']} Severity — ${tier['cost']:,.0f} per finding</h3>
      <table style="font-size:13px;margin:8px 0 12px 0">
        <tr><td style="padding:3px 12px 3px 0;color:var(--text-secondary)">Findings in this tier:</td>
            <td style="font-weight:700">{d['count']}</td></tr>
        <tr><td style="padding:3px 12px 3px 0;color:var(--text-secondary)">Estimated breach probability:</td>
            <td style="font-weight:600">{tier['probability']}</td></tr>
        <tr><td style="padding:3px 12px 3px 0;color:var(--text-secondary)">Tier exposure:</td>
            <td style="font-weight:700;font-family:monospace">${d['cost']:,.0f}</td></tr>
      </table>
      <p style="font-size:13px;line-height:1.7;color:var(--text-secondary);margin:8px 0">{tier['rationale']}</p>
      <h4 style="margin:12px 0 6px 0;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;color:var(--text-secondary)">Authoritative References</h4>
      <ul style="margin:0;padding-left:20px;font-size:12px;color:var(--text-secondary);line-height:1.8">{refs_html}</ul>
    </div>"""

    level_color = {"critical": "#D13438", "high": "#F7630C", "medium": "#FFB900", "low": "#107C10"}.get(level, "#A8A6A3")

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Cost of Non-Compliance Methodology — EnterpriseSecurityIQ</title>
<style>
:root {{
  --bg: #1a1a2e; --bg-card: #22223a; --bg-elevated: #2a2a45;
  --text: #e8e8f0; --text-secondary: #a8a8c0; --text-muted: #7a7a95;
  --border: #3a3a55; --accent: #0078D4;
  --font-sans: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  --font-mono: 'Cascadia Code', 'Fira Code', Consolas, monospace;
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: var(--font-sans); background: var(--bg); color: var(--text); line-height: 1.6; }}
.container {{ max-width: 960px; margin: 0 auto; padding: 32px 24px; }}
h1 {{ font-size: 24px; border-bottom: 2px solid var(--accent); padding-bottom: 12px; margin-bottom: 8px; }}
h2 {{ font-size: 18px; color: var(--accent); margin: 32px 0 12px 0; padding-bottom: 6px; border-bottom: 1px solid var(--border); }}
h3 {{ font-size: 15px; }}
h4 {{ font-size: 13px; }}
a {{ color: #6CB4EE; text-decoration: underline; text-underline-offset: 2px; }}
a:visited {{ color: #B39DDB; }}
a:hover {{ color: #90CAF9; }}
table {{ width: 100%; border-collapse: collapse; }}
th {{ text-align: left; padding: 8px 16px; font-size: 11px; text-transform: uppercase;
     letter-spacing: 0.5px; color: var(--text-secondary); border-bottom: 2px solid var(--border); }}
td {{ padding: 8px 16px; border-bottom: 1px solid var(--border); font-size: 13px; }}
.hero {{ text-align: center; padding: 32px 16px; margin: 24px 0; background: var(--bg-card);
         border-radius: 12px; border: 1px solid var(--border); }}
.hero .amount {{ font-size: 42px; font-weight: 800; color: #D13438; font-family: var(--font-mono); }}
.hero .subtitle {{ font-size: 13px; color: var(--text-secondary); margin-top: 4px; }}
.formula {{ background: var(--bg-elevated); padding: 16px 20px; border-radius: 8px; margin: 16px 0;
            font-family: var(--font-mono); font-size: 14px; border-left: 3px solid var(--accent); }}
.meta {{ color: var(--text-muted); font-size: 12px; margin-bottom: 24px; }}
.back-link {{ display: inline-block; margin-bottom: 20px; padding: 6px 16px; background: var(--bg-card);
              border: 1px solid var(--border); border-radius: 6px; font-size: 12px; text-decoration: none; color: var(--text); }}
.back-link:hover {{ background: var(--bg-elevated); color: var(--accent); }}
.disclaimer {{ margin-top: 32px; padding: 16px 20px; background: var(--bg-card); border-radius: 10px;
               border-left: 4px solid #FFB900; font-size: 12px; color: var(--text-secondary); line-height: 1.8; }}
.ref-list {{ margin: 12px 0; padding-left: 20px; }}
.ref-list li {{ margin: 6px 0; font-size: 13px; color: var(--text-secondary); line-height: 1.6; }}
.nav {{ position: sticky; top: 0; z-index: 100; background: var(--bg-elevated); border-bottom: 1px solid var(--border);
        padding: 10px 24px; display: flex; gap: 16px; align-items: center; font-size: 13px; }}
.nav a {{ color: var(--text-secondary); text-decoration: none; }}
.nav a:hover {{ color: var(--accent); }}
footer {{ margin-top: 40px; padding-top: 16px; border-top: 1px solid var(--border);
          font-size: 11px; color: var(--text-muted); text-align: center; }}
@media print {{
  body {{ background: #fff; color: #1a1a1a; }}
  .nav, .back-link {{ display: none; }}
  .hero {{ border: 2px solid #ccc; }}
}}
</style>
</head>
<body>
<nav class="nav">
  <strong>&#128274; EnterpriseSecurityIQ</strong>
  <a href="#summary">Summary</a>
  <a href="#breakdown">Breakdown</a>
  <a href="#methodology">Methodology</a>
  <a href="#tiers">Tier Details</a>
  <a href="#sources">Sources</a>
  <a href="data-security-assessment.html">&#8592; Back to Assessment</a>
</nav>

<div class="container">
<a href="data-security-assessment.html" class="back-link">&#8592; Back to Data Security Assessment</a>

<h1>&#128176; Cost of Non-Compliance — Methodology Report</h1>
<div class="meta">
  Tenant: <strong>{esc(tenant)}</strong> &nbsp;|&nbsp;
  Assessed: {esc(assessed_at[:19] if assessed_at else 'N/A')} UTC &nbsp;|&nbsp;
  Security Score: <span style="color:{level_color};font-weight:700">{overall:.0f}/100 ({level.upper()})</span> &nbsp;|&nbsp;
  Findings: {len(findings)}
</div>

<!-- ── Executive Summary ── -->
<section id="summary">
<h2>Executive Summary</h2>
<div class="hero">
  <div class="amount">${cost_total:,.0f}</div>
  <div class="subtitle">Estimated Annual Non-Compliance Exposure</div>
</div>
<p style="font-size:14px;color:var(--text-secondary);margin:12px 0;line-height:1.8">
  This report details how EnterpriseSecurityIQ estimates the financial exposure arising from unresolved
  data security findings. The model assigns a cost multiplier to each severity tier based on
  <strong>breach probability</strong>, <strong>average incident cost</strong>, and
  <strong>regulatory fine exposure</strong>. All multipliers are grounded in authoritative
  industry research and compliance standards.
</p>
<div class="formula">
  <strong>Formula:</strong>&nbsp; Total Exposure = &Sigma; (Findings<sub>severity</sub> &times; Cost-per-Finding<sub>severity</sub>)
</div>
</section>

<!-- ── Cost Breakdown ── -->
<section id="breakdown">
<h2>Cost Breakdown by Severity</h2>
<table>
  <thead><tr>
    <th style="text-align:left">Severity</th>
    <th style="text-align:center">Findings</th>
    <th style="text-align:right">Cost / Finding</th>
    <th style="text-align:center">Calculation</th>
    <th style="text-align:right">Total</th>
  </tr></thead>
  <tbody>{cost_rows_html}</tbody>
  <tfoot><tr style="border-top:2px solid var(--border)">
    <td colspan="4" style="text-align:right;padding:10px 16px;font-weight:700">Estimated Annual Exposure</td>
    <td style="text-align:right;padding:10px 16px;font-family:var(--font-mono);font-weight:800;font-size:16px;color:#D13438">${cost_total:,.0f}</td>
  </tr></tfoot>
</table>
</section>

<!-- ── Methodology ── -->
<section id="methodology">
<h2>Cost Model Methodology</h2>
<p style="font-size:13px;color:var(--text-secondary);line-height:1.8;margin:8px 0">
  The cost model uses a <strong>severity-weighted multiplier approach</strong>. Each finding is assigned
  to a severity tier during the data security assessment based on its potential impact, exploitability,
  and compliance implications. The cost-per-finding multiplier for each tier is derived as:
</p>
<div class="formula">
  Cost<sub>tier</sub> = P(breach | finding) &times; E(cost | breach) + P(audit finding) &times; E(remediation cost)
</div>
<p style="font-size:13px;color:var(--text-secondary);line-height:1.8;margin:12px 0">Where:</p>
<ul style="font-size:13px;color:var(--text-secondary);line-height:2;margin:0 0 0 20px">
  <li><strong>P(breach | finding)</strong> — Probability that this type of gap leads to a security incident, derived from industry breach statistics</li>
  <li><strong>E(cost | breach)</strong> — Expected cost per incident including direct costs (forensics, notification, legal) and indirect costs (reputation, customer churn)</li>
  <li><strong>P(audit finding)</strong> — Probability that this gap is flagged during a compliance audit (SOC 2, PCI-DSS QSA, HIPAA audit, etc.)</li>
  <li><strong>E(remediation cost)</strong> — Expected cost of emergency remediation under audit pressure vs. planned remediation</li>
</ul>
</section>

<!-- ── Tier-by-Tier Details ── -->
<section id="tiers">
<h2>Severity Tier Details &amp; Rationale</h2>
<p style="font-size:13px;color:var(--text-secondary);margin:8px 0">
  Each tier below explains the cost multiplier derivation, breach probability estimate,
  and the authoritative standards that inform the calculation.
</p>
{tier_sections}
</section>

<!-- ── Authoritative Sources ── -->
<section id="sources">
<h2>Authoritative Sources &amp; References</h2>
<p style="font-size:13px;color:var(--text-secondary);margin:8px 0 16px 0">
  The following authoritative documents were used to derive and validate the cost model.
  These are widely recognized by auditors, regulators, and compliance professionals.
</p>

<h3 style="margin:20px 0 8px 0;font-size:14px">Industry Research</h3>
<ul class="ref-list">
  <li><strong>IBM Security &amp; Ponemon Institute</strong> — <em>Cost of a Data Breach Report 2024</em>. Average global breach cost $4.88M; healthcare $10.93M. Per-record cost $165 (regulated industries $239). <a href="https://www.ibm.com/reports/data-breach" target="_blank" rel="noopener">ibm.com/reports/data-breach</a></li>
  <li><strong>Ponemon Institute</strong> — <em>Cost of Insider Threats Global Report 2024</em>. Average annual cost of insider threat incidents $16.2M per organization.</li>
  <li><strong>Verizon</strong> — <em>Data Breach Investigations Report (DBIR) 2024</em>. Analysis of 30,458 security incidents and 10,626 confirmed breaches across 94 countries.</li>
</ul>

<h3 style="margin:20px 0 8px 0;font-size:14px">NIST Standards</h3>
<ul class="ref-list">
  <li><strong>NIST SP 800-184</strong> — <em>Guide for Cybersecurity Event Recovery</em>. Framework for recovery planning and cost estimation after cyber events. <a href="https://csrc.nist.gov/pubs/sp/800/184/final" target="_blank" rel="noopener">csrc.nist.gov</a></li>
  <li><strong>NIST SP 800-53 Rev. 5</strong> — <em>Security and Privacy Controls for Information Systems</em>. Comprehensive control catalog (SC-28, SC-7, SC-8, SC-12, AU-6, SI-4). <a href="https://csrc.nist.gov/pubs/sp/800/53/r5/upd1/final" target="_blank" rel="noopener">csrc.nist.gov</a></li>
  <li><strong>NIST Cybersecurity Framework v2.0</strong> — Risk assessment and continuous improvement functions. <a href="https://www.nist.gov/cyberframework" target="_blank" rel="noopener">nist.gov/cyberframework</a></li>
</ul>

<h3 style="margin:20px 0 8px 0;font-size:14px">Regulatory Frameworks &amp; Penalty Structures</h3>
<ul class="ref-list">
  <li><strong>PCI DSS v4.0</strong> — Payment Card Industry Data Security Standard. Non-compliance penalties $5,000–$100,000/month; breach liability includes card reissuance costs ($3–$10/card) and forensic investigation ($20K–$500K). <a href="https://www.pcisecuritystandards.org/document_library/" target="_blank" rel="noopener">pcisecuritystandards.org</a></li>
  <li><strong>GDPR (EU Regulation 2016/679)</strong> — Article 83: fines up to €20M or 4% of annual global turnover. Article 82: right to compensation for data subjects. <a href="https://gdpr-info.eu/art-83-gdpr/" target="_blank" rel="noopener">gdpr-info.eu</a></li>
  <li><strong>HIPAA</strong> — Breach Notification Rule (45 CFR §164.404–408). HHS Office for Civil Rights enforcement: Tier 1 $100–$50,000/violation; Tier 4 $50,000/violation with $2.067M annual cap. <a href="https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/" target="_blank" rel="noopener">hhs.gov</a></li>
  <li><strong>SOX (Sarbanes-Oxley Act)</strong> — Section 302/906: penalties up to $5M and 20 years imprisonment for willful non-compliance with IT controls affecting financial reporting.</li>
  <li><strong>CCPA/CPRA</strong> — California Consumer Privacy Act: statutory damages $100–$750 per consumer per incident; AG penalties up to $7,500 per intentional violation.</li>
</ul>

<h3 style="margin:20px 0 8px 0;font-size:14px">Microsoft &amp; Cloud Security Standards</h3>
<ul class="ref-list">
  <li><strong>Microsoft Cloud Security Benchmark (MCSB) v1</strong> — Comprehensive security controls for Azure services. <a href="https://learn.microsoft.com/en-us/security/benchmark/azure/overview" target="_blank" rel="noopener">learn.microsoft.com</a></li>
  <li><strong>Microsoft Well-Architected Framework — Security Pillar</strong> — Design principles and best practices for cloud security posture. <a href="https://learn.microsoft.com/en-us/azure/well-architected/security/" target="_blank" rel="noopener">learn.microsoft.com</a></li>
  <li><strong>CIS Azure Foundations Benchmark v2.0</strong> — Center for Internet Security prescriptive security configuration guidance. <a href="https://www.cisecurity.org/benchmark/azure" target="_blank" rel="noopener">cisecurity.org</a></li>
  <li><strong>ISO/IEC 27001:2022</strong> — Information security management systems. Annex A controls for encryption, access control, and data lifecycle management.</li>
</ul>
</section>

<!-- ── Disclaimer ── -->
<div class="disclaimer">
  <strong style="color:var(--text)">&#9888;&#65039; Important Disclaimer</strong><br><br>
  The cost estimates in this report are <strong>directional indicators</strong> designed to help
  prioritize remediation efforts. They are <strong>not</strong> actuarial predictions or financial guarantees.
  Actual costs of a security incident or compliance failure depend on numerous factors including:
  <ul style="margin:8px 0 0 20px;line-height:2">
    <li>Data sensitivity and volume of records affected</li>
    <li>Regulatory jurisdiction(s) — GDPR, HIPAA, PCI-DSS, CCPA, SOX, etc.</li>
    <li>Industry sector and applicable vertical regulations</li>
    <li>Breach detection and response time (IBM 2024: avg 258 days to identify + contain)</li>
    <li>Existing insurance coverage and contractual obligations</li>
    <li>Organization size and annual revenue (affects fine calculations)</li>
  </ul>
  <br>
  <strong>For precise financial risk modeling</strong>, engage qualified risk management professionals and
  consider formal quantitative risk analysis methodologies such as FAIR (Factor Analysis of Information Risk).
  <br><br>
  <em style="font-size:11px">Cost multipliers last updated: Q1 2026. Based on IBM 2024, Ponemon 2024, NIST 2023 publications.</em>
</div>

<footer>
  Generated by EnterpriseSecurityIQ {VERSION} &nbsp;|&nbsp;
  <a href="data-security-assessment.html">Full Data Security Report</a> &nbsp;|&nbsp;
  <a href="executive-brief.html">Executive Brief</a>
</footer>
</div>
</body></html>"""

    path = out_dir / "cost-methodology.html"
    path.write_text(html, encoding="utf-8")
    log.info("Cost methodology report -> %s", path)
    return path
