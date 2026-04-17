"""
M365 Copilot Readiness Report — Interactive HTML
Full-width professional report showing readiness posture for
Microsoft 365 Copilot adoption with executive summary, category
breakdown, severity distribution, top findings, and remediation.
"""

from __future__ import annotations

import hashlib
import pathlib
from app.reports.shared_theme import (
    get_css, get_js, esc, format_date_short, VERSION,
    SEVERITY_COLORS,
)
from app.logger import log


# ── Category metadata ────────────────────────────────────────────────────

_CATEGORY_META: dict[str, dict] = {
    "oversharing_risk": {
        "icon": "&#128275;",  # 🔓
        "name": "Oversharing Risk",
        "color": "#D13438",
        "description": "Sites with broad membership, Everyone permissions, anonymous links, and external sharing posture that could expose data to Copilot.",
    },
    "label_coverage": {
        "icon": "&#127991;",  # 🏷
        "name": "Label Coverage",
        "color": "#8764B8",
        "description": "Sensitivity label definitions, mandatory labeling, auto-labeling, and site label coverage across the tenant.",
    },
    "dlp_readiness": {
        "icon": "&#128737;",  # 🛡
        "name": "DLP Readiness",
        "color": "#0078D4",
        "description": "Data Loss Prevention policy existence, label integration, and workload coverage for Copilot interactions.",
    },
    "restricted_search": {
        "icon": "&#128269;",  # 🔍
        "name": "Restricted SharePoint Search",
        "color": "#00B7C3",
        "description": "Restricted SharePoint Search (RSS) configuration to limit Copilot access to approved sites.",
    },
    "access_governance": {
        "icon": "&#128101;",  # 👥
        "name": "Access Governance",
        "color": "#F7630C",
        "description": "Conditional Access policies and license scoping for Copilot access control.",
    },
    "content_lifecycle": {
        "icon": "&#128197;",  # 📅
        "name": "Content Lifecycle",
        "color": "#FFB900",
        "description": "Stale content risk for Copilot grounding and retention policy coverage.",
    },
    "audit_monitoring": {
        "icon": "&#128202;",  # 📊
        "name": "Audit & Monitoring",
        "color": "#107C10",
        "description": "Unified audit logging and Copilot-specific interaction logging.",
    },
    "copilot_security": {
        "icon": "&#129302;",  # 🤖
        "name": "Copilot-Specific Security",
        "color": "#8764B8",
        "description": "Plugin restrictions, data residency, eDiscovery, DSPM for AI, Graph grounding, and insider risk.",
    },
    "zero_trust": {
        "icon": "&#128737;",  # 🛡
        "name": "Zero Trust Posture",
        "color": "#005A9E",
        "description": "Continuous access evaluation, token protection, phishing-resistant MFA, authentication contexts, and workload identity protection.",
    },
    "shadow_ai": {
        "icon": "&#128065;",  # 👁
        "name": "Shadow AI Governance",
        "color": "#C239B3",
        "description": "Unauthorized AI app registrations, consent grants to AI services, shadow Copilot agents, and AI data exfiltration prevention.",
    },
}

_READINESS_META: dict[str, dict] = {
    "READY":     {"color": "#107C10", "label": "Ready",     "icon": "&#10003;"},
    "NEEDS WORK": {"color": "#FFB900", "label": "Needs Work", "icon": "&#9888;&#65039;"},
    "NOT READY": {"color": "#D13438", "label": "Not Ready", "icon": "&#128680;"},
}


# ── SVG helpers ──────────────────────────────────────────────────────────

def _ring_score_svg(score: float, size: int = 140) -> str:
    r = size // 2 - 8
    circ = 2 * 3.14159 * r
    pct = min(score, 100) / 100
    dash = circ * pct
    gap = circ - dash
    cx = cy = size // 2
    color = "#107C10" if score >= 75 else "#FFB900" if score >= 50 else "#F7630C" if score >= 25 else "#D13438"
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" class="ring" role="img">'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="var(--ring-track)" stroke-width="10"/>'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="10" '
        f'stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-dashoffset="{circ * 0.25:.2f}" '
        f'stroke-linecap="round" style="transition:stroke-dasharray 1s ease"/>'
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" font-size="28" font-weight="700" '
        f'fill="{color}" font-family="var(--font-mono)">{score:.0f}</text>'
        f'<text x="{cx}" y="{cy + 18}" text-anchor="middle" font-size="10" fill="var(--text-muted)">/100</text>'
        f'</svg>'
    )


def _donut_svg(slices: list[tuple[str, float, str]], size: int = 140, center_text: str | None = None) -> str:
    total = sum(v for _, v, _ in slices) or 1
    r = size // 2 - 4
    circ = 2 * 3.14159 * r
    cx = cy = size // 2
    parts = [f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" role="img">']
    offset = 0
    for label, val, color in slices:
        if val <= 0:
            continue
        pct = val / total
        dash = circ * pct
        gap = circ - dash
        parts.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" '
            f'stroke-width="{r * 0.4}" stroke-dasharray="{dash:.2f} {gap:.2f}" '
            f'stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})">'
            f'<title>{esc(label)}: {int(val)} ({pct*100:.0f}%)</title></circle>'
        )
        offset += dash
    ct = center_text if center_text is not None else str(int(total))
    parts.append(
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" '
        f'font-size="22" font-weight="700" fill="var(--text)" '
        f'font-family="var(--font-mono)">{esc(ct)}</text>'
    )
    parts.append("</svg>")
    return "".join(parts)


# ── Severity badge ───────────────────────────────────────────────────────

def _severity_badge(sev: str) -> str:
    color = SEVERITY_COLORS.get(sev.lower(), "#A8A6A3")
    return f'<span class="badge" style="background:{color}">{esc(sev.upper())}</span>'


def _compliance_badge(status: str) -> str:
    colors = {"compliant": "#107C10", "gap": "#D13438", "partial": "#FFB900"}
    color = colors.get(status.lower(), "#A8A6A3")
    return f'<span class="badge" style="background:{color}">{esc(status.upper())}</span>'


def _ms_ref_link(label: str, url: str) -> str:
    """Render remediation guidance as inline detailed text for the row context."""
    if not label:
        return '<span style="color:var(--text-muted)">—</span>'
    explanation = _MS_REFERENCE_EXPLANATIONS.get(label, "")
    if explanation:
        return (
            f'<div style="font-size:11px;line-height:1.5">'
            f'<div style="font-weight:600;color:var(--primary);margin-bottom:3px">{esc(label)}</div>'
            f'<div style="color:var(--text-secondary)">{esc(explanation)}</div>'
            f'</div>'
        )
    return f'<span style="font-size:11px;color:var(--text-muted)">{esc(label)}</span>'


_MS_REFERENCE_EXPLANATIONS: dict[str, str] = {
    "MS Copilot Readiness — Oversharing Assessment": (
        "Run a SharePoint site access review: Open SharePoint Admin Center > Sites > Active sites. "
        "For each site, click Permissions to audit who has access. Remove 'Everyone' and 'Everyone except external users' "
        "from site members/visitors. Use Get-SPOSite -Limit All | Select Url,SharingCapability to list sharing posture. "
        "Prioritize sites with SharingCapability set to ExternalUserAndGuestSharing."
    ),
    "MS Copilot Readiness — Permission Cleanup": (
        "Audit SharePoint permissions: Run Get-SPOSite -Limit All | ForEach { Get-SPOSiteGroup -Site $_.Url } to list all groups. "
        "Remove stale users from site groups. Set site-level sharing to 'Only people in your organization'. "
        "Use Set-SPOSite -Identity <url> -SharingCapability Disabled for sensitive sites. "
        "Review and remove broken permission inheritance on document libraries."
    ),
    "MS Copilot Readiness — Sharing Controls": (
        "Set org-wide sharing limits: SharePoint Admin Center > Policies > Sharing. "
        "Set external sharing to 'Least permissive' (Existing guests or Only people in your organization). "
        "PowerShell: Set-SPOTenant -SharingCapability ExistingExternalUserSharingOnly. "
        "Disable 'Anyone' links at org level. Set default link type to 'Specific people'. "
        "Set default link permission to 'View' not 'Edit'."
    ),
    "MS Copilot Readiness — External Sharing": (
        "Restrict external sharing per-site: SharePoint Admin Center > Sites > select site > Sharing. "
        "Set each sensitive site to 'Only people in your organization'. "
        "PowerShell: Set-SPOSite -Identity <url> -SharingCapability Disabled. "
        "For sites that need external sharing, set to 'Existing guests' and require authentication. "
        "Block sharing to specific domains with Set-SPOTenant -SharingBlockedDomainList."
    ),
    "MS Copilot Readiness — Least Privilege": (
        "Implement least-privilege access model: Remove broad group memberships from SharePoint sites. "
        "Replace 'Everyone' permissions with specific Entra ID security groups. "
        "Use SharePoint Advanced Management access reviews to identify over-permissioned sites. "
        "PowerShell: Get-SPOSiteGroup -Site <url> to list groups, then Remove-SPOUser to clean up. "
        "Enable 'Site owners can share files and folders' only — not members."
    ),
    "MS SharePoint — Advanced Management": (
        "Enable SharePoint Advanced Management: M365 Admin Center > Settings > Org settings > SharePoint. "
        "Required SKU: Microsoft Syntex - SharePoint Advanced Management. "
        "Once enabled, use Site access reviews (Admin Center > SharePoint > Access reviews) to audit permissions. "
        "Configure Data access governance reports to identify overshared content. "
        "Set up Restricted Content Discoverability to limit Copilot indexing scope."
    ),
    "MS SharePoint — SAM Restricted Access": (
        "Configure Restricted Access Control: SharePoint admin center > Sites > select site > Settings. "
        "Enable Restricted Access Control and specify which security groups can access the site. "
        "PowerShell: Set-SPOSite -Identity <url> -RestrictedAccessControl $true. "
        "Then: Set-SPOSite -Identity <url> -RestrictedAccessControlGroups <groupId>. "
        "Apply to all sites containing sensitive or confidential content before Copilot rollout."
    ),
    "MS SharePoint — SAM Site Lifecycle": (
        "Configure site lifecycle policies: SharePoint admin center > Policies > Site lifecycle management. "
        "Enable inactive site detection with a threshold (e.g., 180 days of no activity). "
        "Configure email notifications to site owners when sites become inactive. "
        "Set escalation actions: archive or delete sites with no owner response. "
        "This prevents Copilot from indexing outdated content in abandoned sites."
    ),
    "MS SharePoint — SAM DAG Reports": (
        "Access DAG reports: SharePoint admin center > Reports > Data access governance. "
        "Review 'Oversharing' report: identifies sites shared with Everyone or large groups. "
        "Review 'Sensitive content' report: identifies sites with files matching sensitive info types. "
        "Review 'Everyone links' report: sites with org-wide sharing links. "
        "Export reports and prioritize remediation of highest-risk sites before Copilot deployment."
    ),
    "MS Purview — Information Protection": (
        "Deploy sensitivity labels: Purview Compliance Portal > Information protection > Labels. "
        "Create a label taxonomy (Public, Internal, Confidential, Highly Confidential). "
        "Publish labels via label policies to all users. Set mandatory labeling in policy settings. "
        "PowerShell: Connect-IPPSSession then New-Label / Set-LabelPolicy. "
        "Enable co-authoring support for encrypted files. Test with a pilot group before org-wide rollout."
    ),
    "MS Purview — Label Policies": (
        "Configure label policies: Purview Portal > Information protection > Label policies > Create policy. "
        "Select target labels, scope to all users/groups, enable 'Require users to apply a label'. "
        "Set a default label for documents (e.g., Internal). Enable 'Require justification to remove a label'. "
        "PowerShell: Set-LabelPolicy -Identity <policy> -AdvancedSettings @{requiredowngradejustification='true'}. "
        "Monitor adoption via Activity Explorer in the Purview portal."
    ),
    "MS Purview — Auto-Labeling": (
        "Set up auto-labeling: Purview Portal > Information protection > Auto-labeling policies. "
        "Create rules based on sensitive information types (SSN, credit card, etc). "
        "Configure simulation mode first to review matches before enforcement. "
        "Scope to specific SharePoint sites and OneDrive accounts. "
        "PowerShell: New-AutoSensitivityLabelPolicy with -ExchangeLocation, -SharePointLocation parameters. "
        "Review simulation results and activate after validation."
    ),
    "MS Purview — Container Labels": (
        "Apply labels to containers: Purview Portal > Information protection > Labels > edit label > "
        "enable 'Groups & sites' scope. Configure privacy setting (Public/Private), external sharing, "
        "and Conditional Access. Publish via label policy. "
        "PowerShell: Set-UnifiedGroup -Identity <group> -SensitivityLabelId <labelId>. "
        "Requires Azure AD Premium P1. Enable support: Set-SPOTenant -EnableAIPIntegration $true."
    ),
    "MS Purview — Default Labels": (
        "Set default labels in Office apps: Purview Portal > Information protection > Label policies > "
        "edit policy > set 'Apply this default label to documents' (e.g., Internal). "
        "Also set default label for emails. Users see the label applied automatically when creating new files. "
        "Enable 'Require justification to change label' for downgrade protection. "
        "Test: Create a new Word document and verify the label auto-applies."
    ),
    "MS Purview — DLP": (
        "Create DLP policies: Purview Portal > Data loss prevention > Policies > Create policy. "
        "Select regulatory template (e.g., U.S. PII, GDPR). Choose locations: Exchange, SharePoint, Teams, OneDrive. "
        "Configure rules: block sharing of content with sensitive info types. Set to 'Test mode' first. "
        "PowerShell: New-DlpCompliancePolicy -Name <name> -SharePointLocation All -Mode TestWithNotifications. "
        "Review DLP alerts dashboard and tune rules to reduce false positives before enforcement."
    ),
    "MS Purview — DLP Label Conditions": (
        "Add label conditions to DLP rules: Edit DLP policy > Rules > Add condition > "
        "'Content contains sensitivity label'. Select the label (e.g., Confidential, Highly Confidential). "
        "Configure action: Block external sharing of labeled content. "
        "This ensures DLP works with your labeling taxonomy — content labeled 'Confidential' triggers DLP. "
        "PowerShell: New-DlpComplianceRule -ContentPropertyContainsWords 'MSIP_Label_<guid>'."
    ),
    "MS Purview — DLP Locations": (
        "Extend DLP to all workloads: Edit DLP policy > Locations > enable all: "
        "Exchange email, SharePoint sites, OneDrive accounts, Teams chat and channel messages, Devices, Power BI. "
        "Key coverage gaps: Teams messages are often missed — enable explicitly. "
        "PowerShell: Set-DlpCompliancePolicy -Identity <name> -TeamsLocation All. "
        "Review covered vs uncovered locations in DLP policy details page."
    ),
    "MS Purview — Endpoint DLP": (
        "Enable Endpoint DLP: Purview Portal > Settings > Endpoint DLP settings > Turn on Endpoint DLP. "
        "Onboard Windows devices through Microsoft Intune compliance policies. "
        "Configure monitored activities: Copy to clipboard, Print, Upload to browser, USB. "
        "Create DLP policy with 'Devices' location enabled. Set enforcement actions per activity type. "
        "Requires Microsoft 365 E5/A5 or Microsoft 365 E5 Compliance license."
    ),
    "MS SharePoint — Restricted Search": (
        "Enable Restricted SharePoint Search: SharePoint Admin Center > Search > Restricted Access. "
        "Toggle on 'Restrict SharePoint Search'. Add approved sites to the allow list. "
        "Only listed sites will be indexed by Copilot for search results. "
        "PowerShell: Set-SPOTenant -IsSearchBoxInNavBarEnabled $true and configure allowed sites list. "
        "Review and update the list quarterly as new sites are created."
    ),
    "MS Entra — Conditional Access": (
        "Create a Conditional Access policy for Copilot: Entra Admin Center > Protection > Conditional Access. "
        "New policy: assign to Copilot users group, target 'Office 365' cloud app. "
        "Grant controls: Require MFA + Require compliant device. Session: set sign-in frequency to 8hrs. "
        "PowerShell: use Microsoft.Graph module — New-MgIdentityConditionalAccessPolicy. "
        "Test in Report-only mode before enabling. Monitor sign-in logs for blocked access."
    ),
    "MS Admin — License Management": (
        "Manage Copilot licenses: M365 Admin Center > Billing > Licenses > Microsoft 365 Copilot. "
        "Assign to specific users or groups — do not assign to all users without readiness review. "
        "Create a security group 'Copilot-Pilot-Users' for phased rollout. "
        "PowerShell: Set-MgUserLicense -UserId <upn> -AddLicenses @{SkuId='<copilot-sku-id>'}. "
        "Review usage reports: M365 Admin Center > Reports > Copilot usage."
    ),
    "MS Admin — Licensing": (
        "Review license inventory: M365 Admin Center > Billing > Licenses. "
        "Verify Copilot prerequisite: users need Microsoft 365 E3/E5/Business Premium. "
        "Check consumed vs available licenses. Reclaim unused licenses with Get-MgUser -Filter 'assignedLicenses/any()'. "
        "Ensure Copilot SKU is available before assigning. "
        "For group-based licensing: Entra Admin Center > Groups > select group > Licenses."
    ),
    "MS Entra — Identity Governance": (
        "Set up access reviews: Entra Admin Center > Identity Governance > Access reviews > New review. "
        "Scope: review group memberships for SharePoint site access groups quarterly. "
        "Reviewers: group owners or managers. Auto-apply results to remove stale access. "
        "PowerShell: New-MgAccessReview with -Scope and -Reviewers parameters. "
        "Also configure entitlement management access packages for Copilot access requests."
    ),
    "MS Purview — Information Barriers": (
        "Configure information barriers: Purview Portal > Information barriers > Segments. "
        "Create user segments based on department/role attributes in Entra ID. "
        "Create barrier policies: New-InformationBarrierPolicy -AssignedSegment <segment> -SegmentsBlocked <list>. "
        "Apply policies: Start-InformationBarrierPoliciesApplication. "
        "This prevents Copilot from surfacing content across restricted organizational boundaries."
    ),
    "MS Entra — Authentication Methods": (
        "Configure MFA: Entra Admin Center > Protection > Authentication methods. "
        "Enable Microsoft Authenticator (push + passwordless). Disable SMS and voice call for high-security users. "
        "Create registration campaign: Admin Center > Protection > Authentication methods > Registration campaign. "
        "PowerShell: Update-MgPolicyAuthenticationMethodPolicy to manage method configurations. "
        "Monitor: Authentication methods activity report shows adoption percentage per method."
    ),
    "MS Entra — PIM": (
        "Enable PIM: Entra Admin Center > Identity Governance > Privileged Identity Management. "
        "Convert permanent Global Admin and SharePoint Admin role assignments to eligible. "
        "Set activation duration: max 8 hours, require MFA and justification for activation. "
        "Configure approval workflow for Global Admin role. Set up notifications for role activations. "
        "PowerShell: New-MgRoleManagementDirectoryRoleEligibilityScheduleRequest."
    ),
    "MS Entra — Named Locations": (
        "Define named locations: Entra Admin Center > Protection > Conditional Access > Named locations. "
        "Add your corporate IP ranges as 'trusted'. Add countries you operate in. "
        "Reference in CA policies: 'Conditions > Locations > Exclude trusted locations'. "
        "Block Copilot access from untrusted locations with a CA policy targeting Office 365 apps. "
        "PowerShell: New-MgIdentityConditionalAccessNamedLocation -DisplayName 'Corp IPs' -IpRanges <list>."
    ),
    "MS Entra — Identity Protection": (
        "Enable risk policies: Entra Admin Center > Protection > Identity Protection. "
        "Sign-in risk policy: Require MFA for medium+ risk sign-ins. "
        "User risk policy: Require password change for high risk users. "
        "Configure risk-based Conditional Access: target all users, grant control = require MFA when risk detected. "
        "Review risky users and sign-ins reports weekly. Investigate and remediate confirmed compromises."
    ),
    "MS Intune — Device Compliance": (
        "Create compliance policy: Intune Admin Center > Devices > Compliance policies > Create. "
        "Settings: Require BitLocker, Require OS version minimum, Require system security (firewall, antivirus). "
        "Set non-compliance actions: Mark device non-compliant immediately, send notification after 3 days. "
        "Reference in CA policy: Grant > 'Require device to be marked as compliant'. "
        "This ensures Copilot is only accessible from devices meeting your security baseline."
    ),
    "MS SharePoint — Site Lifecycle": (
        "Manage stale content: SharePoint Admin Center > Sites > Active sites > sort by 'Last activity'. "
        "Archive or delete sites with no activity for 180+ days. "
        "PowerShell: Get-SPOSite -Limit All | Where { $_.LastContentModifiedDate -lt (Get-Date).AddDays(-180) }. "
        "Configure inactive site policies: Admin Center > Policies > Site lifecycle management. "
        "Set automatic notifications to site owners of stale sites. Move archived sites out of Copilot indexing."
    ),
    "MS Purview — Retention": (
        "Configure retention: Purview Portal > Data lifecycle management > Retention policies. "
        "Create policy for SharePoint: retain 7 years, then delete. Apply to all sites. "
        "Create policy for Teams messages: retain 3 years. Create policy for email: retain 5 years. "
        "PowerShell: New-RetentionCompliancePolicy -SharePointLocation All -RetentionDuration 2555. "
        "Use retention labels for item-level control on sensitive content categories."
    ),
    "MS Admin — M365 Backup": (
        "Enable Microsoft 365 Backup: M365 Admin Center > Settings > Microsoft 365 Backup. "
        "Enable backup protection for Exchange mailboxes, OneDrive accounts, and SharePoint sites. "
        "Configure backup scope — start with critical sites and expand. "
        "Set backup frequency and retention per your RPO/RTO requirements. "
        "Verify restore capability with a test restore before Copilot rollout."
    ),
    "MS Purview — Audit": (
        "Enable unified audit: Purview Portal > Audit > Start recording. "
        "Verify status: Search-UnifiedAuditLog -StartDate <date> -EndDate <date> -RecordType AzureActiveDirectory. "
        "Upgrade to Audit Premium (E5) for 1-year retention and high-value events. "
        "Configure audit log retention: Purview Portal > Audit > Audit retention policies. "
        "Key audit events for Copilot: FileAccessed, FileModified, SearchQueryPerformed."
    ),
    "MS Purview — Copilot Audit Events": (
        "Enable Copilot audit events: Purview Portal > Audit > verify 'CopilotInteraction' record type. "
        "Search for Copilot events: RecordType = CopilotInteraction in audit log search. "
        "Track: Copilot prompts, content grounding sources, file access via Copilot. "
        "Requires Microsoft 365 E5 or E5 Compliance. Events include which files Copilot read to generate answers. "
        "Set up alert policies on Copilot access to Highly Confidential labeled content."
    ),
    "MS Purview — Alert Policies": (
        "Create alert policies: Purview Portal > Policies > Alert policies > New policy. "
        "Configure alerts for: unusual file access volume, external sharing spikes, DLP policy matches. "
        "Set severity and notification recipients (security team distribution list). "
        "For Copilot-specific: create alert on 'FileAccessed' events for Highly Confidential content. "
        "PowerShell: New-ProtectionAlert -Name 'High-vol access' -Category DataLossPrevention -NotifyUser <email>."
    ),
    "MS Defender — Cloud Apps": (
        "Configure Defender for Cloud Apps: security.microsoft.com > Cloud apps > Policies. "
        "Create session policy: monitor and control real-time Copilot access to sensitive files. "
        "Create activity policy: alert on mass file downloads after Copilot interactions. "
        "Enable app governance: Cloud apps > App governance > turn on for all OAuth apps. "
        "Key policy: Block download of Confidential files when accessed from unmanaged devices via Copilot."
    ),
    "MS Admin — Copilot Settings": (
        "Configure Copilot settings: M365 Admin Center > Settings > Copilot. "
        "Manage plugins and connectors: disable third-party plugins if not needed. "
        "Control web content access: choose whether Copilot can search the web. "
        "Configure data access: set which M365 services Copilot can use for grounding. "
        "Review enabled experiences: M365 Admin Center > Org settings > Copilot > manage individual features."
    ),
    "MS Trust Center — Data Residency": (
        "Review data residency: M365 Admin Center > Settings > Org settings > Organization profile > Data location. "
        "Verify tenant data is stored in the expected geography. "
        "For Advanced Data Residency (ADR): M365 Admin Center > Settings > Org settings > Advanced Data Residency. "
        "ADR ensures Copilot processing and storage stays within your selected geo. "
        "Requires Microsoft 365 Advanced Data Residency add-on license."
    ),
    "MS Purview — eDiscovery": (
        "Configure eDiscovery for Copilot: Purview Portal > eDiscovery > Cases > New case. "
        "Add custodians and their M365 data sources (mailbox, OneDrive, Teams). "
        "Search for Copilot interaction data using content search with RecordType:CopilotInteraction. "
        "Place legal holds to preserve Copilot-related data for litigation or audit. "
        "Requires Microsoft 365 E5 or E5 eDiscovery and Audit add-on."
    ),
    "MS Copilot Readiness — Graph Grounding": (
        "Control Graph grounding: M365 Admin Center > Settings > Copilot > manage how Copilot uses org data. "
        "Configure which data sources Copilot can access for response generation. "
        "Use Restricted SharePoint Search to limit which sites are used for grounding. "
        "Review Graph connector settings if custom connectors are enabled. "
        "PowerShell: Use Microsoft Graph API to audit Copilot grounding configuration."
    ),
    "MS Purview — Insider Risk": (
        "Configure Insider Risk: Purview Portal > Insider risk management > Policies > Create policy. "
        "Select template: 'Data leaks' or 'Security policy violations'. "
        "Enable indicators: Copilot activity, file downloads after Copilot queries, unusual content access. "
        "Scope to Copilot-licensed users. Set alert thresholds based on risk level. "
        "Review alerts: Insider risk management > Alerts. Requires Microsoft 365 E5 or E5 Insider Risk add-on."
    ),
    "MS Purview — Communication Compliance": (
        "Set up Communication Compliance: Purview Portal > Communication compliance > Policies > Create. "
        "Select 'Monitor for inappropriate content' or custom policy. Scope to Copilot responses if available. "
        "Configure reviewers from your compliance team. Set review workflow and escalation paths. "
        "Monitor policy matches via the Communication compliance dashboard. "
        "Requires Microsoft 365 E5 or E5 Compliance add-on."
    ),
    "MS Purview — DSPM for AI": (
        "Enable DSPM for AI: Microsoft Purview portal > Data Security Posture Management > DSPM for AI. "
        "Review the AI security dashboard for sensitive data exposure risks. "
        "Run oversharing assessments to identify sites with sensitive content accessible to Copilot. "
        "Configure policies to monitor and alert on sensitive data in AI interactions. "
        "Requires Microsoft 365 E5 or E5 Compliance license."
    ),
    "MS Purview — DSPM Oversharing": (
        "Run DSPM oversharing assessment: Purview portal > DSPM for AI > Oversharing assessment > New assessment. "
        "Select scope (all sites or specific sites) and run the assessment. "
        "Review identified overshared content and follow remediation recommendations. "
        "Set up recurring assessments to maintain visibility as permissions change. "
        "Prioritize sites flagged with sensitive information types before Copilot deployment."
    ),
    # ── Phase 1 — Identity & Licensing ──
    "MS Entra — User Sign-In Activity": (
        "Detect stale accounts: Entra admin center > Users > All users > filter by 'Last sign-in' column. "
        "PowerShell: Get-MgUser -All -Property SignInActivity | Where-Object { $_.SignInActivity.LastSignInDateTime -lt (Get-Date).AddDays(-90) } | "
        "Select DisplayName,UserPrincipalName,@{N='LastSignIn';E={$_.SignInActivity.LastSignInDateTime}}. "
        "Disable or block sign-in for accounts inactive longer than 90 days. "
        "Remove Copilot licenses from stale accounts to prevent unused license consumption."
    ),
    "MS Entra — Global Admin Best Practices": (
        "Reduce Global Admin count: Entra admin center > Roles and administrators > Global Administrator. "
        "Target: no more than 2-4 Global Admins (Microsoft recommendation). "
        "PowerShell: Get-MgDirectoryRoleMember -DirectoryRoleId (Get-MgDirectoryRole -Filter \"displayName eq 'Global Administrator'\").Id | Measure-Object. "
        "Transition excess Global Admins to least-privilege roles (e.g. User Administrator, Groups Administrator). "
        "Enable PIM for all remaining Global Admins: Entra > Identity Governance > Privileged Identity Management."
    ),
    "MS Entra — Shared Accounts": (
        "Identify shared accounts: Entra admin center > Users > filter for generic names (admin@, support@, shared@). "
        "PowerShell: Get-MgUser -All | Where-Object { $_.UserPrincipalName -match 'shared|generic|admin|service' } | Select DisplayName,UserPrincipalName. "
        "Migrate shared accounts to dedicated service accounts or shared mailboxes with proper delegation. "
        "Ensure no Copilot licenses are assigned to shared/generic accounts — Copilot licenses must be per-user. "
        "Enable sign-in logs review for any remaining shared accounts."
    ),
    "MS Entra — Group-Based Licensing": (
        "Configure Group-Based Licensing: Entra admin center > Groups > select Copilot deployment group > Licenses. "
        "Assign Microsoft 365 Copilot license to the group. "
        "PowerShell: Set-MgGroupLicense -GroupId <group-id> -AddLicenses @{SkuId='<copilot-sku-id>'} -RemoveLicenses @(). "
        "Create dynamic groups for auto-assignment: Entra > Groups > New group > Membership type: Dynamic user > "
        "Rule: (user.department -eq \"Sales\") -and (user.accountEnabled -eq true). "
        "This replaces manual per-user license assignment and simplifies onboarding/offboarding."
    ),
    "MS Entra — Session Controls": (
        "Configure Sign-in Frequency: Entra admin center > Protection > Conditional Access > select or create policy > "
        "Session > Sign-in frequency > set to 'Every time' or specific hours (e.g. 4 hours for sensitive apps). "
        "Configure Persistent Browser: same policy > Session > Persistent browser session > 'Never persistent'. "
        "Target: Copilot-licensed users accessing M365 apps. Require re-authentication for sensitive content access. "
        "PowerShell: Use Microsoft Graph API — New-MgIdentityConditionalAccessPolicy with sessionControls block."
    ),
    # ── Phase 2 — Exchange & Agent Governance ──
    "MS Exchange — Mailbox Permissions": (
        "Audit mailbox delegations: Exchange Admin Center > Recipients > Mailboxes > select mailbox > Delegation. "
        "PowerShell: Get-Mailbox -ResultSize Unlimited | Get-MailboxPermission | Where-Object { $_.User -ne 'NT AUTHORITY\\SELF' } | "
        "Select Identity,User,AccessRights. Review Full Access, Send As, and Send on Behalf permissions. "
        "Remove unnecessary delegations — Copilot can surface content from delegated mailboxes. "
        "Focus on executive and sensitive mailboxes first."
    ),
    "MS Exchange — Shared Mailboxes": (
        "Audit shared mailbox sprawl: Exchange Admin Center > Recipients > Shared. "
        "PowerShell: Get-Mailbox -RecipientTypeDetails SharedMailbox | Get-MailboxPermission | "
        "Where-Object { $_.User -ne 'NT AUTHORITY\\SELF' } | Group-Object Identity | "
        "Select Name,@{N='DelegateCount';E={$_.Count}} | Sort DelegateCount -Descending. "
        "Shared mailboxes with many delegates expose content to all delegates via Copilot. "
        "Restrict delegate lists to minimum required and apply sensitivity labels to sensitive shared mailboxes."
    ),
    "MS Copilot Studio — Agent Governance": (
        "Review Copilot agents: Microsoft 365 admin center > Settings > Copilot > Agents. "
        "Inventory all published and draft agents. Verify each agent's data source connections are scoped appropriately. "
        "Power Platform admin center > Environments > select environment > Copilot Studio agents. "
        "Establish an agent approval workflow before new agents are published to production. "
        "Disable creation of new agents by non-admin users if not needed: Settings > Copilot > Agent settings."
    ),
    "MS Entra — App Permissions": (
        "Audit app permissions: Entra admin center > Applications > Enterprise applications > Permissions. "
        "PowerShell: Get-MgServicePrincipal -All | ForEach-Object { Get-MgServicePrincipalOAuth2PermissionGrant -ServicePrincipalId $_.Id } | "
        "Where-Object { $_.ConsentType -eq 'AllPrincipals' } | Select ResourceId,Scope. "
        "Review apps with 'admin consent granted' for broad Microsoft Graph permissions (Sites.Read.All, Mail.Read). "
        "Restrict Copilot agent/plugin permissions to minimum Microsoft Graph scopes required. "
        "Revoke unused OAuth grants and disable apps no longer needed."
    ),
    "MS Purview — IB Segments": (
        "Verify Information Barrier segments: Purview portal > Information barriers > Segments. "
        "PowerShell: Get-InformationBarrierSegment | Select Name,UserGroupFilter. "
        "Ensure all segments have correct member filters. Check active policies: Get-InformationBarrierPolicy | Where-Object State -eq 'Active'. "
        "Verify enforcement: compliance status should show 'Applied' for all targeted users. "
        "IB prevents Copilot from surfacing content across barrier boundaries (e.g. between departments with Chinese walls)."
    ),
    "MS Purview — Label Policy Scope": (
        "Expand mandatory labeling scope: Purview portal > Information protection > Label policies > select policy > Edit. "
        "Under 'Apply this label by default' and 'Require users to apply a label', enable for all M365 workloads: "
        "Word, Excel, PowerPoint, Outlook, Teams, SharePoint, OneDrive. "
        "PowerShell: Set-LabelPolicy -Identity '<policy-name>' -AdvancedSettings @{mandatory='true'} -ExchangeLocation 'All'. "
        "Ensure Exchange Online, Teams messages, and SharePoint/OneDrive files all require a sensitivity label before saving."
    ),
    # ── Phase 3 — Compliance & Regulatory ──
    "MS Purview — Compliance Manager": (
        "Map to regulatory frameworks: Purview portal > Compliance Manager > Assessments > Add assessment. "
        "Select applicable templates: ISO 27001, SOC 2, GDPR, HIPAA, NIST 800-53, etc. "
        "Review Copilot-relevant improvement actions under each framework. "
        "Map Copilot security controls to framework requirements using Compliance Manager's built-in mapping. "
        "Track progress via the compliance score dashboard and assign action owners for each improvement action."
    ),
    "MS Trust Center — Multi-Geo": (
        "Verify Multi-Geo compliance: Microsoft 365 admin center > Settings > Org settings > Organization profile > Data location. "
        "PowerShell: Get-SPOGeoStorageQuota to view data distribution across geos. "
        "Ensure Copilot processing and data storage align with your data residency requirements. "
        "For EU Data Boundary: verify tenant is enrolled in EU Data Boundary program at https://admin.microsoft.com > Settings > Org settings. "
        "Review Copilot data processing locations in your Microsoft Product Terms."
    ),
    "MS Admin — Copilot License Offboarding": (
        "Create offboarding checklist: When removing Copilot access, revoke the license within 24 hours of role change. "
        "PowerShell: Set-MgUserLicense -UserId <user-id> -RemoveLicenses @('<copilot-sku-id>') -AddLicenses @(). "
        "If using Group-Based Licensing: remove the user from the Copilot deployment group. "
        "Review and revoke any Copilot agent connections or Power Platform access for the departing user. "
        "Ensure offboarding is part of your HR/IT workflow — add to existing identity lifecycle automation."
    ),
    "MS Responsible AI": (
        "Establish Responsible AI policy: Review Microsoft's Responsible AI Standard at https://www.microsoft.com/en-us/ai/responsible-ai. "
        "Create an internal AI Use Policy covering: acceptable use cases, prohibited uses, data handling, privacy expectations, and review cadence. "
        "Communicate the policy to all Copilot-licensed users. Include in onboarding training. "
        "Establish an AI ethics review process for new Copilot agents and custom solutions. "
        "Designate a Responsible AI champion or committee to review incidents and policy updates quarterly."
    ),
    "MS Admin — Copilot Usage Reports": (
        "Enable usage analytics: Microsoft 365 admin center > Reports > Usage > Microsoft 365 Copilot. "
        "Review adoption metrics: active users, feature usage (Word, Excel, Teams, Outlook), and usage trends. "
        "PowerShell: Get-MgReportM365AppUserDetail -Period 'D30' for app-level detail. "
        "Cross-reference usage with license assignment to identify unused licenses for re-allocation. "
        "Set up recurring usage review cadence (monthly) and share dashboards with stakeholders."
    ),
    # ── Phase 4 — Advanced Security Operations ──
    "MS SharePoint — Blast Radius": (
        "Assess permission blast radius: SharePoint Admin Center > Sites > Active sites > select site > Permissions. "
        "PowerShell: Get-SPOSite -Limit All | ForEach-Object { Get-SPOSiteGroup -Site $_.Url | Where-Object { $_.Users -match 'everyone' } }. "
        "Identify sites where 'Everyone' or 'Everyone except external users' has access — these represent maximum blast radius. "
        "Prioritize sites containing sensitive/confidential data. Use DSPM for AI oversharing reports to quantify exposure. "
        "Create a remediation plan: remove broad groups, replace with specific security groups, apply sensitivity labels."
    ),
    "MS Purview — Prompt Monitoring": (
        "Monitor Copilot prompts: Purview portal > Audit > search for 'CopilotInteraction' activities. "
        "PowerShell: Search-UnifiedAuditLog -RecordType CopilotInteraction -StartDate (Get-Date).AddDays(-7) -EndDate (Get-Date). "
        "Look for patterns: repeated prompts targeting sensitive topics, attempts to access restricted data, unusual query volumes. "
        "Configure alert policies: Purview > Alert policies > New alert > Activity: CopilotInteraction > configure thresholds. "
        "Cross-reference with DLP policy matches and Insider Risk alerts for correlated investigation."
    ),
    "MS SharePoint — External Sharing Governance": (
        "Create external sharing scorecard: SharePoint Admin Center > Policies > Sharing. "
        "PowerShell: Get-SPOSite -Limit All | Select Url,SharingCapability,SharingAllowedDomainList | "
        "Group-Object SharingCapability | Select Name,Count. "
        "Target: reduce 'ExternalUserAndGuestSharing' to minimum sites. Set org default to 'New and existing guests'. "
        "Enable sharing expiration: Set-SPOTenant -DefaultSharingLinkType DirectAccess -RequireAnonymousLinksExpireInDays 30. "
        "Review quarterly: generate sharing activity reports and revoke stale external access."
    ),
    # ── Content Lifecycle ──
    "MS Purview — Legal Hold": (
        "Verify legal hold compatibility: Purview portal > eDiscovery > Cases > review active holds. "
        "PowerShell: Get-CaseHoldPolicy | Get-CaseHoldRule | Select Name,ContentMatchQuery,Disabled. "
        "Ensure Copilot interaction data (audit logs, chat transcripts) is included in preservation scope. "
        "Create a new hold for Copilot data if not covered: eDiscovery > New case > Hold > add Exchange and SharePoint locations. "
        "Copilot responses stored in user mailboxes and SharePoint are subject to existing holds — verify coverage."
    ),
    # ── Phase 5 — App Protection, Label Encryption, RCD ──
    "MS Intune — App Protection Policies": (
        "Configure App Protection: Microsoft Intune admin center > Apps > App protection policies > Create policy. "
        "Select platform (iOS/Android). Target: M365 apps (Outlook, Teams, Word, Excel, PowerPoint, OneDrive). "
        "Settings: Require PIN/biometric, encrypt app data, block copy/paste to unmanaged apps, block screenshots. "
        "PowerShell: Use Microsoft Graph — New-MgDeviceAppManagementManagedAppPolicy. "
        "Assign to Copilot-licensed users group. This prevents Copilot-accessed content from leaking to unmanaged apps on mobile."
    ),
    "MS Purview — Label Encryption": (
        "Configure label encryption: Purview portal > Information protection > Labels > select label > Encryption. "
        "Enable 'Apply encryption' and configure permissions: assign specific users/groups with View, Edit, or Full Control. "
        "PowerShell: Set-Label -Identity '<label-name>' -EncryptionEnabled $true -EncryptionProtectionType Template. "
        "Verify site/group container settings: Labels > Container labels > select label > configure site-level settings. "
        "Note: encryption restricts Copilot's ability to process content — use intentionally for highly sensitive data only."
    ),
    "MS SharePoint — Restricted Content Discoverability": (
        "Enable RCD: SharePoint Admin Center > Settings > Restricted Content Discoverability. "
        "PowerShell: Set-SPOTenant -RestrictedContentDiscoverability $true. "
        "Add specific sites to the restricted list: Set-SPOSite -Identity <site-url> -RestrictContentOrgWideSearch $true. "
        "RCD prevents content from appearing in organization-wide search results and Copilot grounding for non-members. "
        "Use for sensitive sites (HR, legal, M&A) that should not appear in Copilot responses for unauthorized users."
    ),
    # ── Phase 6 — Checklist Gap Closure ──
    "MS Entra — Cross-Tenant Access": (
        "Review cross-tenant access: Entra admin center > External Identities > Cross-tenant access settings. "
        "Audit inbound and outbound trust configurations. Verify B2B collaboration policies. "
        "PowerShell: Get-MgPolicyCrossTenantAccessPolicyPartner | Select TenantId,InboundTrust,OutboundTrust. "
        "Ensure Copilot does not inadvertently surface content shared via cross-tenant collaboration. "
        "Restrict B2B access to named partner tenants only — block all by default, allow explicitly."
    ),
    "MS Entra — Hybrid Identity": (
        "Monitor hybrid identity sync: Entra admin center > Hybrid management > Entra Connect > Connect Sync. "
        "Check sync status, last sync time, and error count. PowerShell: Get-MgDirectorySyncStatus (via Graph). "
        "Verify password hash sync or pass-through authentication is configured correctly. "
        "Ensure on-premises directory clean-up: remove duplicate accounts, fix UPN mismatches, resolve sync errors. "
        "Unhealthy sync can cause stale identities to retain Copilot access after on-premises disablement."
    ),
    "MS Admin — Copilot License Segmentation": (
        "Segment Copilot licenses by business unit: Microsoft 365 admin center > Billing > Licenses > assign via groups. "
        "Create separate groups per department or sensitivity tier. "
        "Consider phased rollout: start with a pilot group of 50-100 users, expand after verifying security posture. "
        "PowerShell: New-MgGroup -DisplayName 'Copilot-Pilot' -MailEnabled:$false -SecurityEnabled:$true; then assign licenses to group. "
        "Track license utilization per segment via usage reports."
    ),
    "MS Copilot — Agent Approval": (
        "Configure agent approval workflow: Microsoft 365 admin center > Settings > Copilot > Agent governance. "
        "Enable 'Require admin approval before agents are published'. "
        "Set up a review process: designate approvers, define approval criteria (data scope, permissions, sensitivity). "
        "Power Platform admin center > Settings > Features > toggle 'Agent publishing requires approval'. "
        "Review pending agent submissions weekly. Reject agents with overly broad data source connections."
    ),
    "MS Graph — External Connectors": (
        "Audit external connectors: Microsoft 365 admin center > Settings > Search & intelligence > Connectors. "
        "Review all Microsoft Graph connectors and their data source permissions. "
        "PowerShell: Get-MgExternalConnection | Select Name,State,Description. "
        "Verify each connector indexes only approved content. Check connector permissions and service principal scopes. "
        "Disable unused connectors — indexed external content is available to Copilot via Microsoft Search."
    ),
    "MS Defender — Copilot Incidents": (
        "Monitor Copilot security incidents: Microsoft Defender portal > Incidents & alerts > filter for 'Copilot'. "
        "Set up detection rules: Defender > Custom detections > create rule targeting CopilotInteraction audit events. "
        "PowerShell: Search-UnifiedAuditLog -RecordType CopilotInteraction | Where-Object { $_.Operations -match 'Suspicious' }. "
        "Configure automated response: assign incidents to SOC team, set severity-based escalation. "
        "Integrate with Microsoft Sentinel for cross-signal correlation with identity and endpoint alerts."
    ),
    "MS Copilot — Prompt Guardrails": (
        "Configure prompt guardrails: Microsoft 365 admin center > Settings > Copilot > Safety and security. "
        "Enable content filtering for sensitive topics. Configure organizational prohibited topics list. "
        "Review guardrail effectiveness via audit logs: search for blocked or modified Copilot responses. "
        "Set up Responsible AI monitoring dashboards. "
        "Train users on acceptable prompt patterns and escalation procedures for guardrail blocks."
    ),
    # ── Zero Trust ──
    "MS Entra — Continuous Access Evaluation": (
        "Enable CAE: Entra admin center > Protection > Conditional Access > select policy > Session > "
        "Continuous access evaluation > 'Strictly enforce location policies'. "
        "CAE enables near-real-time token revocation when user risk changes, IP changes, or account is disabled. "
        "Verify: Get-MgIdentityConditionalAccessPolicy | Where-Object { $_.SessionControls.ContinuousAccessEvaluation } | Select DisplayName. "
        "Critical for Copilot: ensures revoked users lose access to Copilot immediately rather than waiting for token expiry."
    ),
    "MS Entra — Token Protection": (
        "Enable Token Protection: Entra admin center > Protection > Conditional Access > create new policy > "
        "Grant > Require token protection for sign-in sessions (Preview). "
        "Token binding cryptographically ties tokens to the device, preventing token theft and replay attacks. "
        "Currently supported for Windows sign-in sessions on specific apps. "
        "Monitor: Entra sign-in logs > filter for tokenProtectionStatus. Blocks access from stolen tokens used on different devices."
    ),
    "MS Entra — Phishing-Resistant MFA": (
        "Deploy phishing-resistant MFA: Entra admin center > Protection > Authentication methods > "
        "enable FIDO2 security keys and/or Windows Hello for Business. "
        "Create CA policy: Protection > Conditional Access > new policy > Grant > Require authentication strength > 'Phishing-resistant MFA'. "
        "PowerShell: New-MgIdentityConditionalAccessPolicy with grantControls.authenticationStrength.id = '<phishing-resistant-id>'. "
        "Target Copilot users accessing sensitive data. Migrating from SMS/phone MFA to FIDO2/WHfB eliminates phishing risk."
    ),
    "MS Entra — Authentication Context": (
        "Configure authentication context: Entra admin center > Protection > Authentication context > New authentication context. "
        "Create context for sensitive actions (e.g. 'Access confidential data via Copilot'). "
        "Link to Conditional Access policy: require step-up authentication when this context is triggered. "
        "Apply to sensitivity labels: Purview > Labels > select label > configure authentication context requirement. "
        "This enforces additional verification when Copilot accesses content tagged with high-sensitivity labels."
    ),
    "MS Entra — Workload Identities": (
        "Protect workload identities: Entra admin center > Applications > Enterprise applications > audit service principals. "
        "PowerShell: Get-MgServicePrincipal -All | Where-Object { $_.AppOwnerOrganizationId -ne '<your-tenant-id>' } | Select DisplayName,AppId. "
        "Review all multi-tenant and third-party workload identities. Apply Conditional Access for workload identities (Preview). "
        "Create CA policy targeting workload identities: Protection > Conditional Access > assign to 'Workload identities'. "
        "Monitor: Entra > Identity Protection > Risky workload identities for anomalous app behavior."
    ),
    "MS Entra — Global Secure Access": (
        "Configure Compliant Network Check: Entra admin center > Global Secure Access > Connect > Traffic forwarding. "
        "Enable Microsoft 365 traffic profile. Create Conditional Access policy: Grant > Require compliant network. "
        "This ensures only traffic routed through Global Secure Access (formerly Entra Internet Access) can reach M365 services. "
        "PowerShell: Use Graph API to configure trafficForwardingProfiles. "
        "Blocks Copilot access from unmanaged networks while allowing corporate and GSA-tunneled connections."
    ),
    # ── Shadow AI ──
    "MS Entra — App Registrations": (
        "Detect unauthorized AI apps: Entra admin center > Applications > App registrations > All applications. "
        "Filter for apps with AI-related names or Microsoft Graph permissions (e.g. Sites.Read.All, Chat.Read). "
        "PowerShell: Get-MgApplication -All | Where-Object { $_.DisplayName -match 'AI|GPT|Copilot|OpenAI|Claude|Gemini' } | Select DisplayName,AppId. "
        "Review API permissions granted to flagged apps. Revoke admin consent for unauthorized AI applications. "
        "Configure: Entra > User settings > 'Users can consent to apps' = No — require admin consent for new apps."
    ),
    "MS Entra — Consent Grants": (
        "Audit OAuth consent grants: Entra admin center > Applications > Enterprise applications > Consent and permissions > Admin consent requests. "
        "PowerShell: Get-MgOauth2PermissionGrant -All | Where-Object { $_.ConsentType -eq 'AllPrincipals' } | "
        "Select ClientId,ResourceId,Scope | ForEach-Object { [PSCustomObject]@{App=(Get-MgServicePrincipal -ServicePrincipalId $_.ClientId).DisplayName; Scope=$_.Scope} }. "
        "Revoke overly broad grants. Focus on apps with Mail.Read, Files.Read.All, Chat.Read, Sites.Read.All. "
        "Enable: Entra > User settings > Admin consent workflow > require admin approval for new consent grants."
    ),
    "MS Copilot — Agent Governance": (
        "Audit shadow Copilot agents: Microsoft 365 admin center > Settings > Copilot > Agents. "
        "Review agents created outside official IT governance. Check Power Platform admin center for unauthorized Copilot Studio agents. "
        "PowerShell (Power Platform): Get-AdminPowerApp | Where-Object { $_.DisplayName -match 'Copilot|Agent' }. "
        "Establish agent creation policies: restrict who can publish to production. "
        "Enable DLP policies in Power Platform to prevent agents from connecting to unauthorized data sources."
    ),
    "MS Defender — App Governance": (
        "Configure App Governance: Microsoft Defender portal > Cloud Apps > App governance. "
        "Enable monitoring for OAuth-enabled apps with high-privilege permissions. "
        "Create governance policies: Cloud Apps > Policies > App governance > Create new policy > "
        "target apps with Graph permissions like Mail.ReadWrite, Sites.FullControl, or any AI-service permissions. "
        "Set alerts for new apps requesting sensitive permissions. Review and remediate policy violations weekly."
    ),
}


def _effort_badge(effort: str) -> str:
    """Render effort estimate as a styled badge."""
    colors = {"quick_win": "#107C10", "moderate": "#FFB900", "major": "#D13438"}
    labels = {"quick_win": "Quick Win", "moderate": "Moderate", "major": "Major"}
    color = colors.get(effort, "#A8A6A3")
    label = labels.get(effort, effort.replace("_", " ").title())
    return f'<span style="display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600;background:{color}20;color:{color};border:1px solid {color}40">{label}</span>'


def _tip(text: str, help_text: str) -> str:
    """Wrap text in a span with a hover tooltip."""
    if not help_text:
        return text
    return f'<span class="tip" data-tip="{esc(help_text)}">{text}</span>'


def _th_tip(label: str, help_text: str, style: str = "") -> str:
    """Render a <th> with a tooltip on the header text."""
    style_attr = f' style="{style}"' if style else ""
    return f'<th{style_attr}>{_tip(label, help_text)}</th>'


def _copilot_risk_pill(level: str, reason: str = "", why: str = "", how: str = "") -> str:
    """Render a clickable Copilot Risk pill that opens a modal with Why / How details."""
    colors = {
        "high": ("#EF5350", "rgba(209,52,56,.15)", "rgba(209,52,56,.3)"),
        "medium": ("#FFB300", "rgba(255,185,0,.15)", "rgba(255,185,0,.3)"),
        "low": ("#4CAF50", "rgba(16,124,16,.15)", "rgba(16,124,16,.3)"),
    }
    fg, bg, border = colors.get(level, colors["low"])
    title_attr = f' title="{esc(reason)} — click for details"' if reason else ""
    data_attrs = ""
    if why or how:
        data_attrs = (
            f' data-level="{esc(level)}" data-why="{esc(why)}" data-how="{esc(how)}"'
            f' onclick="showRiskDetail(this)"'
        )
    return (
        f'<span class="status-pill" style="background:{bg};color:{fg};border:1px solid {border}'
        f'{(";cursor:pointer" if data_attrs else "")}"{title_attr}{data_attrs}>'
        f'{level.upper()}</span>'
    )


def _risk_why_how(why: str, how: str) -> tuple[str, str]:
    """Legacy helper — now returns empty strings since Why/How moved to modal."""
    return "", ""


def _inject_risk_modal(risk_html: str, why: str, how: str) -> str:
    """Inject modal data attributes into an existing risk pill for clickable Why/How popup."""
    if not why:
        return risk_html
    for tag in ('HIGH', 'MEDIUM', 'LOW'):
        if f'>{tag}<' in risk_html:
            level = tag.lower()
            attrs = (
                f' data-level="{esc(level)}" data-why="{esc(why)}" data-how="{esc(how)}"'
                f' onclick="showRiskDetail(this)"'
            )
            return risk_html.replace(f'>{tag}<', f'{attrs}>{tag}<', 1)
    return risk_html


def _domain_pill(cat_key: str) -> str:
    """Render a compact domain pill using _CATEGORY_META colours."""
    meta = _CATEGORY_META.get(cat_key)
    if not meta:
        return ""
    c = meta["color"]
    return (
        f'<span style="display:inline-block;padding:2px 7px;border-radius:4px;font-size:10px;'
        f'font-weight:600;background:{c}18;color:{c};border:1px solid {c}30;'
        f'white-space:nowrap">{meta["icon"]} {meta["name"]}</span>'
    )


def _domain_pills(cat_keys: list[str]) -> str:
    """Render one or more domain pills side-by-side."""
    return " ".join(_domain_pill(k) for k in cat_keys if k)


# ── SVG bar chart ────────────────────────────────────────────────────────

def _bar_chart_svg(items: list[tuple[str, int, str]], width: int = 480, bar_h: int = 22) -> str:
    if not items:
        return '<p class="empty">No data</p>'
    max_val = max(v for _, v, _ in items) or 1
    gap = 6
    h = (bar_h + gap) * len(items) + 10
    parts = [f'<svg width="{width}" height="{h}" role="img" aria-label="Bar chart">']
    for i, (label, val, color) in enumerate(items):
        y = i * (bar_h + gap) + 4
        bw = int((val / max_val) * (width - 200))
        parts.append(
            f'<text x="0" y="{y + bar_h - 5}" font-size="12" fill="var(--text-secondary)" '
            f'font-family="var(--font-primary)">{esc(label)}</text>'
        )
        parts.append(
            f'<rect x="130" y="{y}" width="{max(bw, 2)}" height="{bar_h}" rx="3" fill="{color}" opacity="0.85">'
            f'<title>{esc(label)}: {val}</title></rect>'
        )
        parts.append(
            f'<text x="{135 + bw}" y="{y + bar_h - 5}" font-size="12" fill="var(--text)" '
            f'font-family="var(--font-mono)">{val}</text>'
        )
    parts.append("</svg>")
    return "".join(parts)


# ── Rendering helpers ────────────────────────────────────────────────────

def _render_finding(f: dict) -> str:
    sev = f.get("Severity", "medium").lower()
    cat = f.get("Category", "unknown")
    subcat = f.get("Subcategory", "")
    title = f.get("Title", "Untitled")
    desc = f.get("Description", "")
    affected = f.get("AffectedCount", 0)
    affected_resources = f.get("AffectedResources", [])
    remediation = f.get("Remediation", {})
    compliance_status = f.get("ComplianceStatus", "gap")

    rem_html = ""
    if remediation:
        rem_parts = []
        rem_desc = remediation.get("Description", "")
        if rem_desc:
            rem_parts.append(f'<div class="rem-desc">{esc(rem_desc)}</div>')
        for key, label in [("AzureCLI", "Azure CLI"), ("PowerShell", "PowerShell")]:
            cmd = remediation.get(key, "")
            if cmd:
                rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">{label}:</strong><pre>{esc(cmd)}</pre></div>')
        steps = remediation.get("PortalSteps", [])
        if steps:
            step_items = "".join(f"<li>{esc(s)}</li>" for s in steps)
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">Portal Steps:</strong><ol class="portal-steps">{step_items}</ol></div>')
        if rem_parts:
            rem_html = f'<div class="remediation-box"><h4>&#128736; Remediation</h4>{"".join(rem_parts)}</div>'

    affected_html = ""
    if affected_resources:
        _STD = {"Name", "name", "Type", "type", "ResourceId", "resource_id", "id"}
        detail_keys = [k for k in (affected_resources[0] if affected_resources else {}) if k not in _STD] if affected_resources else []
        header = '<th>Resource</th><th>Type</th><th>ID</th>'
        for dk in detail_keys:
            header += f'<th>{esc(dk.replace("_", " ").title())}</th>'
        rows = ""
        for ar in affected_resources[:20]:
            if isinstance(ar, dict):
                rows += (f'<tr><td class="res-name">{esc(str(ar.get("Name", ar.get("name", "—"))))}</td>'
                         f'<td class="res-type">{esc(str(ar.get("Type", ar.get("type", "—"))))}</td>'
                         f'<td class="res-id">{esc(str(ar.get("ResourceId", ar.get("resource_id", "—"))))}</td>')
                for dk in detail_keys:
                    rows += f'<td class="res-detail">{esc(str(ar.get(dk, "—")))}</td>'
                rows += '</tr>'
        if len(affected_resources) > 20:
            rows += f'<tr class="more-row"><td colspan="{3+len(detail_keys)}">… and {len(affected_resources)-20} more</td></tr>'
        affected_html = (
            f'<details class="affected-details" open>'
            f'<summary>{_severity_badge(sev)} {affected} affected</summary>'
            f'<table class="resource-table"><thead><tr>{header}</tr></thead><tbody>{rows}</tbody></table></details>'
        )

    cat_meta = _CATEGORY_META.get(cat, {})
    cat_name = cat_meta.get("name", cat.replace("_", " ").title())

    return (
        f'<div class="finding-card {sev}" data-severity="{sev}" data-category="{esc(cat)}" '
        f'data-subcategory="{esc(subcat)}" data-affected="{affected}">'
        f'<div class="finding-title">{_severity_badge(sev)} {_compliance_badge(compliance_status)} {esc(title)}</div>'
        f'<div class="finding-meta">'
        f'<span>&#128193; {esc(cat_name)}</span>'
        f'<span>&#128196; {esc(subcat.replace("_", " ").title())}</span>'
        f'<span>&#128202; {affected} affected</span>'
        f'</div>'
        f'<div class="finding-desc">{esc(desc)}</div>'
        f'{affected_html}{rem_html}'
        f'</div>'
    )


def _render_finding_row(idx: int, f: dict) -> str:
    """Render a finding as a clickable summary row + hidden detail row for collapsible table."""
    sev = f.get("Severity", "medium").lower()
    cat = f.get("Category", "unknown")
    subcat = f.get("Subcategory", "")
    effort = f.get("Effort", "moderate")
    title = f.get("Title", "")
    desc = f.get("Description", "")
    affected = f.get("AffectedCount", 0)
    affected_resources = f.get("AffectedResources", [])
    remediation = f.get("Remediation", {})
    cat_meta = _CATEGORY_META.get(cat, {})
    cat_name = cat_meta.get("name", cat.replace("_", " ").title())

    # Build detail content
    detail_parts: list[str] = []
    if desc:
        detail_parts.append(
            f'<div style="font-size:13px;color:var(--text-secondary);line-height:1.6;margin-bottom:10px">{esc(desc)}</div>'
        )

    if affected_resources:
        _STD = {"Name", "name", "Type", "type", "ResourceId", "resource_id", "id"}
        detail_keys = [k for k in (affected_resources[0] if affected_resources else {}) if k not in _STD]
        header = '<th>Resource</th><th>Type</th><th>ID</th>'
        for dk in detail_keys:
            header += f'<th>{esc(dk.replace("_", " ").title())}</th>'
        res_rows = ""
        for ar in affected_resources[:20]:
            if isinstance(ar, dict):
                res_rows += (
                    f'<tr><td class="res-name">{esc(str(ar.get("Name", ar.get("name", "—"))))}</td>'
                    f'<td class="res-type">{esc(str(ar.get("Type", ar.get("type", "—"))))}</td>'
                    f'<td class="res-id">{esc(str(ar.get("ResourceId", ar.get("resource_id", "—"))))}</td>'
                )
                for dk in detail_keys:
                    res_rows += f'<td class="res-detail">{esc(str(ar.get(dk, "—")))}</td>'
                res_rows += '</tr>'
        if len(affected_resources) > 20:
            res_rows += f'<tr class="more-row"><td colspan="{3 + len(detail_keys)}">… and {len(affected_resources) - 20} more</td></tr>'
        detail_parts.append(
            f'<div style="margin:8px 0"><strong style="font-size:12px;color:var(--text-secondary)">'
            f'Affected Resources ({affected})</strong>'
            f'<table class="resource-table"><thead><tr>{header}</tr></thead>'
            f'<tbody>{res_rows}</tbody></table></div>'
        )

    if remediation:
        rem_parts: list[str] = []
        rem_desc = remediation.get("Description", "")
        if rem_desc:
            rem_parts.append(f'<div class="rem-desc">{esc(rem_desc)}</div>')
        for key, label in [("AzureCLI", "Azure CLI"), ("PowerShell", "PowerShell")]:
            cmd = remediation.get(key, "")
            if cmd:
                rem_parts.append(
                    f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">'
                    f'{label}:</strong><pre>{esc(cmd)}</pre></div>'
                )
        steps = remediation.get("PortalSteps", [])
        if steps:
            step_items = "".join(f"<li>{esc(s)}</li>" for s in steps)
            rem_parts.append(
                f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">'
                f'Portal Steps:</strong><ol class="portal-steps">{step_items}</ol></div>'
            )
        if rem_parts:
            detail_parts.append(
                f'<div class="remediation-box"><h4>&#128736; Remediation</h4>'
                f'{"".join(rem_parts)}</div>'
            )

    detail_html = "".join(detail_parts)

    rem_text = esc(remediation.get("Description", "")) if remediation else ""

    summary = (
        f'<tr class="finding-summary-row {sev}" data-severity="{sev}" '
        f'data-category="{esc(cat)}" data-subcategory="{esc(subcat)}" '
        f'onclick="toggleFindingDetail(this)" style="cursor:pointer">'
        f'<td style="text-align:center;font-family:var(--font-mono);font-weight:700;color:var(--text-muted)">{idx}</td>'
        f'<td>{_severity_badge(sev)}</td>'
        f'<td>{_effort_badge(effort)}</td>'
        f'<td style="font-weight:600;font-size:12px">{esc(title)}</td>'
        f'<td style="font-size:11px;color:var(--text-secondary)">{esc(cat_name)}</td>'
        f'<td style="text-align:center;font-family:var(--font-mono);font-size:12px">{affected}</td>'
        f'<td style="font-size:11px;color:var(--text-secondary);max-width:280px">{rem_text}</td>'
        f'</tr>'
    )

    detail = (
        f'<tr class="finding-detail-row" style="display:none">'
        f'<td colspan="7"><div class="finding-detail-content">{detail_html}</div></td></tr>'
    )

    return summary + detail


# ── Report-specific CSS ─────────────────────────────────────────────────

def _cr_css() -> str:
    return """
.top-nav{position:sticky;top:0;z-index:500;display:flex;align-items:center;gap:16px;padding:8px 24px;
  background:var(--bg-elevated);border-bottom:1px solid var(--border);font-size:13px;flex-wrap:wrap}
.top-nav .brand{font-weight:700;color:var(--primary);font-size:14px;margin-right:12px}
.top-nav a{color:var(--text-secondary);text-decoration:none;padding:6px 10px;border-radius:6px;transition:all .2s}
.top-nav a:hover{color:var(--text);background:var(--bg-card)}
.nav-dropdown{position:relative;display:inline-block}
.nav-dropdown>a::after{content:'\\25BE';font-size:9px;margin-left:4px;opacity:.5}
.nav-dropdown-menu{display:none;position:absolute;top:100%;left:0;min-width:220px;background:var(--bg-elevated);border:1px solid var(--border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.35);padding:6px 0;z-index:510;max-height:60vh;overflow-y:auto}
.nav-dropdown:hover .nav-dropdown-menu{display:block}
.nav-dropdown-menu a{display:block;padding:6px 14px;font-size:12px;color:var(--text-secondary);border-radius:0;white-space:nowrap}
.nav-dropdown-menu a:hover{color:var(--text);background:var(--bg-card)}
.full-width-content{padding:32px 40px;max-width:1200px;margin:0 auto}
.exec-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;margin:16px 0}
.exec-panel{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px}
.exec-panel h3{font-size:14px;color:var(--text-secondary);margin-bottom:12px;border:none;padding:0}
.category-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin:12px 0}
.category-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px;text-align:center;transition:all .3s;cursor:pointer}
a.category-card-link{text-decoration:none;color:inherit;display:block}
.category-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md)}
.category-icon{font-size:32px;margin-bottom:8px}
.category-name{font-size:13px;color:var(--text-secondary);margin-bottom:4px}
.category-score{font-size:28px;font-weight:700;font-family:var(--font-mono)}
.category-level{font-size:11px;text-transform:uppercase;font-weight:600;letter-spacing:.5px;margin-top:2px}
.category-findings{font-size:11px;color:var(--text-muted);margin-top:4px}
.score-display{display:flex;align-items:center;gap:40px;flex-wrap:wrap;margin:20px 0}
.level-badge{display:inline-block;padding:4px 12px;border-radius:6px;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.sev-bars{display:flex;flex-direction:column;gap:8px;margin:12px 0}
.sev-row{display:flex;align-items:center;gap:12px}
.sev-label{width:70px;font-size:12px;text-transform:uppercase;color:var(--text-secondary);font-weight:600}
.sev-track{flex:1;height:10px;background:var(--bar-bg);border-radius:5px;overflow:hidden}
.sev-fill{height:100%;border-radius:5px;transition:width .6s ease}
.sev-count{width:30px;text-align:right;font-family:var(--font-mono);font-size:13px}
.finding-card{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:20px;margin-bottom:12px;transition:all .2s}
.finding-card:hover{background:var(--bg-card-hover)}
.finding-card.critical{border-left:4px solid #D13438}
.finding-card.high{border-left:4px solid #F7630C}
.finding-card.medium{border-left:4px solid #FFB900}
.finding-card.low{border-left:4px solid #107C10}
.finding-card.informational{border-left:4px solid #A8A6A3}
.finding-title{font-size:15px;font-weight:600;margin-bottom:6px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.finding-desc{font-size:13px;color:var(--text-secondary);line-height:1.6;margin-bottom:8px}
.finding-meta{display:flex;gap:16px;flex-wrap:wrap;font-size:12px;color:var(--text-muted);margin-bottom:8px}
.remediation-box{margin-top:10px;padding:14px;background:var(--remediation-bg);border-left:3px solid var(--remediation-border);border-radius:6px}
.remediation-box h4{font-size:12px;color:var(--remediation-border);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}
.remediation-box .rem-desc{font-size:13px;color:#A5D6A7;margin-bottom:8px;line-height:1.5}
.remediation-box pre{font-family:var(--font-mono);font-size:12px;background:var(--code-bg);border:1px solid var(--code-border);border-radius:4px;padding:10px;overflow-x:auto;color:var(--text);white-space:pre-wrap}
.remediation-box .portal-steps{margin:6px 0 0;padding-left:20px;font-size:12px;color:var(--text-secondary)}
.remediation-box .portal-steps li{margin-bottom:3px}
.affected-details summary{cursor:pointer;color:var(--primary);font-weight:500;font-size:12px;padding:6px 0}
.resource-table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;margin:8px 0;border:1px solid var(--border);border-radius:8px;overflow:visible}
.resource-table thead tr:first-child th:first-child{border-top-left-radius:8px}.resource-table thead tr:first-child th:last-child{border-top-right-radius:8px}
.resource-table tbody tr:last-child td:first-child{border-bottom-left-radius:8px}.resource-table tbody tr:last-child td:last-child{border-bottom-right-radius:8px}
.resource-table thead{background:var(--bg-elevated)}
.resource-table th{padding:8px 12px;text-align:left;font-weight:600;color:var(--text-secondary);text-transform:uppercase;font-size:11px;border-bottom:2px solid var(--border)}
.resource-table td{padding:8px 12px;border-bottom:1px solid var(--border-light,var(--border));color:var(--text)}
.resource-table .res-name{font-weight:600}.resource-table .res-type{font-size:11px;color:var(--text-muted)}
.resource-table .res-id{font-family:var(--font-mono);font-size:11px;color:var(--text-secondary);word-break:break-all;max-width:320px}
.resource-table .res-detail{font-size:11px;color:var(--text-secondary)}
.filter-bar{display:flex;align-items:center;gap:8px;margin-bottom:16px;flex-wrap:wrap;font-size:13px}
.filter-bar input[type="search"]{min-width:240px;padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar select{padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}

.readiness-banner{text-align:center;padding:24px;border-radius:12px;margin:16px 0}
.readiness-banner .status{font-size:48px;font-weight:700;font-family:var(--font-mono)}
.top-finding-row{display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid var(--border-light)}
.top-finding-row:last-child{border-bottom:none}
.top-finding-rank{font-size:20px;font-weight:700;font-family:var(--font-mono);color:var(--text-muted);width:30px;text-align:center}
.top-finding-info{flex:1}
.top-finding-title{font-size:13px;font-weight:600}
.top-finding-meta{font-size:11px;color:var(--text-muted);margin-top:2px}
.zoom-controls{display:flex;align-items:center;gap:4px;margin-left:auto}
.zoom-controls button{padding:4px 10px;border:1px solid var(--border);border-radius:4px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:14px;min-height:32px;transition:all .2s}
.zoom-controls button:hover{border-color:var(--primary);color:var(--primary)}
#zoom-label{font-size:12px;font-family:var(--font-mono);width:40px;text-align:center}
.export-bar{display:flex;gap:4px}
/* Security Controls Table */
.controls-table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;border:1px solid var(--border);border-radius:10px;overflow:hidden;table-layout:auto;word-wrap:break-word}
.controls-table thead tr:first-child th:first-child{border-top-left-radius:10px}.controls-table thead tr:first-child th:last-child{border-top-right-radius:10px}
.controls-table tbody tr:last-child td:first-child{border-bottom-left-radius:10px}.controls-table tbody tr:last-child td:last-child{border-bottom-right-radius:10px}
.controls-table thead{background:var(--bg-elevated)}
.controls-table th{padding:8px 10px;text-align:left;font-weight:700;color:var(--text-secondary);text-transform:uppercase;font-size:11px;letter-spacing:.3px;border-bottom:2px solid var(--border);white-space:nowrap}
.controls-table td{padding:8px 10px;border-bottom:1px solid var(--border-light,var(--border));color:var(--text);vertical-align:top;white-space:normal;word-wrap:break-word;overflow-wrap:break-word}
.controls-table td:first-child{max-width:220px}
.controls-table tbody tr{transition:background .15s}
.controls-table tbody tr:hover{background:var(--bg-card-hover,rgba(255,255,255,.03))}
.controls-table .ctrl-id{font-family:var(--font-mono);font-size:11px;color:var(--text-muted);white-space:nowrap}
.controls-table .ctrl-name{font-weight:600;font-size:12px}
.controls-table .ctrl-desc{font-size:11px;color:var(--text-secondary);max-width:280px}
.controls-table .ctrl-resource{font-size:11px;color:var(--text-secondary)}
.controls-table .ctrl-ref{font-size:10px;color:var(--text-muted);font-style:italic}
.status-pill{display:inline-block;padding:2px 8px;border-radius:12px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.3px;white-space:nowrap}
.status-pill.pass{background:rgba(16,124,16,.15);color:#4CAF50;border:1px solid rgba(16,124,16,.3)}
.status-pill.fail{background:rgba(209,52,56,.15);color:#EF5350;border:1px solid rgba(209,52,56,.3)}
.status-pill.partial{background:rgba(255,185,0,.15);color:#FFB300;border:1px solid rgba(255,185,0,.3)}
.status-pill.not-assessed{background:rgba(168,166,163,.12);color:var(--text-muted);border:1px solid rgba(168,166,163,.3)}
.status-pill[onclick]{cursor:pointer}
.ctrl-cat-header td{background:var(--bg-elevated);font-weight:700;font-size:12px;color:var(--text);padding:6px 10px;border-bottom:2px solid var(--border)}
.ctrl-summary-bar{display:flex;gap:20px;margin:16px 0;flex-wrap:wrap}
.ctrl-summary-item{display:flex;align-items:center;gap:8px;font-size:13px}
.ctrl-summary-count{font-size:22px;font-weight:700;font-family:var(--font-mono)}
/* Findings summary table */
.findings-summary-table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;border:1px solid var(--border);border-radius:10px;overflow:visible;margin-bottom:20px}
.findings-summary-table thead tr:first-child th:first-child{border-top-left-radius:10px}.findings-summary-table thead tr:first-child th:last-child{border-top-right-radius:10px}
.findings-summary-table tbody tr:last-child td:first-child{border-bottom-left-radius:10px}.findings-summary-table tbody tr:last-child td:last-child{border-bottom-right-radius:10px}
.findings-summary-table thead{background:var(--bg-elevated)}
.findings-summary-table th{padding:10px 14px;text-align:left;font-weight:700;color:var(--text-secondary);text-transform:uppercase;font-size:11px;border-bottom:2px solid var(--border);white-space:nowrap}
.findings-summary-table td{padding:10px 14px;border-bottom:1px solid var(--border-light,var(--border));color:var(--text);vertical-align:top}
.findings-summary-table tbody tr{transition:background .15s}
.findings-summary-table tbody tr:hover{background:var(--bg-card-hover,rgba(255,255,255,.03))}
/* Section help hints */
.section-hint{font-size:12px;color:var(--text-muted);margin:-4px 0 16px;line-height:1.5}
/* Inventory tabs */
.inv-tabs{display:flex;gap:2px;border-bottom:2px solid var(--border);margin-bottom:16px;flex-wrap:wrap}
.inv-tab{padding:8px 16px;font-size:12px;font-weight:600;cursor:pointer;border:1px solid transparent;border-bottom:none;border-radius:6px 6px 0 0;background:transparent;color:var(--text-muted);transition:all .2s}
.inv-tab:hover{color:var(--text);background:var(--bg-card)}
.inv-tab.active{color:var(--primary);background:var(--bg-card);border-color:var(--border);border-bottom:2px solid var(--bg-card);margin-bottom:-2px}
.inv-tab .tab-count{font-size:10px;font-family:var(--font-mono);margin-left:4px;opacity:.7}
.inv-panel{display:block;margin-bottom:24px;scroll-margin-top:60px}
.inv-panel-header{font-size:14px;font-weight:700;color:var(--text);margin:0 0 8px;padding:8px 0 6px;border-bottom:1px solid var(--border)}
/* Legacy how-to-read (collection warnings) */
.how-to-read{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:12px 16px;margin-bottom:16px}
.how-to-read h4{margin:0 0 6px;font-size:13px;color:var(--text)}
.how-to-read p,.how-to-read li{font-size:12px;color:var(--text-secondary);line-height:1.5;margin-bottom:4px}
.how-to-read ul{margin:4px 0 6px 18px;padding:0}
/* Tighter section spacing */
.section{margin-bottom:28px!important}
h2{margin-bottom:6px}

/* Scroll progress bar */
.scroll-progress{position:fixed;top:0;left:0;height:3px;background:var(--primary);z-index:700;transition:width .1s;width:0}
/* Collapsible findings detail table */
.findings-detail-table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;border:1px solid var(--border);border-radius:10px;overflow:visible}
.findings-detail-table thead tr:first-child th:first-child{border-top-left-radius:10px}.findings-detail-table thead tr:first-child th:last-child{border-top-right-radius:10px}
.findings-detail-table tbody tr:last-child td:first-child{border-bottom-left-radius:10px}.findings-detail-table tbody tr:last-child td:last-child{border-bottom-right-radius:10px}
.findings-detail-table thead{background:var(--bg-elevated)}
.findings-detail-table th{padding:10px 14px;text-align:left;font-weight:700;color:var(--text-secondary);text-transform:uppercase;font-size:11px;border-bottom:2px solid var(--border);white-space:nowrap}
.findings-detail-table td{padding:10px 14px;border-bottom:1px solid var(--border-light,var(--border));color:var(--text);vertical-align:top}
.finding-summary-row{transition:background .15s}
.finding-summary-row:hover{background:var(--bg-card-hover,rgba(255,255,255,.03))}
.finding-summary-row.expanded{background:var(--bg-card);border-left:3px solid var(--primary)}
.finding-summary-row td:first-child{position:relative}
.finding-summary-row td:first-child::after{content:'\\25B6';position:absolute;right:2px;top:50%;transform:translateY(-50%);font-size:8px;color:var(--text-muted);transition:transform .2s}
.finding-summary-row.expanded td:first-child::after{transform:translateY(-50%) rotate(90deg)}
.finding-detail-content{padding:12px 8px;background:var(--bg-card);border-radius:6px}
.findings-cat-header td{background:var(--bg-elevated);font-weight:700;font-size:12px;color:var(--text);padding:8px 14px;border-bottom:2px solid var(--border)}
.findings-cat-header{scroll-margin-top:60px}
@media(max-width:768px){.full-width-content{padding:16px}.exec-grid{grid-template-columns:1fr}.controls-table{font-size:11px}.controls-table th,.controls-table td{padding:6px 8px}}
/* Tooltip system (JS-managed) */
.tip{position:relative;cursor:help}
.nav-tip{cursor:help}
#ciq-tooltip{position:fixed;z-index:99999;pointer-events:none;opacity:0;transition:opacity .18s ease;max-width:380px;min-width:200px;padding:14px 18px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  color:var(--text);
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  border-radius:10px;
  font-size:12.5px;line-height:1.6;font-weight:400;text-transform:none;letter-spacing:normal;white-space:normal;
  box-shadow:0 2px 6px rgba(0,0,0,.18),0 8px 24px rgba(0,0,0,.32),0 0 0 1px rgba(255,255,255,.06) inset;
}
#ciq-tooltip.visible{opacity:1}
#ciq-tooltip::before{content:'';position:absolute;width:12px;height:12px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  transform:rotate(45deg);z-index:-1}
#ciq-tooltip.arrow-bottom::before{bottom:-7px;left:var(--arrow-x,24px);border-top:none;border-left:none}
#ciq-tooltip.arrow-top::before{top:-7px;left:var(--arrow-x,24px);border-bottom:none;border-right:none}
#ciq-tooltip .t-sep{display:block;border-top:1px solid rgba(255,255,255,.15);margin:8px 0 4px;padding-top:6px;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--primary)}
/* Help icon inline */
.help-icon{display:inline-block;width:14px;height:14px;border-radius:50%;background:rgba(0,120,212,.15);color:var(--primary);font-size:9px;font-weight:700;text-align:center;line-height:14px;margin-left:4px;cursor:help;vertical-align:middle}
/* Risk detail modal */
.risk-modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:9999;align-items:center;justify-content:center}
.risk-modal-overlay.active{display:flex}
.risk-modal{background:var(--bg-card);border:1px solid var(--border);border-radius:14px;width:560px;max-width:92vw;max-height:80vh;overflow-y:auto;box-shadow:0 12px 48px rgba(0,0,0,.5);padding:0}
.risk-modal-header{display:flex;align-items:center;justify-content:space-between;padding:18px 24px 14px;border-bottom:1px solid var(--border)}
.risk-modal-header h3{margin:0;font-size:15px;font-weight:700;color:var(--text)}
.risk-modal-close{background:none;border:none;color:var(--text-muted);font-size:22px;cursor:pointer;padding:0 4px;line-height:1}
.risk-modal-close:hover{color:var(--text)}
.risk-modal-body{padding:20px 24px 24px}
.risk-modal-body .risk-section{margin-bottom:18px}
.risk-modal-body .risk-section:last-child{margin-bottom:0}
.risk-modal-body .risk-section-title{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted);margin-bottom:8px}
.risk-modal-body .risk-section-content{font-size:13px;line-height:1.65;color:var(--text-secondary)}
@media print{.risk-modal-overlay{display:none!important}}
@media print{.top-nav,.filter-bar,.zoom-controls,.export-bar,.back-to-top,.scroll-progress{display:none!important}.full-width-content{padding:16px;max-width:100%}body{background:#fff;color:#000}.status-pill.pass{background:#d4edda;color:#155724;border-color:#c3e6cb}.status-pill.fail{background:#f8d7da;color:#721c24;border-color:#f5c6cb}#ciq-tooltip{display:none!important}}
"""


def _cr_js() -> str:
    return """
// Zoom
var zoomLevel=100;
function zoomIn(){zoomLevel=Math.min(zoomLevel+10,150);applyZoom()}
function zoomOut(){zoomLevel=Math.max(zoomLevel-10,70);applyZoom()}
function zoomReset(){zoomLevel=100;applyZoom()}
function applyZoom(){document.querySelector('.full-width-content').style.zoom=(zoomLevel/100);document.getElementById('zoom-label').textContent=zoomLevel+'%'}

// Filter findings (updated for collapsible table rows)
function filterFindings(){
  var q=(document.getElementById('finding-filter').value||'').toLowerCase();
  var sev=(document.getElementById('filter-severity').value||'').toLowerCase();
  var cat=(document.getElementById('filter-category').value||'').toLowerCase();
  document.querySelectorAll('.finding-summary-row').forEach(function(row){
    var text=row.textContent.toLowerCase();
    var detail=row.nextElementSibling;
    var match=true;
    if(q&&text.indexOf(q)<0)match=false;
    if(sev&&row.getAttribute('data-severity')!==sev)match=false;
    if(cat&&row.getAttribute('data-category')!==cat)match=false;
    row.style.display=match?'table-row':'none';
    if(detail&&detail.classList.contains('finding-detail-row'))detail.style.display='none';
    row.classList.remove('expanded');
  });
}

// Toggle finding detail row
function toggleFindingDetail(row){
  var detail=row.nextElementSibling;
  if(detail&&detail.classList.contains('finding-detail-row')){
    var isHidden=detail.style.display==='none';
    detail.style.display=isHidden?'table-row':'none';
    row.classList.toggle('expanded',isHidden);
  }
}

function switchInvTab(btn,panelId){
  btn.closest('.inv-tabs').querySelectorAll('.inv-tab').forEach(function(t){t.classList.remove('active');t.setAttribute('aria-selected','false')});
  btn.classList.add('active');btn.setAttribute('aria-selected','true');
  var el=document.getElementById(panelId);
  if(el)el.scrollIntoView({behavior:'smooth',block:'start'});
}

// Scroll progress bar
window.addEventListener('scroll',function(){
  var h=document.documentElement;
  var pct=h.scrollTop/(h.scrollHeight-h.clientHeight)*100;
  var bar=document.getElementById('scroll-progress');
  if(bar)bar.style.width=Math.min(pct,100)+'%';
});

// Risk detail modal
function showRiskDetail(el){
  var w=el.getAttribute('data-why')||'';
  var h=el.getAttribute('data-how')||'';
  var lv=el.getAttribute('data-level')||'';
  var overlay=document.getElementById('risk-modal-overlay');
  if(!overlay)return;
  document.getElementById('risk-modal-level').textContent=lv.toUpperCase()+' RISK';
  document.getElementById('risk-modal-why').textContent=w;
  var howEl=document.getElementById('risk-modal-how');
  // Format numbered steps as a list
  var steps=h.split(/(?=\d+\.\s)/).filter(function(s){return s.trim()});
  if(steps.length>1){
    howEl.innerHTML='<ol style="margin:0;padding-left:20px">'+steps.map(function(s){return '<li style="margin-bottom:6px">'+s.replace(/^\d+\.\s*/,'')+'</li>'}).join('')+'</ol>';
  } else {
    howEl.textContent=h;
  }
  overlay.classList.add('active');
}
function closeRiskModal(){
  var o=document.getElementById('risk-modal-overlay');
  if(o)o.classList.remove('active');
}
document.addEventListener('keydown',function(e){if(e.key==='Escape')closeRiskModal()});
document.addEventListener('click',function(e){if(e.target.id==='risk-modal-overlay')closeRiskModal()});

// ── Tooltip engine (viewport-aware positioning) ──
(function(){
  var tip=document.getElementById('ciq-tooltip');
  if(!tip)return;
  var GAP=10,MARGIN=12;
  function show(ev){
    var tgt=ev.target.closest('[data-tip]');
    if(!tgt)return;
    var text=tgt.getAttribute('data-tip');
    if(!text)return;
    var d=document.createElement('span');d.textContent=text;var safe=d.innerHTML;
    safe=safe.replace(/\\n+YOUR TENANT:/g,'<span class="t-sep">&#x1F4CA; Your Tenant</span>');
    safe=safe.replace(/\\n/g,'<br>');
    tip.innerHTML=safe;
    tip.classList.add('visible');
    tip.setAttribute('aria-hidden','false');
    requestAnimationFrame(function(){
      var r=tgt.getBoundingClientRect();
      var tw=tip.offsetWidth,th=tip.offsetHeight;
      var vw=window.innerWidth,vh=window.innerHeight;
      var above=r.top-GAP-th;
      var below=r.bottom+GAP;
      var top,arrow;
      if(above>=MARGIN){top=above;arrow='arrow-bottom';}
      else if(below+th<=vh-MARGIN){top=below;arrow='arrow-top';}
      else{top=Math.max(MARGIN,vh-th-MARGIN);arrow='';}
      var left=r.left+r.width/2-tw/2;
      left=Math.max(MARGIN,Math.min(left,vw-tw-MARGIN));
      var arrowX=r.left+r.width/2-left;
      arrowX=Math.max(16,Math.min(arrowX,tw-16));
      tip.style.top=top+'px';
      tip.style.left=left+'px';
      tip.style.setProperty('--arrow-x',arrowX+'px');
      tip.className='visible'+(arrow?' '+arrow:'');
    });
  }
  function hide(){
    tip.classList.remove('visible');
    tip.setAttribute('aria-hidden','true');
    tip.className='';
  }
  document.addEventListener('mouseenter',show,true);
  document.addEventListener('mouseleave',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
  document.addEventListener('focusin',show,true);
  document.addEventListener('focusout',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
})();
"""


# ── Security Controls Table rendering ────────────────────────────────────

_CTRL_CATEGORY_NAMES: dict[str, tuple[str, str]] = {
    "oversharing_risk": ("&#128275;", "Oversharing Risk"),
    "label_coverage": ("&#127991;", "Sensitivity Label Coverage"),
    "dlp_readiness": ("&#128737;", "Data Loss Prevention"),
    "restricted_search": ("&#128269;", "Restricted SharePoint Search"),
    "access_governance": ("&#128101;", "Access Governance"),
    "content_lifecycle": ("&#128197;", "Content Lifecycle"),
    "audit_monitoring": ("&#128202;", "Audit & Monitoring"),
    "copilot_security": ("&#129302;", "Copilot-Specific Security"),
}


def _status_pill(status: str) -> str:
    s = status.upper()
    css_cls = {"PASS": "pass", "FAIL": "fail", "PARTIAL": "partial"}.get(s, "not-assessed")
    icons = {"PASS": "&#10003;", "FAIL": "&#10007;", "PARTIAL": "&#9888;", "NOT_ASSESSED": "&#8943;"}
    return f'<span class="status-pill {css_cls}">{icons.get(s, "")} {esc(s.replace("_", " "))}</span>'


def _render_controls_table(controls: list[dict]) -> str:
    """Render the full security controls matrix as an HTML table."""
    if not controls:
        return '<p class="empty">No security controls data available.</p>'

    # Summary counts
    n_pass = sum(1 for c in controls if c["Status"] == "PASS")
    n_fail = sum(1 for c in controls if c["Status"] == "FAIL")
    n_partial = sum(1 for c in controls if c["Status"] == "PARTIAL")
    n_na = sum(1 for c in controls if c["Status"] == "NOT_ASSESSED")
    total = len(controls)

    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#4CAF50">{n_pass}</span> Pass</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#EF5350">{n_fail}</span> Fail</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#FFB300">{n_partial}</span> Partial</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:var(--text-muted)">{n_na}</span> Advisory</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Total</div>'
        f'</div>'
    )

    # Group by category
    by_cat: dict[str, list[dict]] = {}
    for c in controls:
        by_cat.setdefault(c["Category"], []).append(c)

    rows = ""
    cat_order = [
        "oversharing_risk", "label_coverage", "dlp_readiness", "restricted_search",
        "access_governance", "content_lifecycle", "audit_monitoring", "copilot_security",
    ]
    for cat in cat_order:
        ctrls = by_cat.get(cat, [])
        if not ctrls:
            continue
        icon, cat_name = _CTRL_CATEGORY_NAMES.get(cat, ("", cat.replace("_", " ").title()))
        cat_pass = sum(1 for c in ctrls if c["Status"] == "PASS")
        rows += (
            f'<tr id="ctrl-{cat}" class="ctrl-cat-header"><td colspan="7">'
            f'{icon} {esc(cat_name)} — {cat_pass}/{len(ctrls)} controls passing</td></tr>'
        )
        for c in ctrls:
            sev = c.get("Severity", "—")
            sev_html = _severity_badge(sev) if sev not in ("—", "") else '<span style="color:var(--text-muted)">—</span>'
            # Build resource column — show all affected resources
            resources = c.get("AffectedResources", [])
            res_primary = esc(c.get("Resource", "—"))
            if len(resources) > 1:
                res_extra = "".join(
                    f'<div style="font-size:10px;color:var(--text-muted);padding:1px 0">{esc(r.get("Name", ""))}</div>'
                    for r in resources[1:6]
                )
                more = f'<div style="font-size:10px;color:var(--text-muted)">+{len(resources) - 6} more</div>' if len(resources) > 6 else ""
                res_html = f'<strong>{res_primary}</strong>{res_extra}{more}'
            else:
                res_html = res_primary
            rows += (
                f'<tr>'
                f'<td class="ctrl-id">{esc(c["ControlId"])}</td>'
                f'<td><div class="ctrl-name">{esc(c["ControlName"])}</div>'
                f'<div class="ctrl-desc">{esc(c["Description"])}</div></td>'
                f'<td>{_status_pill(c["Status"])}</td>'
                f'<td>{sev_html}</td>'
                f'<td class="ctrl-resource">{res_html}</td>'
                f'<td style="font-size:11px;color:var(--text-secondary);max-width:220px">{esc(c.get("Details", ""))}</td>'
                f'<td class="ctrl-ref">{_ms_ref_link(c.get("MicrosoftReference", ""), c.get("MicrosoftReferenceUrl", ""))}</td>'
                f'</tr>'
            )

    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>'
        f'{_th_tip("Control ID", "Unique identifier for this security control check.", "width:7%")}'
        f'{_th_tip("Control", "Name and description of the security control being evaluated.", "width:20%")}'
        f'{_th_tip("Status", "PASS = meets criteria. FAIL = gap found. PARTIAL = partially meets criteria. NOT ASSESSED = data unavailable.", "width:8%")}'
        f'{_th_tip("Severity", "Impact level if this control fails. Critical > High > Medium > Low > Informational.", "width:7%")}'
        f'{_th_tip("Resource", "The M365 resource or configuration area this control applies to.", "width:10%")}'
        f'{_th_tip("Details", "Specific finding details or configuration values detected during assessment.", "width:24%")}'
        f'{_th_tip("Remediation Guidance", "Step-by-step guidance to remediate this control gap. Shows the Microsoft product area and actions for your technical team.", "width:24%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_findings_summary_table(findings: list[dict]) -> str:
    """Render findings as a compact priority-ranked summary table with effort tags."""
    if not findings:
        return '<p class="empty">No gaps found — your tenant is ready for Copilot!</p>'

    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    sorted_f = sorted(findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))

    # Effort summary counts
    n_quick = sum(1 for f in findings if f.get("Effort") == "quick_win")
    n_mod = sum(1 for f in findings if f.get("Effort") == "moderate")
    n_major = sum(1 for f in findings if f.get("Effort") == "major")

    effort_summary = (
        f'<div class="ctrl-summary-bar" style="margin-bottom:12px">'
        f'<div class="ctrl-summary-item">{_effort_badge("quick_win")} <span style="font-size:13px;margin-left:4px"><strong>{n_quick}</strong> Quick Wins</span></div>'
        f'<div class="ctrl-summary-item">{_effort_badge("moderate")} <span style="font-size:13px;margin-left:4px"><strong>{n_mod}</strong> Moderate</span></div>'
        f'<div class="ctrl-summary-item">{_effort_badge("major")} <span style="font-size:13px;margin-left:4px"><strong>{n_major}</strong> Major Projects</span></div>'
        f'</div>'
    )

    rows = ""
    for i, f in enumerate(sorted_f, 1):
        sev = f.get("Severity", "medium").lower()
        cat = f.get("Category", "")
        effort = f.get("Effort", "moderate")
        affected = f.get("AffectedCount", 0)
        rem = f.get("Remediation", {})
        rem_text = rem.get("Description", "")

        cat_meta = _CATEGORY_META.get(cat, {})
        cat_name = cat_meta.get("name", cat.replace("_", " ").title())

        rows += (
            f'<tr>'
            f'<td style="text-align:center;font-family:var(--font-mono);font-weight:700;color:var(--text-muted)">{i}</td>'
            f'<td>{_severity_badge(sev)}</td>'
            f'<td>{_effort_badge(effort)}</td>'
            f'<td style="font-weight:600;font-size:12px">{esc(f.get("Title", ""))}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)">{esc(cat_name)}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono);font-size:12px">{affected}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary);max-width:280px">{esc(rem_text)}</td>'
            f'</tr>'
        )

    return (
        f'{effort_summary}'
        f'<table class="findings-summary-table">'
        f'<thead><tr>'
        f'{_th_tip("Priority", "Ranking position — lower number = higher priority, sorted by severity.", "width:30px")}'
        f'{_th_tip("Severity", "Impact level: Critical (data exposure), High (significant gap), Medium (moderate risk), Low (minor), Info (awareness).", "width:80px")}'
        f'{_th_tip("Effort", "Estimated remediation effort. Quick Win = under 1hr. Moderate = days. Major = weeks.", "width:80px")}'
        f'{_th_tip("Finding", "Title of the security or configuration gap detected.", "min-width:200px")}'
        f'{_th_tip("Category", "Which of the 8 readiness domains this finding belongs to.", "width:120px")}'
        f'{_th_tip("Affected", "Number of resources or configurations affected by this gap.", "width:60px")}'
        f'{_th_tip("Remediation", "Recommended action to resolve this gap.")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_site_inventory(sites: list[dict]) -> str:
    """Render the SharePoint site inventory as a detailed HTML table."""
    if not sites:
        return '<p class="empty">No SharePoint site data collected.</p>'

    total = len(sites)
    stale = sum(1 for s in sites if s.get("IsStale"))
    overshared = sum(1 for s in sites if s.get("IsOvershared"))
    unlabeled = sum(1 for s in sites if s.get("SensitivityLabel") in ("None", ""))
    anon_links = sum(1 for s in sites if s.get("AnonymousLinks", 0) > 0)

    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Sites</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#EF5350">{stale}</span> Stale</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#F7630C">{overshared}</span> Overshared</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#FFB300">{unlabeled}</span> Unlabeled</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#D13438">{anon_links}</span> Anon Links</div>'
        f'</div>'
    )

    rows = ""
    for s in sorted(sites, key=lambda x: (not x.get("IsOvershared"), not x.get("IsStale"), x.get("SiteName", ""))):
        label = s.get("SensitivityLabel", "None")
        label_html = (
            f'<span class="status-pill not-assessed">{esc(label)}</span>'
            if label in ("None", "") else
            f'<span class="status-pill pass">{esc(label)}</span>'
        )
        stale_html = (
            '<span class="status-pill fail">Stale</span>' if s.get("IsStale")
            else '<span class="status-pill pass">Active</span>'
        )
        overshare_html = (
            '<span class="status-pill fail">Yes</span>' if s.get("IsOvershared")
            else '<span class="status-pill pass">No</span>'
        )
        perms = s.get("TotalPermissions", 0)
        guests = s.get("GuestCount", 0)
        ext = s.get("ExternalUserCount", 0)
        anon = s.get("AnonymousLinks", 0)
        org_links = s.get("OrganizationLinks", 0)
        shared_items = s.get("TotalSharedItems", 0)

        perm_detail = (
            f'<span title="Owners: {s.get("OwnerCount", 0)}, Members: {s.get("MemberCount", 0)}, '
            f'Guests: {guests}, External: {ext}">{perms}</span>'
        )
        sharing_detail = (
            f'<span title="Shared items: {shared_items}, Org links: {org_links}, '
            f'External links: {s.get("ExternalLinks", 0)}">'
            f'{"🔴 " + str(anon) if anon > 0 else "✓ 0"}</span>'
        )

        url = s.get("WebUrl", "")
        name = s.get("SiteName", url.split("/")[-1] if url else "—")
        # Copilot risk assessment
        if s.get("IsOvershared") or anon > 0:
            risk_why = (
                f"This site has {'broad permissions (Everyone/large groups)' if s.get('IsOvershared') else ''}"
                f"{' and ' if s.get('IsOvershared') and anon > 0 else ''}"
                f"{f'{anon} anonymous sharing link(s)' if anon > 0 else ''}. "
                f"Microsoft 365 Copilot inherits the signed-in user's permissions, so any content reachable by broad or "
                f"anonymous access will be surfaced in Copilot responses — including sensitive files, HR records, "
                f"and confidential documents that may exist on this site."
            )
            risk_how = (
                "1. Review and remove 'Everyone' / 'Everyone except external users' permissions in SharePoint admin center > Active sites > select site > Permissions. "
                "2. Delete or convert anonymous (Anyone) sharing links to 'People in your organization' or 'Specific people' links via SharePoint admin center > Sharing settings. "
                "3. Apply a sensitivity label with encryption (e.g. Confidential — All Employees) to restrict Copilot from surfacing content to unauthorized users (ref: Microsoft Purview > Information Protection). "
                "4. Enable Restricted SharePoint Search (RSS) in SharePoint admin center > Search settings to exclude this site from Copilot and org-wide search until permissions are remediated. "
                "5. Run a DSPM for AI oversharing assessment in Microsoft Purview > Data Security Posture Management to identify and remediate exposed content."
            )
            risk_html = _copilot_risk_pill("high", "Overshared or anonymous links — click for details", risk_why, risk_how)
        elif s.get("IsStale") or label in ("None", ""):
            reasons = []
            if s.get("IsStale"):
                reasons.append("no activity in 90+ days (stale)")
            if label in ("None", ""):
                reasons.append("no sensitivity label applied")
            risk_why = (
                f"This site has {' and '.join(reasons)}. "
                f"Copilot grounds responses using content the signed-in user can access (per Microsoft 365 Copilot permission model). "
                f"{'Stale sites may contain outdated information that Copilot surfaces as current, leading to inaccurate responses. ' if s.get('IsStale') else ''}"
                f"{'Without a sensitivity label, content cannot be protected by label-based encryption or targeted by label-scoped DLP policies — leaving it fully accessible to any user with site permissions via Copilot. ' if label in ('None', '') else ''}"
            )
            risk_how = (
                "1. " + ("Archive or delete stale content, or update the site to mark it active in SharePoint admin center > Active sites. " if s.get("IsStale") else "")
                + ("Apply a sensitivity label with encryption via Microsoft Purview > Information Protection, or configure auto-labeling policies to classify content at scale. " if label in ("None", "") else "")
                + "2. Configure a retention policy in Microsoft Purview > Data Lifecycle Management to automatically archive or delete stale content. "
                "3. Exclude this site from Copilot and org-wide search using Restricted SharePoint Search (RSS) in SharePoint admin center until content is reviewed."
            )
            risk_html = _copilot_risk_pill("medium", "Stale or unlabeled — click for details", risk_why, risk_how)
        else:
            risk_why = (
                "This site is actively maintained with recent modifications and has a sensitivity label applied. "
                "If the label includes encryption, Copilot will honor those restrictions and only surface content to users "
                "with decryption rights (per Microsoft Purview Information Protection). Labels without encryption provide "
                "classification metadata for DLP policy targeting but do not independently restrict Copilot access."
            )
            risk_how = (
                "No immediate action required. Verify that the applied label includes encryption if content is sensitive. "
                "Monitor label compliance in Microsoft Purview > Content Explorer and review site permissions periodically "
                "to ensure the principle of least privilege is maintained."
            )
            risk_html = _copilot_risk_pill("low", "Labeled and actively managed — click for details", risk_why, risk_how)
        # Domain mapping — dynamic based on site state
        site_doms: list[str] = []
        if s.get("IsOvershared") or anon > 0:
            site_doms.append("oversharing_risk")
        if label in ("None", ""):
            site_doms.append("label_coverage")
        if s.get("IsStale"):
            site_doms.append("content_lifecycle")
        if not site_doms:
            site_doms.append("label_coverage")
        domain_html = _domain_pills(site_doms)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600;font-size:12px" title="{esc(url)}">{esc(name)}</td>'
            f'<td style="font-size:10px;color:var(--text-muted)" title="{esc(url)}">{esc(url)}</td>'
            f'<td>{label_html}</td>'
            f'<td>{stale_html}</td>'
            f'<td>{overshare_html}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono);font-size:12px">{perm_detail}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono);font-size:12px">{sharing_detail}</td>'
            f'<td>{domain_html}</td>'
            f'<td>{risk_html}</td>'
            f'</tr>'
        )

    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>'
        f'{_th_tip("Site Name", "Name of the SharePoint site as shown in the admin center.", "width:14%")}'
        f'{_th_tip("URL", "Full web address of the SharePoint site.", "width:18%")}'
        f'{_th_tip("Label", "Sensitivity label applied to this site. Unlabeled sites may expose data to Copilot.", "width:9%")}'
        f'{_th_tip("Status", "Active = recently modified. Stale = no activity in 90+ days, may contain outdated data.", "width:6%")}'
        f'{_th_tip("Overshared", "Whether this site has broad permissions (Everyone, large groups) that could let Copilot surface restricted content.", "width:8%")}'
        f'{_th_tip("Permissions", "Total unique permission entries. Hover individual values for breakdown by role.", "width:8%")}'
        f'{_th_tip("Anon Links", "Anonymous sharing links that allow access without authentication — high risk for Copilot.", "width:7%")}'
        f'{_th_tip("Domain", "Compliance domain(s) this resource impacts: Oversharing Risk, Label Coverage, Content Lifecycle, etc.", "width:16%")}'
        f'{_th_tip("Copilot Risk", "Click the risk pill for detailed Why and How guidance.", "width:14%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


# ── Main generator ───────────────────────────────────────────────────────

def _render_license_inventory(licenses: list[dict]) -> str:
    """Render the license/SKU inventory as an HTML table."""
    if not licenses:
        return '<p class="empty">No license data collected.</p>'
    total = len(licenses)
    copilot_count = sum(1 for l in licenses if "copilot" in l.get("SkuPartNumber", "").lower())
    rows = ""
    for lic in licenses:
        part = lic.get("SkuPartNumber", "")
        enabled = lic.get("EnabledUnits", 0)
        consumed = lic.get("ConsumedUnits", 0)
        usage_pct = (consumed / enabled * 100) if enabled > 0 else 0
        bar_color = "#107C10" if usage_pct < 80 else "#FFB900" if usage_pct < 95 else "#D13438"
        is_copilot = "copilot" in part.lower()
        name_style = ' style="color:var(--primary);font-weight:700"' if is_copilot else ""
        # Copilot risk: highlight unassigned Copilot licenses or missing Copilot SKU
        if is_copilot and consumed == 0:
            risk_html = _copilot_risk_pill("high", "Copilot licenses purchased but none assigned")
            risk_why = (
                f"The {part} SKU has {enabled:,} purchased licenses but zero are assigned to users. "
                f"Per Microsoft 365 admin center documentation, Copilot requires an assigned license per user to function. "
                f"Unassigned licenses mean zero Copilot usage — security controls (DLP policies, sensitivity labels, CA policies) "
                f"configured for Copilot remain untested against real user interactions."
            )
            risk_how = (
                "1. Assign Copilot licenses to a pilot group in Microsoft 365 admin center > Billing > Licenses > select Copilot SKU > Assign. "
                "2. Ensure pilot users are scoped under Conditional Access policies requiring MFA (Microsoft Entra admin center > Protection > Conditional Access). "
                "3. Verify sensitivity labels and DLP policies are active in Microsoft Purview before broader rollout. "
                "4. Monitor Copilot interactions using DSPM for AI in Microsoft Purview > Data Security Posture Management."
            )
            lic_doms = ["access_governance", "copilot_security"]
        elif is_copilot and usage_pct < 50:
            risk_html = _copilot_risk_pill("medium", "Low Copilot license utilization — review assignment")
            risk_why = (
                f"The {part} SKU shows only {usage_pct:.0f}% utilization ({consumed:,} of {enabled:,} assigned). "
                f"Per Microsoft 365 admin center Usage reports, low adoption may indicate blocked access (CA policies), "
                f"lack of user awareness, or misconfiguration. Underutilized licenses reduce visibility into Copilot "
                f"interactions for security monitoring in Microsoft Purview."
            )
            risk_how = (
                "1. Review unassigned licenses in Microsoft 365 admin center > Billing > Licenses and assign or reclaim unused seats. "
                "2. Check Conditional Access policies in Microsoft Entra admin center to ensure they are not blocking Copilot access. "
                "3. Run the Copilot usage report in Microsoft 365 admin center > Reports > Usage > Microsoft Copilot to identify inactive users. "
                "4. Provide user training and enablement resources (Microsoft Copilot adoption hub) to drive adoption."
            )
            lic_doms = ["access_governance"]
        elif is_copilot:
            risk_html = _copilot_risk_pill("low", "Copilot licenses actively assigned")
            risk_why = (
                f"The {part} SKU is well-utilized at {usage_pct:.0f}% ({consumed:,} of {enabled:,}). "
                f"Active assignment provides real interaction data for security monitoring in Microsoft Purview "
                f"audit logs and the Microsoft 365 admin center Usage reports."
            )
            risk_how = (
                "No immediate action required. Monitor Copilot usage in Microsoft 365 admin center > Reports > Usage > Microsoft Copilot. "
                "Review Copilot interaction audit logs periodically in Microsoft Purview > Audit."
            )
            lic_doms = ["access_governance"]
        else:
            risk_html = '<span style="color:var(--text-muted);font-size:11px">—</span>'
            risk_why = "Non-Copilot license — no direct Copilot data exposure risk from this SKU."
            risk_how = "No Copilot-specific action needed for this license SKU."
            lic_doms = ["access_governance"]
        risk_html = _inject_risk_modal(risk_html, risk_why, risk_how)
        rows += (
            f'<tr>'
            f'<td{name_style}>{esc(part)}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono)">{enabled:,}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono)">{consumed:,}</td>'
            f'<td style="width:120px"><div style="background:var(--bg-elevated);border-radius:4px;height:14px;overflow:hidden">'
            f'<div style="width:{min(usage_pct, 100):.0f}%;height:100%;background:{bar_color};border-radius:4px"></div>'
            f'</div><span style="font-size:10px;color:var(--text-muted)">{usage_pct:.0f}%</span></td>'
            f'<td>{_domain_pills(lic_doms)}</td>'
            f'<td>{risk_html}</td>'
            f'</tr>'
        )
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> SKUs</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:var(--primary)">{copilot_count}</span> Copilot</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("SKU Part Number", "Microsoft license SKU identifier. Copilot-related licenses are highlighted in blue.", "width:28%")}{_th_tip("Enabled", "Total number of licenses purchased or available in this SKU.", "width:8%")}'
        f'{_th_tip("Consumed", "Number of licenses currently assigned to users.", "width:8%")}{_th_tip("Usage", "Percentage of available licenses in use. High usage may indicate need for more licenses.", "width:14%")}'
        f'{_th_tip("Domain", "Compliance domain this license impacts.", "width:22%")}'
        f'{_th_tip("Copilot Risk", "Risk indicator for Copilot-related licenses. Click for details.", "width:18%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_label_inventory(labels: list[dict]) -> str:
    """Render the sensitivity label inventory."""
    if not labels:
        return '<p class="empty">No sensitivity label definitions collected. Ensure InformationProtection.Read.All permission is granted.</p>'
    total = len(labels)
    active = sum(1 for l in labels if l.get("IsActive", True))
    parents = sum(1 for l in labels if not l.get("ParentId"))
    rows = ""
    for lbl in labels:
        name = lbl.get("Name", "")
        priority = lbl.get("Priority", 0)
        is_active = lbl.get("IsActive", True)
        parent_id = lbl.get("ParentId", "")
        content_type = lbl.get("ContentType", "")
        color = lbl.get("Color", "")
        is_encrypted = lbl.get("IsEncryptionEnabled", False)
        has_site_group = lbl.get("HasSiteAndGroupSettings", False)
        status_html = (
            '<span class="status-pill pass">Active</span>' if is_active
            else '<span class="status-pill not-assessed">Inactive</span>'
        )
        encryption_html = (
            '<span class="status-pill pass">Yes</span>' if is_encrypted
            else '<span class="status-pill not-assessed">No</span>'
        )
        site_group_html = (
            '<span class="status-pill pass">Yes</span>' if has_site_group
            else '<span class="status-pill not-assessed">No</span>'
        )
        level = "Sub-label" if parent_id else "Parent"
        color_swatch = (
            f'<span style="display:inline-block;width:14px;height:14px;border-radius:3px;background:{esc(color)};border:1px solid var(--border);vertical-align:middle"></span> '
            if color else ""
        )
        # Copilot risk: inactive labels = gap in classification
        if not is_active:
            lbl_risk = _copilot_risk_pill("medium", "Inactive label — not protecting content from Copilot exposure")
            lbl_why = (
                f"The sensitivity label '{name}' is inactive and cannot be applied to content. "
                f"Inactive labels create gaps in the classification taxonomy — content that should be classified under "
                f"this label remains unprotected. Per Microsoft Purview Information Protection, sensitivity labels with "
                f"encryption restrict which users can access content, including through Copilot. Without this label active, "
                f"content cannot be classified or encrypted under its scope, leaving it accessible to any user with permissions."
            )
            lbl_how = (
                "1. Re-activate this label in Microsoft Purview compliance portal > Information Protection > Labels if the classification is still needed. "
                "2. If deprecated, ensure a replacement label exists and migrate content using Content Explorer (Microsoft Purview > Data Classification > Content Explorer). "
                "3. Run a content scan to identify documents referencing this label using Microsoft Purview > Data Classification > Content Explorer. "
                "4. Update auto-labeling policies in Microsoft Purview > Information Protection > Auto-labeling to use active labels."
            )
            lbl_doms = ["label_coverage", "dlp_readiness", "copilot_security"]
        else:
            lbl_risk = _copilot_risk_pill("low", "Active label available for classification")
            lbl_why = (
                f"The sensitivity label '{name}' is active and available for users and auto-labeling policies. "
                f"If this label includes encryption, Copilot honors the access restrictions and only surfaces "
                f"labeled content to users with decryption rights. Active labels also enable DLP policy targeting "
                f"and provide classification metadata for audit and compliance reporting in Microsoft Purview."
            )
            lbl_how = (
                "No immediate action required. Verify this label is included in auto-labeling policies (Microsoft Purview > "
                "Information Protection > Auto-labeling). Monitor label usage in Microsoft Purview > Data Classification > Activity Explorer."
            )
            lbl_doms = ["label_coverage", "dlp_readiness"]
        lbl_risk = _inject_risk_modal(lbl_risk, lbl_why, lbl_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600">{color_swatch}{esc(name)}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono)">{priority}</td>'
            f'<td>{status_html}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)">{esc(level)}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)">{esc(content_type)}</td>'
            f'<td>{encryption_html}</td>'
            f'<td>{site_group_html}</td>'
            f'<td>{_domain_pills(lbl_doms)}</td>'
            f'<td>{lbl_risk}</td>'
            f'</tr>'
        )
    encrypted = sum(1 for l in labels if l.get("IsEncryptionEnabled", False) and l.get("IsActive", True))
    site_group = sum(1 for l in labels if l.get("HasSiteAndGroupSettings", False) and l.get("IsActive", True))
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Labels</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#107C10">{active}</span> Active</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{parents}</span> Parent</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{encrypted}</span> Encrypted</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{site_group}</span> Site&amp;Group</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("Label Name", "Sensitivity label name as defined in Microsoft Purview.", "width:16%")}{_th_tip("Priority", "Processing priority — lower numbers are evaluated first when multiple labels apply.", "width:6%")}'
        f'{_th_tip("Status", "Whether this label is currently active and available for use by users.", "width:7%")}{_th_tip("Level", "Parent = top-level label. Sub-label = nested under a parent label.", "width:7%")}'
        f'{_th_tip("Content Type", "Types of content this label can be applied to (files, emails, sites, etc.).", "width:14%")}'
        f'{_th_tip("Encryption", "Whether this label applies encryption to protect content. Copilot honors encryption rights.", "width:8%")}'
        f'{_th_tip("Site &amp; Group", "Whether this label has site/group settings to control container-level privacy, sharing, and guest access.", "width:8%")}'
        f'{_th_tip("Domain", "Compliance domain(s) this label impacts.", "width:18%")}{_th_tip("Copilot Risk", "Risk to Copilot: inactive labels leave content unclassified. Click for details.", "width:14%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_ca_inventory(policies: list[dict]) -> str:
    """Render the Conditional Access policy inventory."""
    if not policies:
        return '<p class="empty">No Conditional Access policies collected. Ensure Security Reader role is assigned.</p>'
    total = len(policies)
    enabled = sum(1 for p in policies if p.get("State") == "enabled")
    mfa_count = sum(1 for p in policies if p.get("RequiresMFA"))
    rows = ""
    for pol in policies:
        name = pol.get("DisplayName", "")
        state = pol.get("State", "")
        state_html = (
            '<span class="status-pill pass">Enabled</span>' if state == "enabled"
            else '<span class="status-pill not-assessed">Report-only</span>' if "report" in state.lower()
            else '<span class="status-pill fail">Disabled</span>'
        )
        flags = []
        if pol.get("RequiresMFA"):
            flags.append("MFA")
        if pol.get("RequiresCompliantDevice"):
            flags.append("Compliant Device")
        if pol.get("BlocksLegacyAuth"):
            flags.append("Block Legacy")
        if pol.get("HasLocationCondition"):
            flags.append("Location")
        if pol.get("TargetsAllUsers"):
            flags.append("All Users")
        flags_html = " ".join(
            f'<span style="display:inline-block;padding:2px 6px;border-radius:4px;font-size:10px;background:rgba(0,120,212,.12);color:var(--primary);margin-right:2px">{f}</span>'
            for f in flags
        ) if flags else '<span style="color:var(--text-muted);font-size:11px">—</span>'
        grants = pol.get("GrantControls", "")
        # Copilot risk for CA policies
        if state.lower() == "disabled":
            ca_risk = _copilot_risk_pill("high", "Disabled policy — no enforcement on Copilot access")
            ca_why = (
                f"The Conditional Access policy '{name}' is disabled and not enforcing any controls. "
                f"Per Microsoft Entra ID documentation, Conditional Access policies gate access to M365 cloud apps — "
                f"without enforcement, Copilot (accessed via Microsoft 365 apps) can be reached from unmanaged devices, "
                f"untrusted locations, and without MFA. An attacker who compromises credentials gains unrestricted "
                f"Copilot access to query any data the compromised user can reach."
            )
            ca_how = (
                "1. Enable this policy in Microsoft Entra admin center > Protection > Conditional Access > select policy > toggle to 'On'. "
                "2. If deprecated, ensure replacement policies cover the same user scope and enforce MFA (per Microsoft Identity best practices). "
                "3. Create a CA policy targeting the 'Office 365' cloud app (which includes Copilot) requiring MFA + compliant device. "
                "4. Block legacy authentication protocols via a separate CA policy to prevent token replay attacks (Entra ID > Security > Authentication methods)."
            )
            ca_doms = ["access_governance", "copilot_security"]
        elif "report" in state.lower():
            ca_risk = _copilot_risk_pill("medium", "Report-only — not enforcing controls on Copilot")
            ca_why = (
                f"The policy '{name}' is in report-only mode — per Microsoft Entra ID documentation, it logs "
                f"what-if evaluations in sign-in logs but does not block or enforce controls. Users can access "
                f"Copilot (via Office 365 cloud apps) without meeting the policy's security requirements. "
                f"Report-only is intended for impact analysis, not production protection."
            )
            ca_how = (
                "1. Review the policy's impact in Microsoft Entra admin center > Sign-in logs > Conditional Access tab to verify enforcement won't disrupt users. "
                "2. Switch from 'Report-only' to 'On' once impact analysis confirms acceptable user experience. "
                "3. Verify the policy's cloud app condition includes 'Office 365' to scope Copilot access (Copilot runs within M365 apps)."
            )
            ca_doms = ["access_governance"]
        elif pol.get("RequiresMFA"):
            ca_risk = _copilot_risk_pill("low", "Enabled with MFA — Copilot access is protected")
            ca_why = (
                f"The policy '{name}' is enabled and requires MFA, providing strong identity verification "
                f"before granting access to M365 services including Copilot. Per Microsoft Entra ID security defaults "
                f"documentation, MFA blocks over 99.9% of credential-based attacks."
            )
            ca_how = (
                "No immediate action required. Verify this policy covers all Copilot-licensed users. Per Microsoft "
                "guidance, prefer phishing-resistant MFA methods (FIDO2 security keys, Windows Hello for Business, "
                "certificate-based authentication) over SMS/phone call verification."
            )
            ca_doms = ["access_governance"]
        else:
            ca_risk = _copilot_risk_pill("low", "Enabled and enforcing controls")
            ca_why = (
                f"The policy '{name}' is enabled and enforcing access controls. "
                f"Active enforcement means access to M365 services (including Copilot) is gated by this policy's conditions."
            )
            ca_how = (
                "No immediate action required. Consider adding MFA requirement if not already included "
                "(Microsoft Entra admin center > Protection > Conditional Access). Ensure the policy scope covers Copilot-licensed users."
            )
            ca_doms = ["access_governance"]
        ca_risk = _inject_risk_modal(ca_risk, ca_why, ca_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600;font-size:12px">{esc(name)}</td>'
            f'<td>{state_html}</td>'
            f'<td>{flags_html}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)">{esc(grants)}</td>'
            f'<td>{_domain_pills(ca_doms)}</td>'
            f'<td>{ca_risk}</td>'
            f'</tr>'
        )
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Policies</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#107C10">{enabled}</span> Enabled</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:var(--primary)">{mfa_count}</span> Require MFA</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("Policy Name", "Name of the Conditional Access policy in Entra ID.", "width:22%")}{_th_tip("State", "Enabled = actively enforcing. Report-only = monitoring only. Disabled = inactive.", "width:8%")}'
        f'{_th_tip("Controls", "Security requirements enforced by this policy: MFA, compliant device, block legacy auth, etc.", "width:16%")}{_th_tip("Grant Controls", "Specific grant or session controls configured for this policy.", "width:16%")}'
        f'{_th_tip("Domain", "Compliance domain this policy impacts.", "width:20%")}'
        f'{_th_tip("Copilot Risk", "Risk to Copilot access: disabled/report-only policies leave Copilot unprotected. Click for details.", "width:16%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_groups_inventory(groups: list[dict]) -> str:
    """Render the M365 Groups & Teams inventory."""
    if not groups:
        return '<p class="empty">No M365 Groups collected. Ensure Group.Read.All permission is granted.</p>'
    total = len(groups)
    teams_count = sum(1 for g in groups if g.get("IsTeam"))
    public_count = sum(1 for g in groups if (g.get("Visibility") or "").lower() == "public")
    rows = ""
    for grp in groups:
        name = grp.get("DisplayName", "")
        is_team = grp.get("IsTeam", False)
        tipo = (
            '<span class="status-pill pass">Team</span>' if is_team
            else '<span class="status-pill not-assessed">Group</span>'
        )
        vis = grp.get("Visibility", "Private")
        vis_html = (
            '<span class="status-pill fail">Public</span>' if vis.lower() == "public"
            else f'<span class="status-pill pass">{esc(vis)}</span>'
        )
        mail = grp.get("Mail", "")
        created = grp.get("CreatedDate", "")
        # Copilot risk for groups
        if vis.lower() == "public":
            grp_risk = _copilot_risk_pill("high", "Public group — all org users can access content via Copilot")
            grp_domain = _domain_pills(["oversharing_risk", "copilot_security"])
            grp_why = (
                f"The {'Team' if is_team else 'M365 Group'} '{name}' has Public visibility. "
                f"Per Microsoft 365 group settings, any user in the organization can self-join and access all content — "
                f"files, conversations, and shared resources. Since Copilot respects the signed-in user's permissions, "
                f"and any user can join a public group, Copilot can effectively surface this group's content to any "
                f"Copilot-licensed user who joins — including confidential discussions, shared documents, and uploaded files."
            )
            grp_how = (
                "1. Change visibility from Public to Private in Microsoft 365 admin center > Teams & groups > select group > Settings, or via Teams admin center. "
                "2. Review current membership and remove unauthorized users after switching to Private. "
                "3. Apply a sensitivity label with a privacy setting to enforce Private visibility (Microsoft Purview > Information Protection > Labels > Group & site settings). "
                "4. Audit the group's shared files and channels for sensitive content that may have been exposed while Public."
            )
        else:
            grp_risk = _copilot_risk_pill("low", "Private group — membership-controlled access")
            grp_domain = _domain_pill("access_governance")
            grp_why = (
                f"The {'Team' if is_team else 'M365 Group'} '{name}' is Private — only approved members can access content. "
                f"Copilot respects the signed-in user's permissions and group membership boundaries, so content from this "
                f"group is only surfaced to its members in Copilot responses."
            )
            grp_how = (
                "No immediate action required. Periodically review group membership in M365 admin center > Teams & groups. "
                "Consider applying a sensitivity label with a privacy setting to enforce Private visibility at the label level."
            )
        grp_risk = _inject_risk_modal(grp_risk, grp_why, grp_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600;font-size:12px">{esc(name)}</td>'
            f'<td>{tipo}</td>'
            f'<td>{vis_html}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)" title="{esc(mail)}">{esc(mail)}</td>'
            f'<td style="font-size:11px;font-family:var(--font-mono);color:var(--text-muted)">{esc(created)}</td>'
            f'<td>{grp_domain}</td>'
            f'<td>{grp_risk}</td>'
            f'</tr>'
        )
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Groups</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#107C10">{teams_count}</span> Teams</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#D13438">{public_count}</span> Public</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("Group Name", "Display name of the M365 Group or Team.", "width:22%")}{_th_tip("Type", "Team = Teams-enabled group. Group = M365 group without Teams.", "width:7%")}'
        f'{_th_tip("Visibility", "Public = anyone in the org can join and view content. Private = membership required.", "width:7%")}{_th_tip("Mail", "Email address associated with this group.", "width:20%")}'
        f'{_th_tip("Created", "Date the group was created.", "width:9%")}{_th_tip("Domain", "Compliance domain this group impacts.", "width:18%")}{_th_tip("Copilot Risk", "Risk to Copilot: public groups expose content to all users. Click for details.", "width:15%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_dlp_inventory(policies: list[dict]) -> str:
    """Render the DLP policy inventory."""
    if not policies:
        return (
            '<p class="empty">No DLP policy data available via Graph API. '
            'DLP policies require Security &amp; Compliance PowerShell for full visibility. '
            'DLP readiness is still scored from available evidence.</p>'
        )
    rows = ""
    for pol in policies:
        name = pol.get("PolicyName", "")
        state = pol.get("State", "")
        state_html = (
            '<span class="status-pill pass">Enabled</span>' if state.lower() == "enabled"
            else '<span class="status-pill fail">Disabled</span>'
        )
        workloads = pol.get("Workloads", "")
        # Copilot risk for DLP
        if state.lower() != "enabled":
            dlp_risk = _copilot_risk_pill("high", "Disabled DLP policy — sensitive data not protected from Copilot")
            dlp_why = (
                f"The DLP policy '{name}' is disabled and not scanning content across its configured workloads ({workloads}). "
                f"Per Microsoft Purview DLP documentation, DLP policies protect data referenced in Copilot interactions "
                f"for supported workloads (Microsoft Teams, SharePoint, OneDrive, Exchange). Without enforcement, Copilot can "
                f"surface sensitive data (PII, financial records, health data) in user responses without triggering any policy action."
            )
            dlp_how = (
                "1. Enable this policy in Microsoft Purview compliance portal > Data Loss Prevention > Policies > select policy > turn on. "
                "2. Verify the policy's sensitive information types (SITs) cover the data categories required by your compliance framework (GDPR, HIPAA, PCI-DSS). "
                "3. Ensure the policy covers Microsoft Teams (where DLP applies to Copilot chat interactions), plus SharePoint and OneDrive for file-level protection. "
                "4. Test with a subset of users in simulation mode before broad enforcement to minimize disruption. "
                "5. Monitor DLP policy matches in Microsoft Purview > Activity Explorer to validate effectiveness."
            )
            dlp_doms = ["dlp_readiness", "copilot_security", "label_coverage"]
        else:
            dlp_risk = _copilot_risk_pill("low", "DLP actively protecting workloads")
            dlp_why = (
                f"The DLP policy '{name}' is enabled and actively scanning workloads: {workloads}. "
                f"Per Microsoft Purview DLP documentation, active policies protect data in Copilot interactions "
                f"for supported workloads — blocking, warning, or logging sensitive data exposure based on configured rules."
            )
            dlp_how = (
                "No immediate action required. Monitor DLP policy matches and false positives in Microsoft Purview > Activity Explorer. "
                "Periodically review and update sensitive information types (SITs) to keep pace with evolving data patterns."
            )
            dlp_doms = ["dlp_readiness", "label_coverage"]
        dlp_risk = _inject_risk_modal(dlp_risk, dlp_why, dlp_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600">{esc(name)}</td>'
            f'<td>{state_html}</td>'
            f'<td style="font-size:12px">{esc(workloads)}</td>'
            f'<td>{_domain_pills(dlp_doms)}</td>'
            f'<td>{dlp_risk}</td>'
            f'</tr>'
        )
    return (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{len(policies)}</span> DLP Policies</div>'
        f'</div>'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("Policy Name", "Name of the Data Loss Prevention policy.", "width:28%")}{_th_tip("State", "Whether this DLP policy is currently enabled and enforcing rules.", "width:8%")}'
        f'{_th_tip("Workloads", "M365 workloads covered by this policy: Exchange, SharePoint, Teams, OneDrive, etc.", "width:24%")}{_th_tip("Domain", "Compliance domain this policy impacts.", "width:22%")}{_th_tip("Copilot Risk", "Risk to Copilot: disabled DLP policies cannot prevent sensitive data from surfacing. Click for details.", "width:16%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_entra_apps_inventory(apps: list[dict]) -> str:
    """Render the Entra ID application registration inventory."""
    if not apps:
        return '<p class="empty">No Entra application registrations collected. Ensure Application.Read.All permission is granted.</p>'
    total = len(apps)
    graph_apps = sum(1 for a in apps if a.get("HasGraphAccess"))
    high_perm = sum(1 for a in apps if a.get("ApplicationPermissions", 0) >= 5)
    rows = ""
    for app in apps:
        name = app.get("DisplayName", "")
        audience = app.get("SignInAudience", "")
        audience_labels = {
            "AzureADMyOrg": "Single tenant",
            "AzureADMultipleOrgs": "Multi-tenant",
            "AzureADandPersonalMicrosoftAccount": "Multi + Personal",
            "PersonalMicrosoftAccount": "Personal only",
        }
        audience_display = audience_labels.get(audience, audience)
        has_graph = app.get("HasGraphAccess", False)
        graph_html = (
            '<span class="status-pill pass">Yes</span>' if has_graph
            else '<span class="status-pill not-assessed">No</span>'
        )
        n_del = app.get("DelegatedPermissions", 0)
        n_app = app.get("ApplicationPermissions", 0)
        total_p = app.get("TotalPermissions", 0)
        perm_color = "#D13438" if n_app >= 5 else "#FFB900" if total_p >= 10 else "var(--text)"
        certs = app.get("CertificateCount", 0)
        secrets = app.get("SecretCount", 0)
        cred_html = ""
        if certs > 0:
            cred_html += f'<span style="font-size:10px;color:#107C10" title="Certificates">&#128274; {certs}</span> '
        if secrets > 0:
            cred_html += f'<span style="font-size:10px;color:#FFB900" title="Secrets">&#128273; {secrets}</span>'
        if not cred_html:
            cred_html = '<span style="color:var(--text-muted)">—</span>'
        created = app.get("CreatedDate", "")
        # Copilot risk for Entra apps
        if n_app >= 5:
            app_risk = _copilot_risk_pill("high", "High app-only permissions — can access tenant data without user context")
            app_why = (
                f"The application '{name}' has {n_app} application-level (app-only) permissions. "
                f"Per Microsoft Entra ID documentation, app-only permissions (Application type in API permissions) "
                f"operate without a signed-in user context, granting this application direct programmatic access to "
                f"tenant data — including mail, files, calendar, and directory data. While Copilot uses user-delegated "
                f"access, this app represents a parallel data access path to the same tenant data Copilot grounds on."
            )
            app_how = (
                "1. Audit each application permission in Microsoft Entra admin center > App registrations > API permissions and remove unnecessary ones. "
                "2. Replace broad app-only permissions (e.g. Mail.Read, Files.Read.All) with delegated permissions where possible (ref: Microsoft identity platform least-privilege guidance). "
                "3. Use certificate credentials instead of client secrets for stronger authentication (Entra ID > App registrations > Certificates & secrets). "
                "4. Scope app access using application access policies (e.g. Exchange Online application access policy for mail). "
                "5. Monitor app activity in Microsoft Entra admin center > Sign-in logs > Service principal sign-ins."
            )
            app_doms = ["copilot_security", "access_governance", "oversharing_risk"]
        elif has_graph and total_p >= 10:
            app_risk = _copilot_risk_pill("medium", "Graph access with many permissions — broad data access surface")
            app_why = (
                f"The application '{name}' has Microsoft Graph API access with {total_p} total permissions "
                f"({n_del} delegated, {n_app} application). "
                f"Microsoft Graph is the unified API for M365 data. A high permission count increases the data access "
                f"surface — each Graph permission grants access to a specific resource type (users, mail, files, etc.). "
                f"This app can read tenant data via Graph, creating a parallel data access pathway alongside Copilot."
            )
            app_how = (
                "1. Review all permissions in Microsoft Entra admin center > App registrations > API permissions and remove unused ones. "
                "2. Apply least-privilege: replace broad permissions (e.g. User.Read.All → User.Read) per Microsoft identity platform guidance. "
                "3. Monitor the app's activity in Entra admin center > Sign-in logs > Service principal sign-ins. "
                "4. Configure admin consent workflow in Entra ID > Enterprise applications > Consent and permissions to control permission grants."
            )
            app_doms = ["copilot_security", "access_governance"]
        elif has_graph:
            app_risk = _copilot_risk_pill("medium", "Graph access — can read tenant data accessible to Copilot")
            app_why = (
                f"The application '{name}' has Microsoft Graph API access. "
                f"Microsoft Graph provides access to M365 tenant data (users, mail, files, calendar). While Copilot uses "
                f"user-delegated access for grounding, this app has an independent data access pathway to tenant resources."
            )
            app_how = (
                "1. Verify the app's Graph permissions are minimal and appropriate for its function in Microsoft Entra admin center > App registrations > API permissions. "
                "2. Monitor API call patterns in Entra admin center > Sign-in logs > Service principal sign-ins. "
                "3. Ensure consent is properly scoped — review in Entra ID > Enterprise applications > Permissions."
            )
            app_doms = ["copilot_security", "access_governance"]
        else:
            app_risk = _copilot_risk_pill("low", "Limited permissions — low data exposure risk")
            app_why = (
                f"The application '{name}' has limited permissions without direct Microsoft Graph API access. "
                f"Minimal permission scope means low data exposure risk per Microsoft identity platform least-privilege guidance."
            )
            app_how = "No immediate action required. Review permissions during regular access recertification in Microsoft Entra admin center > App registrations."
            app_doms = ["access_governance"]
        app_risk = _inject_risk_modal(app_risk, app_why, app_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600;font-size:12px">{esc(name)}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)">{esc(audience_display)}</td>'
            f'<td>{graph_html}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono);color:{perm_color}" title="Delegated: {n_del}, Application: {n_app}">{total_p}</td>'
            f'<td style="text-align:center;font-size:11px;font-family:var(--font-mono);color:{"#D13438" if n_app >= 5 else "var(--text-muted)"}">{n_app}</td>'
            f'<td style="text-align:center">{cred_html}</td>'
            f'<td>{_domain_pills(app_doms)}</td>'
            f'<td>{app_risk}</td>'
            f'</tr>'
        )
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Apps</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:var(--primary)">{graph_apps}</span> Graph Access</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#D13438">{high_perm}</span> High-Permission</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("App Name", "Display name of the registered application in Entra ID.", "width:20%")}{_th_tip("Audience", "Who can sign in: single tenant, multi-tenant, or personal accounts.", "width:10%")}'
        f'{_th_tip("Graph", "Whether this app has Microsoft Graph API access, meaning it can read tenant data.", "width:6%")}{_th_tip("Perms", "Total API permissions assigned (delegated + application). Hover for breakdown.", "width:5%")}'
        f'{_th_tip("App-Only", "Application-level permissions that work without a signed-in user — higher privilege.", "width:7%")}{_th_tip("Credentials", "Certificates and client secrets configured for authentication.", "width:8%")}'
        f'{_th_tip("Domain", "Compliance domain(s) this application impacts.", "width:22%")}'
        f'{_th_tip("Copilot Risk", "Risk to Copilot data security. Click for details.", "width:18%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_service_principal_inventory(sps: list[dict]) -> str:
    """Render the Entra service principal (enterprise app) inventory."""
    if not sps:
        return '<p class="empty">No service principals collected. Ensure Application.Read.All permission is granted.</p>'
    total = len(sps)
    enterprise_count = sum(1 for s in sps if s.get("IsEnterprise"))
    disabled_count = sum(1 for s in sps if not s.get("Enabled", True))
    rows = ""
    for sp in sps:
        name = sp.get("DisplayName", "")
        sp_type = sp.get("Type", "")
        is_enterprise = sp.get("IsEnterprise", False)
        enabled = sp.get("Enabled", True)
        type_html = (
            '<span class="status-pill pass">Enterprise</span>' if is_enterprise
            else f'<span class="status-pill not-assessed">{esc(sp_type)}</span>'
        )
        enabled_html = (
            '<span class="status-pill pass">Enabled</span>' if enabled
            else '<span class="status-pill fail">Disabled</span>'
        )
        role_count = sp.get("AppRoleAssignmentCount", 0)
        # Copilot risk for service principals
        if role_count >= 5 and enabled:
            sp_risk = _copilot_risk_pill("high", "Many role assignments — broad tenant data access surface")
            sp_why = (
                f"The service principal '{name}' is enabled with {role_count} app role assignments. "
                f"Per Microsoft Entra ID documentation, each app role assignment grants the SP access to a specific "
                f"application's API surface. With {role_count} assignments, this SP has broad access across multiple services. "
                f"While Copilot uses user-delegated access, a compromised SP with broad roles represents a parallel "
                f"data access path to tenant resources that Copilot also surfaces."
            )
            sp_how = (
                "1. Review all role assignments in Microsoft Entra admin center > Enterprise applications > select this SP > Permissions. "
                "2. Remove unnecessary role assignments per the principle of least privilege (Microsoft identity platform guidance). "
                "3. Check if this SP is used by an active application — if orphaned, consider disabling or deleting it. "
                "4. Monitor sign-in activity in Entra admin center > Sign-in logs > Service principal sign-ins. "
                "5. Replace SPs with managed identities where possible for Azure resource access (per Microsoft security best practices)."
            )
            sp_doms = ["copilot_security", "access_governance", "oversharing_risk"]
        elif not enabled:
            sp_risk = _copilot_risk_pill("medium", "Disabled SP — review for orphaned permissions")
            sp_why = (
                f"The service principal '{name}' is disabled but still exists in the directory"
                f"{f' with {role_count} role assignment(s)' if role_count > 0 else ''}. "
                f"Per Microsoft Entra ID documentation, disabling a SP prevents sign-in but does not remove its "
                f"role assignments or credentials. If re-enabled (accidentally or maliciously), all previous permissions "
                f"immediately become active — creating a latent access risk to tenant resources."
            )
            sp_how = (
                "1. Verify whether this SP is still needed — if not, delete it in Microsoft Entra admin center > Enterprise applications. "
                "2. If keeping it disabled, remove all role assignments in the SP's Permissions tab to eliminate orphaned access. "
                "3. Revoke credentials (certificates and secrets) in the associated App registration > Certificates & secrets. "
                "4. Document the reason for disabling in a change management record for audit compliance."
            )
            sp_doms = ["copilot_security", "access_governance"]
        else:
            sp_risk = _copilot_risk_pill("low", "Active with limited role assignments")
            sp_why = (
                f"The service principal '{name}' is enabled with {role_count} role assignment(s). "
                f"Limited role assignments indicate a narrow, well-scoped access surface per least-privilege principles."
            )
            sp_how = (
                "No immediate action required. Review role assignments during regular access recertification cycles. "
                "Rotate credentials on schedule per Microsoft Entra ID security best practices."
            )
            sp_doms = ["copilot_security", "access_governance"]
        sp_risk = _inject_risk_modal(sp_risk, sp_why, sp_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600;font-size:12px" title="{esc(name)}">{esc(name)}</td>'
            f'<td>{type_html}</td>'
            f'<td>{enabled_html}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono);font-size:12px">{role_count}</td>'
            f'<td>{_domain_pills(sp_doms)}</td>'
            f'<td>{sp_risk}</td>'
            f'</tr>'
        )
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Service Principals</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:var(--primary)">{enterprise_count}</span> Enterprise</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#D13438">{disabled_count}</span> Disabled</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("Name", "Display name of the service principal (enterprise application).", "width:22%")}{_th_tip("Type", "Enterprise = third-party or first-party app. Other types may be managed identities.", "width:14%")}'
        f'{_th_tip("Status", "Whether this service principal is currently enabled for sign-in.", "width:8%")}{_th_tip("Roles", "Number of app role assignments granted to this service principal.", "width:6%")}'
        f'{_th_tip("Domain", "Compliance domain(s) this service principal impacts.", "width:24%")}'
        f'{_th_tip("Copilot Risk", "Risk to Copilot: SPs with many role assignments have broad data access. Click for details.", "width:18%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def _render_app_protection_inventory(policies: list[dict]) -> str:
    """Render the Intune App Protection Policy (MAM) inventory."""
    if not policies:
        return '<p class="empty">No App Protection Policies collected. Ensure Intune DeviceManagementApps.Read.All permission is granted.</p>'
    total = len(policies)
    active = sum(1 for p in policies if p.get("IsActive", True))
    platforms = set(p.get("Platform", "") for p in policies if p.get("IsActive", True))
    rows = ""
    for pol in policies:
        name = pol.get("PolicyName", "")
        platform = pol.get("Platform", "")
        is_active = pol.get("IsActive", True)
        assigned_apps = pol.get("AssignedApps", 0)
        scope = pol.get("Scope", "")
        encryption_req = pol.get("EncryptionRequired", False)
        pin_req = pol.get("PinRequired", False)
        copy_blocked = pol.get("CopyPasteBlocked", False)
        status_html = (
            '<span class="status-pill pass">Active</span>' if is_active
            else '<span class="status-pill fail">Inactive</span>'
        )
        enc_html = (
            '<span class="status-pill pass">Yes</span>' if encryption_req
            else '<span class="status-pill not-assessed">No</span>'
        )
        pin_html = (
            '<span class="status-pill pass">Yes</span>' if pin_req
            else '<span class="status-pill not-assessed">No</span>'
        )
        copy_html = (
            '<span class="status-pill pass">Blocked</span>' if copy_blocked
            else '<span class="status-pill not-assessed">Allowed</span>'
        )
        # Copilot risk assessment
        if not is_active:
            pol_risk = _copilot_risk_pill("medium", "Inactive policy — not protecting Copilot data")
            pol_why = (
                f"The App Protection Policy '{name}' for {platform} is inactive. "
                f"Users on this platform can copy Copilot-generated content to unmanaged apps."
            )
            pol_how = (
                "1. Re-activate this policy in Intune admin center > Apps > App protection policies. "
                "2. Verify the policy targets the correct user groups with Copilot licenses."
            )
            pol_doms = ["access_governance", "copilot_security"]
        elif not encryption_req or not copy_blocked:
            pol_risk = _copilot_risk_pill("medium", "Weak data protection settings")
            pol_why = (
                f"The App Protection Policy '{name}' for {platform} is active but missing "
                f"{'encryption' if not encryption_req else ''}"
                f"{' and ' if not encryption_req and not copy_blocked else ''}"
                f"{'copy/paste blocking' if not copy_blocked else ''}. "
                f"This allows Copilot-generated sensitive content to leave managed apps."
            )
            pol_how = (
                "1. Edit this policy in Intune admin center > Apps > App protection policies. "
                "2. Enable 'Encrypt org data' and set 'Restrict cut, copy, paste' to 'Policy managed apps'."
            )
            pol_doms = ["access_governance", "copilot_security"]
        else:
            pol_risk = _copilot_risk_pill("low", "Strong data protection configured")
            pol_why = (
                f"The App Protection Policy '{name}' for {platform} has encryption and "
                f"copy/paste restrictions enabled — Copilot data is protected on this platform."
            )
            pol_how = "No action required. Monitor compliance in Intune > Apps > Monitor > App protection status."
            pol_doms = ["access_governance"]
        pol_risk = _inject_risk_modal(pol_risk, pol_why, pol_how)
        rows += (
            f'<tr>'
            f'<td style="font-weight:600">{esc(name)}</td>'
            f'<td>{esc(platform)}</td>'
            f'<td>{status_html}</td>'
            f'<td style="text-align:center;font-family:var(--font-mono)">{assigned_apps}</td>'
            f'<td style="font-size:11px;color:var(--text-secondary)">{esc(scope)}</td>'
            f'<td>{enc_html}</td>'
            f'<td>{pin_html}</td>'
            f'<td>{copy_html}</td>'
            f'<td>{_domain_pills(pol_doms)}</td>'
            f'<td>{pol_risk}</td>'
            f'</tr>'
        )
    platform_str = ", ".join(sorted(platforms)) if platforms else "None"
    summary = (
        f'<div class="ctrl-summary-bar">'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{total}</span> Policies</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count" style="color:#107C10">{active}</span> Active</div>'
        f'<div class="ctrl-summary-item"><span class="ctrl-summary-count">{len(platforms)}</span> Platforms ({platform_str})</div>'
        f'</div>'
    )
    return (
        f'{summary}'
        f'<table class="controls-table">'
        f'<thead><tr>{_th_tip("Policy Name", "Name of the App Protection Policy as defined in Intune.", "width:18%")}'
        f'{_th_tip("Platform", "Target platform (iOS, Android, Windows).", "width:7%")}'
        f'{_th_tip("Status", "Whether this policy is currently active and enforced.", "width:6%")}'
        f'{_th_tip("Apps", "Number of managed apps covered by this policy.", "width:5%")}'
        f'{_th_tip("Scope", "Assignment scope — which user groups this policy targets.", "width:14%")}'
        f'{_th_tip("Encryption", "Whether org data encryption is required on the device.", "width:8%")}'
        f'{_th_tip("PIN", "Whether a PIN or biometric is required to access org data.", "width:5%")}'
        f'{_th_tip("Copy/Paste", "Whether cut/copy/paste to unmanaged apps is blocked.", "width:8%")}'
        f'{_th_tip("Domain", "Compliance domain(s) this policy impacts.", "width:16%")}'
        f'{_th_tip("Copilot Risk", "Risk to Copilot data leakage on mobile. Click for details.", "width:13%")}'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


# ── Excel export ─────────────────────────────────────────────────────────

def _export_excel(results: dict, output_dir: pathlib.Path) -> pathlib.Path:
    """Export all assessment data to a multi-tab Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.warning("[CopilotReadinessReport] openpyxl not installed — skipping Excel export")
        return None  # type: ignore[return-value]

    wb = Workbook()
    thin = Side(style="thin", color="444444")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _add_sheet(name: str, headers: list[str], rows: list[list], col_widths: list[int] | None = None):
        ws = wb.create_sheet(title=name)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = border
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = wrap
                c.border = border
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        if col_widths:
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + ci) if ci <= 26 else "A"].width = w

    # -- Tab 1: Executive Summary
    scores = results.get("CopilotReadinessScores", {})
    cat_scores = scores.get("CategoryScores", {})
    compliance = scores.get("ComplianceBreakdown", {})
    sev_dist = scores.get("SeverityDistribution", {})

    summary_rows = [
        ["Overall Score", f'{scores.get("OverallScore", 0):.0f}/100'],
        ["Readiness Status", scores.get("ReadinessStatus", "N/A")],
        ["Total Findings", results.get("FindingCount", 0)],
        ["Subscriptions", results.get("SubscriptionCount", 0)],
        ["Compliant Controls", compliance.get("compliant", 0)],
        ["Gap Controls", compliance.get("gap", 0)],
        ["Partial Controls", compliance.get("partial", 0)],
        ["", ""],
        ["Severity Distribution", ""],
        ["Critical", sev_dist.get("critical", 0)],
        ["High", sev_dist.get("high", 0)],
        ["Medium", sev_dist.get("medium", 0)],
        ["Low", sev_dist.get("low", 0)],
        ["Informational", sev_dist.get("informational", 0)],
        ["", ""],
        ["Category Scores", ""],
    ]
    for cat_key, meta in _CATEGORY_META.items():
        cs = cat_scores.get(cat_key, {})
        summary_rows.append([meta["name"], f'{cs.get("Score", 0):.0f}/100 — {cs.get("Level", "N/A").replace("_", " ").upper()} ({cs.get("FindingCount", 0)} gaps)'])
    _add_sheet("Executive Summary", ["Metric", "Value"], summary_rows, [30, 60])

    # -- Tab 2: All Findings (full detail)
    findings = results.get("Findings", [])
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    sorted_findings = sorted(findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))
    finding_rows = []
    for i, f in enumerate(sorted_findings, 1):
        rem = f.get("Remediation", {})
        cat_meta = _CATEGORY_META.get(f.get("Category", ""), {})
        resources = f.get("AffectedResources", [])
        resource_names = "; ".join(
            str(r.get("Name", r.get("name", ""))) for r in resources[:50] if isinstance(r, dict)
        )
        finding_rows.append([
            i,
            f.get("Severity", "").upper(),
            {"quick_win": "Quick Win", "moderate": "Moderate", "major": "Major"}.get(f.get("Effort", ""), f.get("Effort", "")),
            cat_meta.get("name", f.get("Category", "")),
            f.get("Subcategory", ""),
            f.get("Title", ""),
            f.get("Description", ""),
            f.get("AffectedCount", 0),
            resource_names,
            rem.get("Description", ""),
            rem.get("PowerShell", ""),
            rem.get("AzureCLI", ""),
            "; ".join(rem.get("PortalSteps", [])),
        ])
    _add_sheet(
        "Findings",
        ["#", "Severity", "Effort", "Category", "Subcategory", "Title", "Description",
         "Affected Count", "Affected Resources", "Remediation", "PowerShell", "Azure CLI", "Portal Steps"],
        finding_rows,
        [5, 10, 12, 20, 20, 40, 50, 12, 40, 40, 40, 40, 40],
    )

    # -- Tab 3: Security Controls Matrix
    controls = results.get("SecurityControlsMatrix", [])
    ctrl_rows = []
    for c in controls:
        ctrl_rows.append([
            c.get("ControlId", ""),
            c.get("ControlName", ""),
            c.get("Category", ""),
            c.get("Status", ""),
            c.get("Severity", ""),
            c.get("MicrosoftReference", ""),
            _MS_REFERENCE_EXPLANATIONS.get(c.get("MicrosoftReference", ""), ""),
            c.get("Description", ""),
            c.get("Evidence", ""),
        ])
    _add_sheet(
        "Security Controls",
        ["Control ID", "Control Name", "Category", "Status", "Severity",
         "Remediation Guidance Area", "Detailed Guidance", "Description", "Evidence"],
        ctrl_rows,
        [12, 30, 20, 10, 10, 30, 50, 40, 30],
    )

    # -- Tab 4: SharePoint Sites (with Copilot Risk columns matching HTML)
    sites = results.get("SiteInventory", [])
    site_rows = []
    for s in sites:
        label = s.get("SensitivityLabel", "None")
        anon = s.get("AnonymousLinks", 0)
        # Compute Copilot Risk — mirrors HTML _copilot_risk_pill logic
        if s.get("IsOvershared") or anon > 0:
            risk_level = "HIGH"
            risk_why = (
                f"This site has {'broad permissions (Everyone/large groups)' if s.get('IsOvershared') else ''}"
                f"{' and ' if s.get('IsOvershared') and anon > 0 else ''}"
                f"{f'{anon} anonymous sharing link(s)' if anon > 0 else ''}. "
                "Copilot inherits the signed-in user's permissions, so any content reachable by broad or "
                "anonymous access will be surfaced in Copilot responses."
            )
            risk_how = (
                "1. Remove 'Everyone' permissions in SharePoint admin > Active sites > Permissions. "
                "2. Delete or convert anonymous sharing links to 'Specific people'. "
                "3. Apply a sensitivity label with encryption. "
                "4. Enable Restricted SharePoint Search (RSS) to exclude this site. "
                "5. Run DSPM for AI oversharing assessment in Purview."
            )
        elif s.get("IsStale") or label in ("None", ""):
            risk_level = "MEDIUM"
            reasons = []
            if s.get("IsStale"):
                reasons.append("no activity in 90+ days (stale)")
            if label in ("None", ""):
                reasons.append("no sensitivity label applied")
            risk_why = f"This site has {' and '.join(reasons)}. " + (
                "Stale sites may contain outdated information that Copilot surfaces as current. " if s.get("IsStale") else ""
            ) + (
                "Without a label, content cannot be protected by label-based encryption." if label in ("None", "") else ""
            )
            risk_how = (
                "1. Archive or delete stale content in SharePoint admin. "
                "2. Apply a sensitivity label via Purview > Information Protection. "
                "3. Configure retention policies for stale content. "
                "4. Exclude from Copilot via RSS until reviewed."
            )
        else:
            risk_level = "LOW"
            risk_why = "This site is actively maintained and has a sensitivity label applied."
            risk_how = "No immediate action. Verify label includes encryption if content is sensitive."
        # Domain mapping
        domains = []
        if s.get("IsOvershared") or anon > 0:
            domains.append("Oversharing Risk")
        if label in ("None", ""):
            domains.append("Label Coverage")
        if s.get("IsStale"):
            domains.append("Content Lifecycle")
        if not domains:
            domains.append("Compliant")
        site_rows.append([
            s.get("SiteName", ""),
            s.get("SiteUrl", ""),
            s.get("SiteTemplate", ""),
            label,
            "Yes" if s.get("IsStale") else "No",
            "Yes" if s.get("IsOvershared") else "No",
            s.get("TotalPermissions", 0),
            f"Owners: {s.get('OwnerCount', 0)}, Members: {s.get('MemberCount', 0)}, Guests: {s.get('GuestCount', 0)}, External: {s.get('ExternalUserCount', 0)}",
            s.get("AnonymousLinks", 0),
            s.get("OrganizationLinks", 0),
            s.get("TotalSharedItems", 0),
            s.get("SharingCapability", ""),
            str(s.get("LastModified", "")),
            "; ".join(domains),
            risk_level,
            risk_why,
            risk_how,
        ])
    _add_sheet(
        "SharePoint Sites",
        ["Site Name", "URL", "Template", "Sensitivity Label", "Stale", "Overshared",
         "Permissions", "Permission Detail", "Anon Links", "Org Links", "Shared Items",
         "Sharing Capability", "Last Modified", "Domain", "Copilot Risk", "Risk Reason", "Remediation"],
        site_rows,
        [25, 35, 12, 15, 8, 10, 10, 35, 10, 10, 10, 15, 18, 20, 10, 45, 45],
    )

    # -- Tab 5: Teams & Groups (with Copilot Risk)
    groups = results.get("GroupsInventory", [])
    group_rows = []
    for g in groups:
        vis = (g.get("Visibility", "") or "").lower()
        members = g.get("MemberCount", 0)
        guests = g.get("GuestCount", 0)
        label_g = g.get("SensitivityLabel", "")
        if vis == "public":
            g_risk = "HIGH"
            g_why = "Public group — any org member can join and access all content. Copilot can surface this content to any user with self-service join."
            g_how = "1. Change visibility to Private in Teams admin. 2. Apply a sensitivity label with container settings. 3. Review membership."
        elif guests > 0 and not label_g:
            g_risk = "MEDIUM"
            g_why = f"Group has {guests} guest(s) without a sensitivity label. External users can access group content surfaced by Copilot."
            g_how = "1. Apply a sensitivity label. 2. Review guest access. 3. Configure expiration for guest accounts."
        else:
            g_risk = "LOW"
            g_why = "Private group with controlled membership."
            g_how = "No immediate action. Review membership periodically."
        domains_g = []
        if vis == "public":
            domains_g.append("Oversharing Risk")
        if not label_g:
            domains_g.append("Label Coverage")
        if guests > 0:
            domains_g.append("Access Governance")
        if not domains_g:
            domains_g.append("Compliant")
        group_rows.append([
            g.get("DisplayName", ""),
            g.get("GroupType", ""),
            members,
            guests,
            g.get("Visibility", ""),
            g.get("HasTeam", ""),
            label_g,
            str(g.get("CreatedDateTime", "")),
            "; ".join(domains_g),
            g_risk,
            g_why,
            g_how,
        ])
    _add_sheet(
        "Teams & Groups",
        ["Name", "Type", "Members", "Guests", "Visibility", "Has Team",
         "Sensitivity Label", "Created", "Domain", "Copilot Risk", "Risk Reason", "Remediation"],
        group_rows,
        [25, 12, 10, 10, 12, 10, 15, 18, 20, 10, 45, 45],
    )

    # -- Tab 6: Licenses (with Usage % and Copilot Risk)
    licenses = results.get("LicenseInventory", [])
    lic_rows = []
    copilot_sku_keywords = ("copilot", "m365_copilot", "microsoft_365_copilot")
    for l in licenses:
        enabled = l.get("EnabledUnits", 0)
        consumed = l.get("ConsumedUnits", 0)
        available = enabled - consumed
        usage_pct = f"{(consumed / enabled * 100):.0f}%" if enabled > 0 else "0%"
        sku = (l.get("SkuPartNumber", "") or "").lower().replace(" ", "_").replace("-", "_")
        is_copilot = any(kw in sku for kw in copilot_sku_keywords)
        if is_copilot and consumed == 0:
            l_risk = "HIGH"
            l_why = "Copilot licenses purchased but 0 assigned — wasted investment and no Copilot adoption."
            l_how = "1. Assign licenses to target users in M365 admin center > Billing > Licenses. 2. Start with pilot group."
        elif is_copilot and enabled > 0 and consumed / enabled < 0.5:
            l_risk = "MEDIUM"
            l_why = f"Low Copilot license utilization ({usage_pct}). Licenses may be unassigned or users not adopted."
            l_how = "1. Review assignment in M365 admin > Billing. 2. Drive adoption with training and enablement."
        else:
            l_risk = "LOW"
            l_why = "License utilization is healthy." if not is_copilot else f"Copilot license utilization at {usage_pct}."
            l_how = "No action needed."
        lic_rows.append([
            l.get("SkuPartNumber", ""),
            enabled,
            consumed,
            available,
            usage_pct,
            "Yes" if is_copilot else "",
            l_risk if is_copilot else "",
            l_why if is_copilot else "",
            l_how if is_copilot else "",
        ])
    _add_sheet(
        "Licenses",
        ["SKU", "Enabled", "Consumed", "Available", "Usage %", "Copilot SKU",
         "Copilot Risk", "Risk Reason", "Remediation"],
        lic_rows,
        [35, 10, 10, 10, 10, 10, 10, 45, 45],
    )

    # -- Tab 7: Sensitivity Labels (with Encryption, Site & Group, Risk)
    labels = results.get("LabelInventory", [])
    label_rows = []
    for lb in labels:
        is_active = lb.get("IsActive", True)
        has_encrypt = lb.get("HasEncryption", lb.get("Encryption", ""))
        has_site_group = lb.get("HasSiteGroupSettings", lb.get("SiteGroupSettings", ""))
        if not is_active:
            lb_risk = "MEDIUM"
            lb_why = "Inactive label — gap in classification taxonomy. Content tagged with this label may lose protection if the label is removed."
            lb_how = "1. Re-activate the label or migrate tagged content to an active label. 2. Update auto-labeling policies."
        elif not has_encrypt:
            lb_risk = "MEDIUM"
            lb_why = "Label does not include encryption. Content with this label can be surfaced by Copilot to any user with site access."
            lb_how = "1. Add encryption to the label in Purview > Information Protection. 2. Choose 'Assign permissions now' targeting appropriate groups."
        else:
            lb_risk = "LOW"
            lb_why = "Label is active with encryption protection."
            lb_how = "No action needed. Verify encryption configuration periodically."
        label_rows.append([
            lb.get("DisplayName", ""),
            lb.get("LabelId", ""),
            lb.get("Priority", ""),
            "Yes" if is_active else "No",
            lb.get("ContentType", ""),
            lb.get("ParentLabel", ""),
            "Yes" if has_encrypt else "No",
            "Yes" if has_site_group else "No",
            lb_risk,
            lb_why,
            lb_how,
        ])
    _add_sheet(
        "Sensitivity Labels",
        ["Label Name", "Label ID", "Priority", "Active", "Content Type", "Parent Label",
         "Encryption", "Site & Group Settings", "Copilot Risk", "Risk Reason", "Remediation"],
        label_rows,
        [25, 36, 10, 8, 15, 20, 10, 15, 10, 45, 45],
    )

    # -- Tab 8: Conditional Access (with Controls flags and Risk)
    ca_policies = results.get("CAPolicyInventory", [])
    ca_rows = []
    for ca in ca_policies:
        state = (ca.get("State", "") or "").lower()
        grant = str(ca.get("GrantControls", ""))
        # Parse control flags matching HTML badges
        controls_flags = []
        grant_lower = grant.lower()
        if "mfa" in grant_lower or "multifactor" in grant_lower:
            controls_flags.append("MFA")
        if "compliant" in grant_lower:
            controls_flags.append("Compliant Device")
        if "block" in grant_lower:
            controls_flags.append("Block")
        if "domainjoined" in grant_lower or "hybrid" in grant_lower:
            controls_flags.append("Hybrid Joined")
        if not controls_flags and grant:
            controls_flags.append(grant[:50])
        if state in ("disabled",):
            ca_risk = "HIGH"
            ca_why = "Policy is disabled — enforcement rules are not applied. Users bypass this security control when accessing Copilot."
            ca_how = "1. Enable the policy (start with Report-only). 2. Test with 'What If'. 3. Move to enforced after validation."
        elif state in ("enabledforreportingbutntenforced", "reportonly"):
            ca_risk = "MEDIUM"
            ca_why = "Policy is in Report-only mode — monitoring but not enforcing. Security controls are not blocking risky access."
            ca_how = "1. Review sign-in logs for policy impact. 2. Move to Enabled when ready. 3. Monitor for disruption."
        else:
            ca_risk = "LOW"
            ca_why = "Policy is enabled and enforcing security controls."
            ca_how = "No action needed. Review policy scope periodically."
        ca_rows.append([
            ca.get("DisplayName", ""),
            ca.get("State", ""),
            grant,
            "; ".join(controls_flags),
            ca.get("SessionControls", ""),
            ca.get("Conditions", ""),
            str(ca.get("CreatedDateTime", "")),
            ca_risk,
            ca_why,
            ca_how,
        ])
    _add_sheet(
        "Conditional Access",
        ["Policy Name", "State", "Grant Controls", "Control Flags", "Session Controls",
         "Conditions", "Created", "Copilot Risk", "Risk Reason", "Remediation"],
        ca_rows,
        [30, 12, 25, 20, 25, 30, 18, 10, 45, 45],
    )

    # -- Tab 9: DLP Policies (with Risk)
    dlp = results.get("DLPInventory", [])
    dlp_rows = []
    for d in dlp:
        d_mode = (d.get("Mode", d.get("State", "")) or "").lower()
        if d_mode in ("disabled", "off"):
            d_risk = "HIGH"
            d_why = "DLP policy is disabled — no data loss prevention enforcement. Sensitive data can flow freely through Copilot-accessible channels."
            d_how = "1. Enable the policy in Purview > DLP. 2. Verify SIT definitions. 3. Cover Teams, Exchange, SPO. 4. Test and monitor."
        elif d_mode in ("testwithoutnotes", "testwithnotes", "simulate"):
            d_risk = "MEDIUM"
            d_why = "DLP policy is in test/simulation mode — monitoring but not blocking. Sensitive data leakage via Copilot is not prevented."
            d_how = "1. Review test results. 2. Move to enforcement mode when validated."
        else:
            d_risk = "LOW"
            d_why = "DLP policy is active and enforcing."
            d_how = "No action needed. Review policy coverage periodically."
        dlp_rows.append([
            d.get("PolicyName", d.get("Name", "")),
            d.get("Mode", d.get("State", "")),
            d.get("Priority", ""),
            d.get("Workloads", d.get("Locations", "")),
            d.get("RuleCount", ""),
            str(d.get("CreatedDate", d.get("CreatedDateTime", ""))),
            d_risk,
            d_why,
            d_how,
        ])
    _add_sheet(
        "DLP Policies",
        ["Policy Name", "Mode", "Priority", "Workloads/Locations", "Rule Count", "Created",
         "Copilot Risk", "Risk Reason", "Remediation"],
        dlp_rows,
        [30, 12, 10, 25, 10, 18, 10, 45, 45],
    )

    # -- Tab 10: Entra Apps (with Graph Access, Perm breakdown, Credentials, Risk)
    apps = results.get("EntraAppsInventory", [])
    app_rows = []
    for a in apps:
        has_graph = a.get("HasGraphAccess", False)
        delegated = a.get("DelegatedPermissions", 0)
        app_perms = a.get("ApplicationPermissions", 0)
        total = a.get("TotalPermissions", delegated + app_perms)
        certs = a.get("CertificateCount", 0)
        secrets = a.get("SecretCount", 0)
        cred_str = []
        if certs > 0:
            cred_str.append(f"{certs} cert(s)")
        if secrets > 0:
            cred_str.append(f"{secrets} secret(s)")
        if app_perms >= 5 and has_graph:
            a_risk = "HIGH"
            a_why = f"App has {app_perms} application-level Graph permissions — highest privilege. Can read org data without user context."
            a_how = "1. Review permissions in Entra > App registrations. 2. Remove unnecessary app permissions. 3. Convert to delegated. 4. Require admin consent."
        elif has_graph and (total >= 10 or app_perms > 0):
            a_risk = "MEDIUM"
            a_why = f"App has Graph access with {total} permissions ({app_perms} application-level). Can access org data Copilot also indexes."
            a_how = "1. Audit permissions. 2. Apply least privilege. 3. Review delegated vs application permissions."
        elif has_graph:
            a_risk = "MEDIUM"
            a_why = "App has Graph API access. Review data access scope."
            a_how = "1. Verify permissions are minimal. 2. Monitor sign-in activity."
        else:
            a_risk = "LOW"
            a_why = "App does not access Microsoft Graph."
            a_how = "No action needed for Copilot risk."
        app_rows.append([
            a.get("DisplayName", ""),
            a.get("AppId", ""),
            a.get("SignInAudience", ""),
            a.get("PublisherDomain", ""),
            "Yes" if has_graph else "No",
            delegated,
            app_perms,
            total,
            "; ".join(cred_str) if cred_str else "None",
            "; ".join(a.get("RequiredPermissions", [])) if isinstance(a.get("RequiredPermissions"), list) else str(a.get("RequiredPermissions", "")),
            str(a.get("CreatedDateTime", "")),
            a_risk,
            a_why,
            a_how,
        ])
    _add_sheet(
        "Entra Apps",
        ["App Name", "App ID", "Sign-In Audience", "Publisher Domain", "Graph Access",
         "Delegated Perms", "App Perms", "Total Perms", "Credentials",
         "Permission Details", "Created", "Copilot Risk", "Risk Reason", "Remediation"],
        app_rows,
        [25, 36, 18, 20, 10, 12, 10, 10, 15, 35, 18, 10, 45, 45],
    )

    # -- Tab 11: Service Principals (with Risk)
    sps = results.get("ServicePrincipalInventory", [])
    sp_rows = []
    for sp in sps:
        sp_enabled = sp.get("AccountEnabled", sp.get("Enabled", True))
        sp_roles = sp.get("AppRoleAssignmentCount", sp.get("RoleAssignments", 0))
        sp_type = sp.get("ServicePrincipalType", sp.get("Type", ""))
        if sp_roles >= 5 and sp_enabled:
            sp_risk = "HIGH"
            sp_why = f"Service principal has {sp_roles} role assignments while enabled — high blast radius if compromised."
            sp_how = "1. Review role assignments in Entra admin > Enterprise apps. 2. Remove unnecessary roles. 3. Apply least privilege."
        elif not sp_enabled and sp_roles > 0:
            sp_risk = "MEDIUM"
            sp_why = f"Disabled service principal with {sp_roles} orphaned role assignment(s). May indicate stale permissions."
            sp_how = "1. Clean up orphaned permissions. 2. Delete the service principal if no longer needed."
        else:
            sp_risk = "LOW"
            sp_why = "Service principal has appropriate permissions."
            sp_how = "No action needed."
        sp_rows.append([
            sp.get("DisplayName", ""),
            sp_type,
            "Yes" if sp_enabled else "No",
            sp_roles,
            sp_risk,
            sp_why,
            sp_how,
        ])
    _add_sheet(
        "Service Principals",
        ["Name", "Type", "Enabled", "Role Assignments", "Copilot Risk", "Risk Reason", "Remediation"],
        sp_rows,
        [30, 18, 10, 15, 10, 45, 45],
    )

    # -- Tab 12: App Protection Policies (NEW — was HTML only)
    app_prot = results.get("AppProtectionInventory", [])
    ap_rows = []
    for ap in app_prot:
        ap_status = ap.get("Status", ap.get("IsActive", ""))
        ap_platform = ap.get("Platform", "")
        ap_encrypt = ap.get("RequiresEncryption", ap.get("Encryption", ""))
        ap_pin = ap.get("RequiresPin", ap.get("PIN", ""))
        ap_copypaste = ap.get("BlockCopyPaste", ap.get("CopyPaste", ""))
        if str(ap_status).lower() in ("inactive", "disabled", "false"):
            ap_risk = "MEDIUM"
            ap_why = "App protection policy is inactive — managed apps on this platform are unprotected."
            ap_how = "1. Activate the policy in Intune admin > Apps > App protection. 2. Verify assignment scope."
        elif not ap_encrypt and not ap_pin:
            ap_risk = "MEDIUM"
            ap_why = "App protection policy lacks both encryption and PIN requirements — weak mobile data protection."
            ap_how = "1. Enable org data encryption. 2. Require PIN/biometric for app access. 3. Block copy/paste to unmanaged apps."
        else:
            ap_risk = "LOW"
            ap_why = "App protection policy has strong settings."
            ap_how = "No action needed."
        ap_rows.append([
            ap.get("PolicyName", ap.get("DisplayName", "")),
            ap_platform,
            str(ap_status),
            ap.get("AppCount", ap.get("Apps", "")),
            ap.get("Scope", ap.get("AssignmentScope", "")),
            "Yes" if ap_encrypt else "No",
            "Yes" if ap_pin else "No",
            "Blocked" if ap_copypaste else "Allowed",
            ap_risk,
            ap_why,
            ap_how,
        ])
    if ap_rows:
        _add_sheet(
            "App Protection",
            ["Policy Name", "Platform", "Status", "Apps", "Scope",
             "Encryption", "PIN", "Copy/Paste", "Copilot Risk", "Risk Reason", "Remediation"],
            ap_rows,
            [25, 10, 10, 8, 25, 10, 8, 10, 10, 45, 45],
        )

    # -- Tab 13: Category Detail (per-category summary)
    cat_detail_rows = []
    for cat_key, meta in _CATEGORY_META.items():
        cs = cat_scores.get(cat_key, {})
        cat_findings = [f for f in findings if f.get("Category") == cat_key]
        for f in sorted(cat_findings, key=lambda x: sev_order.get(x.get("Severity", "medium").lower(), 5)):
            cat_detail_rows.append([
                meta["name"],
                f'{cs.get("Score", 0):.0f}/100',
                cs.get("Level", "").replace("_", " ").upper(),
                f.get("Severity", "").upper(),
                f.get("Title", ""),
                f.get("Description", ""),
                f.get("Remediation", {}).get("Description", ""),
            ])
        if not cat_findings:
            cat_detail_rows.append([
                meta["name"],
                f'{cs.get("Score", 0):.0f}/100',
                cs.get("Level", "").replace("_", " ").upper(),
                "",
                "No gaps in this category",
                "",
                "",
            ])
    _add_sheet(
        "Category Detail",
        ["Domain", "Score", "Status", "Severity", "Finding", "Description", "Remediation"],
        cat_detail_rows,
        [22, 10, 14, 10, 35, 45, 40],
    )

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    xlsx_path = output_dir / "copilot-readiness-assessment.xlsx"
    wb.save(str(xlsx_path))
    log.info("[CopilotReadinessReport] Excel export: %s (%d KB)", xlsx_path, xlsx_path.stat().st_size // 1024)
    return xlsx_path


def generate_copilot_readiness_report(results: dict, output_dir: str | pathlib.Path) -> pathlib.Path:
    """Generate the M365 Copilot Readiness Assessment HTML report."""
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "copilot-readiness-assessment.html"

    ts = format_date_short()
    scores = results.get("CopilotReadinessScores", {})
    findings = results.get("Findings", [])
    finding_count = results.get("FindingCount", len(findings))
    sub_count = results.get("SubscriptionCount", 0)
    assessed_at = results.get("AssessedAt", "")
    assessment_id = results.get("AssessmentId", "")
    tenant_id = results.get("TenantId", "")
    tenant_display = results.get("TenantDisplayName", "")
    evidence_count = results.get("EvidenceRecordCount", 0)
    evidence_source = results.get("EvidenceSource", "targeted_collection")
    report_id = f"CIQ-CR-{assessed_at[:10].replace('-', '')}-{assessed_at[11:19].replace(':', '')}" if len(assessed_at) >= 19 else f"CIQ-CR-{ts.replace('-', '').replace(':', '').replace(' ', '-')}"

    overall_score = scores.get("OverallScore", 0)
    readiness_status = scores.get("ReadinessStatus", "NOT READY")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})
    compliance = scores.get("ComplianceBreakdown", {})

    n_crit = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_med = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)

    rmeta = _READINESS_META.get(readiness_status, _READINESS_META["NOT READY"])
    status_color = rmeta["color"]

    score_ring = _ring_score_svg(overall_score, size=160)

    sev_donut = _donut_svg([
        ("Critical", n_crit, "#D13438"), ("High", n_high, "#F7630C"),
        ("Medium", n_med, "#FFB900"), ("Low", n_low, "#107C10"), ("Info", n_info, "#A8A6A3"),
    ], size=140)

    max_sev = max(n_crit, n_high, n_med, n_low, n_info, 1)
    sev_bars = ""
    for name, count, color in [("Critical", n_crit, "#D13438"), ("High", n_high, "#F7630C"),
                                ("Medium", n_med, "#FFB900"), ("Low", n_low, "#107C10"), ("Info", n_info, "#A8A6A3")]:
        pct = (count / max_sev) * 100 if max_sev > 0 else 0
        sev_bars += (f'<div class="sev-row"><span class="sev-label">{name}</span>'
                     f'<div class="sev-track"><div class="sev-fill" style="width:{pct:.0f}%;background:{color}"></div></div>'
                     f'<span class="sev-count">{count}</span></div>')

    # Category cards
    cats_html = ""
    for cat_key, meta in _CATEGORY_META.items():
        cs = cat_scores.get(cat_key, {"Score": 100.0, "Level": "ready", "FindingCount": 0})
        c_score = cs.get("Score", 100.0)
        c_level = cs.get("Level", "ready")
        c_count = cs.get("FindingCount", 0)
        level_colors = {"not_ready": "#D13438", "needs_work": "#F7630C", "mostly_ready": "#FFB900", "ready": "#107C10"}
        c_color = level_colors.get(c_level, "#107C10")
        cats_html += (
            f'<a href="#ctrl-{cat_key}" class="category-card-link"><div id="cat-{cat_key}" class="category-card tip" data-tip="{esc(meta["description"])}&#10;YOUR TENANT:&#10;&#8226; Score: {c_score:.0f}/100 ({c_level.replace("_", " ").upper()})&#10;&#8226; {c_count} gap{"s" if c_count != 1 else ""} detected"><div class="category-icon">{meta["icon"]}</div>'
            f'<div class="category-name">{esc(meta["name"])}</div>'
            f'<div class="category-score" style="color:{c_color}">{c_score:.0f}</div>'
            f'<div class="category-level" style="color:{c_color}">{esc(c_level.replace("_", " ").upper())}</div>'
            f'<div class="category-findings">{c_count} gap{"s" if c_count != 1 else ""}</div></div></a>'
        )

    # Compliance breakdown
    n_compliant = compliance.get("compliant", 0)
    n_gap = compliance.get("gap", 0)
    n_partial = compliance.get("partial", 0)
    compliance_donut = _donut_svg([
        ("Compliant", n_compliant, "#107C10"), ("Gap", n_gap, "#D13438"), ("Partial", n_partial, "#FFB900"),
    ], size=140)

    # Findings HTML grouped by category
    findings_by_cat: dict[str, list[dict]] = {}
    for f in findings:
        findings_by_cat.setdefault(f.get("Category", "unknown"), []).append(f)

    # Build collapsible findings detail table
    cat_options = ""
    detail_rows = ""
    fidx = 0
    _nav_findings_items_list: list[str] = []
    for cat_key, cat_findings in sorted(findings_by_cat.items()):
        meta = _CATEGORY_META.get(cat_key, {"name": cat_key.title(), "icon": "&#128196;"})
        cat_options += f'<option value="{esc(cat_key)}">{esc(meta["name"])} ({len(cat_findings)})</option>'
        sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
        sorted_f = sorted(cat_findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))
        detail_rows += (
            f'<tr class="findings-cat-header" id="find-{cat_key}"><td colspan="7">'
            f'{meta["icon"]} {esc(meta["name"])} ({len(cat_findings)} gaps)</td></tr>'
        )
        _nav_findings_items_list.append(
            f'<a href="#find-{cat_key}">{meta["icon"]} {esc(meta["name"])} ({len(cat_findings)})</a>'
        )
        for f in sorted_f:
            fidx += 1
            detail_rows += _render_finding_row(fidx, f)
    _nav_findings_items = "".join(_nav_findings_items_list)

    # Effort summary counts
    n_quick = sum(1 for f in findings if f.get("Effort") == "quick_win")
    n_mod = sum(1 for f in findings if f.get("Effort") == "moderate")
    n_major = sum(1 for f in findings if f.get("Effort") == "major")
    effort_summary = (
        f'<div class="ctrl-summary-bar" style="margin-bottom:12px">'
        f'<div class="ctrl-summary-item">{_effort_badge("quick_win")} <span style="font-size:13px;margin-left:4px"><strong>{n_quick}</strong> Quick Wins</span></div>'
        f'<div class="ctrl-summary-item">{_effort_badge("moderate")} <span style="font-size:13px;margin-left:4px"><strong>{n_mod}</strong> Moderate</span></div>'
        f'<div class="ctrl-summary-item">{_effort_badge("major")} <span style="font-size:13px;margin-left:4px"><strong>{n_major}</strong> Major Projects</span></div>'
        f'</div>'
    )

    all_findings_detail_html = (
        f'{effort_summary}'
        f'<table class="findings-detail-table">'
        f'<thead><tr>'
        f'{_th_tip("#", "Row number for reference.", "width:30px")}'
        f'{_th_tip("Severity", "Impact level of this gap. Critical > High > Medium > Low > Info.", "width:80px")}'
        f'{_th_tip("Effort", "Estimated remediation effort. Quick Win = under 1hr admin change. Moderate = days. Major = organizational rollout.", "width:80px")}'
        f'{_th_tip("Finding", "Click any row to expand details, affected resources, and remediation commands.", "min-width:200px")}'
        f'{_th_tip("Category", "Which readiness domain this finding belongs to.", "width:120px")}'
        f'{_th_tip("Affected", "Number of resources or configurations affected.", "width:60px")}'
        f'{_th_tip("Remediation", "Recommended action to resolve this gap. Click row for full details with commands.")}'
        f'</tr></thead><tbody>{detail_rows}</tbody></table>'
    ) if detail_rows else '<p class="empty">No gaps found — your tenant is ready for Copilot!</p>'

    # Category bar chart — REMOVED (duplicate of category cards)

    # Security controls matrix
    controls_matrix = results.get("SecurityControlsMatrix", [])
    controls_table_html = _render_controls_table(controls_matrix)

    # Site inventory
    site_inventory = results.get("SiteInventory", [])
    site_inventory_html = _render_site_inventory(site_inventory)

    # Additional resource inventories
    license_inventory = results.get("LicenseInventory", [])
    license_inv_html = _render_license_inventory(license_inventory)
    label_inventory = results.get("LabelInventory", [])
    label_inv_html = _render_label_inventory(label_inventory)
    ca_inventory = results.get("CAPolicyInventory", [])
    ca_inv_html = _render_ca_inventory(ca_inventory)
    groups_inventory = results.get("GroupsInventory", [])
    groups_inv_html = _render_groups_inventory(groups_inventory)
    dlp_inventory = results.get("DLPInventory", [])
    dlp_inv_html = _render_dlp_inventory(dlp_inventory)
    entra_apps_inventory = results.get("EntraAppsInventory", [])
    entra_apps_inv_html = _render_entra_apps_inventory(entra_apps_inventory)
    sp_inventory = results.get("ServicePrincipalInventory", [])
    sp_inv_html = _render_service_principal_inventory(sp_inventory)
    app_protection_inventory = results.get("AppProtectionInventory", [])
    app_protection_inv_html = _render_app_protection_inventory(app_protection_inventory)



    # D4: Collection warnings panel
    collection_warnings = [
        f for f in findings if f.get("Subcategory") == "unable_to_assess"
    ]
    warnings_html = ""
    if collection_warnings:
        warning_items = ""
        for w in collection_warnings:
            rem = w.get("Remediation", {})
            warning_items += (
                f'<div style="padding:10px;border-left:3px solid #FFB900;margin-bottom:8px;background:rgba(255,185,0,.06);border-radius:0 6px 6px 0">'
                f'<div style="font-weight:600;font-size:13px">&#9888; {esc(w.get("Title", ""))}</div>'
                f'<div style="font-size:12px;color:var(--text-secondary);margin-top:4px">{esc(rem.get("Description", ""))}</div>'
                f'</div>'
            )
        warnings_html = (
            f'<section class="section" aria-labelledby="warnings-heading">'
            f'<h2 id="warnings-heading" class="tip" data-tip="Items that could not be assessed due to missing permissions or API errors. Fix these and re-run.">&#9888;&#65039; Collection Warnings</h2>'
            f'<div class="how-to-read"><h4>Some data could not be collected</h4>'
            f'<p>The following issues were detected during evidence collection. '
            f'Fix the underlying permission gaps and re-run for a complete assessment.</p></div>'
            f'{warning_items}</section>'
        )

    # D2: Trend rendering
    trend_data = results.get("Trend")
    trend_html = ""
    if trend_data:
        prev_score = trend_data.get("PreviousScore", 0)
        curr_score = trend_data.get("CurrentScore", 0)
        delta = trend_data.get("ScoreDelta", 0)
        new_count = trend_data.get("NewCount", 0)
        resolved_count = trend_data.get("ResolvedCount", 0)
        arrow = "&#8593;" if delta > 0 else "&#8595;" if delta < 0 else "&#8594;"
        delta_color = "#107C10" if delta > 0 else "#D13438" if delta < 0 else "#A8A6A3"

        new_findings_html = ""
        for f in trend_data.get("NewFindings", [])[:5]:
            new_findings_html += f'<div style="padding:4px 0;font-size:12px">{_severity_badge(f.get("Severity", ""))} {esc(f.get("Title", ""))}</div>'
        resolved_findings_html = ""
        for f in trend_data.get("ResolvedFindings", [])[:5]:
            resolved_findings_html += f'<div style="padding:4px 0;font-size:12px">&#10003; {esc(f.get("Title", ""))}</div>'

        trend_html = (
            f'<section class="section" aria-labelledby="trend-heading">'
            f'<h2 id="trend-heading" class="tip" data-tip="Comparison with a previous assessment showing score change, new gaps, and resolved items.">&#128200; Trend Comparison</h2>'
            f'<div class="exec-grid">'
            f'<div class="exec-panel">'
            f'<h3>Score Change</h3>'
            f'<div style="font-size:36px;font-weight:700;font-family:var(--font-mono);color:{delta_color}">{arrow} {abs(delta):.1f}</div>'
            f'<div style="font-size:13px;color:var(--text-secondary);margin-top:4px">{prev_score:.0f} &rarr; {curr_score:.0f}</div>'
            f'</div>'
            f'<div class="exec-panel">'
            f'<h3>New Gaps ({new_count})</h3>'
            f'{new_findings_html if new_findings_html else "<p class=empty>No new gaps</p>"}'
            f'</div>'
            f'<div class="exec-panel">'
            f'<h3>Resolved ({resolved_count})</h3>'
            f'{resolved_findings_html if resolved_findings_html else "<p class=empty>No resolved gaps</p>"}'
            f'</div></div></section>'
        )

    # D3: Suppressed findings section
    suppressed = results.get("SuppressedFindings", [])
    suppressed_html = ""
    if suppressed:
        sup_rows = ""
        for f in suppressed:
            sup_rows += (
                f'<div style="padding:8px 12px;border-bottom:1px solid var(--border);display:flex;gap:12px;align-items:center;font-size:12px;opacity:.7">'
                f'<span>{_severity_badge(f.get("Severity", ""))}</span>'
                f'<span style="flex:1">{esc(f.get("Title", ""))}</span>'
                f'<span style="color:var(--text-muted)">{esc(f.get("Category", ""))}</span>'
                f'</div>'
            )
        suppressed_html = (
            f'<section class="section" aria-labelledby="suppressed-heading">'
            f'<h2 id="suppressed-heading" class="tip" data-tip="Findings suppressed via accepted-risk rules. These do not count toward the readiness score.">&#128275; Suppressed Findings ({len(suppressed)})</h2>'
            f'<details><summary style="cursor:pointer;color:var(--primary);font-weight:600;font-size:13px;padding:8px 0">'
            f'Show {len(suppressed)} suppressed (accepted-risk) findings</summary>'
            f'<div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-top:8px">'
            f'{sup_rows}</div></details></section>'
        )

    # Executive summary
    if readiness_status == "READY":
        exec_text = f"Your tenant is <strong>ready for Microsoft 365 Copilot</strong> with a readiness score of {overall_score:.0f}/100."
    elif readiness_status == "NEEDS WORK":
        exec_text = (f"Your tenant <strong>needs additional preparation</strong> before M365 Copilot rollout. "
                     f"Readiness score: {overall_score:.0f}/100 with {n_crit} critical and {n_high} high gaps.")
    else:
        exec_text = (f"Your tenant is <strong>not ready for M365 Copilot</strong>. "
                     f"Readiness score: {overall_score:.0f}/100 — address {n_crit} critical and {n_high} high gaps first.")

    # Build nav dropdown items for Controls → category sub-links
    _nav_ctrl_items = "".join(
        f'<a href="#ctrl-{k}">{icon} {name}</a>'
        for k, (icon, name) in _CTRL_CATEGORY_NAMES.items()
    )
    _nav_res_items = (
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-spo]\'),\'inv-spo\')},50)">&#127760; SharePoint</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-groups]\'),\'inv-groups\')},50)">&#128101; Teams &amp; Groups</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-licenses]\'),\'inv-licenses\')},50)">&#128179; Licenses</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-labels]\'),\'inv-labels\')},50)">&#127991; Sensitivity Labels</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-ca]\'),\'inv-ca\')},50)">&#128272; Conditional Access</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-dlp]\'),\'inv-dlp\')},50)">&#128737; DLP Policies</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-entra-apps]\'),\'inv-entra-apps\')},50)">&#128221; Entra Apps</a>'
        '<a href="#resources" onclick="setTimeout(function(){switchInvTab(document.querySelector(\'[data-panel=inv-sp]\'),\'inv-sp\')},50)">&#128274; Service Principals</a>'
    )

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>M365 Copilot Readiness — EnterpriseSecurityIQ</title>
<style>{get_css()}{_cr_css()}</style>
</head>
<body>
<div class="scroll-progress" id="scroll-progress"></div>
<a href="#main" class="skip-nav">Skip to content</a>

<nav class="top-nav" aria-label="Report sections">
  <span class="brand">&#129302; EnterpriseSecurityIQ Copilot Readiness</span>
  <a href="#doc-control" class="nav-tip" data-tip="Jump to document control, audit attestation, and report metadata">Document Control</a>
  <a href="#summary" class="nav-tip" data-tip="Jump to KPI cards, readiness score, and executive summary">Summary</a>
  <a href="#categories" class="nav-tip" data-tip="Jump to the 8 readiness domains scored 0-100">Categories</a>
  <span class="nav-dropdown"><a href="#controls" class="nav-tip" data-tip="Jump to the full security controls matrix with pass/fail status">Controls</a><div class="nav-dropdown-menu">{_nav_ctrl_items}</div></span>
  <span class="nav-dropdown"><a href="#resources" class="nav-tip" data-tip="Jump to resource inventory tabs (SharePoint, Teams, Labels, DLP, etc.)">Resources</a><div class="nav-dropdown-menu">{_nav_res_items}</div></span>
  <span class="nav-dropdown"><a href="#findings" class="nav-tip" data-tip="Jump to priority-ranked gaps with remediation guidance">Findings</a><div class="nav-dropdown-menu">{_nav_findings_items}</div></span>
  {'<a href="#warnings-heading" class="nav-tip" data-tip="Jump to collection warnings where data could not be assessed">Warnings</a>' if collection_warnings else ''}
  {'<a href="#trend-heading" class="nav-tip" data-tip="Jump to trend comparison showing score change over time">Trend</a>' if trend_data else ''}
  <div class="zoom-controls" aria-label="Page zoom">
    <button onclick="zoomOut()" aria-label="Zoom out" data-tip="Decrease page zoom by 10%">&minus;</button>
    <span id="zoom-label">100%</span>
    <button onclick="zoomIn()" aria-label="Zoom in" data-tip="Increase page zoom by 10%">&plus;</button>
    <button onclick="zoomReset()" aria-label="Reset zoom" data-tip="Reset zoom to 100%" style="font-size:11px">Reset</button>
  </div>
  <button class="theme-btn" onclick="toggleTheme()" style="margin:0;padding:6px 14px" data-tip="Switch between dark and light theme for this report"
          aria-label="Toggle dark and light theme">Switch to Light</button>
</nav>

<main id="main" class="full-width-content">

<section id="doc-control" class="section">
  <h1 class="page-title">EnterpriseSecurityIQ &mdash; M365 Copilot Readiness Assessment Report</h1>
  <table class="doc-control-table">
    <tr><th>Report Identifier</th><td>{esc(report_id)}</td></tr>
    <tr><th>Assessment Name</th><td>EnterpriseSecurityIQ M365 Copilot Readiness Assessment</td></tr>
    <tr><th>Date Generated</th><td>{esc(ts)}</td></tr>
    <tr><th>Tenant ID</th><td><code>{esc(tenant_id) if tenant_id else 'N/A'}</code></td></tr>
    <tr><th>Tenant Name</th><td>{esc(tenant_display) if tenant_display else 'Unknown'}</td></tr>
    <tr><th>Frameworks Evaluated</th><td>M365 Copilot Readiness (10 domains)</td></tr>
    <tr><th>Classification</th><td>CONFIDENTIAL &mdash; Authorized Recipients Only</td></tr>
    <tr><th>Tool</th><td>EnterpriseSecurityIQ AI Agent v{VERSION}</td></tr>
    <tr><th>Collection Method</th><td>Microsoft Graph API + SharePoint Admin API (Read-Only)</td></tr>
  </table>
  <div class="conf-notice">
    <strong>CONFIDENTIALITY NOTICE:</strong> This document contains sensitive security and compliance
    information about the assessed environment. Distribution is restricted to authorized personnel only.
  </div>
  <h3>Audit Attestation</h3>
  <table class="doc-control-table">
    <tr><th>Assessment Scope</th><td>M365 Copilot readiness review of Microsoft 365, Entra ID, and SharePoint configuration</td></tr>
    <tr><th>Data Integrity</th><td>All evidence collected via read-only API calls; no tenant modifications were made</td></tr>
    <tr><th>Evidence Records</th><td>{evidence_count:,} records collected and evaluated</td></tr>
    <tr><th>Report Hash (SHA-256)</th><td><code id="report-hash">Computed at render</code></td></tr>
    <tr><th>Assessment Period</th><td>{esc(assessed_at[:19]) if assessed_at else esc(ts)} (point-in-time snapshot)</td></tr>
  </table>
</section>

<section id="summary" class="section" aria-labelledby="summary-heading">
  <h2 id="summary-heading" class="page-title">Executive Summary</h2>
  <div style="color:var(--text-secondary);font-size:13px;line-height:1.5;max-width:960px;margin:8px 0 18px;display:flex;flex-direction:column;gap:8px">
    <div style="display:flex;flex-wrap:wrap;gap:12px;font-size:12px;align-items:center">
      <span>Score 0&ndash;100 &rarr;</span>
      <span class="tip" data-tip="Score 75 or above means your tenant meets readiness criteria for Copilot deployment." style="color:#107C10;font-weight:600">&ge;75 READY</span>
      <span class="tip" data-tip="Score between 50 and 74 indicates gaps that should be addressed before broad rollout." style="color:#FFB900;font-weight:600">50&ndash;74 NEEDS WORK</span>
      <span class="tip" data-tip="Score below 50 means significant security gaps exist. Copilot deployment is not recommended." style="color:#D13438;font-weight:600">&lt;50 NOT READY</span>
      <span style="opacity:.5">|</span>
      <span class="tip" data-tip="Severity ranking from most to least urgent. Critical and High gaps should be resolved first.">Severity: Critical &gt; High &gt; Medium &gt; Low &gt; Info</span>
    </div>
  </div>
  <div class="meta-bar" style="display:flex;gap:24px;color:var(--text-secondary);font-size:13px;margin-bottom:16px;flex-wrap:wrap">
    <span>Generated: {esc(ts)}</span>
    <span>Assessed: {esc(assessed_at[:19]) if assessed_at else "N/A"}</span>
    {f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_display)} ({esc(tenant_id[:8])}\u2026)</span>' if tenant_id and tenant_display else f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_id)}</span>' if tenant_id else ''}
    <span>EnterpriseSecurityIQ v{VERSION}</span>
  </div>

  <div class="readiness-banner tip" data-tip="Overall readiness verdict based on weighted score and critical/high gap analysis.\nYOUR TENANT:\n\u2022 Score: {overall_score:.0f}/100\n\u2022 {n_crit} critical + {n_high} high gaps" style="background:{'rgba(16,124,16,.1)' if readiness_status == 'READY' else 'rgba(209,52,56,.1)' if readiness_status == 'NOT READY' else 'rgba(255,185,0,.1)'};border:1px solid {status_color}">
    <div class="status" style="color:{status_color}">{readiness_status}</div>
    <div style="color:var(--text-secondary);font-size:14px;margin-top:8px">{exec_text}</div>
  </div>

  <div class="stat-grid" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin:16px 0">
    <div class="stat-card tip" data-tip="Weighted score (0-100) measuring overall M365 Copilot deployment readiness. 75+ = Ready, 50-74 = Needs Work, below 50 = Not Ready.\nYOUR TENANT:\n\u2022 Score: {overall_score:.0f}/100\n\u2022 Status: {readiness_status}\n\u2022 {n_crit + n_high} critical/high gaps require attention" style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:{status_color};font-family:var(--font-mono)">{overall_score:.0f}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Readiness Score /100</div>
    </div>
    <div class="stat-card tip" data-tip="Total number of security or configuration gaps found that could impact Copilot data exposure or governance.\nYOUR TENANT:\n\u2022 {finding_count} total gaps detected\n\u2022 {n_crit} critical, {n_high} high, {n_med} medium\n\u2022 {n_low} low, {n_info} informational" style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;font-family:var(--font-mono)">{finding_count}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Total Gaps</div>
    </div>
    <div class="stat-card tip" data-tip="Gaps with critical severity \u2014 must be resolved before enabling Copilot to prevent serious data exposure.\nYOUR TENANT:\n\u2022 {n_crit} critical gaps found\n\u2022 {'No critical issues \u2014 ready for Copilot' if n_crit == 0 else 'Address these before Copilot rollout'}" style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:#D13438;font-family:var(--font-mono)">{n_crit}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Critical</div>
    </div>
    <div class="stat-card tip" data-tip="High severity gaps \u2014 should be addressed before broad Copilot rollout to minimize risk.\nYOUR TENANT:\n\u2022 {n_high} high severity gaps found\n\u2022 {'No high-severity issues detected' if n_high == 0 else 'Prioritize these for early remediation'}" style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:#F7630C;font-family:var(--font-mono)">{n_high}</div>
      <div style="font-size:12px;color:var(--text-secondary)">High</div>
    </div>
    <div class="stat-card tip" data-tip="Number of security controls that fully pass all assessment criteria with no gaps detected.\nYOUR TENANT:\n\u2022 {n_compliant} controls fully compliant\n\u2022 {n_gap} with gaps, {n_partial} partial\n\u2022 Target: move all controls to compliant" style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;color:#107C10;font-family:var(--font-mono)">{n_compliant}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Compliant</div>
    </div>
    <div class="stat-card tip" data-tip="Number of M365 license SKUs detected in this tenant. Includes Copilot and non-Copilot subscriptions.\nYOUR TENANT:\n\u2022 {sub_count} license SKUs detected" style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center">
      <div style="font-size:28px;font-weight:700;font-family:var(--font-mono)">{sub_count}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Subscriptions</div>
    </div>
  </div>

  <div class="exec-grid">
    <div class="exec-panel">
      <h3>{_tip('Readiness Score', f'Overall score from 0-100. Green (75+) = Ready for Copilot. Yellow (50-74) = Needs Work. Red (below 50) = Not Ready.\nYOUR TENANT:\n\u2022 Score: {overall_score:.0f}/100\n\u2022 Status: {readiness_status}')}</h3>
      <div class="score-display">{score_ring}</div>
    </div>
    <div class="exec-panel">
      <h3>{_tip('Severity Distribution', f'Breakdown of all gaps by severity level. Critical and High gaps pose the greatest risk and should be prioritized.\nYOUR TENANT:\n\u2022 {n_crit} critical, {n_high} high\n\u2022 {n_med} medium, {n_low} low, {n_info} info')}</h3>
      <div style="text-align:center">{sev_donut}</div>
      <div style="display:flex;gap:16px;justify-content:center;margin-top:12px;font-size:12px;color:var(--text-secondary);flex-wrap:wrap">
        <span class="tip" data-tip="Gaps that pose an immediate threat to data security. Must be resolved before Copilot deployment.\nYOUR TENANT:\n\u2022 {n_crit} critical gaps"><span style="color:#D13438">&#9679;</span> Critical ({n_crit})</span>
        <span class="tip" data-tip="Significant security risks that should be addressed urgently.\nYOUR TENANT:\n\u2022 {n_high} high gaps"><span style="color:#F7630C">&#9679;</span> High ({n_high})</span>
        <span class="tip" data-tip="Moderate risks. Address after critical and high gaps.\nYOUR TENANT:\n\u2022 {n_med} medium gaps"><span style="color:#FFB900">&#9679;</span> Medium ({n_med})</span>
        <span class="tip" data-tip="Minor improvements. Low urgency.\nYOUR TENANT:\n\u2022 {n_low} low gaps"><span style="color:#107C10">&#9679;</span> Low ({n_low})</span>
        <span class="tip" data-tip="Informational observations. No action required.\nYOUR TENANT:\n\u2022 {n_info} informational"><span style="color:#A8A6A3">&#9679;</span> Info ({n_info})</span>
      </div>
    </div>
    <div class="exec-panel">
      <h3>{_tip('Compliance Breakdown', f'How many controls are fully compliant (pass), non-compliant (gap), or partially meeting criteria.\nYOUR TENANT:\n\u2022 {n_compliant} compliant, {n_gap} gaps, {n_partial} partial')}</h3>
      <div style="text-align:center">{compliance_donut}</div>
      <div style="display:flex;gap:16px;justify-content:center;margin-top:12px;font-size:12px;color:var(--text-secondary)">
        <span class="tip" data-tip="Controls that fully pass all assessment criteria with no gaps.\nYOUR TENANT:\n\u2022 {n_compliant} compliant controls">&#9989; Compliant ({n_compliant})</span>
        <span class="tip" data-tip="Controls that fail one or more assessment criteria. These need remediation.\nYOUR TENANT:\n\u2022 {n_gap} controls with gaps">&#10060; Gap ({n_gap})</span>
        <span class="tip" data-tip="Controls that partially meet criteria. Some improvements needed.\nYOUR TENANT:\n\u2022 {n_partial} partially compliant controls">&#9888; Partial ({n_partial})</span>
      </div>
    </div>
  </div>

  <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px 20px;margin-top:8px;font-size:12px;line-height:1.7;color:var(--text-secondary)">
    <strong style="color:var(--text-primary);font-size:13px">How to read this dashboard</strong>
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:12px 24px;margin-top:8px">
      <div><strong style="color:var(--text-primary)">Status Banner</strong> &mdash; READY (green), NEEDS WORK (yellow), or NOT READY (red) based on the overall score and the presence of critical/high gaps.</div>
      <div><strong style="color:var(--text-primary)">Stat Cards</strong> &mdash; key numbers at a glance: readiness score, total gaps, breakdown by severity (critical and high need immediate attention), compliant controls, and license SKUs.</div>
      <div><strong style="color:var(--text-primary)">Readiness Score</strong> &mdash; overall weighted score (0&ndash;100). Critical gaps deduct more than low-severity ones. A score &ge;75 means the tenant is ready for broad Copilot rollout.</div>
      <div><strong style="color:var(--text-primary)">Severity Distribution</strong> &mdash; donut chart showing how gaps split across severity levels. Red and orange slices indicate urgent work.</div>
      <div><strong style="color:var(--text-primary)">Compliance Breakdown</strong> &mdash; donut chart showing controls that fully pass, partially pass, or have gaps. The goal is to move all controls to Compliant.</div>
      <div><strong style="color:var(--text-primary)">Scoring</strong> &mdash; evidence collected via Microsoft Graph, SharePoint Admin &amp; Entra ID APIs. Score = 100 minus severity-weighted gaps. Controls marked NOT ASSESSED need additional permissions or manual review.</div>
    </div>
  </div>
</section>

<section id="categories" class="section" aria-labelledby="categories-heading">
  <h2 id="categories-heading" class="tip" data-tip="8 readiness domains scored 0-100. Higher = fewer gaps. Each domain maps to a set of security controls.">&#128202; Category Breakdown</h2>
  <p class="section-hint">Your tenant is evaluated across <strong>8 readiness domains</strong>. Each card scores one domain from 0 to 100 (higher = fewer gaps). Click a domain below to jump to its card.</p>
  <div style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:14px">
    {"".join(f'<a href="#cat-{k}" style="background:var(--bg-card);border:1px solid var(--border);border-radius:4px;padding:2px 8px;font-size:12px;color:var(--text-primary);text-decoration:none;transition:border-color .2s" onmouseover="this.style.borderColor=&apos;var(--primary)&apos;" onmouseout="this.style.borderColor=&apos;var(--border)&apos;">{esc(m["name"])}</a>' for k, m in _CATEGORY_META.items())}
  </div>
  <div class="category-grid">{cats_html}</div>
</section>

<section id="controls" class="section" aria-labelledby="controls-heading">
  <h2 id="controls-heading" class="tip" data-tip="Full matrix of security controls evaluated for Copilot. Each control is assessed as PASS, FAIL, PARTIAL, or NOT ASSESSED.">&#128737;&#65039; M365 Copilot Security Controls</h2>
  <p class="section-hint">All security controls evaluated for Copilot deployment. Focus on <span style="color:#EF5350;font-weight:600">FAIL</span> rows by severity. Hover any column header or MS Reference label for explanations.</p>
  {controls_table_html}
</section>

<section id="resources" class="section" aria-labelledby="resources-heading">
  <h2 id="resources-heading" class="tip" data-tip="Inventory of M365 resources Copilot can access. Review each tab to identify oversharing, missing labels, and governance gaps.">&#128451;&#65039; Resource Inventory</h2>
  <p class="section-hint">Browse data sources Copilot can access. Each tab shows a different resource class with its security posture. Hover column headers for explanations of each field.</p>
  <div class="inv-tabs" role="tablist">
    <button class="inv-tab active" role="tab" aria-selected="true" data-panel="inv-spo" onclick="switchInvTab(this,'inv-spo')" data-tip="SharePoint sites Copilot can access. Check for oversharing, external sharing, and missing sensitivity labels.&#10;YOUR TENANT:&#10;&#8226; {len(site_inventory)} sites discovered">&#127760; SharePoint<span class="tab-count">({len(site_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-groups" onclick="switchInvTab(this,'inv-groups')" data-tip="Microsoft 365 Groups and Teams. Review membership, guest access, and privacy settings.&#10;YOUR TENANT:&#10;&#8226; {len(groups_inventory)} groups discovered">&#128101; Teams &amp; Groups<span class="tab-count">({len(groups_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-licenses" onclick="switchInvTab(this,'inv-licenses')" data-tip="M365 license SKUs including Copilot entitlements and feature availability.&#10;YOUR TENANT:&#10;&#8226; {len(license_inventory)} license SKUs">&#128179; Licenses<span class="tab-count">({len(license_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-labels" onclick="switchInvTab(this,'inv-labels')" data-tip="Sensitivity labels for classifying and protecting content. Critical for Copilot data governance.&#10;YOUR TENANT:&#10;&#8226; {len(label_inventory)} labels configured">&#127991; Sensitivity Labels<span class="tab-count">({len(label_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-ca" onclick="switchInvTab(this,'inv-ca')" data-tip="Conditional Access policies controlling authentication requirements for Copilot access.&#10;YOUR TENANT:&#10;&#8226; {len(ca_inventory)} policies found">&#128272; Conditional Access<span class="tab-count">({len(ca_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-dlp" onclick="switchInvTab(this,'inv-dlp')" data-tip="Data Loss Prevention policies preventing sensitive data leaks through Copilot responses.&#10;YOUR TENANT:&#10;&#8226; {len(dlp_inventory)} DLP policies">&#128737; DLP Policies<span class="tab-count">({len(dlp_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-entra-apps" onclick="switchInvTab(this,'inv-entra-apps')" data-tip="Entra ID application registrations. Review permissions and consent grants for data exposure risk.&#10;YOUR TENANT:&#10;&#8226; {len(entra_apps_inventory)} apps registered">&#128221; Entra Apps<span class="tab-count">({len(entra_apps_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-sp" onclick="switchInvTab(this,'inv-sp')" data-tip="Service principals with API permissions. Check for over-privileged app identities.&#10;YOUR TENANT:&#10;&#8226; {len(sp_inventory)} service principals">&#128274; Service Principals<span class="tab-count">({len(sp_inventory)})</span></button>
    <button class="inv-tab" role="tab" aria-selected="false" data-panel="inv-mam" onclick="switchInvTab(this,'inv-mam')" data-tip="App protection policies (MAM) for mobile devices accessing Copilot and M365 data.&#10;YOUR TENANT:&#10;&#8226; {len(app_protection_inventory)} policies configured">&#128241; App Protection<span class="tab-count">({len(app_protection_inventory)})</span></button>
  </div>
  <div id="inv-spo" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="All SharePoint sites discoverable by Copilot. Review external sharing, sensitivity labels, and permission levels.">&#127760; SharePoint Sites ({len(site_inventory)})</h3>{site_inventory_html}</div>
  <div id="inv-groups" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="M365 Groups and Teams with membership and guest access details.">&#128101; Teams &amp; Groups ({len(groups_inventory)})</h3>{groups_inv_html}</div>
  <div id="inv-licenses" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="License SKUs and their enabled service plans.">&#128179; Licenses ({len(license_inventory)})</h3>{license_inv_html}</div>
  <div id="inv-labels" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="Sensitivity labels available for classifying documents and emails.">&#127991; Sensitivity Labels ({len(label_inventory)})</h3>{label_inv_html}</div>
  <div id="inv-ca" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="Conditional Access policies enforcing authentication and device compliance.">&#128272; Conditional Access ({len(ca_inventory)})</h3>{ca_inv_html}</div>
  <div id="inv-dlp" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="DLP policies protecting sensitive information from being shared externally.">&#128737; DLP Policies ({len(dlp_inventory)})</h3>{dlp_inv_html}</div>
  <div id="inv-entra-apps" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="Entra ID application registrations with their API permissions and consent status.">&#128221; Entra Apps ({len(entra_apps_inventory)})</h3>{entra_apps_inv_html}</div>
  <div id="inv-sp" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="Service principals and their granted API permissions.">&#128274; Service Principals ({len(sp_inventory)})</h3>{sp_inv_html}</div>
  <div id="inv-mam" class="inv-panel" role="tabpanel"><h3 class="inv-panel-header tip" data-tip="Intune app protection policies for mobile data protection.">&#128241; App Protection Policies ({len(app_protection_inventory)})</h3>{app_protection_inv_html}</div>
</section>

<section id="findings" class="section" aria-labelledby="findings-heading">
  <h2 id="findings-heading" class="tip" data-tip="All gaps ranked by severity with remediation steps. Click any row to expand details and PowerShell commands.\nYOUR TENANT:\n\u2022 {finding_count} total findings\n\u2022 {n_crit} critical, {n_high} high">&#128270; Findings ({finding_count})</h2>
  <p class="section-hint">Priority-ranked gaps with remediation. Click any row to expand details, affected resources, and commands. Hover column headers for field descriptions. Use search and filters to narrow results.</p>
  <div class="filter-bar">
    <label for="finding-filter">Search:</label>
    <input id="finding-filter" type="search" placeholder="Search…" oninput="filterFindings()" data-tip="Type to filter findings by title, description, or category">
    <label for="filter-severity">Severity:</label>
    <select id="filter-severity" onchange="filterFindings()" data-tip="Filter findings by severity level">
      <option value="">All</option>
      <option value="critical">Critical</option><option value="high">High</option>
      <option value="medium">Medium</option><option value="low">Low</option>
      <option value="informational">Informational</option>
    </select>
    <label for="filter-category">Category:</label>
    <select id="filter-category" onchange="filterFindings()" data-tip="Filter findings by readiness category">
      <option value="">All</option>{cat_options}
    </select>
  </div>
  {all_findings_detail_html}
</section>

{warnings_html}
{trend_html}
{suppressed_html}



</main>

<!-- Risk detail modal -->
<div class="risk-modal-overlay" id="risk-modal-overlay">
<div class="risk-modal">
<div class="risk-modal-header"><h3 id="risk-modal-level">RISK</h3><button class="risk-modal-close" onclick="closeRiskModal()" aria-label="Close">&times;</button></div>
<div class="risk-modal-body">
<div class="risk-section"><div class="risk-section-title">Why this is a risk</div><div class="risk-section-content" id="risk-modal-why"></div></div>
<div class="risk-section"><div class="risk-section-title">How to remediate</div><div class="risk-section-content" id="risk-modal-how"></div></div>
</div>
</div>
</div>

<div id="ciq-tooltip" role="tooltip" aria-hidden="true"></div>

<button class="back-to-top" aria-label="Back to top">&#8593;</button>

<script>{get_js()}</script>
<script>{_cr_js()}</script>
</body>
</html>"""

    report_hash = hashlib.sha256(html.encode("utf-8")).hexdigest()
    html = html.replace("Computed at render", report_hash)
    out_path.write_text(html, encoding="utf-8")
    log.info("[CopilotReadinessReport] Written to %s (%d KB)", out_path, len(html) // 1024)

    # Excel export
    _export_excel(results, output_dir)

    return out_path
