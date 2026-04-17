"""
Master Report — Single flat-table report per framework with ALL findings.

Produces three files:
  - master-report.html  (interactive, searchable/sortable)
  - master-report.md    (markdown table)
  - master-report.xlsx  (single sheet, auto-filter, colour-coded)

Every finding appears as one row.  Resource-level findings show the resource
name/type extracted from SupportingEvidence; tenant-level findings show
"Tenant-Level" in the Scope column with "—" for resource fields.
"""

from __future__ import annotations

import hashlib
import html as _html
import pathlib
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.logger import log
from app.reports.shared_theme import (
    esc, get_css, get_js, SEVERITY_COLORS, STATUS_LABELS, DOMAIN_META,
    score_color, severity_badge, status_badge, sev_order, format_ts, format_date_short,
    humanize_framework, humanize_domain, VERSION,
)

# ── Constants ───────────────────────────────────────────────────────────────

_SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
_STATUS_ORDER = {"non_compliant": 0, "partial": 1, "missing_evidence": 2, "not_assessed": 3, "compliant": 4}
_PRIORITY_MAP = {
    "critical": "P0 — Immediate",
    "high": "P1 — Short-term",
    "medium": "P2 — Medium-term",
    "low": "P3 — Long-term",
    "informational": "P4 — Informational",
}

# Excel styling
_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Segoe UI", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="EDEBE9"),
    right=Side(style="thin", color="EDEBE9"),
    top=Side(style="thin", color="EDEBE9"),
    bottom=Side(style="thin", color="EDEBE9"),
)
_STATUS_FILLS = {
    "compliant": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
    "non_compliant": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "partial": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "missing_evidence": PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid"),
    "not_assessed": PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid"),
}
_SEVERITY_FILLS = {
    "critical": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "high": PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),
    "medium": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "low": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
    "informational": PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid"),
}

_COLUMNS = [
    "#", "Control ID", "Control Title", "Domain", "Severity", "Status",
    "Scope", "Resource Name", "Resource Type", "Finding", "Remediation", "Priority",
]


# ── Shared helpers ──────────────────────────────────────────────────────────

def _build_rows(findings: list[dict], missing_evidence: list[dict] | None = None) -> list[dict]:
    """Normalise findings + missing-evidence into flat row dicts, sorted by severity → status → control."""
    rows: list[dict] = []
    for f in findings:
        status = f.get("Status", "not_assessed")
        severity = f.get("Severity", "medium")

        # Extract resource info from SupportingEvidence
        se = f.get("SupportingEvidence", [])
        resource_name = ""
        resource_type = f.get("ResourceType", "")
        if se and isinstance(se, list) and se:
            resource_name = se[0].get("ResourceName", "")
            if not resource_type:
                resource_type = se[0].get("ResourceType", "")

        scope = "Resource-Level" if resource_name else "Tenant-Level"

        # A5: contextual labels for tenant-level rows
        if not resource_name:
            domain_key = f.get("Domain", "")
            if domain_key in ("identity", "access"):
                resource_name = "Tenant Configuration"
                resource_type = resource_type or "Entra ID"
            elif domain_key == "network":
                resource_name = "Tenant Configuration"
                resource_type = resource_type or "Azure Networking"
            elif domain_key == "governance":
                resource_name = "Tenant Configuration"
                resource_type = resource_type or "Azure Policy / Governance"
            else:
                resource_name = "Tenant Configuration"
                resource_type = resource_type or "Azure Platform"

        rows.append({
            "control_id": f.get("ControlId", ""),
            "control_title": f.get("ControlTitle", ""),
            "domain": humanize_domain(f.get("Domain", "")),
            "severity": severity,
            "status": status,
            "scope": scope,
            "resource_name": resource_name or "—",
            "resource_type": resource_type or "—",
            "finding": f.get("Description", ""),
            "remediation": f.get("Recommendation", "") if status != "compliant" else "No action required",
            "priority": _PRIORITY_MAP.get(severity, "P3 — Long-term"),
        })

    # Add missing-evidence entries as rows
    for m in (missing_evidence or []):
        severity = m.get("Severity", "medium")
        missing_types = ", ".join(m.get("MissingTypes", []))
        domain_key = m.get("Domain", "")
        rows.append({
            "control_id": m.get("ControlId", ""),
            "control_title": m.get("ControlTitle", ""),
            "domain": humanize_domain(domain_key),
            "severity": severity,
            "status": "missing_evidence",
            "scope": "Tenant-Level",
            "resource_name": "—",
            "resource_type": "—",
            "finding": f"Missing evidence: {missing_types}" if missing_types else "Required evidence not collected",
            "remediation": "",
            "priority": _PRIORITY_MAP.get(severity, "P3 — Long-term"),
        })

    rows.sort(key=lambda r: (
        _SEV_ORDER.get(r["severity"], 5),
        _STATUS_ORDER.get(r["status"], 5),
        r["control_id"],
    ))

    # Add row numbers
    for i, r in enumerate(rows, 1):
        r["num"] = i

    return rows


def _summary_stats(rows: list[dict]) -> dict:
    """Compute summary counters for the header section."""
    total = len(rows)
    compliant = sum(1 for r in rows if r["status"] == "compliant")
    non_compliant = sum(1 for r in rows if r["status"] == "non_compliant")
    partial = sum(1 for r in rows if r["status"] == "partial")
    resource_level = sum(1 for r in rows if r["scope"] == "Resource-Level")
    tenant_level = total - resource_level
    critical = sum(1 for r in rows if r["severity"] == "critical")
    high = sum(1 for r in rows if r["severity"] == "high")
    return {
        "total": total,
        "compliant": compliant,
        "non_compliant": non_compliant,
        "partial": partial,
        "resource_level": resource_level,
        "tenant_level": tenant_level,
        "critical": critical,
        "high": high,
    }


# ── HTML generation ─────────────────────────────────────────────────────────

_SEV_BADGE_COLORS = {
    "critical": ("#D13438", "#FDE7E9"),
    "high": ("#C4710D", "#FFF0E0"),
    "medium": ("#8A7400", "#FFF4CE"),
    "low": ("#107C10", "#E6F5E6"),
    "informational": ("#616161", "#F3F2F1"),
}

_STATUS_DISPLAY = {
    "compliant": ("Compliant", "#107C10"),
    "non_compliant": ("Non-Compliant", "#D13438"),
    "partial": ("Partial", "#C4710D"),
    "missing_evidence": ("Missing Evidence", "#8A7400"),
    "not_assessed": ("Not Assessed", "#616161"),
}


def _html_report(rows: list[dict], stats: dict, framework: str, summary: dict,
                 tenant_info: dict | None = None, evidence_count: int = 0,
                 access_denied: list | None = None) -> str:
    """Build the full HTML master report with shared dark/light theme and attestation."""
    ts = datetime.now(timezone.utc)
    score = summary.get("ComplianceScore", 0)
    total_controls = summary.get("TotalControls", 0)
    compliant_controls = summary.get("Compliant", 0)
    sc = score_color(score)
    tenant = tenant_info or {}
    fw_label = humanize_framework(framework) if framework else framework
    report_id = f"CIQ-{ts.strftime('%Y%m%d-%H%M%S')}"

    missing_rows = [r for r in rows if r["status"] == "missing_evidence"]

    # Table rows (A3: full text with expandable details for long content)
    tbody = []
    for r in rows:
        sev_fg, sev_bg = _SEV_BADGE_COLORS.get(r["severity"], ("#616161", "#F3F2F1"))
        scope_color = "#0078D4" if r["scope"] == "Resource-Level" else "#8764B8"

        # A3: show full text; use <details> for content over 200 chars
        finding_full = esc(r["finding"])
        if len(r["finding"]) > 200:
            finding_text = f'{esc(r["finding"][:200])}<details class="inline-expand"><summary>...more</summary>{finding_full}</details>'
        else:
            finding_text = finding_full

        remediation_full = esc(r["remediation"])
        if len(r["remediation"]) > 200:
            remediation_text = f'{esc(r["remediation"][:200])}<details class="inline-expand"><summary>...more</summary>{remediation_full}</details>'
        else:
            remediation_text = remediation_full

        # A11: compliant rows get green checkmark
        if r["status"] == "compliant" and (not r["remediation"] or r["remediation"] == "No action required"):
            remediation_text = '<span style="color:#107C10">&#10003; No action required</span>'

        tbody.append(
            f'<tr data-severity="{r["severity"]}" data-status="{r["status"]}" data-scope="{r["scope"]}">'
            f'<td class="num">{r["num"]}</td>'
            f'<td><code>{esc(r["control_id"])}</code></td>'
            f'<td>{esc(r["control_title"])}</td>'
            f'<td>{esc(r["domain"])}</td>'
            f'<td>{severity_badge(r["severity"])}</td>'
            f'<td>{status_badge(r["status"])}</td>'
            f'<td><span style="color:{scope_color};font-weight:500">{r["scope"]}</span></td>'
            f'<td>{esc(r["resource_name"])}</td>'
            f'<td><code style="font-size:11px">{esc(r["resource_type"])}</code></td>'
            f'<td class="desc">{finding_text}</td>'
            f'<td class="desc">{remediation_text}</td>'
            f'<td><span class="badge" style="background:{sev_bg};color:{sev_fg}">{r["priority"]}</span></td>'
            f'</tr>'
        )

    # A9: Access-denied section
    ad_rows_html = ""
    if access_denied:
        ad_items = "".join(
            f'<tr><td>{esc(ad.get("collector", ""))}</td><td>{esc(ad.get("source", ""))}</td>'
            f'<td>{esc(ad.get("api", ""))}</td><td>HTTP {ad.get("status_code", 403)}</td></tr>'
            for ad in access_denied
        )
        ad_rows_html = (
            f'<section id="access-denied" class="section"><h2>Access Denied ({len(access_denied)})</h2>'
            f'<p style="margin-bottom:12px;color:var(--text-secondary)">The following collectors were blocked due to insufficient permissions. '
            f'Resolve these to improve assessment coverage.</p>'
            f'<table class="data-table"><thead><tr><th>Collector</th><th>Source</th><th>API</th><th>Status</th></tr></thead>'
            f'<tbody>{ad_items}</tbody></table></section>'
        )
    else:
        ad_rows_html = '<section id="access-denied" class="section"><h2>Access Denied</h2><p class="empty">All collectors had sufficient permissions.</p></section>'

    extra_css = """
.master-controls{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;align-items:center}
.master-controls input{flex:1;min-width:200px;padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-card);color:var(--text);font-size:13px;font-family:var(--font-primary)}
.master-controls select{padding:8px 10px;border:1px solid var(--border);border-radius:6px;background:var(--bg-card);color:var(--text);font-size:13px}
.master-controls button{padding:8px 14px;border:1px solid var(--border);border-radius:6px;background:var(--bg-card);color:var(--text);cursor:pointer;font-size:13px;transition:all .2s}
.master-controls button:hover{border-color:var(--primary);color:var(--primary)}
.master-table{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:16px}
.master-table thead{position:sticky;top:0;z-index:1}
.master-table th{background:var(--primary);color:#fff;padding:10px 8px;text-align:left;font-size:11px;white-space:nowrap;cursor:pointer;user-select:none;text-transform:uppercase;letter-spacing:.3px}
.master-table th:hover{opacity:.9}
.master-table td{padding:8px;border-bottom:1px solid var(--border-light);vertical-align:top}
.master-table tr:hover td{background:var(--bg-card-hover)}
.master-table .num{text-align:center;color:var(--text-secondary);font-size:11px;width:40px}
.master-table .desc{max-width:300px;font-size:11px;line-height:1.4}
.master-table code{font-family:var(--font-mono);font-size:11px;color:var(--primary)}
.count-display{font-size:13px;color:var(--text-secondary);margin-bottom:10px}
.inline-expand{display:inline;font-size:11px}
.inline-expand summary{display:inline;cursor:pointer;color:var(--primary);font-size:11px}
"""

    filter_js = """
function applyFilters(){
  var q=document.getElementById('search').value.toLowerCase();
  var sev=document.getElementById('fSeverity').value;
  var st=document.getElementById('fStatus').value;
  var sc=document.getElementById('fScope').value;
  var rows=document.querySelectorAll('#masterTable tbody tr');
  var shown=0;var visNum=1;
  rows.forEach(function(r){
    var text=r.textContent.toLowerCase();
    var match=(!q||text.indexOf(q)>=0)&&(!sev||r.dataset.severity===sev)&&(!st||r.dataset.status===st)&&(!sc||r.dataset.scope===sc);
    r.style.display=match?'':'none';
    if(match){shown++;r.cells[0].textContent=visNum++;}
    if(match)shown++;
  });
  document.getElementById('countDisplay').textContent='Showing '+shown+' of '+rows.length+' findings';
}
function resetFilters(){
  document.getElementById('search').value='';
  document.getElementById('fSeverity').value='';
  document.getElementById('fStatus').value='';
  document.getElementById('fScope').value='';
  applyFilters();
}
var sortDir={};
function sortTable(col){
  var tb=document.querySelector('#masterTable tbody');
  var rows=Array.from(tb.rows);
  sortDir[col]=!sortDir[col];
  rows.sort(function(a,b){
    var va=a.cells[col].textContent.trim(),vb=b.cells[col].textContent.trim();
    var na=parseFloat(va),nb=parseFloat(vb);
    if(!isNaN(na)&&!isNaN(nb))return sortDir[col]?na-nb:nb-na;
    return sortDir[col]?va.localeCompare(vb):vb.localeCompare(va);
  });
  rows.forEach(function(r){tb.appendChild(r)});
}
function exportCSV(){
  var table=document.getElementById('masterTable');
  var rows=table.querySelectorAll('tr');
  var csv=[];
  rows.forEach(function(r){
    if(r.style.display==='none')return;
    var cells=r.querySelectorAll('th,td');
    csv.push(Array.from(cells).map(function(c){return '"'+c.textContent.replace(/"/g,'""').trim()+'"'}).join(','));
  });
  var blob=new Blob([csv.join('\\n')],{type:'text/csv'});
  var a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download='master-report-filtered.csv';
  a.click();
}
"""

    return f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Master Compliance Report &mdash; {esc(fw_label)}</title>
<style>{get_css()}{extra_css}</style>
</head>
<body>
<a href="#main-content" class="skip-nav">Skip to content</a>
<div class="layout">
  <nav class="sidebar" role="navigation" aria-label="Report sections">
    <h2>EnterpriseSecurityIQ</h2>
    <span class="version">Master Report &middot; v{VERSION}</span>
    <div class="nav-group">Report</div>
    <nav>
      <a href="#doc-control" class="active">Document Control</a>
      <a href="#how-to-read">How to Read This Report</a>
      <a href="#summary">Summary</a>
      <a href="#findings">All Findings ({stats['total']})</a>
      <a href="#missing">Missing Evidence ({len(missing_rows)})</a>
      <a href="#access-denied">Access Denied ({len(access_denied or [])})</a>
    </nav>
    <button class="theme-btn" onclick="toggleTheme()">Switch to Light</button>
  </nav>
  <main class="content" id="main-content">

    <h1 class="page-title">Master Compliance Report &mdash; {esc(fw_label)}</h1>
    <div class="meta-bar">
      <span>Tenant: <code>{esc(tenant.get('TenantId', tenant.get('tenant_id', '')))}</code></span>
      <span>{format_date_short(ts)}</span>
      <span>Score: <strong style="color:{sc}">{score:.1f}%</strong> ({compliant_controls}/{total_controls})</span>
    </div>

    <!-- Document Control -->
    <section id="doc-control" class="section">
      <h2>Document Control</h2>
      <table class="doc-control-table">
        <tr><th>Report ID</th><td><code>{report_id}</code></td></tr>
        <tr><th>Assessment Name</th><td>EnterpriseSecurityIQ {esc(fw_label)} Master Report</td></tr>
        <tr><th>Date Generated</th><td>{format_ts(ts)}</td></tr>
        <tr><th>Tenant ID</th><td><code>{esc(tenant.get('TenantId', tenant.get('tenant_id', 'N/A')))}</code></td></tr>
        <tr><th>Tenant Name</th><td>{esc(tenant.get('DisplayName', tenant.get('display_name', 'Unknown')))}</td></tr>
        <tr><th>Framework</th><td>{esc(fw_label)}</td></tr>
        <tr><th>Framework Version</th><td>{esc(summary.get('FrameworkVersion', 'See framework documentation'))}</td></tr>
        <tr><th>Classification</th><td>CONFIDENTIAL &mdash; Authorized Recipients Only</td></tr>
        <tr><th>Tool</th><td>EnterpriseSecurityIQ AI Agent v{VERSION}</td></tr>
      </table>
      <div class="conf-notice">
        <strong>CONFIDENTIALITY NOTICE:</strong> This document contains sensitive security and compliance
        information about the assessed environment. Distribution is restricted to authorized personnel only.
      </div>
      <h3>Audit Attestation</h3>
      <table class="doc-control-table">
        <tr><th>Assessment Scope</th><td>Control-plane configuration review of Azure and Microsoft Entra ID resources</td></tr>
        <tr><th>Data Integrity</th><td>All evidence collected via read-only API calls; no tenant modifications were made</td></tr>
        <tr><th>Evidence Records</th><td>{evidence_count:,} records collected and evaluated</td></tr>
        <tr><th>Report Hash (SHA-256)</th><td><code id="report-hash">Computed at render</code></td></tr>
        <tr><th>Assessment Period</th><td>{format_date_short(ts)} (point-in-time snapshot)</td></tr>
      </table>
    </section>

    <!-- How to Read This Report -->
    <section id="how-to-read" class="section">
      <h2>How to Read This Report</h2>
      <div class="how-to-read">
        <p><strong>Purpose:</strong> This master report provides a single flat table of <em>every</em> compliance finding
        for the <strong>{esc(fw_label)}</strong> framework. It is designed for engineers, security analysts, and
        remediation teams who need actionable, resource-level detail.</p>
        <p><strong>Columns:</strong> Each row represents one finding. <em>Control ID</em> maps to the framework control,
        <em>Severity</em> indicates risk level, <em>Status</em> shows compliance state, and <em>Remediation</em>
        provides specific corrective actions. <em>Priority</em> assigns P0&ndash;P4 based on severity.</p>
        <p><strong>Filtering:</strong> Use the search box and dropdowns above the table to filter by severity, status,
        scope, or free text. Click any column header to sort. Use <em>Export CSV</em> to download filtered data.</p>
        <h4>Severity Legend</h4>
        <div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:6px">
          <span>{severity_badge('critical')} Immediate risk &mdash; remediate within 0-14 days</span>
          <span>{severity_badge('high')} Urgent &mdash; remediate within 0-30 days</span>
          <span>{severity_badge('medium')} Short-term &mdash; remediate within 30-90 days</span>
          <span>{severity_badge('low')} Long-term &mdash; remediate within 90-180 days</span>
        </div>
      </div>
    </section>

    <!-- Summary -->
    <section id="summary" class="section">
      <h2>Compliance Summary</h2>
      <p style="margin-bottom:16px;color:var(--text-secondary)">This assessment evaluated
      <strong>{total_controls}</strong> controls from the <strong>{esc(fw_label)}</strong> framework across
      tenant <code>{esc(tenant.get('TenantId', tenant.get('tenant_id', '')))}</code>.
      The overall compliance score is <strong style="color:{sc}">{score:.1f}%</strong> with
      <strong style="color:#D13438">{stats['critical']} critical</strong> and
      <strong style="color:#F7630C">{stats['high']} high</strong> severity findings requiring attention.</p>
      <div class="stat-grid">
        <div class="stat-card"><div class="stat-value" style="color:{sc}">{score:.1f}%</div><div class="stat-label">Compliance Score</div></div>
        <div class="stat-card"><div class="stat-value">{total_controls}</div><div class="stat-label">Total Controls</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#107C10">{compliant_controls}</div><div class="stat-label">Compliant</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#D13438">{summary.get('NonCompliant', 0)}</div><div class="stat-label">Non-Compliant</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#FFB900">{summary.get('Partial', 0)}</div><div class="stat-label">Partial</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#A8A6A3">{summary.get('MissingEvidence', 0)}</div><div class="stat-label">Missing Evidence</div></div>
      </div>
      <h3>Findings Breakdown</h3>
      <div class="stat-grid">
        <div class="stat-card"><div class="stat-value">{stats['total']}</div><div class="stat-label">Total Finding Rows</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#D13438">{stats['critical']}</div><div class="stat-label">Critical</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#F7630C">{stats['high']}</div><div class="stat-label">High</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#0078D4">{stats['resource_level']}</div><div class="stat-label">Resource-Level</div></div>
        <div class="stat-card"><div class="stat-value" style="color:#8764B8">{stats['tenant_level']}</div><div class="stat-label">Tenant-Level</div></div>
      </div>
    </section>

    <!-- All Findings -->
    <section id="findings" class="section">
      <h2>All Findings ({stats['total']})</h2>
      <div class="master-controls">
        <input type="text" id="search" placeholder="Search by control ID, resource name, domain, finding text..." oninput="applyFilters()">
        <select id="fSeverity" onchange="applyFilters()"><option value="">All Severities</option><option value="critical">Critical</option><option value="high">High</option><option value="medium">Medium</option><option value="low">Low</option></select>
        <select id="fStatus" onchange="applyFilters()"><option value="">All Statuses</option><option value="non_compliant">Non-Compliant</option><option value="partial">Partial</option><option value="compliant">Compliant</option><option value="missing_evidence">Missing Evidence</option></select>
        <select id="fScope" onchange="applyFilters()"><option value="">All Scopes</option><option value="Resource-Level">Resource-Level</option><option value="Tenant-Level">Tenant-Level</option></select>
        <button onclick="resetFilters()">Reset</button>
        <button onclick="exportCSV()">Export CSV</button>
      </div>
      <div class="count-display" id="countDisplay">Showing {stats['total']} of {stats['total']} findings</div>
      <table class="master-table" id="masterTable">
      <thead><tr>{''.join(f'<th onclick="sortTable({i})">{c}</th>' for i, c in enumerate(_COLUMNS))}</tr></thead>
      <tbody>
      {''.join(tbody)}
      </tbody>
      </table>
    </section>

    <!-- Missing Evidence -->
    <section id="missing" class="section">
      <h2>Missing Evidence ({len(missing_rows)})</h2>
      {'<p class="empty">All evidence collected &mdash; no missing items.</p>' if not missing_rows else ''.join(f'<div class="missing-card"><code>{esc(r["control_id"])}</code> {severity_badge(r["severity"])} &mdash; {esc(r["finding"])}</div>' for r in missing_rows)}
    </section>

    <!-- Access Denied -->
    {ad_rows_html}

    <footer style="text-align:center;color:var(--text-muted);font-size:12px;padding:32px 0;border-top:1px solid var(--border)">
      Generated by EnterpriseSecurityIQ AI Agent v{VERSION} &mdash; {format_date_short(ts)} &mdash; CONFIDENTIAL
    </footer>
  </main>
</div>
<button class="back-to-top" title="Back to top" aria-label="Back to top">&#8593;</button>
<script>{get_js()}</script>
<script>
{filter_js}
</script>
</body>
</html>"""


# ── Markdown generation ─────────────────────────────────────────────────────

def _md_report(rows: list[dict], stats: dict, framework: str, summary: dict) -> str:
    """Build the Markdown master report."""
    score = summary.get("ComplianceScore", 0)
    total_controls = summary.get("TotalControls", 0)
    compliant_controls = summary.get("Compliant", 0)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    lines = [
        f"# Master Compliance Report — {framework}",
        "",
        f"Generated: {ts} | Score: {score:.1f}% ({compliant_controls}/{total_controls} compliant)",
        "",
        "## Summary",
        "",
        f"| Metric | Count |",
        f"|--------|------:|",
        f"| Total Findings | {stats['total']} |",
        f"| Non-Compliant | {stats['non_compliant']} |",
        f"| Compliant | {stats['compliant']} |",
        f"| Partial | {stats['partial']} |",
        f"| Critical | {stats['critical']} |",
        f"| High | {stats['high']} |",
        f"| Resource-Level | {stats['resource_level']} |",
        f"| Tenant-Level | {stats['tenant_level']} |",
        "",
        "## All Findings",
        "",
        "| # | Control ID | Control Title | Domain | Severity | Status | Scope | Resource Name | Resource Type | Finding | Remediation | Priority |",
        "|---|-----------|---------------|--------|----------|--------|-------|---------------|---------------|---------|-------------|----------|",
    ]

    for r in rows:
        status_display = r["status"].replace("_", " ").title()
        # A4: Truncate with ellipsis for markdown readability
        finding = r["finding"][:150].replace("|", "\\|").replace("\n", " ")
        if len(r["finding"]) > 150:
            finding += "..."
        remediation = r["remediation"][:150].replace("|", "\\|").replace("\n", " ")
        if len(r["remediation"]) > 150:
            remediation += "..."
        rname = r["resource_name"].replace("|", "\\|")
        rtype = r["resource_type"].replace("|", "\\|")
        lines.append(
            f"| {r['num']} "
            f"| {r['control_id']} "
            f"| {r['control_title'].replace('|', '\\|')} "
            f"| {r['domain']} "
            f"| {r['severity'].title()} "
            f"| {status_display} "
            f"| {r['scope']} "
            f"| {rname} "
            f"| {rtype} "
            f"| {finding} "
            f"| {remediation} "
            f"| {r['priority']} |"
        )

    lines.append("")
    return "\n".join(lines)


# ── Excel generation ────────────────────────────────────────────────────────

def _excel_report(rows: list[dict], stats: dict, framework: str, summary: dict, filepath: pathlib.Path):
    """Write an .xlsx with a single Master Findings sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Findings"

    # Header row
    for col, h in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"
    ws.freeze_panes = "A2"

    for r in rows:
        row_num = r["num"] + 1  # +1 for header
        severity = r["severity"]
        status = r["status"]
        fills: dict[int, PatternFill] = {}
        if severity in _SEVERITY_FILLS:
            fills[5] = _SEVERITY_FILLS[severity]
        if status in _STATUS_FILLS:
            fills[6] = _STATUS_FILLS[status]

        values = [
            r["num"], r["control_id"], r["control_title"], r["domain"],
            severity.title(), status.replace("_", " ").title(),
            r["scope"], r["resource_name"], r["resource_type"],
            r["finding"], r["remediation"], r["priority"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in fills:
                cell.fill = fills[col]

    # Auto-width
    col_widths = [5, 14, 30, 16, 10, 14, 14, 28, 22, 50, 50, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(str(filepath))


# ── Public API ──────────────────────────────────────────────────────────────

def generate_master_report(
    results: dict[str, Any],
    output_dir: str,
    framework: str = "",
    evidence: list[dict] | None = None,
    tenant_info: dict | None = None,
    access_denied: list[dict] | None = None,
    **_kwargs: Any,
) -> dict[str, str]:
    """Generate master-report.html, .md, and .xlsx for one framework.

    Returns dict of ``{format: filepath}`` for generated files.
    """
    findings = results.get("findings", [])
    missing_evidence = results.get("missing_evidence", [])
    summary = results.get("summary", {})
    fw_label = humanize_framework(framework) if framework else "Compliance"

    out = pathlib.Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    rows = _build_rows(findings, missing_evidence)
    stats = _summary_stats(rows)
    evidence_count = len(evidence) if evidence else summary.get("TotalEvidence", 0)
    paths: dict[str, str] = {}

    # HTML (with SHA-256 hash)
    html_path = out / "master-report.html"
    html = _html_report(rows, stats, fw_label, summary,
                        tenant_info=tenant_info, evidence_count=evidence_count,
                        access_denied=access_denied)
    report_hash = hashlib.sha256(html.encode("utf-8")).hexdigest()
    html = html.replace("Computed at render", report_hash)
    html_path.write_text(html, encoding="utf-8")
    paths["html"] = str(html_path)
    log.info("Master report (HTML): %s", html_path)

    # Markdown
    md_path = out / "master-report.md"
    md_path.write_text(_md_report(rows, stats, fw_label, summary), encoding="utf-8")
    paths["md"] = str(md_path)
    log.info("Master report (MD): %s", md_path)

    # Excel
    xlsx_path = out / "master-report.xlsx"
    _excel_report(rows, stats, fw_label, summary, xlsx_path)
    paths["xlsx"] = str(xlsx_path)
    log.info("Master report (Excel): %s", xlsx_path)

    log.info("Master report: %d findings (%d resource-level, %d tenant-level)",
             stats["total"], stats["resource_level"], stats["tenant_level"])

    return paths
