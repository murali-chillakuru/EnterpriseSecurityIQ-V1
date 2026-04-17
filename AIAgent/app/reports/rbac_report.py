"""
RBAC Report — Interactive HTML
Full-width expandable tree showing the management-group → subscription → resource-group
→ resource hierarchy with role assignments, PIM status, principal resolution,
group membership expansion, executive summary, and risk analysis.
"""

from __future__ import annotations

import hashlib
import pathlib
from datetime import datetime, timezone
from app.reports.shared_theme import (
    get_css, get_js, esc, score_color, format_date_short, VERSION,
    SEVERITY_COLORS,
)
from app.logger import log


# ── SVG helpers ──────────────────────────────────────────────────────────

def _donut_svg(slices: list[tuple[str, float, str]], size: int = 160, hole: float = 0.6) -> str:
    """Render a donut chart from (label, value, color) tuples."""
    total = sum(v for _, v, _ in slices) or 1
    r = size // 2 - 4
    circ = 2 * 3.14159 * r
    cx = cy = size // 2
    parts = [f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" role="img" aria-label="Donut chart">']
    offset = 0
    for label, val, color in slices:
        pct = val / total
        dash = circ * pct
        gap = circ - dash
        parts.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" '
            f'stroke-width="{r * (1 - hole)}" stroke-dasharray="{dash:.2f} {gap:.2f}" '
            f'stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})">'
            f'<title>{esc(label)}: {int(val)} ({pct*100:.0f}%)</title></circle>'
        )
        offset += dash
    # Centre label
    parts.append(
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" '
        f'font-size="22" font-weight="700" fill="var(--text)" '
        f'font-family="var(--font-mono)">{int(total)}</text>'
    )
    parts.append("</svg>")
    return "".join(parts)


def _bar_chart_svg(items: list[tuple[str, int, str]], width: int = 480, bar_h: int = 22) -> str:
    """Horizontal bar chart. items: (label, value, color). Responsive via viewBox."""
    if not items:
        return ""
    max_val = max(v for _, v, _ in items) or 1
    gap = 6
    label_w = 170          # wider label area for long names
    val_w = 40             # reserved space for the value text
    bar_area = width - label_w - val_w
    h = (bar_h + gap) * len(items) + 10
    parts = [f'<svg width="100%" viewBox="0 0 {width} {h}" preserveAspectRatio="xMinYMin meet" role="img" aria-label="Bar chart">']
    for i, (label, val, color) in enumerate(items):
        y = i * (bar_h + gap) + 4
        bw = int((val / max_val) * bar_area)
        # Truncate long labels to ~22 chars
        display_label = label if len(label) <= 22 else label[:20] + "…"
        parts.append(
            f'<text x="0" y="{y + bar_h - 5}" font-size="12" fill="var(--text-secondary)" '
            f'font-family="var(--font-primary)">'
            f'<title>{esc(label)}</title>{esc(display_label)}</text>'
        )
        parts.append(
            f'<rect x="{label_w}" y="{y}" width="{max(bw, 2)}" height="{bar_h}" rx="3" fill="{color}" opacity="0.85">'
            f'<title>{esc(label)}: {val}</title></rect>'
        )
        parts.append(
            f'<text x="{label_w + bw + 5}" y="{y + bar_h - 5}" font-size="12" fill="var(--text)" '
            f'font-family="var(--font-mono)">{val}</text>'
        )
    parts.append("</svg>")
    return "".join(parts)


# ── Tree rendering ───────────────────────────────────────────────────────

_ICONS = {
    "ManagementGroup": "&#127970;",   # 🏢
    "Subscription": "&#128187;",       # 💻
    "ResourceGroup": "&#128193;",      # 📁
    "Resource": "&#128196;",           # 📄
}

_TYPE_COLORS = {
    "ManagementGroup": "#0078D4",
    "Subscription": "#00B7C3",
    "ResourceGroup": "#F7630C",
    "Resource": "#A8A6A3",
}

_STATUS_COLORS = {
    "Active": "#107C10",
    "Eligible": "#FFB900",
}

_NODE_TYPE_TIPS = {
    "ManagementGroup": "Management Group \u2014 top-level container. Roles here cascade to ALL child subscriptions and resources.",
    "Subscription": "Subscription \u2014 billing and access boundary. Roles here apply to every RG and resource within.",
    "ResourceGroup": "Resource Group \u2014 logical container for related resources. Roles apply to all resources in this group.",
    "Resource": "Individual Azure resource. Roles at this scope have the narrowest blast radius.",
}

_PRINCIPAL_ICONS = {
    "User": "&#128100;",           # 👤
    "Group": "&#128101;",          # 👥
    "ServicePrincipal": "&#9881;", # ⚙
    "Unknown": "&#10067;",         # ❓
}

_SUBTYPE_ICONS = {
    "Member": "&#128100;",         # 👤
    "Guest": "&#127758;",          # 🌎
    "Group": "&#128101;",          # 👥
    "Application": "&#128187;",    # 🖻 (desktop)
    "ManagedIdentity": "&#128272;",# 🔐
    "Legacy": "&#9888;",           # ⚠
    "Unknown": "&#10067;",         # ❓
}

_SUBTYPE_COLORS = {
    "Member": "#0078D4",
    "Guest": "#8764B8",
    "Group": "#F7630C",
    "Application": "#00B7C3",
    "ManagedIdentity": "#107C10",
    "Legacy": "#FFB900",
    "Unknown": "#A8A6A3",
}

_SUBTYPE_LABELS = {
    "Member": "Member User",
    "Guest": "Guest User",
    "Group": "Security Group",
    "Application": "App Registration",
    "ManagedIdentity": "Managed Identity",
    "Legacy": "Legacy SP",
    "Unknown": "Unknown",
}


def _principal_sub_type(pid: str, principals: dict) -> str:
    """Get the sub_type for a principal, falling back to base type."""
    info = principals.get(pid, {})
    return info.get("sub_type", info.get("type", "Unknown"))


def _principal_name(pid: str, principals: dict) -> str:
    """Get display name for a principal ID."""
    info = principals.get(pid, {})
    return info.get("display_name", pid) or pid


def _build_scope_names(node: dict) -> dict[str, str]:
    """Walk tree and map scope ID (lowercased) → display_name."""
    names: dict[str, str] = {}
    nid = node.get("id", "")
    nname = node.get("display_name", "") or node.get("name", "")
    if nid and nname:
        names[nid.lower().rstrip("/")] = nname
    for rg in node.get("resource_groups", []):
        rid = rg.get("id", "")
        rname = rg.get("display_name", "") or rg.get("name", "")
        if rid and rname:
            names[rid.lower().rstrip("/")] = rname
        for res in rg.get("resources", []):
            resid = res.get("id", "")
            resname = res.get("display_name", "") or res.get("name", "")
            if resid and resname:
                names[resid.lower().rstrip("/")] = resname
    for child in node.get("children", []):
        names.update(_build_scope_names(child))
    return names


def _scope_friendly(scope: str, scope_names: dict[str, str]) -> str:
    """Return a human-friendly name for a scope, falling back to last path segment."""
    if not scope:
        return ""
    name = scope_names.get(scope.lower().rstrip("/"), "")
    if name:
        return name
    return scope.rsplit("/", 1)[-1] or scope


def _assignment_count(node: dict) -> int:
    """Count direct assignments + eligible at this node."""
    return len(node.get("assignments", [])) + len(node.get("eligible", []))


def _build_risk_lookup(risks: list[dict]) -> dict[tuple[str, str, str], list[dict]]:
    """Build a lookup from (principal_id, scope_lower, role_name) → list of risk dicts."""
    lookup: dict[tuple[str, str, str], list[dict]] = {}
    for r in risks:
        key = (r.get("principal_id", ""), r.get("scope", "").lower().rstrip("/"), r.get("role_name", ""))
        lookup.setdefault(key, []).append(r)
    return lookup


def _render_assignments(assignments: list[dict], eligible: list[dict], principals: dict, node_scope: str = "", scope_names: dict[str, str] | None = None, risk_lookup: dict | None = None) -> str:
    """Render assignment rows as an accessible sortable table with inherited/direct badges."""
    if scope_names is None:
        scope_names = {}
    if risk_lookup is None:
        risk_lookup = {}
    all_a = [(a, False) for a in assignments] + [(a, True) for a in eligible]
    if not all_a:
        return '<div class="rbac-empty" role="status">No direct role assignments</div>'
    rows = []
    for a, is_elig in all_a:
        status = "Eligible" if is_elig else a.get("status", "Active")
        s_color = _STATUS_COLORS.get(status, "#A8A6A3")
        role = esc(a.get("role_name", "Unknown"))
        pid = a.get("principal_id", "")
        ptype = principals.get(pid, {}).get("type", a.get("principal_type", "Unknown"))
        psub = _principal_sub_type(pid, principals)
        pname = esc(_principal_name(pid, principals))
        picon = _SUBTYPE_ICONS.get(psub, _PRINCIPAL_ICONS.get(ptype, ""))
        priv_badge = ""
        if a.get("is_privileged"):
            priv_badge = (' <span class="badge badge-priv" aria-label="Privileged role"'
                          ' data-tip="High-impact role that can modify resources, manage access, or escalate privileges. '
                          'Should be PIM-eligible with MFA and time-limited.">PRIVILEGED</span>')
        custom_badge = ""
        if a.get("is_custom"):
            custom_badge = (' <span class="badge badge-custom" aria-label="Custom role"'
                           ' data-tip="Tenant-defined custom role — not a built-in Azure role. '
                           'Review the role definition to ensure it follows least-privilege.">CUSTOM</span>')

        # Inherited vs Direct
        a_scope = a.get("scope", "")
        is_inherited = bool(node_scope and a_scope.lower().rstrip("/") != node_scope.lower().rstrip("/"))
        if is_inherited:
            origin = esc(_scope_friendly(a_scope, scope_names))
            inherit_badge = (f' <span class="badge badge-inherited" aria-label="Inherited from {origin}"'
                            f' data-tip="This role was assigned at a parent scope ({origin}) and cascades down. '
                            f'The principal has this access here because of inheritance, not a direct assignment at this level.">INHERITED</span>')
        else:
            inherit_badge = (' <span class="badge badge-direct" aria-label="Directly assigned at this scope"'
                            ' data-tip="This role was explicitly assigned at this exact scope level. '
                            'It does not cascade from a parent — it was intentionally granted here.">DIRECT</span>')

        # Group member expansion
        member_html = ""
        pinfo = principals.get(pid, {})
        if pinfo.get("type") == "Group" and pinfo.get("members"):
            member_items = []
            for m in pinfo["members"]:
                mi = _PRINCIPAL_ICONS.get(m.get("type", ""), "")
                mn = esc(m.get("display_name", m.get("id", "")))
                mu = esc(m.get("upn", ""))
                label = f"{mn} ({mu})" if mu else mn
                # Nested sub-group with its own members
                if m.get("type") == "Group" and m.get("members"):
                    sub_items = []
                    for sm in m["members"]:
                        smi = _PRINCIPAL_ICONS.get(sm.get("type", ""), "")
                        smn = esc(sm.get("display_name", sm.get("id", "")))
                        smu = esc(sm.get("upn", ""))
                        slabel = f"{smn} ({smu})" if smu else smn
                        sub_items.append(f"<li>{smi} {slabel}</li>")
                    sub_html = (
                        f'<details class="rbac-members rbac-sub-group"><summary>{mi} {label} — {len(m["members"])} members</summary>'
                        f'<ul>{"".join(sub_items)}</ul></details>'
                    )
                    member_items.append(f"<li>{sub_html}</li>")
                else:
                    member_items.append(f"<li>{mi} {label}</li>")
            member_html = (
                f'<details class="rbac-members"><summary>{len(pinfo["members"])} members</summary>'
                f'<ul>{"".join(member_items)}</ul></details>'
            )

        scope_display = esc(_scope_friendly(a_scope, scope_names)) if a_scope else ""

        # Risk lookup for this assignment
        risk_cell = ""
        a_scope_key = a_scope.lower().rstrip("/") if a_scope else ""
        a_role = a.get("role_name", "")
        matched_risks = risk_lookup.get((pid, a_scope_key, a_role), [])
        if matched_risks:
            risk_parts = []
            for mr in matched_risks:
                rsev = mr.get("severity", "medium")
                rcolor = SEVERITY_COLORS.get(rsev, "#A8A6A3")
                rtitle = esc(mr.get("title", ""))
                rremediation = esc(mr.get("remediation", ""))
                part = (
                    f'<div style="margin-bottom:6px">'
                    f'<span class="badge" style="background:{rcolor};min-width:52px;text-align:center;font-size:9px">'
                    f'{esc(rsev.upper())}</span> '
                    f'<span style="font-size:11px">{rtitle}</span>'
                )
                if rremediation:
                    part += (
                        f'<div style="font-size:11px;color:var(--text-secondary);margin-top:4px;line-height:1.4">'
                        f'<strong style="color:var(--primary);font-size:10px">Remediation:</strong> {rremediation}</div>'
                    )
                part += '</div>'
                risk_parts.append(part)
            risk_cell = ''.join(risk_parts)
        else:
            # Explain WHY no risk was flagged for this assignment
            nr_reasons: list[str] = []
            a_scope_lvl = a.get("scope_level", "")
            a_is_priv = a.get("is_privileged", False)
            a_is_custom = a.get("is_custom", False)
            a_ptype = principals.get(pid, {}).get("type", a.get("principal_type", ""))
            role_lower = a_role.lower()
            if is_elig:
                nr_reasons.append("PIM-eligible \u2014 requires just-in-time activation with MFA")
            if role_lower.endswith("reader") or role_lower == "reader":
                nr_reasons.append("Read-only role \u2014 cannot modify or delete resources")
            elif not a_is_priv:
                nr_reasons.append("Non-privileged role \u2014 limited permission scope")
            if a_scope_lvl in ("ResourceGroup", "Resource"):
                nr_reasons.append("Narrow scope \u2014 limited to resource group or resource")
            if psub == "ManagedIdentity":
                nr_reasons.append("Managed identity \u2014 Azure-managed credential, no stored secrets")
            # Privileged + Active + broad scope but didn't match specific risk rules
            if a_is_priv and not is_elig and a_scope_lvl in ("ManagementGroup", "Subscription"):
                if a_role not in ("Owner", "User Access Administrator", "Contributor"):
                    nr_reasons.append(
                        f"Role is not Owner/UAA/Contributor \u2014 risk checks target those roles at broad scope")
                elif a_role == "Contributor" and a_scope_lvl == "Subscription" and a_ptype != "ServicePrincipal":
                    nr_reasons.append(
                        "Human Contributor at subscription \u2014 risk checks target Service Principals at this level")
            if a_is_priv and not is_elig and a_ptype == "ServicePrincipal" and a_scope_lvl in ("ResourceGroup", "Resource"):
                nr_reasons.append("Service principal scoped to RG/resource \u2014 blast radius is contained")
            if not nr_reasons:
                nr_reasons.append("No matching risk pattern for this role/scope/identity combination")
            nr_html = "<br>".join(nr_reasons)
            risk_cell = (
                f'<span style="color:#16a34a;font-size:11px">\u2713 No risk detected</span>'
                f'<div style="font-size:10px;color:var(--text-secondary);margin-top:2px;line-height:1.4">{nr_html}</div>'
            )

        status_tip = ("Always-on access \u2014 no activation required. Consider converting privileged roles to PIM Eligible."
                      if status == "Active" else
                      "PIM Eligible \u2014 requires just-in-time activation with MFA and justification before access is granted.")
        psub_label = _SUBTYPE_LABELS.get(psub, psub)
        psub_color = _SUBTYPE_COLORS.get(psub, "#A8A6A3")
        mi_detail = ""
        mi_type = principals.get(pid, {}).get("mi_type", "")
        if mi_type:
            mi_detail = f' <span style="font-size:10px;color:var(--text-secondary)">({mi_type})</span>'
        rows.append(
            f'<tr data-status="{status}" data-privileged="{"1" if a.get("is_privileged") else "0"}" data-inherited="{"1" if is_inherited else "0"}">'  
            f'<td data-sort="{status}"><span class="badge" style="background:{s_color};min-width:52px;text-align:center" data-tip="{status_tip}">{status}</span></td>'
            f'<td data-sort="{role}"><strong>{role}</strong>{priv_badge}{custom_badge}</td>'
            f'<td data-sort="{pname}">{picon} <span class="rbac-principal" title="{esc(pid)}">{pname}</span></td>'
            f'<td data-sort="{esc(psub_label)}"><span class="badge" style="background:{psub_color};min-width:64px;text-align:center;font-size:10px">{esc(psub_label)}</span>{mi_detail}{member_html}</td>'
            f'<td>{inherit_badge}</td>'
            f'<td class="scope-cell" title="{esc(a_scope)}">{scope_display}</td>'
            f'<td class="risk-cell">{risk_cell}</td>'
            f'</tr>'
        )
    return (
        '<table class="rbac-assign-table sortable" role="table">'
        '<thead><tr>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,0)" aria-sort="none">'
        '<span data-tip="Active = always-on access. Eligible = requires PIM activation with MFA.">Status</span> <span class="sort-arrow"></span></th>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,1)" aria-sort="none">'
        '<span data-tip="Azure RBAC role name. PRIVILEGED badge marks high-impact roles (Owner, Contributor, UAA). CUSTOM badge marks tenant-defined roles.">Role</span> <span class="sort-arrow"></span></th>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,2)" aria-sort="none">'
        '<span data-tip="The user, group, or service principal that holds this role assignment.">Principal</span> <span class="sort-arrow"></span></th>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,3)" aria-sort="none">'
        '<span data-tip="Identity sub-type: Member User, Guest User, App Registration, Managed Identity (System/User Assigned), Security Group, or Legacy SP.">Type</span> <span class="sort-arrow"></span></th>'
        '<th scope="col">'
        '<span data-tip="DIRECT = assigned at this exact scope. INHERITED = granted at a parent scope and cascading down.">Origin</span></th>'
        '<th scope="col">'
        '<span data-tip="The ARM scope path where the role was assigned. Hover to see the full resource ID.">Scope</span></th>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,6)" aria-sort="none">'
        '<span data-tip="Security risks detected for this specific principal + role + scope combination. Shows severity, title, and recommended remediation.">Risk</span> <span class="sort-arrow"></span></th>'
        '</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table>'
    )


def _render_tree_node(node: dict, principals: dict, depth: int = 0, scope_names: dict[str, str] | None = None, risk_lookup: dict | None = None) -> str:
    """Recursively render a tree node as <details>/<summary> with ARIA attributes."""
    if scope_names is None:
        scope_names = {}
    if risk_lookup is None:
        risk_lookup = {}
    ntype = node["type"]
    icon = _ICONS.get(ntype, "")
    color = _TYPE_COLORS.get(ntype, "#A8A6A3")
    name = esc(node.get("display_name", "") or node.get("name", ""))
    node_scope = node.get("id", "")
    count = _assignment_count(node)
    count_badge = f' <span class="rbac-count" aria-label="{count} assignments" data-tip="{count} role assignments (active + eligible) at this scope">{count}</span>' if count else ""
    ntype_tip = _NODE_TYPE_TIPS.get(ntype, "")
    type_badge = f'<span class="rbac-type-badge" style="background:{color}" data-tip="{ntype_tip}">{esc(ntype)}</span>'
    aria_label = f"{ntype}: {name}, {count} direct assignments"

    # Child elements
    inner = []
    assignments_html = _render_assignments(
        node.get("assignments", []), node.get("eligible", []), principals, node_scope, scope_names, risk_lookup,
    )
    inner.append(f'<div class="rbac-assignments">{assignments_html}</div>')

    # Resource groups (for subscriptions)
    for rg in node.get("resource_groups", []):
        rg_count = len(rg.get("assignments", []))
        res_count = len(rg.get("resources", []))
        rg_name = esc(rg.get("display_name", rg.get("name", "")))
        rg_icon = _ICONS["ResourceGroup"]
        rg_badge = f' <span class="rbac-count">{rg_count}</span>' if rg_count else ""
        rg_type = f'<span class="rbac-type-badge" style="background:{_TYPE_COLORS["ResourceGroup"]}">ResourceGroup</span>'

        rg_scope = rg.get("id", "")
        rg_inner = _render_assignments(rg.get("assignments", []), [], principals, rg_scope, scope_names, risk_lookup)

        # Build a combined list of inherited assignments for leaf resources:
        # subscription-level + RG-level assignments cascade to all resources within the RG.
        inherited_for_resources = list(node.get("assignments", [])) + list(node.get("eligible", [])) + list(rg.get("assignments", []))

        # Resources within this RG
        res_parts = []
        for res in rg.get("resources", []):
            res_name = esc(res.get("display_name", res.get("name", "")))
            rtype = esc(res.get("resource_type", ""))
            res_scope = res.get("id", "")
            res_a_count = len(res.get("assignments", []))
            if res_a_count:
                res_badge = f' <span class="rbac-count">{res_a_count}</span>'
                res_aa = _render_assignments(res.get("assignments", []), [], principals, res_scope, scope_names, risk_lookup)
                res_parts.append(
                    f'<details class="rbac-node rbac-depth-{depth+3}" data-node-type="Resource" open>'
                    f'<summary>{_ICONS["Resource"]} {res_name} '
                    f'<span class="rbac-type-badge" style="background:{_TYPE_COLORS["Resource"]}">{rtype}</span>'
                    f'{res_badge}</summary>'
                    f'<div class="rbac-body"><div class="rbac-assignments">{res_aa}</div></div>'
                    f'</details>'
                )
            elif inherited_for_resources:
                # Leaf resource with 0 direct assignments but has inherited parent assignments
                inh_count = len(inherited_for_resources)
                inh_badge = f' <span class="rbac-count rbac-count-inherited" data-tip="{inh_count} inherited assignments from parent scopes (subscription + resource group)">&#x2193;{inh_count}</span>'
                inh_aa = _render_assignments(inherited_for_resources, [], principals, res_scope, scope_names, risk_lookup)
                res_parts.append(
                    f'<details class="rbac-node rbac-depth-{depth+3}" data-node-type="Resource">'
                    f'<summary>{_ICONS["Resource"]} {res_name} '
                    f'<span class="rbac-type-badge" style="background:{_TYPE_COLORS["Resource"]}">{rtype}</span>'
                    f'{inh_badge}</summary>'
                    f'<div class="rbac-body"><div class="rbac-assignments">{inh_aa}</div></div>'
                    f'</details>'
                )
            else:
                res_parts.append(
                    f'<div class="rbac-leaf">{_ICONS["Resource"]} {res_name} '
                    f'<span class="rbac-type-badge" style="background:{_TYPE_COLORS["Resource"]}">{rtype}</span></div>'
                )

        res_summary = f' <span class="rbac-res-info">({res_count} resources)</span>' if res_count else ""
        if rg_count or res_parts:
            inner.append(
                f'<details class="rbac-node rbac-depth-{depth+2}" data-node-type="ResourceGroup" open>'
                f'<summary>{rg_icon} {rg_name} {rg_type}{rg_badge}{res_summary}</summary>'
                f'<div class="rbac-body">'
                f'<div class="rbac-assignments">{rg_inner}</div>'
                f'{"".join(res_parts)}'
                f'</div></details>'
            )
        else:
            inner.append(
                f'<div class="rbac-leaf">{rg_icon} {rg_name} {rg_type}{res_summary}</div>'
            )

    # Children (child MGs or subscriptions)
    for child in node.get("children", []):
        inner.append(_render_tree_node(child, principals, depth + 1, scope_names, risk_lookup))

    return (
        f'<details class="rbac-node rbac-depth-{depth}" data-node-type="{esc(ntype)}" open>'
        f'<summary aria-label="{esc(aria_label)}">{icon} {name} {type_badge}{count_badge}</summary>'
        f'<div class="rbac-body">{"".join(inner)}</div>'
        f'</details>'
    )


# ── Risk section ─────────────────────────────────────────────────────────

def _render_risks(risks: list[dict], principals: dict, scope_names: dict[str, str] | None = None) -> str:
    if scope_names is None:
        scope_names = {}
    if not risks:
        return '<p class="empty">No high-severity RBAC risks detected.</p>'
    rows = []
    for r in risks:
        sev = r.get("severity", "medium")
        color = SEVERITY_COLORS.get(sev, "#A8A6A3")
        pid = r.get("principal_id", "")
        pname = esc(_principal_name(pid, principals))
        scope_raw = r.get("scope", "")
        scope_display = esc(_scope_friendly(scope_raw, scope_names)) if scope_raw else ""
        remediation = esc(r.get("remediation", ""))
        rows.append(
            f'<tr>'
            f'<td data-sort="{esc(sev)}"><span class="badge" style="background:{color};min-width:58px;text-align:center">{esc(sev.upper())}</span></td>'
            f'<td data-sort="{esc(r.get("title", ""))}">{esc(r.get("title", ""))}</td>'
            f'<td data-sort="{pname}">{pname}</td>'
            f'<td class="scope-cell" title="{esc(scope_raw)}">{scope_display}</td>'
            f'<td style="font-size:12px">{esc(r.get("detail", ""))}</td>'
            f'<td style="font-size:12px;color:var(--text-secondary)">{remediation}</td>'
            f'</tr>'
        )
    return (
        '<div style="overflow-x:auto">'
        '<table class="rbac-assign-table sortable" id="risks-table" role="table"'
        f' aria-label="Risk analysis table ({len(risks)} findings)">'
        '<colgroup><col style="width:7%"><col style="width:15%"><col style="width:14%"><col style="width:10%"><col style="width:27%"><col style="width:27%"></colgroup>'
        '<thead><tr>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,0)" aria-sort="none">'
        '<span data-tip="Critical = immediate action needed. High = address soon. Medium = review. Low = informational.">Severity</span> <span class="sort-arrow"></span></th>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,1)" aria-sort="none">'
        '<span data-tip="Short description of the detected risk condition and the scope level at which it was found.">Title</span> <span class="sort-arrow"></span></th>'
        '<th scope="col" class="sortable-th" onclick="sortTable(this,2)" aria-sort="none">'
        '<span data-tip="The user, group, or service principal that triggers this risk finding.">Principal</span> <span class="sort-arrow"></span></th>'
        '<th scope="col">'
        '<span data-tip="The Azure scope (MG, subscription, RG) where the risky assignment exists.">Scope</span></th>'
        '<th scope="col">'
        '<span data-tip="Detailed explanation of why this is a risk and its potential impact.">Detail</span></th>'
        '<th scope="col">'
        '<span data-tip="Step-by-step guidance to remediate: typically convert to PIM, reduce scope, or replace with least-privilege role.">Remediation</span></th>'
        '</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table></div>'
    )


# ── Appendix: ID mapping table ───────────────────────────────────────────

def _render_appendix(principals: dict) -> str:
    if not principals:
        return '<p class="empty" role="status">No principals resolved.</p>'
    rows = []
    for pid in sorted(principals, key=lambda p: principals[p].get("display_name", p).lower()):
        info = principals[pid]
        dname = esc(info.get("display_name", pid))
        ptype = esc(info.get("type", "Unknown"))
        psub = info.get("sub_type", ptype)
        psub_label = _SUBTYPE_LABELS.get(psub, psub)
        psub_color = _SUBTYPE_COLORS.get(psub, "#A8A6A3")
        mi_type = info.get("mi_type", "")
        mi_tag = f" ({mi_type})" if mi_type else ""
        upn = esc(info.get("upn", ""))
        app_id = esc(info.get("app_id", ""))
        extra = upn or app_id
        members = info.get("members", [])
        mcell = f"{len(members)} members" if members else ""
        rows.append(
            f"<tr><td><code>{esc(pid)}</code></td><td>{dname}</td>"
            f'<td><span class="badge" style="background:{psub_color};min-width:64px;text-align:center;font-size:10px">'
            f'{esc(psub_label)}</span>{esc(mi_tag)}</td>'
            f"<td>{extra}</td><td>{mcell}</td></tr>"
        )
    return (
        '<div style="overflow-x:auto">'
        '<table class="data-table" id="appendix-table" role="table"'
        f' aria-label="Principal ID reference table ({len(principals)} principals)">'
        '<thead><tr>'
        '<th scope="col"><span data-tip="Azure AD / Entra ID object GUID. Use this to look up the identity in Entra admin portal.">Object ID</span></th>'
        '<th scope="col"><span data-tip="The human-readable name of the user, group, or service principal.">Display Name</span></th>'
        '<th scope="col"><span data-tip="Identity sub-type: Member User, Guest User, App Registration, Managed Identity (System/User Assigned), Security Group, or Legacy SP.">Type</span></th>'
        '<th scope="col"><span data-tip="User Principal Name for users, or Application (client) ID for service principals.">UPN / App ID</span></th>'
        '<th scope="col"><span data-tip="Number of members in a group. Groups with many members and privileged roles expand the blast radius.">Group Members</span></th>'
        '</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table></div>'
    )


# ── Custom CSS for tree ──────────────────────────────────────────────────

def _tree_css() -> str:
    return """
/* ── Tree nodes with connector lines ── */
.rbac-node{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;margin:6px 0;position:relative}
.rbac-node summary{padding:10px 14px;cursor:pointer;display:flex;align-items:center;gap:8px;flex-wrap:wrap;font-weight:600;font-size:14px;min-height:44px;list-style:none}
.rbac-node summary::-webkit-details-marker{display:none}
/* Expand/collapse toggle icon */
.rbac-node summary::before{content:'\\2795';font-size:12px;color:#0078D4;transition:all .2s;flex-shrink:0;width:20px;height:20px;line-height:20px;text-align:center;border-radius:4px;background:rgba(0,120,212,.12)}
.rbac-node[open]>summary::before{content:'\\2796'}
.rbac-node summary:hover{background:var(--bg-card-hover)}
/* WCAG 2.4.7: Visible focus indicator (min 2px, offset for clarity) */
.rbac-node summary:focus-visible{outline:2px solid var(--primary);outline-offset:2px;border-radius:4px}
.rbac-body{padding:4px 12px 12px 32px;position:relative}
/* Blue tree connector lines — vertical and horizontal */
.rbac-body::before{content:'';position:absolute;left:18px;top:0;bottom:12px;width:2px;background:#0078D4;border-radius:1px;opacity:.55}
.rbac-body>.rbac-node,.rbac-body>.rbac-leaf{position:relative;margin-left:8px}
.rbac-body>.rbac-node::before,.rbac-body>.rbac-leaf::before{content:'';position:absolute;left:-22px;top:20px;width:18px;height:2px;background:#0078D4;border-radius:1px;opacity:.55}
.rbac-depth-0>summary{font-size:16px;padding:14px 18px}
.rbac-depth-1>summary{font-size:15px}
.rbac-leaf{padding:6px 14px;font-size:13px;color:var(--text-secondary);border-left:2px solid var(--border-light);margin:2px 0 2px 8px}
.rbac-type-badge{display:inline-block;padding:2px 8px;border-radius:3px;color:#fff;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.3px;vertical-align:middle}
.rbac-count{display:inline-flex;align-items:center;justify-content:center;background:var(--primary);color:#fff;border-radius:10px;font-size:11px;font-weight:700;min-width:22px;height:20px;padding:0 6px;margin-left:4px}
.rbac-count-inherited{background:#8764B8;font-size:10px}

/* ── Assignment table ── */
.rbac-assignments{margin:6px 0}
.rbac-assignments.hidden-by-collapse{display:none}
.rbac-assign-table{width:100%;border-collapse:separate;border-spacing:0;font-size:13px;table-layout:fixed}
.rbac-assign-table th{text-align:left;padding:10px 12px;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted);border-bottom:2px solid var(--border);background:var(--th-bg)}
.rbac-assign-table td{padding:10px 12px;border-bottom:1px solid var(--border-light);vertical-align:top;line-height:1.5;word-wrap:break-word;overflow-wrap:break-word}
.rbac-assign-table tbody tr:nth-child(even) td{background:rgba(255,255,255,.03)}
.rbac-assign-table tbody tr:nth-child(odd) td{background:rgba(0,0,0,.08)}
.rbac-assign-table tr:last-child td{border-bottom:none}
.rbac-assign-table tr:hover td{background:var(--bg-card-hover)}
.scope-cell{max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px;color:var(--text-muted)}
.risk-cell{min-width:180px;max-width:320px;font-size:12px;vertical-align:top}
.rbac-risk-details{margin-top:2px}
.rbac-risk-details summary{list-style:none;font-size:10px;color:var(--primary);cursor:pointer}
.rbac-risk-details summary::-webkit-details-marker{display:none}
.rbac-principal{color:var(--primary);font-weight:500}
.rbac-ptype{color:var(--text-muted);font-size:11px}
.rbac-empty{color:var(--text-muted);font-style:italic;font-size:12px;padding:4px 0}
.rbac-res-info{font-size:11px;color:var(--text-muted);font-weight:normal}
.rbac-members{margin:2px 0 2px 12px;font-size:12px}
.rbac-members summary{cursor:pointer;color:var(--primary);font-weight:500;font-size:12px;min-height:28px}

/* ── Inherited / Direct / Privileged / Custom badges ── */
.badge-direct{background:#0078D4;font-size:9px}
.badge-inherited{background:#8764B8;font-size:9px}
.badge-priv{background:#D13438;font-size:9px}
.badge-custom{background:#9B59B6;font-size:9px}

/* ── Sortable column headers ── */
.sortable-th{cursor:pointer;user-select:none;white-space:nowrap}
.sortable-th:hover{color:var(--primary)}
.sortable-th:focus-visible{outline:2px solid var(--primary);outline-offset:2px}
.sort-arrow{font-size:10px;margin-left:2px;color:var(--text-muted)}
.sort-arrow.asc::after{content:' \\25B2'}
.sort-arrow.desc::after{content:' \\25BC'}

/* ── Zoom ── */
.zoom-controls{display:flex;align-items:center;gap:4px;margin-left:8px}
.zoom-controls button{padding:2px 10px;border:1px solid var(--border);border-radius:4px;background:var(--bg-elevated);color:var(--text);font-size:14px;cursor:pointer;min-width:32px;min-height:32px;font-weight:700;line-height:1}
.zoom-controls button:hover{border-color:var(--primary);color:var(--primary)}
.zoom-controls button:focus-visible{outline:2px solid var(--primary);outline-offset:2px}
.zoom-controls span{font-size:12px;min-width:42px;text-align:center;color:var(--text-secondary)}

/* ── Export toolbar ── */
.export-bar{display:flex;gap:8px;align-items:center}
.rbac-members summary:focus-visible{outline:2px solid var(--primary);outline-offset:2px}
.rbac-members ul{margin:4px 0 0 18px;color:var(--text-secondary);list-style:disc}
.rbac-members li{padding:2px 0}
.rbac-sub-group{margin:4px 0 4px 0}
.rbac-sub-group summary{color:var(--text);font-weight:600}

/* ── Top navigation bar ── */
.top-nav{position:sticky;top:0;z-index:500;display:flex;align-items:center;gap:16px;padding:8px 24px;
  background:var(--bg-elevated);border-bottom:1px solid var(--border);font-size:13px;flex-wrap:wrap}
[id^="cat-"],section[id]{scroll-margin-top:56px}
.top-nav .brand{font-weight:700;color:var(--primary);font-size:14px;margin-right:12px}
.top-nav a{color:var(--text-secondary);text-decoration:none;padding:6px 10px;border-radius:6px;transition:all .2s}
.top-nav a:hover{color:var(--text);background:var(--bg-card)}
.top-nav a:focus-visible{outline:2px solid var(--primary);outline-offset:2px}
.nav-dropdown{position:relative}
.nav-dropdown>.nav-toggle{cursor:pointer;user-select:none;padding:6px 10px;border-radius:6px;color:var(--text-secondary);font-size:13px;display:inline-flex;align-items:center;gap:4px;transition:all .2s;min-height:36px;border:none;background:none;font-family:inherit}
.nav-dropdown>.nav-toggle:hover,.nav-dropdown:focus-within>.nav-toggle{color:var(--text);background:var(--bg-card)}
.nav-dropdown>.nav-toggle::after{content:'\25BE';font-size:10px;margin-left:2px}
.nav-menu{display:none;position:absolute;top:100%;left:0;min-width:220px;background:var(--bg-elevated);border:1px solid var(--border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.3);padding:6px 0;z-index:600;margin-top:4px;max-height:70vh;overflow-y:auto}
.nav-dropdown:hover>.nav-menu,.nav-dropdown:focus-within>.nav-menu{display:block}
.nav-menu a{display:flex;padding:8px 16px;color:var(--text-secondary);font-size:12px;border-radius:0;min-height:auto;white-space:nowrap}
.nav-menu a:hover{color:var(--text);background:var(--bg-card)}
.nav-menu .nav-sep{height:1px;background:var(--border);margin:4px 12px}

/* ── Layout ── */
.full-width-content{padding:32px 40px;max-width:1440px;margin:0 auto}
.exec-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:20px;margin:24px 0}
.exec-panel{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px}
.exec-panel h3{font-size:14px;color:var(--text-secondary);margin-bottom:12px;border:none;padding:0}
.legend{display:flex;flex-wrap:wrap;gap:12px;margin-top:12px}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--text-secondary)}
.legend-dot{width:12px;height:12px;border-radius:3px;display:inline-block}

/* ── Filter bar ── */
.filter-bar{display:flex;align-items:center;gap:8px;margin-bottom:16px;flex-wrap:wrap;font-size:13px}
.filter-bar label{font-size:13px;font-weight:600;color:var(--text-secondary)}
.filter-bar input[type="search"]{min-width:240px;padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar input[type="search"]:focus-visible{outline:2px solid var(--primary);outline-offset:1px;border-color:var(--primary)}
.filter-bar input[type="checkbox"]{width:14px;height:14px;margin:0;cursor:pointer;accent-color:var(--primary)}
.filter-bar select{padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar button{padding:8px 16px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px;cursor:pointer;min-height:32px;font-weight:600}
.filter-bar button:hover{border-color:var(--primary);color:var(--primary)}
.filter-bar button:focus-visible{outline:2px solid var(--primary);outline-offset:2px}

/* ── ARIA live region ── */
.sr-only{position:absolute;width:1px;height:1px;padding:0;margin:-1px;overflow:hidden;clip:rect(0,0,0,0);white-space:nowrap;border:0}

/* ── Hover tooltips ── */
[data-tip]{cursor:help}
/* Underline hint for inline text elements only */
th [data-tip],h3 [data-tip],.legend [data-tip]{border-bottom:1px dotted var(--text-muted)}
/* No dotted underline on badges/cards */
.stat-card[data-tip],.badge[data-tip],.rbac-type-badge[data-tip],.rbac-count[data-tip]{border-bottom:none;cursor:help}
.exec-panel-title{cursor:help;display:inline-block}
.score-ring-wrap{cursor:help;display:inline-block}
.nav-tip{cursor:help}

/* JS-managed tooltip element */
#ciq-tooltip{position:fixed;z-index:99999;pointer-events:none;opacity:0;transition:opacity .18s ease;max-width:520px;min-width:220px;padding:14px 18px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  color:var(--text);
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  border-radius:10px;
  font-size:12.5px;line-height:1.6;font-weight:400;text-transform:none;letter-spacing:normal;white-space:normal;
  box-shadow:0 2px 6px rgba(0,0,0,.18),0 8px 24px rgba(0,0,0,.32),0 0 0 1px rgba(255,255,255,.06) inset;
}
#ciq-tooltip.visible{opacity:1}
#ciq-tooltip::before{content:'';position:absolute;width:12px;height:12px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  transform:rotate(45deg);z-index:-1}
#ciq-tooltip.arrow-bottom::before{bottom:-7px;left:var(--arrow-x,24px);border-top:none;border-left:none}
#ciq-tooltip.arrow-top::before{top:-7px;left:var(--arrow-x,24px);border-bottom:none;border-right:none}
#ciq-tooltip .t-sep{display:block;border-top:1px solid rgba(255,255,255,.15);margin:8px 0 4px;padding-top:6px;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--primary)}

/* ── Print stylesheet ── */
@media print{
  .top-nav,.filter-bar,.theme-btn,.back-to-top,.skip-nav,.zoom-controls,.export-bar{display:none!important}
  body{zoom:1!important}
  .full-width-content{padding:0}
  .rbac-node{break-inside:avoid;border:1px solid #ccc}
  .rbac-node[open]{break-inside:auto}
  .rbac-node summary{background:#f8f8f8!important;color:#000!important}
  .rbac-type-badge{border:1px solid #999;color:#000!important}
  .stat-card{border:1px solid #ccc}
  body{background:#fff;color:#000}
  .exec-panel{border:1px solid #ccc;break-inside:avoid}
}

/* ── High Contrast (forced-colors) ── */
@media (forced-colors:active){
  .rbac-node summary{border:1px solid ButtonText}
  .rbac-type-badge,.badge,.rbac-count{border:1px solid ButtonText;forced-color-adjust:none}
  .rbac-node summary:focus-visible{outline:2px solid Highlight}
}
"""


# ── Custom JS for tree ───────────────────────────────────────────────────

def _tree_js() -> str:
    return """
// ── Expand / Collapse all ──
function expandAll(){
  document.querySelectorAll('.rbac-node').forEach(function(d){d.open=true;
    var s=d.querySelector(':scope>summary');if(s)s.setAttribute('aria-expanded','true');});
  document.querySelectorAll('.rbac-assignments').forEach(function(a){a.classList.remove('hidden-by-collapse');});
  _announceTree('All nodes expanded');
}
function collapseAll(){
  document.querySelectorAll('.rbac-node').forEach(function(d){
    var t=d.getAttribute('data-node-type')||'';
    var keep=(t==='ManagementGroup'||t==='Subscription');
    d.open=keep;
    var s=d.querySelector(':scope>summary');if(s)s.setAttribute('aria-expanded',keep?'true':'false');
  });
  document.querySelectorAll('.rbac-assignments').forEach(function(a){a.classList.add('hidden-by-collapse');});
  _announceTree('Collapsed to subscription level');
}
// ── ARIA live announcement helper ──
function _announceTree(msg){
  var el=document.getElementById('tree-live');
  if(el){el.textContent=msg;setTimeout(function(){el.textContent=''},2000)}
}
// ── Sync aria-expanded on toggle ──
document.addEventListener('toggle',function(e){
  if(e.target.classList&&e.target.classList.contains('rbac-node')){
    var s=e.target.querySelector(':scope>summary');
    if(s)s.setAttribute('aria-expanded',e.target.open?'true':'false');
    if(e.target.open){var a=e.target.querySelector(':scope>.rbac-body>.rbac-assignments');if(a)a.classList.remove('hidden-by-collapse');}
  }
},true);
// ── Filter tree by text ──
var _filterTimer;
function filterTree(){
  clearTimeout(_filterTimer);
  _filterTimer=setTimeout(_doFilterTree,200);
}
function _doFilterTree(){
  var q=(document.getElementById('tree-filter').value||'').toLowerCase();
  var statusF=document.getElementById('filter-status').value;
  var privF=document.getElementById('filter-privileged').checked;
  var hideInhF=document.getElementById('filter-hide-inherited').checked;
  var nodes=document.querySelectorAll('.rbac-node');
  var count=0;
  if(!q&&!statusF&&!privF&&!hideInhF){
    nodes.forEach(function(n){n.style.display='';n.removeAttribute('hidden');
      n.open=true;
      var s=n.querySelector(':scope>summary');if(s)s.setAttribute('aria-expanded','true');
    });
    document.querySelectorAll('.rbac-assignments').forEach(function(a){a.classList.remove('hidden-by-collapse');});
    document.querySelectorAll('.rbac-assign-table tbody tr').forEach(function(r){r.style.display='';});
    _announceTree('All filters cleared');return;
  }
  // Ensure assignment tables are visible
  document.querySelectorAll('.rbac-assignments').forEach(function(a){a.classList.remove('hidden-by-collapse');});
  // Filter every assignment row by ALL active criteria (status, privileged, AND text query)
  document.querySelectorAll('.rbac-assign-table tbody tr').forEach(function(r){
    var show=true;
    if(statusF&&r.getAttribute('data-status')!==statusF)show=false;
    if(show&&privF&&r.getAttribute('data-privileged')!=='1')show=false;
    if(show&&hideInhF&&r.getAttribute('data-inherited')==='1')show=false;
    if(show&&q&&r.textContent.toLowerCase().indexOf(q)===-1)show=false;
    r.style.display=show?'':'none';
  });
  // Open all nodes and reset visibility, then prune leaf-first
  nodes.forEach(function(n){n.style.display='';n.removeAttribute('hidden');n.open=true;
    var s=n.querySelector(':scope>summary');if(s)s.setAttribute('aria-expanded','true');
  });
  var all=Array.from(nodes).reverse();
  all.forEach(function(n){
    // Check if node summary text matches the query (so you can search by scope name too)
    var summaryEl=n.querySelector(':scope>summary');
    var summaryMatch=q&&summaryEl&&summaryEl.textContent.toLowerCase().indexOf(q)!==-1;
    // Count visible rows in this node's own (direct) assignment table
    var ownRows=0;
    n.querySelectorAll(':scope>.rbac-body>.rbac-assignments .rbac-assign-table tbody tr').forEach(function(r){
      if(r.style.display!=='none')ownRows++;
    });
    // Check if any descendant node is still visible (children already processed thanks to reverse order)
    var hasVisibleChild=false;
    n.querySelectorAll('.rbac-node').forEach(function(c){
      if(c!==n&&c.style.display!=='none')hasVisibleChild=true;
    });
    if(ownRows===0&&!hasVisibleChild&&!summaryMatch){
      n.style.display='none';n.setAttribute('hidden','');
    }else{count++}
  });
  _announceTree(count+' matching nodes found');
}
// ── Keyboard navigation for tree (WCAG 2.1.1) ──
document.addEventListener('keydown',function(e){
  var s=document.activeElement;
  if(!s||!s.matches||!s.matches('.rbac-node>summary'))return;
  var node=s.parentElement;
  var key=e.key;
  if(key==='ArrowRight'){
    if(!node.open){node.open=true;s.setAttribute('aria-expanded','true');e.preventDefault();return}
    var firstChild=node.querySelector(':scope>.rbac-body .rbac-node>summary');
    if(firstChild){firstChild.focus();e.preventDefault()}
  }else if(key==='ArrowLeft'){
    if(node.open){node.open=false;s.setAttribute('aria-expanded','false');e.preventDefault();return}
    var parent=node.parentElement&&node.parentElement.closest('.rbac-node');
    if(parent){var ps=parent.querySelector(':scope>summary');if(ps){ps.focus();e.preventDefault()}}
  }else if(key==='ArrowDown'){
    e.preventDefault();
    var allSummaries=Array.from(document.querySelectorAll('.rbac-node:not([hidden])>summary'));
    var idx=allSummaries.indexOf(s);
    for(var i=idx+1;i<allSummaries.length;i++){
      if(allSummaries[i].offsetParent!==null){allSummaries[i].focus();return}
    }
  }else if(key==='ArrowUp'){
    e.preventDefault();
    var allSummaries=Array.from(document.querySelectorAll('.rbac-node:not([hidden])>summary'));
    var idx=allSummaries.indexOf(s);
    for(var i=idx-1;i>=0;i--){
      if(allSummaries[i].offsetParent!==null){allSummaries[i].focus();return}
    }
  }else if(key==='Home'){
    e.preventDefault();
    var first=document.querySelector('.rbac-node:not([hidden])>summary');
    if(first)first.focus();
  }else if(key==='End'){
    e.preventDefault();
    var all=document.querySelectorAll('.rbac-node:not([hidden])>summary');
    if(all.length)all[all.length-1].focus();
  }
});
// ── Sortable assignment tables ──
function sortTable(th,colIdx){
  var table=th.closest('table');
  if(!table)return;
  var tbody=table.querySelector('tbody');
  var rows=Array.from(tbody.querySelectorAll('tr'));
  // Determine sort direction
  var arrow=th.querySelector('.sort-arrow');
  var dir='asc';
  if(arrow.classList.contains('asc')){dir='desc';}
  else if(arrow.classList.contains('desc')){dir='none';}
  // Reset all arrows in this table
  table.querySelectorAll('.sort-arrow').forEach(function(a){a.className='sort-arrow';});
  table.querySelectorAll('.sortable-th').forEach(function(h){h.setAttribute('aria-sort','none');});
  if(dir==='none'){
    // Restore original order — re-append in DOM order (use data-idx)
    rows.sort(function(a,b){return (parseInt(a.dataset.origIdx)||0)-(parseInt(b.dataset.origIdx)||0);});
    rows.forEach(function(r){tbody.appendChild(r);});
    return;
  }
  arrow.classList.add(dir);
  th.setAttribute('aria-sort',dir==='asc'?'ascending':'descending');
  rows.sort(function(a,b){
    var av=(a.cells[colIdx]&&a.cells[colIdx].getAttribute('data-sort'))||'';
    var bv=(b.cells[colIdx]&&b.cells[colIdx].getAttribute('data-sort'))||'';
    var cmp=av.localeCompare(bv,undefined,{sensitivity:'base'});
    return dir==='desc'?-cmp:cmp;
  });
  rows.forEach(function(r){tbody.appendChild(r);});
}
// Store original row order for reset
(function(){
  document.querySelectorAll('.rbac-assign-table tbody').forEach(function(tbody){
    Array.from(tbody.querySelectorAll('tr')).forEach(function(r,i){r.dataset.origIdx=i;});
  });
})();
// ── CSV Export ──
function exportCSV(){
  var tables=document.querySelectorAll('.rbac-assign-table');
  var csv=['Status,Role,Principal,Type,Origin,Scope'];
  tables.forEach(function(t){
    t.querySelectorAll('tbody tr').forEach(function(r){
      if(r.style.display==='none')return;
      var cells=r.querySelectorAll('td');
      if(cells.length<6)return;
      var row=[];
      for(var i=0;i<6;i++){
        var txt=(cells[i].getAttribute('data-sort')||cells[i].textContent||'').trim();
        row.push('"'+txt.replace(/"/g,'""')+'"');
      }
      csv.push(row.join(','));
    });
  });
  var blob=new Blob([csv.join('\\n')],{type:'text/csv;charset=utf-8'});
  var a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download='rbac-assignments.csv';
  a.click();URL.revokeObjectURL(a.href);
  _announceTree('CSV exported');
}
// ── Zoom controls ──
var _zoomLevel=100;
function _applyZoom(){
  document.body.style.zoom=(_zoomLevel/100);
  var lbl=document.getElementById('zoom-label');
  if(lbl)lbl.textContent=_zoomLevel+'%';
  try{localStorage.setItem('rbac-zoom',_zoomLevel)}catch(e){}
}
function zoomIn(){if(_zoomLevel<150){_zoomLevel+=10;_applyZoom()}}
function zoomOut(){if(_zoomLevel>70){_zoomLevel-=10;_applyZoom()}}
function zoomReset(){_zoomLevel=100;_applyZoom();}
(function(){
  try{var z=parseInt(localStorage.getItem('rbac-zoom'));if(z>=70&&z<=150){_zoomLevel=z;_applyZoom()}}catch(e){}
})();
// ── Appendix table filter ──
function filterAppendix(){
  var q=(document.getElementById('appendix-filter').value||'').toLowerCase();
  var rows=document.querySelectorAll('#appendix-table tbody tr');
  var count=0;
  rows.forEach(function(r){
    var vis=r.textContent.toLowerCase().indexOf(q)!==-1;
    r.style.display=vis?'':'none';
    if(vis)count++;
  });
  var live=document.getElementById('appendix-live');
  if(live){live.textContent=count+' principals shown';setTimeout(function(){live.textContent=''},2000)}
}
// ── Top-nav scroll active state ──
(function(){
  var secs=document.querySelectorAll('section[id]');
  var links=document.querySelectorAll('.top-nav a[href^=\"#\"]');
  if(!secs.length||!links.length)return;
  var obs=new IntersectionObserver(function(entries){
    entries.forEach(function(e){
      if(e.isIntersecting){
        var id=e.target.getAttribute('id');
        links.forEach(function(l){
          var isActive=l.getAttribute('href')==='#'+id;
          if(isActive)l.setAttribute('aria-current','true');else l.removeAttribute('aria-current');
        });
      }
    });
  },{rootMargin:'-10% 0px -80% 0px'});
  secs.forEach(function(s){obs.observe(s)});
})();
// ── Init: set aria-expanded on all summaries ──
(function(){
  document.querySelectorAll('.rbac-node').forEach(function(d){
    var s=d.querySelector(':scope>summary');
    if(s){s.setAttribute('aria-expanded',d.open?'true':'false');s.setAttribute('role','button')}
  });
  // Apply default filters (e.g. Hide Inherited) on page load
  if(typeof filterTree==='function')filterTree();
})();
// ── Tooltip engine (viewport-aware positioning) ──
(function(){
  var tip=document.getElementById('ciq-tooltip');
  if(!tip)return;
  var GAP=10,MARGIN=12;
  function show(ev){
    var tgt=ev.target.closest('[data-tip]');
    if(!tgt)return;
    var text=tgt.getAttribute('data-tip');
    if(!text)return;
    var d=document.createElement('span');d.textContent=text;var safe=d.innerHTML;
    safe=safe.replace(/\\n+YOUR TENANT:/g,'<span class="t-sep">&#x1F4CA; Your Tenant</span>');
    safe=safe.replace(/\\n/g,'<br>');
    tip.innerHTML=safe;
    tip.classList.add('visible');
    tip.setAttribute('aria-hidden','false');
    // measure after visible so dimensions are correct
    requestAnimationFrame(function(){
      var r=tgt.getBoundingClientRect();
      var tw=tip.offsetWidth,th=tip.offsetHeight;
      var vw=window.innerWidth,vh=window.innerHeight;
      // vertical: prefer above; fall back to below
      var above=r.top-GAP-th;
      var below=r.bottom+GAP;
      var top,arrow;
      if(above>=MARGIN){top=above;arrow='arrow-bottom';}
      else if(below+th<=vh-MARGIN){top=below;arrow='arrow-top';}
      else{top=Math.max(MARGIN,vh-th-MARGIN);arrow='';}
      // horizontal: centre on trigger, clamp to viewport
      var left=r.left+r.width/2-tw/2;
      left=Math.max(MARGIN,Math.min(left,vw-tw-MARGIN));
      // arrow x relative to tooltip left
      var arrowX=r.left+r.width/2-left;
      arrowX=Math.max(16,Math.min(arrowX,tw-16));
      tip.style.top=top+'px';
      tip.style.left=left+'px';
      tip.style.setProperty('--arrow-x',arrowX+'px');
      tip.className='visible'+(arrow?' '+arrow:'');
    });
  }
  function hide(){
    tip.classList.remove('visible');
    tip.setAttribute('aria-hidden','true');
    tip.className='';
  }
  document.addEventListener('mouseenter',show,true);
  document.addEventListener('mouseleave',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
  document.addEventListener('focusin',show,true);
  document.addEventListener('focusout',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
})();
"""


# ── Main report generator ───────────────────────────────────────────────

def generate_rbac_report(data: dict, output_dir: str | pathlib.Path) -> pathlib.Path:
    """Generate the RBAC HTML report.

    Args:
        data: Result dict from collect_rbac_data (keys: tree, principals, stats, risks).
        output_dir: Directory to write the report into.

    Returns:
        Path to the generated HTML file.
    """
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "rbac-report.html"

    tree = data.get("tree", {})
    principals = data.get("principals", {})
    stats = data.get("stats", {})
    risks = data.get("risks", [])
    tenant_id = data.get("TenantId", "")
    tenant_display = data.get("TenantDisplayName", "")
    ts = format_date_short()
    report_id = f"CIQ-RBAC-{ts.replace('-', '').replace(':', '').replace(' ', '-')}"

    # ── RBAC Health Score ──
    rbac_score = stats.get("rbac_score", 100)
    sc = score_color(rbac_score)
    score_level = (
        "EXCELLENT" if rbac_score >= 90 else
        "GOOD" if rbac_score >= 70 else
        "NEEDS WORK" if rbac_score >= 50 else
        "AT RISK" if rbac_score >= 25 else
        "CRITICAL"
    )

    # Score ring SVG
    ring_r = 54
    ring_circ = 2 * 3.14159 * ring_r
    ring_dash = ring_circ * rbac_score / 100
    ring_gap = ring_circ - ring_dash
    score_ring = (
        f'<svg width="140" height="140" viewBox="0 0 140 140" role="img" aria-label="RBAC score {rbac_score}">'
        f'<circle cx="70" cy="70" r="{ring_r}" fill="none" stroke="var(--border)" stroke-width="10"/>'
        f'<circle cx="70" cy="70" r="{ring_r}" fill="none" stroke="{sc}" stroke-width="10" '
        f'stroke-dasharray="{ring_dash:.1f} {ring_gap:.1f}" stroke-linecap="round" '
        f'transform="rotate(-90 70 70)"/>'
        f'<text x="70" y="66" text-anchor="middle" font-size="28" font-weight="700" '
        f'fill="{sc}" font-family="var(--font-mono)">{rbac_score}</text>'
        f'<text x="70" y="86" text-anchor="middle" font-size="11" fill="var(--text-secondary)">/ 100</text>'
        f'</svg>'
    )

    # ── Executive summary stats ──
    total_a = stats.get("total_assignments", 0)
    active_a = stats.get("active_assignments", 0)
    eligible_a = stats.get("eligible_assignments", 0)
    priv_active = stats.get("privileged_active", 0)
    priv_eligible = stats.get("privileged_eligible", 0)
    n_mg = stats.get("management_groups", 0)
    n_sub = stats.get("subscriptions", 0)
    n_rg = stats.get("resource_groups", 0)
    n_res = stats.get("resources", 0)
    n_principals = stats.get("unique_principals", 0)
    n_groups = stats.get("groups_with_roles", 0)

    # Principal sub-type breakdown cards
    sub_type_counts = stats.get("principal_sub_types", {})
    sub_type_cards_html = ""
    if sub_type_counts:
        total_principals = sum(sub_type_counts.values()) or 1

        # Count assignments per sub-type for context
        all_assignments_for_subtypes = _collect_all_assignments(tree)
        subtype_assign_counts: dict[str, int] = {}
        subtype_priv_counts: dict[str, int] = {}
        for a in all_assignments_for_subtypes:
            pid = a.get("principal_id", "")
            psub = _principal_sub_type(pid, principals)
            subtype_assign_counts[psub] = subtype_assign_counts.get(psub, 0) + 1
            if a.get("is_privileged"):
                subtype_priv_counts[psub] = subtype_priv_counts.get(psub, 0) + 1

        cards = []
        for st in ("Member", "Guest", "Application", "ManagedIdentity", "Group", "Legacy", "Unknown"):
            cnt = sub_type_counts.get(st, 0)
            if cnt == 0:
                continue
            label = _SUBTYPE_LABELS.get(st, st)
            color = _SUBTYPE_COLORS.get(st, "#A8A6A3")
            icon = _SUBTYPE_ICONS.get(st, "")
            pct = cnt * 100 / total_principals
            n_assign = subtype_assign_counts.get(st, 0)
            n_priv = subtype_priv_counts.get(st, 0)
            desc = {
                "Member": "Internal Entra ID user accounts — employees and contractors within your tenant.",
                "Guest": "External B2B guest users invited from other tenants. Audit these for least-privilege access.",
                "Application": "App registrations and enterprise apps with client credentials. Rotate secrets regularly.",
                "ManagedIdentity": "Azure-managed identities (System or User Assigned). Preferred for automation — no credential management needed.",
                "Group": "Security groups whose members inherit the group's role assignments. Large groups widen blast radius.",
                "Legacy": "Legacy service principals — older identity format. Consider migrating to modern App Registrations.",
                "Unknown": "Principals whose type could not be resolved — may be deleted or have restricted visibility.",
            }.get(st, "")
            avg_assign = n_assign / cnt if cnt else 0
            priv_assign_pct = n_priv * 100 / n_assign if n_assign and n_priv else 0
            # Contextual insight based on data
            if st == "Guest" and cnt > 0:
                insight = f"&#10;&#10;\u26A0 {cnt} external {'identity has' if cnt == 1 else 'identities have'} access — review periodically to prevent stale guest accounts."
            elif st == "Group" and cnt > 0:
                insight = f"&#10;&#10;\u26A0 Group-based assignments affect all members. {cnt} {'group carries' if cnt == 1 else 'groups carry'} {n_assign:,} total {'assignment' if n_assign == 1 else 'assignments'} — verify group membership is current."
            elif n_priv and priv_assign_pct >= 50:
                insight = f"&#10;&#10;\u26A0 {priv_assign_pct:.0f}% of assignments are privileged — consider tightening with PIM or scoped roles."
            elif avg_assign >= 10:
                insight = f"&#10;&#10;\u26A0 High avg assignments ({avg_assign:.1f}/identity) — review for over-provisioned access."
            elif avg_assign <= 1 and cnt > 0:
                insight = f"&#10;&#10;\u2705 Low avg assignments ({avg_assign:.1f}/identity) — good least-privilege posture."
            else:
                insight = ""
            priv_line = f"&#10;\u2022 Privileged: {n_priv:,} of {n_assign:,} assignments ({priv_assign_pct:.0f}%)" if n_priv else ""
            tip = (
                f"{desc}"
                f"&#10;&#10;YOUR TENANT:"
                f"&#10;\u2022 {cnt} unique {'identity' if cnt == 1 else 'identities'} ({pct:.0f}% of all {total_principals} principals)"
                f"&#10;\u2022 Holding {n_assign:,} role {'assignment' if n_assign == 1 else 'assignments'} total"
                f"&#10;\u2022 Avg {avg_assign:.1f} assignments per identity"
                f"{priv_line}"
                f"{insight}"
            )
            cards.append(
                f'<div class="stat-card" data-tip="{tip}">'
                f'<div class="stat-value" style="color:{color}">{icon} {cnt}</div>'
                f'<div class="stat-label">{label}</div>'
                f'</div>'
            )
        sub_type_cards_html = (
            '<h3 style="font-size:14px;color:var(--text-secondary);margin:0 0 10px;font-weight:600" '
            'data-tip="Shows how many distinct identities hold role assignments, grouped by type. '
            'Each card counts unique principals \u2014 not assignments. Hover a card for details.">'
            'Identity Breakdown (unique principals)</h3>\n'
            '<div class="stat-grid" style="grid-template-columns:repeat(auto-fit,minmax(140px,1fr));margin-bottom:32px">\n'
            + '\n'.join(cards)
            + '\n</div>'
        )

    # Donut: Active vs Eligible
    status_donut = _donut_svg([
        ("Active", active_a, "#107C10"),
        ("Eligible (PIM)", eligible_a, "#FFB900"),
    ], size=140)

    # Donut: Privileged vs Standard
    priv_total = priv_active + priv_eligible
    std_total = total_a - priv_total
    priv_donut = _donut_svg([
        ("Privileged", priv_total, "#D13438"),
        ("Standard", std_total, "#4EA8DE"),
    ], size=140)

    # Bar chart: top roles
    role_counts: dict[str, int] = {}
    all_assignments = _collect_all_assignments(tree)
    for a in all_assignments:
        role_counts[a.get("role_name", "Unknown")] = role_counts.get(a.get("role_name", "Unknown"), 0) + 1
    top_roles = sorted(role_counts.items(), key=lambda x: -x[1])[:10]
    role_bar = _bar_chart_svg([
        (name, count, "#D13438" if name in _PRIVILEGED_SET else "#4EA8DE")
        for name, count in top_roles
    ], width=520)

    # Bar chart: principal types (sub-types)
    ptype_counts: dict[str, int] = {}
    for a in all_assignments:
        pid = a.get("principal_id", "")
        psub = _principal_sub_type(pid, principals)
        label = _SUBTYPE_LABELS.get(psub, psub)
        ptype_counts[label] = ptype_counts.get(label, 0) + 1
    ptype_bar = _bar_chart_svg([
        (k, v, _SUBTYPE_COLORS.get(
            next((sk for sk, sl in _SUBTYPE_LABELS.items() if sl == k), k), "#A8A6A3"))
        for k, v in sorted(ptype_counts.items(), key=lambda x: -x[1])
    ], width=520)

    # Scope distribution bar
    scope_counts: dict[str, int] = {}
    for a in all_assignments:
        sl = a.get("scope_level", "Unknown")
        scope_counts[sl] = scope_counts.get(sl, 0) + 1
    scope_bar = _bar_chart_svg([
        (k, v, _TYPE_COLORS.get(k, "#A8A6A3"))
        for k, v in sorted(scope_counts.items(), key=lambda x: -x[1])
    ], width=520)

    # Risk severity counts
    risk_counts: dict[str, int] = {}
    for r in risks:
        risk_counts[r.get("severity", "medium")] = risk_counts.get(r.get("severity", "medium"), 0) + 1
    n_critical = risk_counts.get("critical", 0)
    n_high = risk_counts.get("high", 0)
    n_medium = risk_counts.get("medium", 0)
    n_low = risk_counts.get("low", 0)

    # Severity donut
    sev_donut = _donut_svg([
        ("Critical", n_critical, SEVERITY_COLORS.get("critical", "#D13438")),
        ("High", n_high, SEVERITY_COLORS.get("high", "#F7630C")),
        ("Medium", n_medium, SEVERITY_COLORS.get("medium", "#FFB900")),
        ("Low", n_low, SEVERITY_COLORS.get("low", "#107C10")),
    ], size=140) if risks else ""

    # ── Tooltip text constants (after all data is computed) ──
    TIP_SCORE = (
        f"RBAC Health Score (0\u2013100) measures how well your tenant follows "
        f"least-privilege principles. Deductions apply for standing Owner/UAA "
        f"at broad scopes, SPs with Owner, large privileged groups, and custom "
        f"roles at high scopes."
        f"&#10;&#10;90+ Excellent \u00b7 70\u201389 Good \u00b7 50\u201369 Needs Work \u00b7 "
        f"25\u201349 At Risk \u00b7 &lt;25 Critical"
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Score: {rbac_score}/100 ({score_level})"
        f"&#10;\u2022 Critical/high risks: {n_critical + n_high}"
    )
    TIP_TOTAL_ASSIGNMENTS = (
        f"Total count of all Azure RBAC role assignments across the "
        f"tenant \u2014 both Active (always-on) and PIM Eligible (just-in-time). "
        f"A high number may indicate role sprawl."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Total: {total_a:,}"
        f"&#10;\u2022 Active: {active_a:,}"
        f"&#10;\u2022 PIM eligible: {eligible_a:,}"
    )
    TIP_ACTIVE = (
        f"Permanently active role assignments that are always in effect. "
        f"These provide standing access without requiring activation. "
        f"Prefer PIM Eligible for privileged roles."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Active assignments: {active_a:,} ({active_a*100//max(total_a,1)}% of total)"
        f"&#10;\u2022 Privileged standing access: {priv_active:,}"
    )
    TIP_PIM_ELIGIBLE = (
        f"Assignments configured through Privileged Identity Management "
        f"(PIM). Users must activate these just-in-time with MFA and "
        f"justification, reducing the window of standing access."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 PIM eligible: {eligible_a:,} ({eligible_a*100//max(total_a,1)}% of total)"
        f"&#10;\u2022 Privileged roles behind PIM: {priv_eligible:,}"
    )
    TIP_PRIVILEGED = (
        f"Assignments granting high-impact permissions: Owner, Contributor, "
        f"User Access Administrator, Security Admin, Key Vault Admin, and "
        f"similar. These should be tightly controlled."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Privileged total: {priv_active + priv_eligible:,} ({(priv_active+priv_eligible)*100//max(total_a,1)}% of all)"
        f"&#10;\u2022 Standing active: {priv_active:,}"
        f"&#10;\u2022 PIM-eligible: {priv_eligible:,}"
    )
    TIP_PRINCIPALS = (
        f"Distinct identities (users, service principals, groups) that hold "
        f"at least one role assignment. Review for orphaned or inactive "
        f"principals."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Total principals: {n_principals:,}"
        f"&#10;\u2022 Member users: {sub_type_counts.get('Member', 0)}"
        f"&#10;\u2022 Guest users: {sub_type_counts.get('Guest', 0)}"
        f"&#10;\u2022 App registrations: {sub_type_counts.get('Application', 0)}"
        f"&#10;\u2022 Managed identities: {sub_type_counts.get('ManagedIdentity', 0)}"
        f"&#10;\u2022 Groups: {sub_type_counts.get('Group', 0)}"
    )
    TIP_GROUPS = (
        f"Security or Microsoft 365 groups that have been assigned Azure "
        f"RBAC roles. All group members inherit the role. Large groups with "
        f"privileged roles create a wide blast radius."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Groups with roles: {n_groups:,}"
    )
    TIP_CRITICAL_RISKS = (
        f"Critical risk findings \u2014 standing Owner or User Access "
        f"Administrator at Management Group or Subscription scope. "
        f"Convert to PIM-eligible immediately."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Critical findings: {n_critical:,}"
    )
    TIP_HIGH_RISKS = (
        f"High risk findings \u2014 service principals with Owner, standing "
        f"Contributor at MG scope, or large groups with privileged roles. "
        f"Address after Critical items."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 High findings: {n_high:,}"
    )
    TIP_MG = (
        f"Management Groups are containers above subscriptions used to organize "
        f"and apply policy at scale. Role assignments here cascade to all child "
        f"subscriptions and resources."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Management groups: {n_mg:,}"
    )
    TIP_SUB = (
        f"Azure Subscriptions are billing and access boundaries. Roles at this "
        f"scope apply to every resource group and resource within."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Subscriptions: {n_sub:,}"
    )
    TIP_RG = (
        f"Resource Groups are logical containers for related Azure resources. "
        f"Roles at this scope apply to all resources in the group."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Resource groups: {n_rg:,}"
    )
    TIP_RES = (
        f"Individual Azure resources (VMs, storage accounts, databases, etc.). "
        f"Roles at this scope have the narrowest blast radius \u2014 prefer this "
        f"for least privilege."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Resources with assignments: {n_res:,}"
    )
    active_pct = active_a * 100 // max(total_a, 1)
    eligible_pct = eligible_a * 100 // max(total_a, 1)
    priv_standing_pct = priv_active * 100 // max(active_a, 1)
    _ae_insight = (
        f"&#10;&#10;\u26A0 {priv_active:,} privileged assignments are always-on "
        f"\u2014 convert to PIM Eligible for just-in-time access."
        if eligible_a < priv_active
        else f"&#10;&#10;\u2705 PIM covers more than standing privileged access \u2014 good posture."
    )
    TIP_CHART_ACTIVE_ELIGIBLE = (
        f"Active assignments are always-on; PIM Eligible require activation "
        f"and expire automatically. Best practice: convert privileged Active "
        f"roles to Eligible for just-in-time access."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Active: {active_a:,} ({active_pct}% of total)"
        f"&#10;\u2022 PIM Eligible: {eligible_a:,} ({eligible_pct}% of total)"
        f"&#10;\u2022 Privileged always-on: {priv_active:,} ({priv_standing_pct}% of active)"
        f"{_ae_insight}"
    )
    priv_pct = priv_total * 100 // max(total_a, 1)
    std_pct = std_total * 100 // max(total_a, 1)
    _ps_insight = (
        f"&#10;&#10;\u26A0 {priv_pct}% of assignments are privileged \u2014 aim for &lt;20%. "
        f"Replace Owner/Contributor with scoped or reader roles where possible."
        if priv_pct > 20
        else f"&#10;&#10;\u2705 Privileged ratio ({priv_pct}%) is within healthy range (&lt;20%)."
    )
    TIP_CHART_PRIV_STD = (
        f"Privileged roles (Owner, Contributor, User Access Admin, etc.) grant "
        f"broad write or management access. Standard roles are read-only or "
        f"tightly scoped. Aim for &lt;20% privileged."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Privileged: {priv_total:,} ({priv_pct}%)"
        f"&#10;\u2022 Standard: {std_total:,} ({std_pct}%)"
        f"&#10;\u2022 Priv active: {priv_active:,} \u00b7 Priv eligible: {priv_eligible:,}"
        f"{_ps_insight}"
    )
    n_priv_roles_in_top = sum(1 for name, _ in top_roles if name in _PRIVILEGED_SET)
    _top1 = top_roles[0] if top_roles else ("", 0)
    _tr_insight = (
        f"&#10;&#10;\u26A0 {n_priv_roles_in_top} of top 10 roles are privileged (red bars) "
        f"\u2014 review if reader or custom roles can replace them."
        if n_priv_roles_in_top >= 3
        else f"&#10;&#10;\u2705 Only {n_priv_roles_in_top} of top 10 are privileged \u2014 reasonable distribution."
    )
    TIP_CHART_TOP_ROLES = (
        f"The 10 most assigned roles across all scopes. Red = privileged, "
        f"blue = standard. High counts of Owner or Contributor signal "
        f"over-permissioning."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Most assigned: {_top1[0]} ({_top1[1]:,})"
        f"&#10;\u2022 Distinct roles in use: {len(role_counts):,}"
        f"&#10;\u2022 Privileged in top 10: {n_priv_roles_in_top}"
        f"{_tr_insight}"
    )
    # Assignment counts per principal sub-type (matches bar chart values)
    _pt_top = sorted(ptype_counts.items(), key=lambda x: -x[1])
    _pt_lines = ''.join(f"&#10;\u2022 {k}: {v:,} assignments" for k, v in _pt_top[:5])
    _guest_assign = ptype_counts.get('Guest User', 0)
    _pt_insight = (
        f"&#10;&#10;\u26A0 {_guest_assign:,} assignments held by external guests "
        f"\u2014 review for stale or over-privileged access."
        if _guest_assign > 0
        else ""
    )
    TIP_CHART_PTYPES = (
        f"Role assignments grouped by identity type. Shows which types hold "
        f"the most access in your tenant. Managed Identities are preferred "
        f"for automation; Guest Users need periodic review."
        f"&#10;&#10;YOUR TENANT:"
        f"{_pt_lines}"
        f"{_pt_insight}"
    )
    TIP_CHART_SCOPE = (
        f"Where assignments are granted in the hierarchy. ManagementGroup-level "
        f"roles cascade to everything below \u2014 minimize these. "
        f"Prefer ResourceGroup or Resource scope for least privilege."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Management groups: {n_mg}"
        f"&#10;\u2022 Subscriptions: {n_sub}"
        f"&#10;\u2022 Resource groups: {n_rg}"
        f"&#10;\u2022 Resources: {n_res}"
    )
    n_total_risks = len(risks)
    TIP_CHART_SEV = (
        f"Distribution of detected risks by severity. Critical = standing "
        f"Owner/UAA at broad scope. High = SP with Owner, large privileged "
        f"groups, or Contributor at MG. Medium = custom roles at high scopes. "
        f"Address Critical first."
        f"&#10;&#10;YOUR TENANT:"
        f"&#10;\u2022 Total risks: {n_total_risks:,}"
        f"&#10;\u2022 Critical: {n_critical}"
        f"&#10;\u2022 High: {n_high}"
        f"&#10;\u2022 Medium: {n_medium}"
        f"&#10;\u2022 Low: {n_low}"
    )

    # Tree HTML
    scope_names = _build_scope_names(tree)
    risk_lookup = _build_risk_lookup(risks)
    tree_html = _render_tree_node(tree, principals, 0, scope_names, risk_lookup)

    # Risks HTML
    risks_html = _render_risks(risks, principals, scope_names)

    # Appendix
    appendix_html = _render_appendix(principals)

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Azure RBAC Hierarchy &amp; Access Analysis — EnterpriseSecurityIQ</title>
<style>{get_css()}{_tree_css()}</style>
</head>
<body>
<a href="#main" class="skip-nav">Skip to content</a>

<!-- Top Navigation -->
<nav class="top-nav" aria-label="Report sections">
  <span class="brand" data-tip="EnterpriseSecurityIQ Azure RBAC Assessment.\n• Automated RBAC posture analysis for your tenant\n• Covers role assignments, PIM eligibility, and privileged access\n• Read-only: no changes are made to your environment">&#128274; EnterpriseSecurityIQ RBAC Analysis</span>
  <div class="nav-dropdown">
    <button class="nav-toggle" data-tip="Report metadata, confidentiality notice, and audit attestation.">Document</button>
    <div class="nav-menu">
      <a href="#doc-control" data-tip="Report ID, assessment scope, classification, and data integrity details.">Report Metadata</a>
      <a href="#doc-control" data-tip="Data handling and distribution restrictions for this report." onclick="setTimeout(function(){{document.querySelector('#doc-control .conf-notice').scrollIntoView({{behavior:'smooth'}})}},50)">Confidentiality Notice</a>
      <a href="#doc-control" data-tip="Attestation that this report was generated from read-only API data." onclick="setTimeout(function(){{document.querySelector('#doc-control h3').scrollIntoView({{behavior:'smooth'}})}},50)">Audit Attestation</a>
    </div>
  </div>
  <a href="#summary" data-tip="Score, KPI cards, scope summary, and visual breakdown charts.">Summary</a>
  <a href="#risks" data-tip="Table of all detected RBAC risks ranked by severity with remediation steps.">Risk Analysis</a>
  <a href="#tree" data-tip="Interactive drill-down tree of every role assignment organized by MG \u2192 Sub \u2192 RG \u2192 Resource.">Hierarchy Tree</a>
  <a href="#appendix" data-tip="Lookup table mapping principal Object IDs to display names, types, and UPNs.">Appendix</a>
  <div class="zoom-controls" aria-label="Page zoom">
    <button onclick="zoomOut()" aria-label="Zoom out" data-tip="Decrease page zoom by 10% (min 70%). Useful for seeing more of the tree at once.">&minus;</button>
    <span id="zoom-label">100%</span>
    <button onclick="zoomIn()" aria-label="Zoom in" data-tip="Increase page zoom by 10% (max 150%). Useful for reading dense assignment tables.">&plus;</button>
    <button onclick="zoomReset()" aria-label="Reset zoom" data-tip="Reset page zoom to 100%." style="font-size:11px">Reset</button>
  </div>
  <button class="theme-btn" onclick="toggleTheme()" style="margin:0;padding:6px 14px"
          aria-label="Toggle dark and light theme"
          data-tip="Switch between dark and light theme. Your preference is saved in browser storage.">
    Switch to Light
  </button>
</nav>

<main id="main" class="full-width-content">

<!-- ── Document Control ── -->
<section id="doc-control" class="section">
  <h1 class="page-title">EnterpriseSecurityIQ &mdash; Azure RBAC Assessment Report</h1>
  <table class="doc-control-table">
    <tr><th>Report Identifier</th><td>{esc(report_id)}</td></tr>
    <tr><th>Assessment Name</th><td>EnterpriseSecurityIQ Azure RBAC Hierarchy &amp; Access Analysis</td></tr>
    <tr><th>Date Generated</th><td>{esc(ts)}</td></tr>
    <tr><th>Tenant ID</th><td><code>{esc(tenant_id) if tenant_id else 'N/A'}</code></td></tr>
    <tr><th>Tenant Name</th><td>{esc(tenant_display) if tenant_display else 'Unknown'}</td></tr>
    <tr><th>Assessment Domains</th><td>RBAC Role Assignments, PIM Eligibility, Privileged Access, Group Expansion</td></tr>
    <tr><th>Classification</th><td>CONFIDENTIAL &mdash; Authorized Recipients Only</td></tr>
    <tr><th>Tool</th><td>EnterpriseSecurityIQ AI Agent v{VERSION}</td></tr>
    <tr><th>Collection Method</th><td>Azure Resource Manager API + Microsoft Graph API (Read-Only)</td></tr>
  </table>
  <div class="conf-notice">
    <strong>CONFIDENTIALITY NOTICE:</strong> This document contains sensitive security and compliance
    information about the assessed environment. Distribution is restricted to authorized personnel only.
  </div>
  <h3>Audit Attestation</h3>
  <table class="doc-control-table">
    <tr><th>Assessment Scope</th><td>Azure RBAC hierarchy review across management groups, subscriptions, and resource groups</td></tr>
    <tr><th>Data Integrity</th><td>All evidence collected via read-only API calls; no tenant modifications were made</td></tr>
    <tr><th>Evidence Records</th><td>{stats.get('total_assignments', 0):,} role assignments and {stats.get('unique_principals', 0):,} principals analyzed</td></tr>
    <tr><th>Report Hash (SHA-256)</th><td><code id="report-hash">Computed at render</code></td></tr>
    <tr><th>Assessment Period</th><td>{esc(ts)} (point-in-time snapshot)</td></tr>
  </table>
</section>

<!-- ── Executive Summary ── -->
<section id="summary" class="section" aria-labelledby="summary-heading">
  <h2 id="summary-heading" class="page-title">Executive Summary</h2>
  <p style="color:var(--text-secondary);font-size:14px;line-height:1.6;max-width:960px;margin:8px 0 20px">
    This report provides a comprehensive view of Azure Role-Based Access Control (RBAC) across your
    tenant&#8217;s management group and subscription hierarchy. It maps every role assignment &#8212;
    active and PIM-eligible &#8212; to the exact scope where it is granted, identifies inherited vs.
    directly assigned roles, flags privileged and custom role usage, and highlights security risks
    such as broad Owner/Contributor grants and standing access without PIM. Use the interactive tree
    to drill into any scope, filter by status or privilege level, and export assignments for audit
    or remediation.
  </p>
  <div class="meta-bar">
    <span>Generated: {esc(ts)}</span>
    {f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_display)} ({esc(tenant_id[:8])}\u2026)</span>' if tenant_id and tenant_display else f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_id)}</span>' if tenant_id else ''}
    <span>EnterpriseSecurityIQ v{VERSION}</span>
  </div>

  <!-- Score ring + KPI cards -->
  <div style="display:flex;gap:32px;align-items:center;margin-bottom:24px;flex-wrap:wrap">
    <div style="text-align:center;min-width:160px" class="score-ring-wrap" data-tip="{TIP_SCORE}">
      {score_ring}
      <div style="margin-top:8px;font-weight:700;font-size:14px;color:{sc}">{score_level}</div>
      <div style="font-size:11px;color:var(--text-muted)">RBAC Health Score</div>
    </div>
    <div class="stat-grid" style="grid-template-columns:repeat(auto-fit,minmax(140px,1fr));flex:1">
    <div class="stat-card" data-tip="{TIP_TOTAL_ASSIGNMENTS}"><div class="stat-value">{total_a}</div><div class="stat-label">Total Assignments</div></div>
    <div class="stat-card" data-tip="{TIP_ACTIVE}"><div class="stat-value" style="color:#107C10">{active_a}</div><div class="stat-label">Active</div></div>
    <div class="stat-card" data-tip="{TIP_PIM_ELIGIBLE}"><div class="stat-value" style="color:#FFB900">{eligible_a}</div><div class="stat-label">PIM Eligible</div></div>
    <div class="stat-card" data-tip="{TIP_PRIVILEGED}"><div class="stat-value" style="color:#D13438">{priv_active + priv_eligible}</div><div class="stat-label">Privileged</div></div>
    <div class="stat-card" data-tip="{TIP_PRINCIPALS}"><div class="stat-value">{n_principals}</div><div class="stat-label">Unique Principals</div></div>
    <div class="stat-card" data-tip="{TIP_GROUPS}"><div class="stat-value">{n_groups}</div><div class="stat-label">Groups with Roles</div></div>
    <div class="stat-card" data-tip="{TIP_CRITICAL_RISKS}"><div class="stat-value" style="color:#D13438">{n_critical}</div><div class="stat-label">Critical Risks</div></div>
    <div class="stat-card" data-tip="{TIP_HIGH_RISKS}"><div class="stat-value" style="color:#F7630C">{n_high}</div><div class="stat-label">High Risks</div></div>
  </div>
  </div>

  <!-- Scope summary -->
  <div class="stat-grid" style="grid-template-columns:repeat(auto-fit,minmax(140px,1fr));margin-bottom:32px">
    <div class="stat-card" data-tip="{TIP_MG}"><div class="stat-value" style="color:#0078D4">{n_mg}</div><div class="stat-label">Management Groups</div></div>
    <div class="stat-card" data-tip="{TIP_SUB}"><div class="stat-value" style="color:#00B7C3">{n_sub}</div><div class="stat-label">Subscriptions</div></div>
    <div class="stat-card" data-tip="{TIP_RG}"><div class="stat-value" style="color:#F7630C">{n_rg}</div><div class="stat-label">Resource Groups</div></div>
    <div class="stat-card" data-tip="{TIP_RES}"><div class="stat-value" style="color:#A8A6A3">{n_res}</div><div class="stat-label">Resources</div></div>
  </div>

  <!-- Identity Breakdown -->
  {sub_type_cards_html}

  <!-- Charts -->
  <div class="exec-grid">
    <div class="exec-panel" data-tip="{TIP_CHART_ACTIVE_ELIGIBLE}">
      <h3><span class="exec-panel-title">Active vs Eligible</span></h3>
      <div style="text-align:center">{status_donut}</div>
      <div class="legend" style="justify-content:center;margin-top:16px">
        <span class="legend-item"><span class="legend-dot" style="background:#107C10"></span> Active ({active_a})</span>
        <span class="legend-item"><span class="legend-dot" style="background:#FFB900"></span> PIM Eligible ({eligible_a})</span>
      </div>
    </div>
    <div class="exec-panel" data-tip="{TIP_CHART_PRIV_STD}">
      <h3><span class="exec-panel-title">Privileged vs Standard</span></h3>
      <div style="text-align:center">{priv_donut}</div>
      <div class="legend" style="justify-content:center;margin-top:16px">
        <span class="legend-item"><span class="legend-dot" style="background:#D13438"></span> Privileged ({priv_total})</span>
        <span class="legend-item"><span class="legend-dot" style="background:#4EA8DE"></span> Standard ({std_total})</span>
      </div>
    </div>
    <div class="exec-panel" data-tip="{TIP_CHART_TOP_ROLES}">
      <h3><span class="exec-panel-title">Top Roles</span></h3>
      {role_bar}
    </div>
    <div class="exec-panel" data-tip="{TIP_CHART_PTYPES}">
      <h3><span class="exec-panel-title">Principal Types</span></h3>
      {ptype_bar}
    </div>
    <div class="exec-panel" data-tip="{TIP_CHART_SCOPE}">
      <h3><span class="exec-panel-title">Assignments by Scope Level</span></h3>
      {scope_bar}
    </div>
    <div class="exec-panel" data-tip="{TIP_CHART_SEV}">
      <h3><span class="exec-panel-title">Risk Severity Distribution</span></h3>
      <div style="text-align:center">{sev_donut if sev_donut else '<p class="empty">No risks detected</p>'}</div>
      <div class="legend" style="justify-content:center;margin-top:16px">
        <span class="legend-item" data-tip="Standing Owner or User Access Admin at management group or subscription scope. Immediate remediation required."><span class="legend-dot" style="background:{SEVERITY_COLORS.get('critical','#D13438')}"></span> Critical ({n_critical})</span>
        <span class="legend-item" data-tip="Service principals with Owner, large privileged groups, or Contributor at management group scope. Address after Critical items."><span class="legend-dot" style="background:{SEVERITY_COLORS.get('high','#F7630C')}"></span> High ({n_high})</span>
        <span class="legend-item" data-tip="Custom roles at high scopes or moderate over-provisioning patterns. Review and scope down."><span class="legend-dot" style="background:{SEVERITY_COLORS.get('medium','#FFB900')}"></span> Medium ({n_medium})</span>
        <span class="legend-item" data-tip="Informational findings \u2014 minor configuration improvements or best-practice deviations."><span class="legend-dot" style="background:{SEVERITY_COLORS.get('low','#107C10')}"></span> Low ({n_low})</span>
      </div>
    </div>
  </div>
</section>

<!-- ── Risk Analysis ── -->
<section id="risks" class="section" aria-labelledby="risks-heading">
  <h2 id="risks-heading" data-tip="Automated risk detection flags standing privileged access, over-broad scope grants, and missing PIM coverage. Address Critical findings first, then High. {len(risks)} findings detected in your tenant.">&#9888;&#65039; Risk Analysis ({len(risks)} findings)</h2>
  <div class="how-to-read">
    <h4>How to read this section</h4>
    <p>Risks are flagged when standing privileged access is detected at broad scopes,
    service principals hold Owner rights, or large groups inherit privileged roles.
    Address <strong>Critical</strong> items first, then <strong>High</strong>.</p>
  </div>
  {risks_html}
</section>

<!-- ── Hierarchy Tree ── -->
<section id="tree" class="section" aria-labelledby="tree-heading">
  <h2 id="tree-heading" data-tip="Interactive drill-down showing every role assignment organized by Management Group \u2192 Subscription \u2192 Resource Group \u2192 Resource. Click nodes to expand; use filters to narrow results.">&#127795; RBAC Hierarchy Tree</h2>
  <div class="how-to-read">
    <h4>How to read this tree</h4>
    <p>Click any node to expand or collapse it. Use <kbd>Arrow</kbd> keys to navigate,
    <kbd>Right</kbd> to expand, <kbd>Left</kbd> to collapse, <kbd>Home</kbd>/<kbd>End</kbd>
    to jump to first/last. Each assignment shows an <span class="badge badge-direct" style="font-size:9px">DIRECT</span>
    badge if assigned at this exact scope, or <span class="badge badge-inherited" style="font-size:9px">INHERITED</span>
    if granted at a parent scope. Use the <strong>Hide Inherited</strong> toggle (on by default) to focus on direct assignments only.
    Resources with no direct assignments show inherited parent-scope assignments with a
    <span class="rbac-count rbac-count-inherited" style="font-size:9px">&darr;N</span> badge.
    Click column headers to sort assignment tables.</p>
  </div>
  <div class="filter-bar" role="search" aria-label="Filter tree">
    <label for="tree-filter">Search:</label>
    <input id="tree-filter" type="search" placeholder="Search tree (role, principal, scope)…"
           oninput="filterTree()" aria-label="Filter tree by keyword"
           aria-describedby="tree-live" autocomplete="off"
           data-tip="Type to filter by role name, principal name, scope, or any text. Only matching assignment rows and their parent nodes are shown.">
    <label for="filter-status">Status:</label>
    <select id="filter-status" onchange="filterTree()" aria-label="Filter by assignment status"
            data-tip="Filter by PIM status. Active = always-on access. Eligible = requires PIM activation with MFA and justification.">
      <option value="">All</option>
      <option value="Active">Active</option>
      <option value="Eligible">Eligible</option>
    </select>
    <label for="filter-privileged" style="display:inline-flex;align-items:center;gap:4px;cursor:pointer;white-space:nowrap"
           data-tip="When checked, shows only privileged role assignments (Owner, Contributor, User Access Admin, etc.). Hides standard/reader roles.">
      <input type="checkbox" id="filter-privileged" onchange="filterTree()"> Privileged only
    </label>
    <label for="filter-hide-inherited" style="display:inline-flex;align-items:center;gap:4px;cursor:pointer;white-space:nowrap"
           data-tip="When checked, hides inherited role assignments and shows only DIRECT assignments made at each scope. Uncheck to see all cascaded parent assignments.">
      <input type="checkbox" id="filter-hide-inherited" onchange="filterTree()" checked> Hide Inherited
    </label>
    <button onclick="expandAll()" aria-label="Expand all tree nodes"
            data-tip="Expand every tree node and reveal all assignment tables. Useful for full review or before CSV export.">Expand All</button>
    <button onclick="collapseAll()" aria-label="Collapse all tree nodes"
            data-tip="Collapse to subscription level for a high-level overview. Hides resource group and resource details.">Collapse All</button>
  </div>
  <div id="tree-live" class="sr-only" aria-live="polite" aria-atomic="true"></div>
  <div id="tree-container" role="tree" aria-label="RBAC hierarchy with {total_a} assignments across {n_sub} subscriptions">
    {tree_html}
  </div>
</section>

<!-- ── Principal Appendix ── -->
<section id="appendix" class="section" aria-labelledby="appendix-heading">
  <h2 id="appendix-heading" data-tip="Reference table mapping Entra ID Object GUIDs to display names, types, and UPNs. Use this to cross-reference principals from the tree or risk table.">&#128218; Principal ID Appendix</h2>
  <div class="filter-bar" role="search" aria-label="Filter principals">
    <label for="appendix-filter">Search:</label>
    <input id="appendix-filter" type="search" placeholder="Search by name, ID, UPN…"
           oninput="filterAppendix()" aria-label="Filter principal table"
           aria-describedby="appendix-live" autocomplete="off"
           data-tip="Filter the principal reference table by display name, Object ID, UPN, app ID, or identity type.">
  </div>
  <div id="appendix-live" class="sr-only" aria-live="polite" aria-atomic="true"></div>
  {appendix_html}
</section>

</main>

<button class="back-to-top" aria-label="Back to top">&#8593;</button>
<div id="ciq-tooltip" role="tooltip" aria-hidden="true"></div>

<script>{get_js()}</script>
<script>{_tree_js()}</script>
</body>
</html>"""

    report_hash = hashlib.sha256(html.encode("utf-8")).hexdigest()
    html = html.replace("Computed at render", report_hash)
    out_path.write_text(html, encoding="utf-8")
    log.info("[RbacReport] Written to %s (%d KB)", out_path, len(html) // 1024)

    # Excel export
    xlsx_path = _export_rbac_excel(data, output_dir)

    return out_path


# ── Internal helpers ─────────────────────────────────────────────────────

_PRIVILEGED_SET = {
    "Owner", "Contributor", "User Access Administrator",
    "Security Admin", "Global Administrator",
    "Key Vault Administrator", "Key Vault Secrets Officer",
    "Storage Account Contributor", "Virtual Machine Contributor",
    "Network Contributor", "SQL Server Contributor",
    "SQL Security Manager", "Monitoring Contributor",
    "Log Analytics Contributor", "Automation Operator",
    "Managed Identity Operator", "Role Based Access Control Administrator",
    "Azure Kubernetes Service Cluster Admin Role",
    "Azure Kubernetes Service RBAC Admin",
}


def _collect_all_assignments(node: dict) -> list[dict]:
    """Recursively collect all assignments from the tree."""
    result: list[dict] = []
    result.extend(node.get("assignments", []))
    result.extend(node.get("eligible", []))
    for rg in node.get("resource_groups", []):
        result.extend(rg.get("assignments", []))
        for res in rg.get("resources", []):
            result.extend(res.get("assignments", []))
    for child in node.get("children", []):
        result.extend(_collect_all_assignments(child))
    return result


# ── Excel Export ─────────────────────────────────────────────────────────

def _export_rbac_excel(data: dict, output_dir: pathlib.Path) -> pathlib.Path:
    """Generate a multi-tab RBAC Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.warning("[RbacReport] openpyxl not installed — skipping Excel export")
        return output_dir / "rbac-report.xlsx"

    wb = Workbook()
    thin_border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
    )

    tree = data.get("tree", {})
    principals = data.get("principals", {})
    stats = data.get("stats", {})
    risks = data.get("risks", [])
    all_assignments = _collect_all_assignments(tree)
    scope_names = _build_scope_names(tree)

    def _add_sheet(title: str, headers: list[str], rows: list[list], widths: list[int] | None = None):
        ws = wb.create_sheet(title=title)
        hdr_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
        hdr_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        val_font = Font(name="Segoe UI", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = val_font
                cell.border = thin_border
                cell.alignment = wrap_align

        if widths:
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(rows)+1}"
        ws.freeze_panes = "A2"

    # ── Tab 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    label_font = Font(name="Segoe UI", bold=True, size=11)
    value_font = Font(name="Segoe UI", size=11)

    rbac_score = stats.get("rbac_score", 100)
    score_level = (
        "EXCELLENT" if rbac_score >= 90 else
        "GOOD" if rbac_score >= 70 else
        "NEEDS WORK" if rbac_score >= 50 else
        "AT RISK" if rbac_score >= 25 else
        "CRITICAL"
    )

    # Risk severity counts
    sev_counts: dict[str, int] = {}
    for r in risks:
        s = r.get("severity", "medium")
        sev_counts[s] = sev_counts.get(s, 0) + 1

    summary_rows = [
        ("Report", "Azure RBAC Hierarchy & Access Analysis"),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("EnterpriseSecurityIQ Version", VERSION),
        ("", ""),
        ("RBAC Health Score", f"{rbac_score}/100"),
        ("Score Level", score_level),
        ("Total Risk Findings", len(risks)),
        ("", ""),
        ("Risk Severity Distribution", ""),
        ("  Critical", sev_counts.get("critical", 0)),
        ("  High", sev_counts.get("high", 0)),
        ("  Medium", sev_counts.get("medium", 0)),
        ("  Low", sev_counts.get("low", 0)),
        ("", ""),
        ("Assignment Summary", ""),
        ("  Total Assignments", stats.get("total_assignments", 0)),
        ("  Active Assignments", stats.get("active_assignments", 0)),
        ("  PIM Eligible", stats.get("eligible_assignments", 0)),
        ("  Privileged (Active)", stats.get("privileged_active", 0)),
        ("  Privileged (Eligible)", stats.get("privileged_eligible", 0)),
        ("  Custom Role Assignments", stats.get("custom_roles", 0)),
        ("", ""),
        ("Identity Summary", ""),
        ("  Unique Principals", stats.get("unique_principals", 0)),
        ("  Groups with Roles", stats.get("groups_with_roles", 0)),
        ("", ""),
        ("Scope Summary", ""),
        ("  Management Groups", stats.get("management_groups", 0)),
        ("  Subscriptions", stats.get("subscriptions", 0)),
        ("  Resource Groups", stats.get("resource_groups", 0)),
        ("  Resources", stats.get("resources", 0)),
    ]
    for ri, (label, val) in enumerate(summary_rows, 1):
        cl = ws.cell(row=ri, column=1, value=label)
        cl.font = label_font
        cv = ws.cell(row=ri, column=2, value=val)
        cv.font = value_font
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45

    # ── Tab 2: All Assignments ──
    assign_rows = []
    for a in all_assignments:
        pid = a.get("principal_id", "")
        pinfo = principals.get(pid, {})
        pname = pinfo.get("display_name", pid)
        psub = pinfo.get("sub_type", pinfo.get("type", a.get("principal_type", "Unknown")))
        psub_label = _SUBTYPE_LABELS.get(psub, psub)
        mi_type = pinfo.get("mi_type", "")
        if mi_type:
            psub_label = f"{psub_label} ({mi_type})"
        status = "Eligible" if a.get("status") == "Eligible" else a.get("status", "Active")
        role = a.get("role_name", "Unknown")
        scope_raw = a.get("scope", "")
        scope_display = _scope_friendly(scope_raw, scope_names)
        scope_level = a.get("scope_level", "Unknown")
        is_priv = "Yes" if a.get("is_privileged") else "No"
        is_custom = "Yes" if a.get("is_custom") else "No"
        assign_rows.append([
            status, role, pname, psub_label, scope_display, scope_level,
            is_priv, is_custom, pid, scope_raw,
        ])
    _add_sheet(
        "All Assignments",
        ["Status", "Role", "Principal", "Principal Type", "Scope Name",
         "Scope Level", "Privileged", "Custom", "Principal ID", "Scope (Full)"],
        assign_rows,
        [10, 30, 30, 15, 30, 15, 10, 10, 36, 50],
    )

    # ── Tab 3: Privileged Access ──
    priv_rows = [r for r in assign_rows if r[6] == "Yes"]
    _add_sheet(
        "Privileged Access",
        ["Status", "Role", "Principal", "Principal Type", "Scope Name",
         "Scope Level", "Privileged", "Custom", "Principal ID", "Scope (Full)"],
        priv_rows,
        [10, 30, 30, 15, 30, 15, 10, 10, 36, 50],
    )

    # ── Tab 4: PIM Eligible ──
    pim_rows = [r for r in assign_rows if r[0] == "Eligible"]
    _add_sheet(
        "PIM Eligible",
        ["Status", "Role", "Principal", "Principal Type", "Scope Name",
         "Scope Level", "Privileged", "Custom", "Principal ID", "Scope (Full)"],
        pim_rows,
        [10, 30, 30, 15, 30, 15, 10, 10, 36, 50],
    )

    # ── Tab 5: Risk Findings ──
    risk_rows = []
    for r in risks:
        pid = r.get("principal_id", "")
        pname = principals.get(pid, {}).get("display_name", pid)
        scope_raw = r.get("scope", "")
        scope_display = _scope_friendly(scope_raw, scope_names) if scope_raw else ""
        risk_rows.append([
            r.get("severity", "").upper(),
            r.get("title", ""),
            pname,
            scope_display,
            r.get("detail", ""),
            r.get("remediation", ""),
            r.get("powershell", ""),
            pid,
            scope_raw,
        ])
    _add_sheet(
        "Risk Findings",
        ["Severity", "Title", "Principal", "Scope", "Detail",
         "Remediation", "PowerShell", "Principal ID", "Scope (Full)"],
        risk_rows,
        [10, 35, 25, 30, 45, 45, 50, 36, 50],
    )

    # ── Tab 6: Principal Directory ──
    principal_rows = []
    for pid in sorted(principals, key=lambda p: principals[p].get("display_name", p).lower()):
        info = principals[pid]
        members = info.get("members", [])
        psub = info.get("sub_type", info.get("type", "Unknown"))
        psub_label = _SUBTYPE_LABELS.get(psub, psub)
        mi_type = info.get("mi_type", "")
        if mi_type:
            psub_label = f"{psub_label} ({mi_type})"
        principal_rows.append([
            pid,
            info.get("display_name", ""),
            psub_label,
            info.get("upn", ""),
            info.get("app_id", ""),
            len(members) if members else 0,
        ])
    _add_sheet(
        "Principal Directory",
        ["Object ID", "Display Name", "Type", "UPN", "App ID", "Group Members"],
        principal_rows,
        [36, 30, 20, 30, 36, 12],
    )

    # ── Tab 7: Group Membership ──
    group_rows = []
    for pid, info in principals.items():
        if info.get("type") != "Group" or not info.get("members"):
            continue
        group_name = info.get("display_name", pid)
        for m in info["members"]:
            group_rows.append([
                group_name,
                m.get("display_name", m.get("id", "")),
                m.get("type", ""),
                m.get("upn", ""),
            ])
            # Expand nested sub-group members
            if m.get("type") == "Group" and m.get("members"):
                for sm in m["members"]:
                    group_rows.append([
                        f"{group_name} > {m.get('display_name', m.get('id', ''))}",
                        sm.get("display_name", sm.get("id", "")),
                        sm.get("type", ""),
                        sm.get("upn", ""),
                    ])
    _add_sheet(
        "Group Membership",
        ["Group", "Member Name", "Member Type", "UPN"],
        group_rows,
        [30, 30, 15, 35],
    )

    # ── Tab 8: Scope Hierarchy ──
    scope_rows = []

    def _walk_scopes(node: dict, depth: int = 0):
        ntype = node.get("type", "")
        nname = node.get("display_name", "") or node.get("name", "")
        n_assign = len(node.get("assignments", [])) + len(node.get("eligible", []))
        n_priv = sum(1 for a in node.get("assignments", []) + node.get("eligible", []) if a.get("is_privileged"))
        scope_rows.append([
            "  " * depth + nname,
            ntype,
            node.get("id", ""),
            n_assign,
            n_priv,
        ])
        for rg in node.get("resource_groups", []):
            rg_name = rg.get("display_name", "") or rg.get("name", "")
            rg_a = len(rg.get("assignments", []))
            rg_p = sum(1 for a in rg.get("assignments", []) if a.get("is_privileged"))
            scope_rows.append([
                "  " * (depth + 1) + rg_name,
                "ResourceGroup",
                rg.get("id", ""),
                rg_a,
                rg_p,
            ])
            for res in rg.get("resources", []):
                res_name = res.get("display_name", "") or res.get("name", "")
                res_a = len(res.get("assignments", []))
                res_p = sum(1 for a in res.get("assignments", []) if a.get("is_privileged"))
                if res_a:
                    scope_rows.append([
                        "  " * (depth + 2) + res_name,
                        "Resource",
                        res.get("id", ""),
                        res_a,
                        res_p,
                    ])
        for child in node.get("children", []):
            _walk_scopes(child, depth + 1)

    _walk_scopes(tree)
    _add_sheet(
        "Scope Hierarchy",
        ["Scope Name", "Type", "Scope ID", "Assignments", "Privileged"],
        scope_rows,
        [45, 18, 60, 12, 12],
    )

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    xlsx_path = output_dir / "rbac-report.xlsx"
    wb.save(str(xlsx_path))
    log.info("[RbacReport] Excel export: %s (%d KB)", xlsx_path, xlsx_path.stat().st_size // 1024)
    return xlsx_path
