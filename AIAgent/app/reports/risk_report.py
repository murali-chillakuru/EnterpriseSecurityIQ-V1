"""
Risk Analysis Report — Interactive HTML
Full-width professional report showing security risk gap analysis results
with executive summary, category breakdown, severity distribution, top risks,
detailed findings with remediation, and methodology.
"""

from __future__ import annotations

import hashlib
import pathlib
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.reports.shared_theme import (
    get_css, get_js, esc, score_color, format_date_short, VERSION,
    SEVERITY_COLORS,
)
from app.logger import log


# ── Category metadata ────────────────────────────────────────────────────

_CATEGORY_META: dict[str, dict] = {
    "identity": {
        "icon": "&#129682;",  # 🪪
        "name": "Identity & Access",
        "color": "#0078D4",
        "description": "Dormant accounts, over-permissioned service principals, credential hygiene, MFA gaps, admin proliferation, guest risks, risky users.",
    },
    "network": {
        "icon": "&#127760;",  # 🌐
        "name": "Network Security",
        "color": "#00B7C3",
        "description": "Open management ports, public storage, web-app transport, SQL firewall exposure.",
    },
    "defender": {
        "icon": "&#128737;",  # 🛡
        "name": "Defender Posture",
        "color": "#F7630C",
        "description": "Disabled Defender plans, security recommendations, secure score gaps.",
    },
    "config": {
        "icon": "&#9881;",  # ⚙
        "name": "Configuration Drift",
        "color": "#FFB900",
        "description": "Missing diagnostics, Azure Policy violations, tag governance.",
    },
    "insider_risk": {
        "icon": "&#128373;",  # 🕵
        "name": "Insider Risk",
        "color": "#9B59B6",
        "description": "Insider risk management posture, data exfiltration signals, policy violations.",
    },
}

_RISK_LEVEL_META: dict[str, dict] = {
    "critical": {"color": "#D13438", "label": "Critical", "icon": "&#128680;"},
    "high":     {"color": "#F7630C", "label": "High",     "icon": "&#9888;&#65039;"},
    "medium":   {"color": "#FFB900", "label": "Medium",   "icon": "&#9432;"},
    "low":      {"color": "#107C10", "label": "Low",      "icon": "&#10003;"},
}

# ── SVG helpers ──────────────────────────────────────────────────────────

def _donut_svg(slices: list[tuple[str, float, str]], size: int = 160, hole: float = 0.6, center_text: str | None = None) -> str:
    total = sum(v for _, v, _ in slices) or 1
    r = size // 2 - 4
    circ = 2 * 3.14159 * r
    cx = cy = size // 2
    parts = [f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" role="img" aria-label="Donut chart">']
    offset = 0
    for label, val, color in slices:
        if val <= 0:
            continue
        pct = val / total
        dash = circ * pct
        gap = circ - dash
        parts.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" '
            f'stroke-width="{r * (1 - hole)}" stroke-dasharray="{dash:.2f} {gap:.2f}" '
            f'stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})">'
            f'<title>{esc(label)}: {int(val)} ({pct*100:.0f}%)</title></circle>'
        )
        offset += dash
    ct = center_text if center_text is not None else str(int(total))
    parts.append(
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" '
        f'font-size="22" font-weight="700" fill="var(--text)" '
        f'font-family="var(--font-mono)">{esc(ct)}</text>'
    )
    parts.append("</svg>")
    return "".join(parts)


def _ring_score_svg(score: float, size: int = 140) -> str:
    r = size // 2 - 8
    circ = 2 * 3.14159 * r
    pct = min(score, 100) / 100
    dash = circ * pct
    gap = circ - dash
    cx = cy = size // 2
    color = "#D13438" if score >= 75 else "#F7630C" if score >= 50 else "#FFB900" if score >= 25 else "#107C10"
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" class="ring" role="img" aria-label="Risk score {score:.0f} out of 100">'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="var(--ring-track)" stroke-width="10"/>'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="10" '
        f'stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-dashoffset="{circ * 0.25:.2f}" '
        f'stroke-linecap="round" style="transition:stroke-dasharray 1s ease"/>'
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" font-size="28" font-weight="700" '
        f'fill="{color}" font-family="var(--font-mono)">{score:.0f}</text>'
        f'<text x="{cx}" y="{cy + 18}" text-anchor="middle" font-size="10" fill="var(--text-muted)">/100</text>'
        f'</svg>'
    )


def _bar_chart_svg(items: list[tuple[str, int, str]], width: int = 480, bar_h: int = 22) -> str:
    if not items:
        return '<p class="empty">No data</p>'
    max_val = max(v for _, v, _ in items) or 1
    gap = 6
    h = (bar_h + gap) * len(items) + 10
    parts = [f'<svg width="{width}" height="{h}" role="img" aria-label="Bar chart">']
    for i, (label, val, color) in enumerate(items):
        y = i * (bar_h + gap) + 4
        bw = int((val / max_val) * (width - 200))
        parts.append(
            f'<text x="0" y="{y + bar_h - 5}" font-size="12" fill="var(--text-secondary)" '
            f'font-family="var(--font-primary)">{esc(label)}</text>'
        )
        parts.append(
            f'<rect x="130" y="{y}" width="{max(bw, 2)}" height="{bar_h}" rx="3" fill="{color}" opacity="0.85">'
            f'<title>{esc(label)}: {val}</title></rect>'
        )
        parts.append(
            f'<text x="{135 + bw}" y="{y + bar_h - 5}" font-size="12" fill="var(--text)" '
            f'font-family="var(--font-mono)">{val}</text>'
        )
    parts.append("</svg>")
    return "".join(parts)


# ── Report-specific CSS ─────────────────────────────────────────────────

def _risk_css() -> str:
    return """
.top-nav{position:sticky;top:0;z-index:500;display:flex;align-items:center;gap:16px;padding:8px 24px;
  background:var(--bg-elevated);border-bottom:1px solid var(--border);font-size:13px;flex-wrap:wrap}
.top-nav .brand{font-weight:700;color:var(--primary);font-size:14px;margin-right:12px}
.top-nav a{color:var(--text-secondary);text-decoration:none;padding:6px 10px;border-radius:6px;transition:all .2s;min-height:36px;display:inline-flex;align-items:center}
.top-nav a:hover{color:var(--text);background:var(--bg-card)}
.full-width-content{padding:32px 40px;max-width:1200px;margin:0 auto}
.exec-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:20px;margin:24px 0}
.exec-panel{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px}
.exec-panel h3{font-size:14px;color:var(--text-secondary);margin-bottom:12px;border:none;padding:0}
.legend{display:flex;gap:16px;flex-wrap:wrap}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--text-secondary)}
.legend-dot{width:10px;height:10px;border-radius:50%;display:inline-block}
.category-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin:16px 0}
.category-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px;text-align:center;transition:all .3s;cursor:default}
.category-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md)}
.category-icon{font-size:32px;margin-bottom:8px}
.category-name{font-size:13px;color:var(--text-secondary);margin-bottom:4px}
.category-score{font-size:28px;font-weight:700;font-family:var(--font-mono)}
.category-level{font-size:11px;text-transform:uppercase;font-weight:600;letter-spacing:.5px;margin-top:2px}
.category-findings{font-size:11px;color:var(--text-muted);margin-top:4px}
.risk-score-display{display:flex;align-items:center;gap:40px;flex-wrap:wrap;margin:20px 0}
.risk-score-info{display:flex;flex-direction:column;gap:6px}
.risk-level-badge{display:inline-block;padding:4px 12px;border-radius:6px;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.sev-bars{display:flex;flex-direction:column;gap:8px;margin:12px 0}
.sev-row{display:flex;align-items:center;gap:12px}
.sev-label{width:70px;font-size:12px;text-transform:uppercase;color:var(--text-secondary);font-weight:600}
.sev-track{flex:1;height:10px;background:var(--bar-bg);border-radius:5px;overflow:hidden}
.sev-fill{height:100%;border-radius:5px;transition:width .6s cubic-bezier(.16,1,.3,1)}
.sev-count{width:30px;text-align:right;font-family:var(--font-mono);font-size:13px}
.finding-card{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:20px;margin-bottom:12px;transition:all .2s}
.finding-card:hover{background:var(--bg-card-hover)}
.finding-card.critical{border-left:4px solid #D13438}
.finding-card.high{border-left:4px solid #F7630C}
.finding-card.medium{border-left:4px solid #FFB900}
.finding-card.low{border-left:4px solid #107C10}
.finding-title{font-size:15px;font-weight:600;margin-bottom:6px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.finding-desc{font-size:13px;color:var(--text-secondary);line-height:1.6;margin-bottom:8px}
.finding-meta{display:flex;gap:16px;flex-wrap:wrap;font-size:12px;color:var(--text-muted);margin-bottom:8px}
.finding-meta span{display:flex;align-items:center;gap:4px}
.remediation-box{margin-top:10px;padding:14px;background:var(--remediation-bg);border-left:3px solid var(--remediation-border);border-radius:6px}
.remediation-box h4{font-size:12px;color:var(--remediation-border);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}
.remediation-box .rem-desc{font-size:13px;color:#A5D6A7;margin-bottom:8px;line-height:1.5}
.remediation-box pre{font-family:var(--font-mono);font-size:12px;background:var(--code-bg);border:1px solid var(--code-border);border-radius:4px;padding:10px;overflow-x:auto;color:var(--text);margin:6px 0;white-space:pre-wrap;word-break:break-all}
.remediation-box .portal-steps{margin:6px 0 0;padding-left:20px;font-size:12px;color:var(--text-secondary)}
.remediation-box .portal-steps li{margin-bottom:3px}
.affected-details{margin-top:8px}
.affected-details summary{cursor:pointer;color:var(--primary);font-weight:500;font-size:12px}
.affected-table{width:100%;border-collapse:collapse;font-size:12px;margin-top:8px}
.affected-table th{background:var(--bg-elevated);color:var(--text-secondary);text-align:left;padding:8px 10px;font-size:11px;text-transform:uppercase;letter-spacing:.5px;border-bottom:2px solid var(--border);white-space:nowrap}
.affected-table td{padding:8px 10px;border-bottom:1px solid var(--border);color:var(--text);vertical-align:top}
.affected-table tr:hover{background:var(--bg-card-hover)}
.filter-bar{display:flex;align-items:center;gap:8px;margin-bottom:16px;flex-wrap:wrap;font-size:13px}
.filter-bar input[type="search"]{min-width:240px;padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar select{padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar button{padding:6px 14px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:12px;min-height:36px;transition:all .2s}
.filter-bar button:hover{border-color:var(--primary);color:var(--primary)}
.method-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px;margin:16px 0}
.method-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:20px}
.method-card .method-icon{font-size:28px;margin-bottom:8px}
.method-card h4{font-size:14px;margin-bottom:8px;color:var(--text)}
.method-card p{font-size:12px;color:var(--text-secondary);line-height:1.6}
.top-risk-row{display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid var(--border-light)}
.top-risk-row:last-child{border-bottom:none}
.top-risk-rank{font-size:20px;font-weight:700;font-family:var(--font-mono);color:var(--text-muted);width:30px;text-align:center}
.top-risk-info{flex:1}
.top-risk-title{font-size:13px;font-weight:600}
.top-risk-meta{font-size:11px;color:var(--text-muted);margin-top:2px}
.how-to-read{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;margin-bottom:24px}
.how-to-read h4{margin-bottom:8px}
.how-to-read p{font-size:13px;color:var(--text-secondary);margin-bottom:6px}
.zoom-controls{display:flex;align-items:center;gap:4px;margin-left:auto}
.zoom-controls button{padding:4px 10px;border:1px solid var(--border);border-radius:4px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:14px;min-height:32px;transition:all .2s}
.zoom-controls button:hover{border-color:var(--primary);color:var(--primary)}
#zoom-label{font-size:12px;font-family:var(--font-mono);width:40px;text-align:center}
.export-bar{display:flex;gap:4px}
/* ── Hover tooltips ── */
[data-tip]{cursor:help}
th [data-tip],h2 [data-tip],h3 [data-tip],.legend [data-tip]{border-bottom:1px dotted var(--text-muted)}
.stat-card[data-tip],.badge[data-tip],.risk-level-badge[data-tip]{border-bottom:none;cursor:help}
.exec-panel-title{cursor:help;display:inline-block}
.score-ring-wrap{cursor:help;display:inline-block}
.nav-tip{cursor:help}
#ciq-tooltip{position:fixed;z-index:99999;pointer-events:none;opacity:0;transition:opacity .18s ease;max-width:380px;min-width:200px;padding:14px 18px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  color:var(--text);
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  border-radius:10px;
  font-size:12.5px;line-height:1.6;font-weight:400;text-transform:none;letter-spacing:normal;white-space:normal;
  box-shadow:0 2px 6px rgba(0,0,0,.18),0 8px 24px rgba(0,0,0,.32),0 0 0 1px rgba(255,255,255,.06) inset;
}
#ciq-tooltip.visible{opacity:1}
#ciq-tooltip::before{content:'';position:absolute;width:12px;height:12px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  transform:rotate(45deg);z-index:-1}
#ciq-tooltip.arrow-bottom::before{bottom:-7px;left:var(--arrow-x,24px);border-top:none;border-left:none}
#ciq-tooltip.arrow-top::before{top:-7px;left:var(--arrow-x,24px);border-bottom:none;border-right:none}
#ciq-tooltip .t-sep{display:block;border-top:1px solid rgba(255,255,255,.15);margin:8px 0 4px;padding-top:6px;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--primary)}
@media(max-width:768px){.full-width-content{padding:16px}.exec-grid{grid-template-columns:1fr}.stat-grid{grid-template-columns:repeat(2,1fr)}.top-nav{padding:8px 12px}}
@media print{.top-nav,.back-to-top,.filter-bar,.zoom-controls,.export-bar{display:none!important}.full-width-content{padding:16px;max-width:100%}body{background:#fff;color:#000;font-size:12px}.finding-card,.category-card,.exec-panel,.stat-card{border:1px solid #ccc;background:#fff}.badge{border:1px solid #333;print-color-adjust:exact;-webkit-print-color-adjust:exact}.remediation-box{background:#f1faf1;border-left-color:#107C10}}
"""


# ── Report-specific JS ──────────────────────────────────────────────────

def _risk_js() -> str:
    return """
// Zoom
var zoomLevel=100;
function zoomIn(){zoomLevel=Math.min(zoomLevel+10,150);applyZoom()}
function zoomOut(){zoomLevel=Math.max(zoomLevel-10,70);applyZoom()}
function zoomReset(){zoomLevel=100;applyZoom()}
function applyZoom(){document.querySelector('.full-width-content').style.zoom=(zoomLevel/100);document.getElementById('zoom-label').textContent=zoomLevel+'%'}

// Filter findings
function filterFindings(){
  var q=(document.getElementById('finding-filter').value||'').toLowerCase();
  var sev=(document.getElementById('filter-severity').value||'').toLowerCase();
  var cat=(document.getElementById('filter-category').value||'').toLowerCase();
  var cards=document.querySelectorAll('.finding-card[data-category]');
  var shown=0;
  cards.forEach(function(card){
    var text=card.textContent.toLowerCase();
    var cSev=card.getAttribute('data-severity')||'';
    var cCat=card.getAttribute('data-category')||'';
    var match=true;
    if(q&&text.indexOf(q)<0)match=false;
    if(sev&&cSev!==sev)match=false;
    if(cat&&cCat!==cat)match=false;
    card.style.display=match?'':'none';
    if(match)shown++;
  });
  var live=document.getElementById('findings-live');
  if(live)live.textContent=shown+' findings shown';
}

// ── Tooltip engine (viewport-aware positioning) ──
(function(){
  var tip=document.getElementById('ciq-tooltip');
  if(!tip)return;
  var GAP=10,MARGIN=12;
  function show(ev){
    var tgt=ev.target.closest('[data-tip]');
    if(!tgt)return;
    var text=tgt.getAttribute('data-tip');
    if(!text)return;
    var d=document.createElement('span');d.textContent=text;var safe=d.innerHTML;
    safe=safe.replace(/\\n+YOUR TENANT:/g,'<span class="t-sep">&#x1F4CA; Your Tenant</span>');
    safe=safe.replace(/\\n/g,'<br>');
    tip.innerHTML=safe;
    tip.classList.add('visible');
    tip.setAttribute('aria-hidden','false');
    requestAnimationFrame(function(){
      var r=tgt.getBoundingClientRect();
      var tw=tip.offsetWidth,th=tip.offsetHeight;
      var vw=window.innerWidth,vh=window.innerHeight;
      var above=r.top-GAP-th;
      var below=r.bottom+GAP;
      var top,arrow;
      if(above>=MARGIN){top=above;arrow='arrow-bottom';}
      else if(below+th<=vh-MARGIN){top=below;arrow='arrow-top';}
      else{top=Math.max(MARGIN,vh-th-MARGIN);arrow='';}
      var left=r.left+r.width/2-tw/2;
      left=Math.max(MARGIN,Math.min(left,vw-tw-MARGIN));
      var arrowX=r.left+r.width/2-left;
      arrowX=Math.max(16,Math.min(arrowX,tw-16));
      tip.style.top=top+'px';
      tip.style.left=left+'px';
      tip.style.setProperty('--arrow-x',arrowX+'px');
      tip.className='visible'+(arrow?' '+arrow:'');
    });
  }
  function hide(){
    tip.classList.remove('visible');
    tip.setAttribute('aria-hidden','true');
    tip.className='';
  }
  document.addEventListener('mouseenter',show,true);
  document.addEventListener('mouseleave',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
  document.addEventListener('focusin',show,true);
  document.addEventListener('focusout',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
})();
"""


# ── Rendering helpers ────────────────────────────────────────────────────

_SEVERITY_TIPS: dict[str, str] = {
    "critical": "Critical severity \u2014 active breach risk or critical data exposure. Requires immediate remediation.",
    "high": "High severity \u2014 significant security risk that could be exploited. Address urgently after critical items.",
    "medium": "Medium severity \u2014 moderate risk. Plan remediation within your next sprint cycle.",
    "low": "Low severity \u2014 minor improvement with limited direct impact. Enhances overall security hygiene.",
    "informational": "Informational \u2014 observation with no direct risk. Consider as a best-practice recommendation.",
}


def _severity_badge(sev: str) -> str:
    color = SEVERITY_COLORS.get(sev.lower(), "#A8A6A3")
    tip = _SEVERITY_TIPS.get(sev.lower(), "")
    return f'<span class="badge" style="background:{color}" data-tip="{esc(tip)}">{esc(sev.upper())}</span>'


def _render_finding(f: dict) -> str:
    sev = f.get("Severity", "medium").lower()
    cat = f.get("Category", "unknown")
    subcat = f.get("Subcategory", "")
    title = f.get("Title", "Untitled")
    desc = f.get("Description", "")
    affected = f.get("AffectedCount", 0)
    affected_resources = f.get("AffectedResources", [])
    remediation = f.get("Remediation", {})
    detected = f.get("DetectedAt", "")

    # Remediation block
    rem_html = ""
    if remediation:
        rem_parts = []
        rem_desc = remediation.get("Description", "")
        if rem_desc:
            rem_parts.append(f'<div class="rem-desc">{esc(rem_desc)}</div>')
        cli = remediation.get("AzureCLI", "")
        if cli:
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">Azure CLI:</strong><pre>{esc(cli)}</pre></div>')
        ps = remediation.get("PowerShell", "")
        if ps:
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">PowerShell:</strong><pre>{esc(ps)}</pre></div>')
        steps = remediation.get("PortalSteps", [])
        if steps:
            step_items = "".join(f"<li>{esc(s)}</li>" for s in steps)
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">Portal Steps:</strong><ol class="portal-steps">{step_items}</ol></div>')
        if rem_parts:
            rem_html = f'<div class="remediation-box"><h4>&#128736; Remediation</h4>{"".join(rem_parts)}</div>'

    # Affected resources — full table, no truncation
    affected_html = ""
    if affected_resources and len(affected_resources) > 0:
        rem_anchor = f'rem-{f.get("RiskFindingId", "")}'
        table_rows = ""
        for ar in affected_resources:
            if isinstance(ar, dict):
                # Resource Name — try many possible name keys
                r_name = (ar.get("Name") or ar.get("PrincipalName") or ar.get("UserName")
                          or ar.get("NsgName") or ar.get("DisplayName") or ar.get("MemberName")
                          or ar.get("Type") or "—")
                # Resource ID
                r_id = (ar.get("ResourceId") or ar.get("PrincipalId") or ar.get("UserId")
                        or ar.get("AppId") or ar.get("id") or "—")
                # Reason for Risk — contextual detail
                reason_parts = []
                for rk in ("ExposedPort", "Status", "RiskLevel", "RiskState",
                           "RoleName", "Scope", "CredentialExpiry", "FirewallRule"):
                    rv = ar.get(rk)
                    if rv:
                        reason_parts.append(f"{rk}: {rv}")
                if ar.get("Issues"):
                    reason_parts.append("; ".join(ar["Issues"]) if isinstance(ar["Issues"], list) else str(ar["Issues"]))
                # Fallback: show Type or any remaining detail
                if not reason_parts:
                    for rk2 in ("Percent", "StaleUsers", "Registered", "Total",
                                "GuestCount", "MfaRegistrationPercent"):
                        rv2 = ar.get(rk2)
                        if rv2 is not None:
                            reason_parts.append(f"{rk2}: {rv2}")
                reason = "; ".join(reason_parts) if reason_parts else ar.get("Type", "—")
                table_rows += (
                    f'<tr>'
                    f'<td>{esc(str(r_name))}</td>'
                    f'<td style="font-family:var(--font-mono);font-size:11px;word-break:break-all">{esc(str(r_id))}</td>'
                    f'<td>{esc(str(reason))}</td>'
                    f'<td>{_severity_badge(sev)}</td>'
                    f'<td><a href="#{rem_anchor}" style="color:var(--primary);font-size:12px">See below &#8595;</a></td>'
                    f'</tr>'
                )
            else:
                table_rows += f'<tr><td colspan="5">{esc(str(ar))}</td></tr>'
        affected_html = (
            f'<details class="affected-details" open><summary>{affected} affected resource(s)</summary>'
            f'<div style="overflow-x:auto"><table class="affected-table">'
            f'<thead><tr><th>Resource Name</th><th>Resource ID</th><th>Reason for Risk</th><th>Risk Level</th><th>Remediation</th></tr></thead>'
            f'<tbody>{table_rows}</tbody></table></div></details>'
        )

    cat_meta = _CATEGORY_META.get(cat, {})
    cat_name = cat_meta.get("name", cat.replace("_", " ").title())

    cat_desc = _CATEGORY_META.get(cat, {}).get("description", "")

    return (
        f'<div class="finding-card {sev}" data-severity="{sev}" data-category="{esc(cat)}" '
        f'data-subcategory="{esc(subcat)}" data-affected="{affected}">'
        f'<div class="finding-title">{_severity_badge(sev)} {esc(title)}</div>'
        f'<div class="finding-meta">'
        f'<span data-tip="Assessment category: {esc(cat_name)}.\n{esc(cat_desc)}">&#128193; {esc(cat_name)}</span>'
        f'<span data-tip="Security subcategory within {esc(cat_name)} that this finding relates to.">&#128196; {esc(subcat.replace("_", " ").title())}</span>'
        f'<span data-tip="Number of Azure resources, identities, or configurations affected by this finding.">&#128202; {affected} affected</span>'
        + (f'<span data-tip="Date and time when this risk was first detected.">&#128337; {esc(detected[:19])}</span>' if detected else "")
        + f'</div>'
        f'<div class="finding-desc">{esc(desc)}</div>'
        f'{affected_html}'
        f'<div id="rem-{f.get("RiskFindingId", "")}">{rem_html}</div>'
        f'</div>'
    )

# ── Main generator ───────────────────────────────────────────────────────

def generate_risk_report(results: dict, output_dir: str | pathlib.Path) -> pathlib.Path:
    """Generate the Risk Analysis HTML report.

    Args:
        results: Result dict from run_risk_analysis (keys: RiskScores, Findings,
                 FindingCount, SubscriptionCount, EvidenceSource, AnalyzedAt, etc.).
        output_dir: Directory to write the report into.

    Returns:
        Path to the generated HTML file.
    """
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "risk-analysis.html"

    ts = format_date_short()
    scores = results.get("RiskScores", {})
    findings = results.get("Findings", [])
    finding_count = results.get("FindingCount", len(findings))
    sub_count = results.get("SubscriptionCount", 0)
    evidence_src = results.get("EvidenceSource", "unknown")
    analyzed_at = results.get("AnalyzedAt", "")
    tenant_id = results.get("TenantId", "")
    tenant_display = results.get("TenantDisplayName", "")
    evidence_record_count = results.get("EvidenceRecordCount", 0)
    report_id = f"CIQ-RA-{ts.replace('-', '').replace(':', '').replace(' ', '-')}"

    overall_score = scores.get("OverallRiskScore", 0)
    overall_level = scores.get("OverallRiskLevel", "low")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})

    n_critical = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_medium = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)
    total_sev = n_critical + n_high + n_medium + n_low + n_info

    # Level badge
    level_meta = _RISK_LEVEL_META.get(overall_level, _RISK_LEVEL_META["medium"])
    level_color = level_meta["color"]
    level_label = level_meta["label"]

    # Score ring
    score_ring = _ring_score_svg(overall_score, size=160)

    # Severity donut
    sev_donut = _donut_svg([
        ("Critical", n_critical, "#D13438"),
        ("High", n_high, "#F7630C"),
        ("Medium", n_medium, "#FFB900"),
        ("Low", n_low, "#107C10"),
        ("Info", n_info, "#A8A6A3"),
    ], size=140)

    # Severity bars
    max_sev = max(n_critical, n_high, n_medium, n_low, n_info, 1)
    sev_bars_html = ""
    _sev_bar_tips = {
        "Critical": "Critical findings requiring immediate remediation to prevent active breach or data exposure.",
        "High": "High severity findings that should be addressed urgently after critical items.",
        "Medium": "Medium severity findings indicating moderate risk. Plan remediation within your next sprint.",
        "Low": "Low severity findings \u2014 minor configuration improvements with limited direct impact.",
        "Info": "Informational observations with no direct risk. Consider as best-practice recommendations.",
    }
    for sev_name, sev_count, sev_color in [
        ("Critical", n_critical, "#D13438"),
        ("High", n_high, "#F7630C"),
        ("Medium", n_medium, "#FFB900"),
        ("Low", n_low, "#107C10"),
        ("Info", n_info, "#A8A6A3"),
    ]:
        pct = (sev_count / max_sev) * 100 if max_sev > 0 else 0
        bar_tip = f"{_sev_bar_tips.get(sev_name, '')}\nYOUR TENANT: {sev_count} {sev_name.lower()} finding(s) ({pct:.0f}% of max)."
        sev_bars_html += (
            f'<div class="sev-row" data-tip="{esc(bar_tip)}">'
            f'<span class="sev-label">{sev_name}</span>'
            f'<div class="sev-track"><div class="sev-fill" style="width:{pct:.0f}%;background:{sev_color}"></div></div>'
            f'<span class="sev-count">{sev_count}</span>'
            f'</div>'
        )

    # Category cards
    cat_cards_html = ""
    for cat_key in ["identity", "network", "defender", "config", "insider_risk"]:
        meta = _CATEGORY_META[cat_key]
        cs = cat_scores.get(cat_key, {"Score": 0, "Level": "low", "FindingCount": 0})
        c_score = cs.get("Score", 0)
        c_level = cs.get("Level", "low")
        c_count = cs.get("FindingCount", 0)
        c_color = _RISK_LEVEL_META.get(c_level, _RISK_LEVEL_META["low"])["color"]
        cat_cards_html += (
            f'<div class="category-card" data-tip="{esc(meta["description"])} Score: {c_score:.0f}/100 ({c_level.upper()}).">'  
            f'<div class="category-icon">{meta["icon"]}</div>'
            f'<div class="category-name">{esc(meta["name"])}</div>'
            f'<div class="category-score" style="color:{c_color}">{c_score:.0f}</div>'
            f'<div class="category-level" style="color:{c_color}">{esc(c_level.upper())}</div>'
            f'<div class="category-findings">{c_count} finding{"s" if c_count != 1 else ""}</div>'
            f'</div>'
        )

    # Top risks HTML — removed (All Findings covers them all)

    # All findings HTML, grouped by category
    findings_by_cat: dict[str, list[dict]] = {}
    for f in findings:
        findings_by_cat.setdefault(f.get("Category", "unknown"), []).append(f)

    all_findings_html = ""
    # Category select options
    cat_options = ""
    for cat_key, cat_findings in sorted(findings_by_cat.items()):
        meta = _CATEGORY_META.get(cat_key, {"name": cat_key.title(), "icon": "&#128196;"})
        cat_options += f'<option value="{esc(cat_key)}">{esc(meta["name"])} ({len(cat_findings)})</option>'

        # Sort by severity
        sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
        sorted_findings = sorted(cat_findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))

        cat_h3_tip = f"{esc(meta.get('description', ''))}\nYOUR TENANT: {len(cat_findings)} finding(s) in this category."
        all_findings_html += f'<h3 id="cat-{esc(cat_key)}" style="margin-top:24px" data-tip="{cat_h3_tip}">{meta["icon"]} {esc(meta["name"])} ({len(cat_findings)} findings)</h3>'
        for f in sorted_findings:
            all_findings_html += _render_finding(f)

    # Assemble HTML
    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Security Risk Gap Analysis — EnterpriseSecurityIQ</title>
<style>{get_css()}{_risk_css()}</style>
</head>
<body>
<a href="#main" class="skip-nav">Skip to content</a>

<!-- Top Navigation -->
<nav class="top-nav" aria-label="Report sections">
  <span class="brand" aria-hidden="true">&#128274; EnterpriseSecurityIQ Risk Analysis</span>
  <a href="#doc-control" class="nav-tip" data-tip="Jump to document control, audit attestation, and report metadata">Document Control</a>
  <a href="#summary" class="nav-tip" data-tip="Risk score, KPI cards, severity breakdown, and category analysis">Executive Summary</a>
  <a href="#categories" class="nav-tip" data-tip="Per-category risk scores across identity, network, defender, and configuration domains">Categories</a>
  <a href="#findings" class="nav-tip" data-tip="Complete table of all findings filterable by severity, category, and status">All Findings</a>
  <div class="zoom-controls" aria-label="Page zoom">
    <button onclick="zoomOut()" aria-label="Zoom out" data-tip="Decrease page zoom by 10% (min 70%).">&minus;</button>
    <span id="zoom-label">100%</span>
    <button onclick="zoomIn()" aria-label="Zoom in" data-tip="Increase page zoom by 10% (max 150%).">&plus;</button>
    <button onclick="zoomReset()" aria-label="Reset zoom" data-tip="Reset page zoom to 100%." style="font-size:11px">Reset</button>
  </div>
  <button class="theme-btn" onclick="toggleTheme()" style="margin:0;padding:6px 14px"
          aria-label="Toggle dark and light theme"
          data-tip="Switch between dark and light theme. Your preference is saved in browser storage.">
    Switch to Light
  </button>
</nav>

<main id="main" class="full-width-content">

<!-- ── Document Control ── -->
<section id="doc-control" class="section">
  <h1 class="page-title" data-tip="EnterpriseSecurityIQ automated security risk gap analysis. This report evaluates identity, network, defender, and configuration risks across your Azure tenant.">EnterpriseSecurityIQ &mdash; Security Risk Gap Analysis Report</h1>
  <table class="doc-control-table">
    <tr><th data-tip="Unique identifier for this report instance. Use this ID when referencing findings in audit trails or remediation tickets.">Report Identifier</th><td>{esc(report_id)}</td></tr>
    <tr><th data-tip="The name of the compliance assessment that generated this report.">Assessment Name</th><td>EnterpriseSecurityIQ Security Risk Gap Analysis</td></tr>
    <tr><th data-tip="Date and time when this report was generated.">Date Generated</th><td>{esc(ts)}</td></tr>
    <tr><th data-tip="Entra ID (Azure AD) tenant identifier for the assessed environment.">Tenant ID</th><td><code>{esc(tenant_id) if tenant_id else 'N/A'}</code></td></tr>
    <tr><th data-tip="Display name of the Entra ID tenant.">Tenant Name</th><td>{esc(tenant_display) if tenant_display else 'Unknown'}</td></tr>
    <tr><th data-tip="The security domains covered in this risk assessment.">Assessment Domains</th><td>Identity &amp; Access, Network Security, Microsoft Defender, Configuration Drift, Insider Risk</td></tr>
    <tr><th data-tip="Data classification level of this report. Handle according to your organization\u2019s information protection policy.">Classification</th><td>CONFIDENTIAL &mdash; Authorized Recipients Only</td></tr>
    <tr><th data-tip="The EnterpriseSecurityIQ tool version and agent that produced this report.">Tool</th><td>EnterpriseSecurityIQ AI Agent v{VERSION}</td></tr>
    <tr><th data-tip="How evidence was collected. Read-only API calls ensure no changes were made to your tenant.">Collection Method</th><td>Azure Resource Manager API + Microsoft Graph API (Read-Only)</td></tr>
  </table>
  <div class="conf-notice" data-tip="This report contains sensitive security posture data. Handle according to your organization\u2019s information protection policy. Do not share outside authorized personnel.">
    <strong>CONFIDENTIALITY NOTICE:</strong> This document contains sensitive security and compliance
    information about the assessed environment. Distribution is restricted to authorized personnel only.
  </div>
  <h3 data-tip="Audit attestation confirms the scope, integrity, and methodology of this assessment. Use this section to validate the report for compliance audits.">Audit Attestation</h3>
  <table class="doc-control-table">
    <tr><th data-tip="Scope of the assessment \u2014 how many subscriptions and which security domains were evaluated.">Assessment Scope</th><td>Security risk gap analysis across {sub_count} subscription(s) covering identity, network, defender, and configuration domains</td></tr>
    <tr><th data-tip="Confirms data collection was non-invasive. All API calls are read-only with no modifications to your environment.">Data Integrity</th><td>All evidence collected via read-only API calls; no tenant modifications were made</td></tr>
    <tr><th data-tip="Total number of evidence records gathered from Azure APIs and evaluated against security controls.">Evidence Records</th><td>{evidence_record_count:,} records collected and evaluated</td></tr>
    <tr><th data-tip="SHA-256 hash of the complete report HTML. Use this to verify the report has not been tampered with.">Report Hash (SHA-256)</th><td><code id="report-hash">Computed at render</code></td></tr>
    <tr><th data-tip="When the data was collected. This is a point-in-time snapshot \u2014 your environment may have changed since.">Assessment Period</th><td>{esc(analyzed_at[:19]) if analyzed_at else esc(ts)} (point-in-time snapshot)</td></tr>
  </table>
</section>

<!-- ── Executive Summary ── -->
<section id="summary" class="section" aria-labelledby="summary-heading">
  <h2 id="summary-heading" class="page-title" data-tip="High-level overview of your tenant\u2019s security risk posture. Includes composite risk score, KPI cards, severity breakdown, and category analysis.">Executive Summary</h2>
  <p style="color:var(--text-secondary);font-size:14px;line-height:1.6;max-width:960px;margin:8px 0 20px">
    This report presents a comprehensive security risk assessment across your Azure tenant,
    analyzing attack surfaces in four key domains: identity &amp; access, network security,
    Microsoft Defender posture, and configuration drift. Each finding includes severity,
    affected resources, and actionable remediation steps with Azure CLI, PowerShell, and
    portal instructions.
  </p>
  <div class="meta-bar">
    <span>Generated: {esc(ts)}</span>
    <span>Analyzed: {esc(analyzed_at[:19]) if analyzed_at else "N/A"}</span>
    {f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_display)} ({esc(tenant_id[:8])}\u2026)</span>' if tenant_id and tenant_display else f'<span class="tip" data-tip="Entra ID tenant identifier">Tenant: {esc(tenant_id)}</span>' if tenant_id else ''}
    <span>Evidence: {esc(evidence_src.replace("_", " ").title())}</span>
    <span>EnterpriseSecurityIQ v{VERSION}</span>
  </div>

  <!-- KPI cards -->
  <div class="stat-grid" style="grid-template-columns:repeat(auto-fit,minmax(140px,1fr))">
    <div class="stat-card" data-tip="Composite risk score (0\u2013100). Higher = more risk. Severity-weighted across identity, network, defender, and configuration domains.\nThresholds: 0\u201324 = Low (green), 25\u201349 = Medium (yellow), 50\u201374 = High (orange), 75\u2013100 = Critical (red).\nYOUR TENANT: Score {overall_score:.0f}/100 ({level_label}), {finding_count} total findings."><div class="stat-value" style="color:{level_color}">{overall_score:.0f}</div><div class="stat-label">Risk Score /100</div></div>
    <div class="stat-card" data-tip="Total security risk findings detected across identity, network, defender, and configuration domains.\nEach finding maps to a specific gap with severity, affected resources, and remediation.\nYOUR TENANT: {finding_count} findings \u2014 {n_critical} critical, {n_high} high, {n_medium} medium, {n_low} low."><div class="stat-value">{finding_count}</div><div class="stat-label">Total Findings</div></div>
    <div class="stat-card" data-tip="Critical findings requiring immediate remediation to prevent active security breach or data exposure.\nThese carry the highest severity weight in the composite score.\nYOUR TENANT: {n_critical} critical finding(s) detected."><div class="stat-value" style="color:#D13438">{n_critical}</div><div class="stat-label">Critical</div></div>
    <div class="stat-card" data-tip="High severity findings that should be addressed urgently after critical items are resolved.\nSignificant security risk that could be exploited.\nYOUR TENANT: {n_high} high-severity finding(s) detected."><div class="stat-value" style="color:#F7630C">{n_high}</div><div class="stat-label">High</div></div>
    <div class="stat-card" data-tip="Medium severity findings indicating moderate risk.\nReview and plan remediation within your next sprint cycle.\nYOUR TENANT: {n_medium} medium finding(s) detected."><div class="stat-value" style="color:#FFB900">{n_medium}</div><div class="stat-label">Medium</div></div>
    <div class="stat-card" data-tip="Low severity findings \u2014 informational or minor configuration improvements with limited direct impact.\nImproves overall security hygiene.\nYOUR TENANT: {n_low} low-severity finding(s) detected."><div class="stat-value" style="color:#107C10">{n_low}</div><div class="stat-label">Low</div></div>
    <div class="stat-card" data-tip="Number of Azure subscriptions scanned in this risk assessment.\nEach subscription is evaluated for identity, network, defender, and configuration risks.\nYOUR TENANT: {sub_count} subscription(s) assessed."><div class="stat-value">{sub_count}</div><div class="stat-label">Subscriptions</div></div>
    <div class="stat-card" data-tip="Number of risk assessment domains evaluated.\nDomains: Identity &amp; Access, Network Security, Microsoft Defender, Configuration Drift, Insider Risk.\nYOUR TENANT: {len(cat_scores)} categories assessed."><div class="stat-value">{len(cat_scores)}</div><div class="stat-label">Categories</div></div>
  </div>

  <!-- Score ring + severity -->
  <div class="exec-grid">
    <div class="exec-panel" data-tip="Visual risk score ring (0\u2013100). 0 = no risk, 100 = maximum risk.\nThresholds: Critical \u226575, High \u226550, Medium \u226525, Low <25. Target: below 25.\nYOUR TENANT: Score {overall_score:.0f}/100 ({level_label}), {finding_count} findings across {len(cat_scores)} categories.">
      <h3>Overall Risk Score</h3>
      <div class="risk-score-display">
        <div style="text-align:center">{score_ring}</div>
        <div class="risk-score-info">
          <div><span class="risk-level-badge" style="background:{level_color};color:#fff" data-tip="Risk level based on composite score.\nCritical (\u226575): Immediate action. High (50\u201374): Address this sprint. Medium (25\u201349): Plan remediation. Low (<25): Healthy posture.\nYOUR TENANT: {level_label} ({overall_score:.0f}/100).">{level_label}</span></div>
          <div style="font-size:13px;color:var(--text-secondary);margin-top:8px">
            Higher score = higher risk.<br>
            Target: below 25 (Low).
          </div>
        </div>
      </div>
    </div>
    <div class="exec-panel" data-tip="Distribution of findings across severity levels. Larger slices indicate more findings at that severity.\nYOUR TENANT: {n_critical} critical, {n_high} high, {n_medium} medium, {n_low} low, {n_info} informational ({total_sev} total).">
      <h3>Severity Distribution</h3>
      <div style="text-align:center">{sev_donut}</div>
      <div class="legend" style="justify-content:center;margin-top:16px">
        <span class="legend-item" data-tip="Findings requiring immediate action \u2014 active breach risk or critical data exposure.\nYOUR TENANT: {n_critical} critical finding(s)."><span class="legend-dot" style="background:#D13438"></span> Critical ({n_critical})</span>
        <span class="legend-item" data-tip="Significant security risks. Address urgently after critical items.\nYOUR TENANT: {n_high} high-severity finding(s)."><span class="legend-dot" style="background:#F7630C"></span> High ({n_high})</span>
        <span class="legend-item" data-tip="Moderate risks that should be scheduled for remediation.\nYOUR TENANT: {n_medium} medium finding(s)."><span class="legend-dot" style="background:#FFB900"></span> Medium ({n_medium})</span>
        <span class="legend-item" data-tip="Minor improvements with limited direct impact. Low urgency.\nYOUR TENANT: {n_low} low finding(s)."><span class="legend-dot" style="background:#107C10"></span> Low ({n_low})</span>
        <span class="legend-item" data-tip="Informational observations with no direct risk. Best-practice recommendations.\nYOUR TENANT: {n_info} informational finding(s)."><span class="legend-dot" style="background:#A8A6A3"></span> Info ({n_info})</span>
      </div>
    </div>
    <div class="exec-panel" style="grid-column:span 2" data-tip="Horizontal bar visualization showing the count of findings at each severity level for quick comparison.\nYOUR TENANT: {n_critical} critical, {n_high} high, {n_medium} medium, {n_low} low, {n_info} informational ({total_sev} total).">
      <h3>Severity Breakdown</h3>
      <div class="sev-bars">
        {sev_bars_html}
      </div>
    </div>
  </div>
</section>

<!-- ── Category Breakdown ── -->
<section id="categories" class="section" aria-labelledby="categories-heading">
  <h2 id="categories-heading" data-tip="Per-category risk scores across identity, network, defender, and configuration domains. Higher score = more risk.">&#128202; Category Breakdown</h2>
  <div class="how-to-read" data-tip="Guidance on interpreting the category breakdown scores and navigating to detailed findings below.">
    <h4>How to read this section</h4>
    <p>Each category scores 0\u2013100 based on severity-weighted findings. Higher = more risk.
    Click any category name in the findings section below to jump to its detailed findings.</p>
  </div>
  <div class="category-grid">
    {cat_cards_html}
  </div>
</section>

<!-- ── All Findings ── -->
<section id="findings" class="section" aria-labelledby="findings-heading">
  <h2 id="findings-heading" data-tip="Complete list of all detected findings with severity, category, remediation guidance, and affected resources.">&#128270; All Findings ({finding_count})</h2>
  <div class="filter-bar" role="search" aria-label="Filter findings">
    <label for="finding-filter">Search:</label>
    <input id="finding-filter" type="search" placeholder="Search findings…"
           oninput="filterFindings()" aria-label="Filter findings by keyword"
           aria-describedby="findings-live" autocomplete="off"
           data-tip="Type to filter findings by title, description, category, or any keyword.">
    <label for="filter-severity">Severity:</label>
    <select id="filter-severity" onchange="filterFindings()" aria-label="Filter by severity"
            data-tip="Filter the findings list to show only a specific severity level.">
      <option value="">All</option>
      <option value="critical">Critical</option>
      <option value="high">High</option>
      <option value="medium">Medium</option>
      <option value="low">Low</option>
    </select>
    <label for="filter-category">Category:</label>
    <select id="filter-category" onchange="filterFindings()" aria-label="Filter by category"
            data-tip="Filter findings by risk assessment category (identity, network, defender, configuration).">
      <option value="">All</option>
      {cat_options}
    </select>
  </div>
  <div id="findings-live" class="sr-only" aria-live="polite" aria-atomic="true"></div>
  {all_findings_html if all_findings_html else '<p class="empty">No findings — your environment has no detected risks. Well done!</p>'}
</section>

</main>

<button class="back-to-top" aria-label="Back to top">&#8593;</button>

<div id="ciq-tooltip" role="tooltip" aria-hidden="true"></div>
<script>{get_js()}</script>
<script>{_risk_js()}</script>
</body>
</html>"""

    report_hash = hashlib.sha256(html.encode("utf-8")).hexdigest()
    html = html.replace("Computed at render", report_hash)
    out_path.write_text(html, encoding="utf-8")
    log.info("[RiskReport] Written to %s (%d KB)", out_path, len(html) // 1024)
    return out_path


# ── Excel styling ────────────────────────────────────────────────────────
_XL_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_XL_HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
_XL_BODY_FONT = Font(name="Segoe UI", size=10)
_XL_THIN_BORDER = Border(
    left=Side(style="thin", color="EDEBE9"),
    right=Side(style="thin", color="EDEBE9"),
    top=Side(style="thin", color="EDEBE9"),
    bottom=Side(style="thin", color="EDEBE9"),
)
_XL_SEV_FILLS = {
    "critical": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "high": PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),
    "medium": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "low": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
}


def _xl_apply_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _XL_HEADER_FONT
        cell.fill = _XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _XL_THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def _xl_write_row(ws, row_num: int, values: list, col_fills: dict[int, PatternFill] | None = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = _XL_BODY_FONT
        cell.border = _XL_THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_fills and col in col_fills:
            cell.fill = col_fills[col]


def _xl_auto_width(ws, headers: list[str], max_width: int = 55):
    for col, h in enumerate(headers, 1):
        best = min(len(h) + 4, max_width)
        for row in range(2, min(ws.max_row + 1, 52)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = best


# ── Excel export ─────────────────────────────────────────────────────────

def generate_risk_excel(results: dict, output_dir: str) -> pathlib.Path:
    """Generate a comprehensive multi-tab Risk Analysis Excel workbook."""

    out_dir = pathlib.Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "risk-analysis.xlsx"

    scores = results.get("RiskScores", {})
    findings = results.get("Findings", [])
    overall = scores.get("Overall", 0)
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}

    level = (
        "Critical" if overall >= 75 else
        "High" if overall >= 50 else
        "Medium" if overall >= 25 else
        "Low"
    )

    # Count severities
    sev_counts: dict[str, int] = {}
    for f in findings:
        s = f.get("Severity", "medium").lower()
        sev_counts[s] = sev_counts.get(s, 0) + 1

    wb = Workbook()

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    label_font = Font(name="Segoe UI", bold=True, size=11)
    value_font = Font(name="Segoe UI", size=11)

    summary_rows = [
        ("Report", "Security Risk Gap Analysis"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("EnterpriseSecurityIQ Version", VERSION),
        ("Evidence Source", results.get("EvidenceSource", "N/A")),
        ("Subscription Count", results.get("SubscriptionCount", "N/A")),
        ("", ""),
        ("Overall Risk Score", f"{overall:.1f} / 100"),
        ("Risk Level", level),
        ("Total Findings", len(findings)),
        ("", ""),
        ("Severity Distribution", ""),
        ("  Critical", sev_counts.get("critical", 0)),
        ("  High", sev_counts.get("high", 0)),
        ("  Medium", sev_counts.get("medium", 0)),
        ("  Low", sev_counts.get("low", 0)),
        ("  Informational", sev_counts.get("informational", 0)),
        ("", ""),
        ("Category Scores", ""),
    ]
    for cat_key in ["identity", "network", "defender", "config"]:
        meta = _CATEGORY_META[cat_key]
        cat_score = scores.get("CategoryScores", {}).get(cat_key, {})
        summary_rows.append((
            f"  {meta['name']}",
            f"{cat_score.get('Score', 0):.1f} / 100 ({cat_score.get('Level', 'N/A')})"
        ))

    # Top 5 findings
    top_5 = sorted(findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))[:5]
    if top_5:
        summary_rows.append(("", ""))
        summary_rows.append(("Top 5 Priority Findings", ""))
        for i, f in enumerate(top_5, 1):
            summary_rows.append((
                f"  {i}. [{f.get('Severity', '').upper()}]",
                f.get("Title", "")
            ))

    for r, (label, value) in enumerate(summary_rows, 1):
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_l.font = label_font
        cell_l.border = _XL_THIN_BORDER
        cell_v = ws.cell(row=r, column=2, value=value)
        cell_v.font = value_font
        cell_v.border = _XL_THIN_BORDER
        cell_v.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    # ── Sheet 2: All Findings ──
    ws2 = wb.create_sheet("All Findings")
    headers2 = [
        "Severity", "Category", "Subcategory", "Title",
        "Affected Resources", "Remediation", "Resources",
    ]
    _xl_apply_header(ws2, headers2)

    sorted_findings = sorted(findings, key=lambda f: (sev_order.get(f.get("Severity", "medium").lower(), 5), f.get("Category", "")))
    for i, f in enumerate(sorted_findings, 2):
        sev = f.get("Severity", "medium").lower()
        cat_key = f.get("Category", "")
        cat_name = _CATEGORY_META.get(cat_key, {}).get("name", cat_key)

        resources = f.get("AffectedResources", [])
        resource_names = []
        for r in (resources if isinstance(resources, list) else []):
            if isinstance(r, dict):
                name = r.get("Name", r.get("SubscriptionId", r.get("ResourceId", str(r))))
                if isinstance(name, dict):
                    name = str(name)
                resource_names.append(name)
            else:
                resource_names.append(str(r))
        resource_str = ", ".join(resource_names)

        remediation = f.get("Remediation", "")
        if isinstance(remediation, dict):
            parts = []
            for k, v in remediation.items():
                if v:
                    parts.append(f"{k}: {v}")
            remediation = "\n".join(parts)

        fills: dict[int, PatternFill] = {}
        if sev in _XL_SEV_FILLS:
            fills[1] = _XL_SEV_FILLS[sev]

        _xl_write_row(ws2, i, [
            sev.capitalize(),
            cat_name,
            f.get("Subcategory", ""),
            f.get("Title", ""),
            f.get("AffectedCount", len(resources) if isinstance(resources, list) else 0),
            remediation,
            resource_str,
        ], fills)

    _xl_auto_width(ws2, headers2)

    # ── Sheet 3: Category Scores ──
    ws3 = wb.create_sheet("Category Scores")
    headers3 = ["Category", "Score", "Level", "Finding Count", "Description"]
    _xl_apply_header(ws3, headers3)

    row = 2
    for cat_key in ["identity", "network", "defender", "config"]:
        meta = _CATEGORY_META[cat_key]
        cat_score = scores.get("CategoryScores", {}).get(cat_key, {})
        cat_findings = [f for f in findings if f.get("Category") == cat_key]
        _xl_write_row(ws3, row, [
            meta["name"],
            round(cat_score.get("Score", 0), 1),
            cat_score.get("Level", "N/A"),
            len(cat_findings),
            meta["description"],
        ])
        row += 1

    _xl_auto_width(ws3, headers3)

    wb.save(out_path)
    log.info("[RiskExcel] Written to %s", out_path)
    return out_path
