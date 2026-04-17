"""
AI Agent Security Report — Interactive HTML
Full-width professional report showing security posture of AI agents
across Copilot Studio, Microsoft Foundry, and custom agents.
"""

from __future__ import annotations

import hashlib
import pathlib
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.reports.shared_theme import (
    get_css, get_js, esc, format_date_short, VERSION,
    SEVERITY_COLORS,
)
from app.logger import log


# ── Category metadata ────────────────────────────────────────────────────

_CATEGORY_META: dict[str, dict] = {
    "cs_authentication": {
        "icon": "&#128272;",  # 🔐
        "name": "CS Authentication",
        "color": "#D13438",
        "description": "Copilot Studio agent authentication enforcement and provider security.",
        "tooltip": "Checks: Whether Copilot Studio bots enforce user authentication before conversations, and whether auth providers (AAD, OAuth) are securely configured.\nEvaluates: Copilot Studio bot definitions from Power Platform environments.\nZero findings: All bots require authentication, or no bots are published in this tenant.",
    },
    "cs_data_connectors": {
        "icon": "&#128279;",  # 🔗
        "name": "CS Data Connectors",
        "color": "#0078D4",
        "description": "DLP policies, managed environments, and security groups for Power Platform.",
        "tooltip": "Checks: DLP policy coverage across Power Platform environments, managed environment enforcement, and Dataverse security group assignments.\nEvaluates: DLP policies, Power Platform environments, and their security configurations.\nZero findings: All environments are covered by DLP policies with proper security groups.",
    },
    "cs_logging": {
        "icon": "&#128221;",  # 📝
        "name": "CS Conversation Logging",
        "color": "#FFB900",
        "description": "Audit trail and conversation logging for Copilot Studio agents.",
        "tooltip": "Checks: Whether Copilot Studio bots have conversation transcription enabled and whether audit logs capture agent interactions.\nEvaluates: Bot configuration settings for transcript logging and Dataverse audit settings.\nZero findings: All bots have conversation logging enabled, or no bots exist in this tenant.",
    },
    "cs_channels": {
        "icon": "&#127760;",  # 🌐
        "name": "CS Channel Security",
        "color": "#F7630C",
        "description": "Web and Teams channel exposure for Copilot Studio agents.",
        "tooltip": "Checks: Whether bots are exposed via public web channels without authentication, and whether Teams channel deployments follow security best practices.\nEvaluates: Channel configurations on Copilot Studio bot definitions.\nZero findings: All channels are properly secured, or no bots with external channels exist.",
    },
    "cs_knowledge_sources": {
        "icon": "&#128218;",  # 📚
        "name": "CS Knowledge Sources",
        "color": "#4A90D9",
        "description": "SharePoint and Dataverse knowledge source security for Copilot Studio agents.",
        "tooltip": "Checks: Whether knowledge sources (SharePoint sites, Dataverse tables, uploaded documents) connected to bots are properly secured with access controls.\nEvaluates: Knowledge source bindings on Copilot Studio bot definitions.\nZero findings: All knowledge sources have proper access restrictions, or no bots use knowledge sources.",
    },
    "cs_generative_ai": {
        "icon": "&#129302;",  # 🤖
        "name": "CS Generative AI",
        "color": "#6B4FBB",
        "description": "Generative answers guardrails and orchestration controls.",
        "tooltip": "Checks: Whether bots using generative answers have content moderation enabled, whether orchestration is constrained to approved topics, and hallucination guardrails.\nEvaluates: Generative AI settings on Copilot Studio bot definitions.\nZero findings: Generative AI features are properly constrained, or no bots use generative answers.",
    },
    "cs_governance": {
        "icon": "&#128220;",  # 📜
        "name": "CS Governance",
        "color": "#2D7D9A",
        "description": "Solution awareness, ALM governance, and draft agent controls.",
        "tooltip": "Checks: Whether bots are deployed in managed solutions (not standalone), whether ALM practices (dev/test/prod) are followed, and whether draft agents are properly controlled.\nEvaluates: Copilot Studio bot solution metadata and environment lifecycle settings.\nZero findings: All bots follow ALM best practices, or no bots exist in this tenant.",
    },
    "cs_connector_security": {
        "icon": "&#128268;",  # 🔌
        "name": "CS Connectors",
        "color": "#E6526C",
        "description": "Custom connector authentication and premium connector DLP coverage.",
        "tooltip": "Checks: Whether custom connectors require authentication (OAuth/API key), and whether premium connectors are covered by DLP policies.\nEvaluates: Custom connector definitions and DLP policy connector classifications.\nZero findings: All custom connectors require auth and premium connectors are DLP-covered, or no custom connectors exist.",
    },
    "cs_dlp_depth": {
        "icon": "&#128274;",  # 🔒
        "name": "CS DLP Depth",
        "color": "#5C2D91",
        "description": "DLP policy depth analysis: authentication connectors, knowledge sources, channels, skills, HTTP, and default-group enforcement.",
        "tooltip": "Checks: Deep DLP policy analysis — whether HTTP, Skills, and external channel connectors are blocked; whether the default group for new connectors is set to Blocked; auth connector classification.\nEvaluates: Per-connector classification in each DLP policy, default data group settings, and connector-level blocking rules.\nZero findings: All DLP policies enforce strict connector blocking with secure defaults.",
    },
    "cs_environment_governance": {
        "icon": "&#127970;",  # 🏢
        "name": "CS Environment Governance",
        "color": "#00B7C3",
        "description": "Power Platform environment governance: managed environments, tenant isolation, and generative AI controls.",
        "tooltip": "Checks: Whether tenant isolation is enabled in Power Platform; whether environments are converted to Managed Environments (required for advanced governance); generative AI feature controls.\nEvaluates: Power Platform tenant settings and environment properties from admin APIs.\nZero findings: Tenant isolation is enabled and all environments are managed with proper generative AI controls.",
    },
    "cs_audit_compliance": {
        "icon": "&#128220;",  # 📜
        "name": "CS Audit Compliance",
        "color": "#F7630C",
        "description": "Unified audit logging in Microsoft Purview, DSPM for AI, and cross-geo data-movement compliance.",
        "tooltip": "Checks: Whether unified audit logging is enabled in Microsoft Purview; whether Data Security Posture Management (DSPM) for AI is configured to monitor agent chat transcripts; cross-geo data movement.\nEvaluates: Microsoft Purview audit settings, DSPM for AI configuration, and environment geo-location data.\nZero findings: Audit logging, DSPM for AI, and geo-compliance are all properly configured.",
    },
    "cs_dataverse_security": {
        "icon": "&#128451;",  # 🗃
        "name": "CS Dataverse Security",
        "color": "#7B68EE",
        "description": "Dataverse security groups, Customer Lockbox, and customer-managed key (CMK) enforcement.",
        "tooltip": "Checks: Whether production Power Platform environments restrict access via Dataverse security groups; Customer Lockbox availability; customer-managed key (CMK) encryption.\nEvaluates: Environment security group assignments and encryption settings from Power Platform admin APIs.\nZero findings: All production environments have security groups assigned with proper encryption.",
    },
    "cs_readiness_crosscheck": {
        "icon": "&#9989;",  # ✅
        "name": "CS Readiness Crosscheck",
        "color": "#107C10",
        "description": "Pre-deployment readiness: managed environment governance, DLP coverage, and cross-tenant isolation.",
        "tooltip": "Checks: Cross-cutting readiness — whether environments are managed, DLP policies cover all environments, tenant isolation is enabled, and overall governance posture is deployment-ready.\nEvaluates: Aggregated results from other Copilot Studio checks to identify readiness gaps.\nZero findings: All prerequisite governance controls are in place for safe agent deployment.",
    },
    "foundry_network": {
        "icon": "&#128274;",  # 🔒
        "name": "Foundry Network Isolation",
        "color": "#00B7C3",
        "description": "Public access, private endpoints, and workspace network isolation for AI services.",
        "tooltip": "Checks: Whether Azure AI service accounts (CognitiveServices/accounts) have public network access disabled and private endpoints configured.\nEvaluates: Network settings on each AI service account — publicNetworkAccess property and privateEndpointConnections array.\nZero findings: All AI service accounts use private endpoints with public access disabled.",
        "arch": "new",
    },
    "foundry_identity": {
        "icon": "&#128100;",  # 👤
        "name": "Foundry Identity",
        "color": "#8764B8",
        "description": "Managed identity and API key management for Azure AI services.",
        "tooltip": "Checks: Whether AI service accounts have local authentication (API keys) disabled, enforcing managed identity; whether ML workspaces use system-assigned managed identity.\nEvaluates: DisableLocalAuth property on AI service accounts (CognitiveServices/accounts) and identity assignment on ML workspaces.\nZero findings: All AI accounts enforce managed identity with local auth disabled — this is SECURE.",
        "arch": "classic & new",
    },
    "foundry_content_safety": {
        "icon": "&#128737;",  # 🛡
        "name": "Content Safety",
        "color": "#C239B3",
        "description": "Content filter coverage and blocking strength for OpenAI deployments.",
        "tooltip": "Checks: Whether OpenAI model deployments have content filter policies assigned; whether filter categories (hate, violence, sexual, self-harm) are set to blocking mode.\nEvaluates: Content filter policies and their assignment to deployments under AI service accounts.\nZero findings: All deployments have content filters with blocking enabled for all categories.",
        "arch": "new",
    },
    "foundry_deployments": {
        "icon": "&#128640;",  # 🚀
        "name": "Deployment Security",
        "color": "#107C10",
        "description": "Model deployment governance, capacity controls, and cost governance.",
        "tooltip": "Checks: Whether OpenAI deployments have RAI policies assigned; high-capacity (>100 TPM) allocations that may indicate cost governance gaps.\nEvaluates: Deployment configurations under Azure OpenAI accounts — RAI policy assignment, capacity (TPM) limits, and model versions.\nZero findings: All deployments have RAI policies with appropriate capacity limits.",
        "arch": "new",
    },
    "foundry_governance": {
        "icon": "&#127970;",  # 🏢
        "name": "Workspace Governance",
        "color": "#0063B1",
        "description": "Hub/project organizational structure and centralized governance.",
        "tooltip": "Checks: Whether AI ML Workspaces follow hub/project structure; CMK encryption on workspaces; network isolation between hub and child projects.\nEvaluates: Azure ML Workspaces (MachineLearningServices/workspaces) — IsHub, IsProject, HasCMK, network isolation properties.\nZero findings: All workspaces follow proper hub/project structure, or no ML workspaces exist in this tenant (common when using new Foundry architecture).",
        "arch": "classic",
    },
    "foundry_compute": {
        "icon": "&#128187;",  # 💻
        "name": "Compute Security",
        "color": "#3B78AB",
        "description": "AI compute instance security: public IP, SSH access, and idle shutdown.",
        "tooltip": "Checks: Whether ML compute instances have public IP enabled, SSH access open, idle shutdown configured, and managed identity assigned.\nEvaluates: Compute instances provisioned inside Azure ML Workspaces (MachineLearningServices/workspaces/computes). These are training/development VMs inside ML workspaces.\nZero findings: All compute is secure, or no ML workspaces/compute instances exist (common when using new Foundry CognitiveServices architecture instead of classic ML).",
        "arch": "classic",
    },
    "foundry_datastores": {
        "icon": "&#128451;",  # 🗃
        "name": "Datastore Security",
        "color": "#567D46",
        "description": "AI datastore credential management and storage encryption.",
        "tooltip": "Checks: Whether ML datastores use stored credentials (account keys, SAS tokens) instead of identity-based auth; storage encryption validation.\nEvaluates: Datastore definitions inside Azure ML Workspaces (MachineLearningServices/workspaces/datastores) — credential type and encryption settings.\nZero findings: All datastores use identity-based auth with encryption, or no ML workspaces/datastores exist (common when using new Foundry architecture).",
        "arch": "classic",
    },
    "foundry_endpoints": {
        "icon": "&#128228;",  # 📤
        "name": "Endpoint Security",
        "color": "#8B5E3C",
        "description": "Online/batch AI endpoint access control and authentication.",
        "tooltip": "Checks: Whether online/batch ML endpoints have public access; whether auth is set to AAD (not key-based); diagnostic coverage on endpoints.\nEvaluates: Online and batch endpoints inside Azure ML Workspaces (MachineLearningServices/workspaces/onlineEndpoints and batchEndpoints).\nZero findings: All endpoints use AAD auth with restricted access, or no ML endpoints are deployed (common when using OpenAI deployments via new Foundry architecture instead).",
        "arch": "classic",
    },
    "foundry_registry": {
        "icon": "&#128230;",  # 📦
        "name": "Registry Security",
        "color": "#6E5494",
        "description": "Model registry network access and RBAC controls.",
        "tooltip": "Checks: Whether AI model registries have public network access enabled; whether explicit RBAC permissions are configured.\nEvaluates: Azure ML Registries (MachineLearningServices/registries) — shared model/component registries for cross-workspace reuse.\nZero findings: All registries are secured, or no ML registries exist in the tenant subscriptions.",
        "arch": "classic",
    },
    "foundry_connections": {
        "icon": "&#128279;",  # 🔗
        "name": "Connection Security",
        "color": "#D35400",
        "description": "Workspace connection credentials, sharing, and expiration management.",
        "tooltip": "Checks: Whether ML workspace connections use static credentials; whether connections are shared to all users; credential expiration management.\nEvaluates: Connections defined inside Azure ML Workspaces (MachineLearningServices/workspaces/connections) — external service links (e.g., to storage, databases, APIs).\nZero findings: All connections use managed identity, or no ML workspaces/connections exist (common when using new Foundry architecture).",
        "arch": "classic",
    },
    "foundry_serverless": {
        "icon": "&#9889;",  # ⚡
        "name": "Serverless Endpoints",
        "color": "#1ABC9C",
        "description": "Serverless (MaaS) model endpoint authentication and content safety.",
        "tooltip": "Checks: Whether serverless (Model-as-a-Service) endpoints use key auth instead of AAD; content safety enforcement; key rotation status.\nEvaluates: Serverless endpoints in ML Workspaces (MachineLearningServices/workspaces/serverlessEndpoints) — pay-per-token model deployments (e.g., Llama, Mistral via MaaS).\nZero findings: All serverless endpoints are secured, or no MaaS model endpoints are deployed in this tenant.",
        "arch": "classic",
    },
    "foundry_ws_diagnostics": {
        "icon": "&#128200;",  # 📈
        "name": "Workspace Diagnostics",
        "color": "#5B2C6F",
        "description": "Diagnostic settings and audit logging for AI Foundry workspaces.",
        "tooltip": "Checks: Whether ML workspaces have diagnostic settings enabled with Log Analytics workspace as destination for audit and operational logs.\nEvaluates: Diagnostic settings (Insights/diagnosticSettings) on Azure ML Workspaces — log categories, retention, and sink configuration.\nZero findings: All workspaces have diagnostics enabled, or no ML workspaces exist (common when using new Foundry CognitiveServices architecture).",
        "arch": "classic",
    },
    "foundry_prompt_shields": {
        "icon": "&#128737;",  # 🛡
        "name": "Prompt Shield Security",
        "color": "#E74C3C",
        "description": "Prompt injection and jailbreak protection for AI deployments.",
        "tooltip": "Checks: Whether content filter policies include jailbreak detection (Prompt Shields); whether custom blocklists are configured for organization-specific terms.\nEvaluates: Content filter policy definitions on AI service accounts — jailbreak filter settings and custom blocklist configurations.\nZero findings: All content policies include jailbreak protection with custom blocklists configured.",
        "arch": "new",
    },
    "foundry_model_catalog": {
        "icon": "&#128218;",  # 📚
        "name": "Model Catalog Governance",
        "color": "#3498DB",
        "description": "Model approval, versioning, and catalog governance for AI deployments.",
        "tooltip": "Checks: Whether deployed models are from the approved Azure AI model catalog or are custom/unapproved; model versioning and deprecation status.\nEvaluates: Model identifiers on OpenAI deployments cross-referenced against the standard Azure AI model catalog.\nZero findings: All deployed models are from the approved catalog with current versions.",
        "arch": "new",
    },
    "foundry_data_exfiltration": {
        "icon": "&#128274;",  # 🔒
        "name": "Data Exfiltration Prevention",
        "color": "#E67E22",
        "description": "Managed network and outbound rules to prevent data exfiltration.",
        "tooltip": "Checks: Whether ML workspaces use managed virtual networks with outbound rules; whether isolation mode prevents unrestricted internet egress (AllowInternetOutbound).\nEvaluates: ML Workspace network isolation settings (MachineLearningServices/workspaces) — managed network, outbound rules, and isolation mode.\nZero findings: All workspaces have managed networks with restricted egress, or no ML workspaces exist (common when using new Foundry architecture).",
        "arch": "classic",
    },
    "custom_api_security": {
        "icon": "&#128477;",  # 🗝
        "name": "API Key Security",
        "color": "#E74856",
        "description": "API key management and network restrictions for custom agents.",
        "tooltip": "Checks: Whether AI service API keys are exposed without network restrictions; whether key rotation policies are in place.\nEvaluates: API key settings and network ACLs across AI service accounts that serve as backends for custom agent applications.\nZero findings: All API keys are properly restricted with network controls, or managed identity is used exclusively.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "custom_data_residency": {
        "icon": "&#127758;",  # 🌎
        "name": "Data Residency",
        "color": "#2E8B57",
        "description": "AI service region sprawl and data residency compliance.",
        "tooltip": "Checks: Whether AI services are deployed in many different regions (region sprawl); whether data-at-rest location complies with organizational geo-requirements.\nEvaluates: Region/location property on all AI service accounts to detect cross-region deployment patterns.\nZero findings: AI services are consolidated in approved regions with no data residency violations.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "custom_content_leakage": {
        "icon": "&#128680;",  # 🚨
        "name": "Content Leakage",
        "color": "#C72C41",
        "description": "Content filter gaps and encryption for preventing sensitive content exposure.",
        "tooltip": "Checks: Overall content filter coverage ratio across all deployments; whether AI services use customer-managed keys (CMK) for data-at-rest encryption.\nEvaluates: Content filter assignment across all OpenAI deployments and CMK configuration on AI service accounts — cross-platform gap analysis.\nZero findings: Full content filter coverage and CMK encryption across all AI services.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "entra_ai_service_principals": {
        "icon": "&#128101;",  # 👥
        "name": "AI Service Principals",
        "color": "#4B6584",
        "description": "Entra service principal permissions, credentials, and multi-tenant exposure for AI apps.",
        "tooltip": "Checks: Credential expiry on AI-related service principals; privileged directory role assignments; owner governance (single-owner vs group-owned); multi-tenant exposure.\nEvaluates: Entra ID (Azure AD) service principals and app registrations that have AI-related permissions (e.g., CognitiveServices, OpenAI scopes).\nZero findings: All AI service principals have valid credentials, no privileged roles, proper ownership, and single-tenant configuration.",
    },
    "entra_ai_conditional_access": {
        "icon": "&#128275;",  # 🔓
        "name": "AI Conditional Access",
        "color": "#7B68EE",
        "description": "Conditional Access policy coverage and token restrictions for AI applications.",
        "tooltip": "Checks: Whether AI applications are covered by Conditional Access policies; token lifetime restrictions; Continuous Access Evaluation (CAE) and sign-in frequency controls.\nEvaluates: Entra ID Conditional Access policies cross-referenced against AI app registrations to find coverage gaps.\nZero findings: All AI applications are covered by Conditional Access with proper session controls.",
    },
    "entra_ai_consent": {
        "icon": "&#9989;",  # ✅
        "name": "AI Consent Grants",
        "color": "#9B59B6",
        "description": "OAuth consent grants to AI applications with sensitive permission scopes.",
        "tooltip": "Checks: Whether AI applications have admin-consented high-privilege permissions (e.g., Directory.ReadWrite.All, Mail.ReadWrite); risky delegated consent patterns.\nEvaluates: OAuth2 permission grants (oauth2PermissionGrants) on AI-related service principals in Entra ID.\nZero findings: No AI applications have excessive admin-consented permissions.",
    },
    "entra_ai_workload_identity": {
        "icon": "&#128736;",  # 🛠
        "name": "AI Workload Identity",
        "color": "#2E86C1",
        "description": "Workload Identity Federation and federated credentials for AI CI/CD pipelines.",
        "tooltip": "Checks: Whether AI app registrations use Workload Identity Federation for CI/CD; whether federated credentials have proper issuer and subject constraints.\nEvaluates: Federated identity credentials on AI-related app registrations — used for GitHub Actions, Azure DevOps, or other external identity providers.\nZero findings: All federated credentials are properly scoped, or no AI apps use workload identity federation.",
    },
    "entra_ai_cross_tenant": {
        "icon": "&#127760;",  # 🌐
        "name": "AI Cross-Tenant Access",
        "color": "#D35400",
        "description": "Cross-tenant access controls for multi-tenant AI service principals.",
        "tooltip": "Checks: Whether AI service principals are configured as multi-tenant (accessible from other tenants); cross-tenant access policy coverage.\nEvaluates: signInAudience and multi-tenant flags on AI app registrations, cross-referenced with Entra cross-tenant access settings.\nZero findings: All AI apps are single-tenant, or multi-tenant apps have proper cross-tenant access policies.",
    },
    "entra_ai_privileged_access": {
        "icon": "&#9733;",  # ★
        "name": "AI Privileged Access",
        "color": "#8E44AD",
        "description": "PIM coverage and just-in-time activation for AI-privileged roles.",
        "tooltip": "Checks: Whether users/groups with AI-privileged roles (Cognitive Services Contributor, etc.) are enrolled in Privileged Identity Management (PIM) for just-in-time activation.\nEvaluates: Azure RBAC role assignments on AI resources cross-referenced with PIM enrollment status.\nZero findings: All privileged AI roles use PIM with just-in-time activation, or no privileged role assignments exist.",
    },
    "ai_diagnostics": {
        "icon": "&#128202;",  # 📊
        "name": "AI Diagnostics",
        "color": "#20B2AA",
        "description": "Diagnostic settings and audit logging for AI service accounts.",
        "tooltip": "Checks: Whether AI service accounts (CognitiveServices/accounts) have diagnostic settings configured with Log Analytics or storage destinations.\nEvaluates: Diagnostic settings on each AI service account via Azure Monitor APIs — log category coverage and sink configuration.\nZero findings: All AI service accounts have diagnostic settings with appropriate log destinations.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "ai_model_governance": {
        "icon": "&#9881;",  # ⚙
        "name": "Model Governance",
        "color": "#708090",
        "description": "Model version management, rate limiting, and deployment controls.",
        "tooltip": "Checks: Whether OpenAI deployments have TPM (tokens-per-minute) rate limits configured; model version currency and deprecation status.\nEvaluates: Deployment properties under Azure OpenAI accounts — capacity limits, model versions, and rate limiting configuration.\nZero findings: All deployments have appropriate rate limits and use current model versions.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "ai_threat_protection": {
        "icon": "&#9888;",  # ⚠
        "name": "AI Threat Protection",
        "color": "#FF4500",
        "description": "Prompt injection, jailbreak, PII detection, and groundedness controls.",
        "tooltip": "Checks: Whether content filter policies include Prompt Shields (prompt injection protection); groundedness detection for hallucination mitigation; PII detection/redaction.\nEvaluates: Content filter policy features across AI service accounts — jailbreak filters, groundedness detection, and PII filter configurations.\nZero findings: All content policies include prompt injection protection, groundedness detection, and PII filtering.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "ai_data_governance": {
        "icon": "&#128196;",  # 📄
        "name": "AI Data Governance",
        "color": "#2F4F4F",
        "description": "Data classification, sensitivity labels, and retention policies for AI workloads.",
        "tooltip": "Checks: Whether AI services have data retention policies configured; sensitivity label coverage on AI data stores; data classification practices.\nEvaluates: Retention settings and data lifecycle configuration on AI service accounts and their associated storage.\nZero findings: All AI services have data retention policies and classification in place.\nApplies to: Foundry Classic & New architecture.",
    },
    "ai_defender_coverage": {
        "icon": "&#128737;",  # 🛡
        "name": "Defender for AI",
        "color": "#4169E1",
        "description": "Microsoft Defender for AI plan coverage and alert configuration.",
        "tooltip": "Checks: Whether subscriptions containing AI service accounts have Microsoft Defender for AI enabled; alert configuration status.\nEvaluates: Microsoft Defender for Cloud pricing tier on AI-containing subscriptions via Azure Security Center APIs.\nZero findings: All subscriptions with AI resources have Defender for AI enabled.\nApplies to: Foundry New architecture (CognitiveServices/accounts).",
    },
    "ai_policy_compliance": {
        "icon": "&#128220;",  # 📜
        "name": "AI Policy Compliance",
        "color": "#556B2F",
        "description": "Azure Policy assignments and compliance state for AI resources.",
        "tooltip": "Checks: Whether Azure Policy assignments exist that govern AI resources (e.g., deny public access, require CMK, enforce diagnostic settings).\nEvaluates: Azure Policy assignment state at subscription/resource group level for AI-related built-in and custom policies.\nZero findings: Azure Policies are assigned and AI resources are compliant.\nApplies to: Foundry Classic & New architecture.",
    },
    "agent_communication": {
        "icon": "&#128172;",  # 💬
        "name": "Agent Communication",
        "color": "#FF6347",
        "description": "Inter-agent authentication, tool scoping, and memory encryption.",
        "tooltip": "Checks: Whether multi-agent orchestration frameworks enforce authentication between agents; tool scoping restrictions; memory/state encryption.\nEvaluates: Agent application configurations in Foundry projects — inter-agent auth, tool access policies, and state management security.\nZero findings: All agent apps enforce mutual authentication with scoped tool access, or no agent applications are deployed.",
    },
    "agent_governance": {
        "icon": "&#127891;",  # 🎓
        "name": "Agent Governance",
        "color": "#8B4513",
        "description": "Agent inventory, human-in-the-loop controls, and shadow agent detection.",
        "tooltip": "Checks: Whether all deployed AI agents are registered in a central inventory; human-in-the-loop controls for high-risk actions; detection of unregistered (shadow) agents.\nEvaluates: Agent application registrations across Foundry projects and Azure subscriptions to identify governance gaps.\nZero findings: All agents are inventoried with proper human-in-the-loop controls, or no agent applications exist.",
    },
}

_PLATFORM_META: dict[str, dict] = {
    "copilot_studio": {"name": "Copilot Studio", "icon": "&#129302;", "color": "#0078D4",
                       "desc": "Microsoft's low-code bot builder on Power Platform for creating conversational AI agents"},
    "foundry": {"name": "Microsoft Foundry", "icon": "&#129504;", "color": "#8764B8",
                "desc": "Microsoft Foundry platform for deploying and managing AI models and agents"},
    "cross-cutting": {"name": "Cross-Cutting", "icon": "&#128279;", "color": "#FFB900",
                      "desc": "Security controls that span multiple platforms (e.g., DLP, content safety, network isolation)"},
    "entra_identity": {"name": "Entra Identity", "icon": "&#128101;", "color": "#4B6584",
                       "desc": "Microsoft Entra ID (Azure AD) service principals, consent, conditional access, workload identity, cross-tenant, and privileged access for AI apps"},
    "ai_infra": {"name": "AI Infrastructure", "icon": "&#9881;", "color": "#20B2AA",
                 "desc": "Underlying Azure AI services, diagnostic settings, Defender coverage, and policy compliance"},
    "agent_orchestration": {"name": "Agent Orchestration", "icon": "&#128172;", "color": "#FF6347",
                            "desc": "Multi-agent communication, tool scoping, governance, and human-in-the-loop controls"},
}

_PLATFORM_CATEGORIES: dict[str, list[str]] = {
    "copilot_studio": [
        "cs_authentication", "cs_data_connectors", "cs_logging", "cs_channels",
        "cs_knowledge_sources", "cs_generative_ai", "cs_governance", "cs_connector_security",
        "cs_dlp_depth", "cs_environment_governance", "cs_audit_compliance",
        "cs_dataverse_security", "cs_readiness_crosscheck",
    ],
    "foundry": [
        "foundry_network", "foundry_identity", "foundry_content_safety",
        "foundry_deployments", "foundry_governance", "foundry_compute",
        "foundry_datastores", "foundry_endpoints", "foundry_registry",
        "foundry_connections", "foundry_serverless", "foundry_ws_diagnostics",
        "foundry_prompt_shields", "foundry_model_catalog", "foundry_data_exfiltration",
    ],
    "cross-cutting": ["custom_api_security", "custom_data_residency", "custom_content_leakage"],
    "entra_identity": [
        "entra_ai_service_principals", "entra_ai_conditional_access", "entra_ai_consent",
        "entra_ai_workload_identity", "entra_ai_cross_tenant", "entra_ai_privileged_access",
    ],
    "ai_infra": ["ai_diagnostics", "ai_model_governance", "ai_threat_protection", "ai_data_governance", "ai_defender_coverage", "ai_policy_compliance"],
    "agent_orchestration": ["agent_communication", "agent_governance"],
}

_LEVEL_META: dict[str, dict] = {
    "critical": {"color": "#D13438", "label": "Critical"},
    "high": {"color": "#F7630C", "label": "High"},
    "medium": {"color": "#FFB900", "label": "Medium"},
    "secure": {"color": "#107C10", "label": "Secure"},
}

# Tooltip text for well-known detail-column keys in affected-resource tables
_DETAIL_KEY_TIPS: dict[str, str] = {
    "Kind": "Azure resource kind (e.g., OpenAI, CognitiveServices, SpeechServices).\n• The 'kind' property on a Microsoft.CognitiveServices/accounts resource\n• Determines which security checks apply to this resource\n• Source: Azure Resource Manager API",
    "IsOpenAI": "Whether this AI service is an OpenAI-type resource (True/False).\n• True when Kind is 'OpenAI' or 'AzureOpenAI'\n• OpenAI resources get extra checks for content filters, encryption keys, and data retention\n• Non-OpenAI AI services (Speech, Vision, etc.) skip those checks",
    "Modelname": "The name of the deployed model (e.g., gpt-4o, gpt-35-turbo).\n• Set at deployment creation time\n• Different models have different safety and compliance profiles\n• Check for deprecated or end-of-life models",
    "Capacity": "Provisioned throughput in Tokens Per Minute (TPM).\n• Higher capacity = higher cost and broader blast radius\n• Review whether the allocated TPM matches actual usage\n• Consider quotas and rate limits for cost governance",
    "Accountid": "The full Azure resource ID of the AI service account.\n• Format: /subscriptions/.../providers/Microsoft.CognitiveServices/accounts/...\n• Use this ID in Azure CLI or PowerShell remediation commands\n• Clickable IDs open the resource in Azure Portal",
    "Accountname": "The display name of the AI service account in Azure.\n• Set when the resource was created\n• Use this to identify the resource in the Azure Portal\n• May differ from the deployment or model name",
    "Filtereddeployments": "Number of model deployments matching the filter criteria.\n• Deployments with missing or weak content filters\n• Each unfiltered deployment is an exposure point\n• Apply content filter policies to reduce this count",
    "Unfiltereddeployments": "Number of model deployments without content safety filters.\n• These deployments have no RAI content filters applied\n• Unfiltered deployments risk generating harmful or off-topic content\n• Apply content filter policies to all production deployments",
    "Totalfilters": "Total number of content filter policies found on the account.\n• Includes both custom and default filter policies\n• Zero means no content safety guardrails are in place\n• Create and assign filter policies via Microsoft Foundry portal",
    "Blockingfilters": "Number of content filter policies configured in blocking mode.\n• 'Blocking' means harmful content is actively rejected, not just logged\n• Non-blocking (annotate-only) filters still allow harmful content through\n• Set all filter categories to 'Block' for production safety",
    "Coverage": "Percentage of deployments protected by content safety filters.\n• 100% = all deployments have filters — fully covered\n• 0% = no deployments have filters — fully exposed\n• Target: 100% coverage across all production deployments",
}

_SEVERITY_TIPS: dict[str, str] = {
    "critical": "CRITICAL — highest-impact security gap.\n• Immediate exploitation risk\n• Adds 10 points per finding to the risk score\n• Fix these first",
    "high": "HIGH — significant risk requiring prompt attention.\n• Exploitation possible with moderate effort\n• Adds 7.5 points per finding to the risk score\n• Plan remediation within your next maintenance window",
    "medium": "MEDIUM — moderate gap, not immediately exploitable.\n• Could escalate if combined with other issues\n• Adds 5 points per finding to the risk score\n• Schedule remediation in your backlog",
    "low": "LOW — minor improvement opportunity.\n• Best-practice recommendation\n• Adds 2.5 points per finding to the risk score\n• Address when capacity allows",
    "informational": "INFO — observation with no direct risk.\n• For awareness and audit trail only\n• Adds 1 point per finding to the risk score\n• No remediation action needed",
}


# ── SVG helpers ──────────────────────────────────────────────────────────

def _ring_score_svg(score: float, size: int = 140) -> str:
    r = size // 2 - 8
    circ = 2 * 3.14159 * r
    pct = min(score, 100) / 100
    dash = circ * pct
    gap = circ - dash
    cx = cy = size // 2
    color = "#D13438" if score >= 75 else "#F7630C" if score >= 50 else "#FFB900" if score >= 25 else "#107C10"
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" role="img">'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="var(--ring-track)" stroke-width="10"/>'
        f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="10" '
        f'stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-dashoffset="{circ * 0.25:.2f}" '
        f'stroke-linecap="round" style="transition:stroke-dasharray 1s ease"/>'
        f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" font-size="28" font-weight="700" '
        f'fill="{color}" font-family="var(--font-mono)">{score:.0f}</text>'
        f'<text x="{cx}" y="{cy + 18}" text-anchor="middle" font-size="10" fill="var(--text-muted)">/100</text>'
        f'</svg>'
    )


def _donut_svg(slices: list[tuple[str, float, str]], size: int = 140, center_text: str | None = None) -> str:
    total = sum(v for _, v, _ in slices) or 1
    r = size // 2 - 4
    circ = 2 * 3.14159 * r
    cx = cy = size // 2
    parts = [f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" role="img">']
    offset = 0
    for label, val, color in slices:
        if val <= 0:
            continue
        pct = val / total
        dash = circ * pct
        parts.append(
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" '
            f'stroke-width="{r * 0.4}" stroke-dasharray="{dash:.2f} {circ - dash:.2f}" '
            f'stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})">'
            f'<title>{esc(label)}: {int(val)}</title></circle>'
        )
        offset += dash
    ct = center_text if center_text is not None else str(int(total))
    parts.append(f'<text x="{cx}" y="{cy}" text-anchor="middle" dy=".35em" font-size="22" font-weight="700" fill="var(--text)" font-family="var(--font-mono)">{esc(ct)}</text></svg>')
    return "".join(parts)


# ── Badges ───────────────────────────────────────────────────────────────

def _severity_badge(sev: str) -> str:
    color = SEVERITY_COLORS.get(sev.lower(), "#A8A6A3")
    tip = _SEVERITY_TIPS.get(sev.lower(), "")
    return f'<span class="sev-text" style="color:{color}" data-tip="{esc(tip)}">{esc(sev.upper())}</span>'


def _platform_badge(platform: str) -> str:
    meta = _PLATFORM_META.get(platform, {"name": platform, "color": "#A8A6A3"})
    ptip = f'{esc(meta["name"])} — {esc(meta.get("desc", "Technology platform for this finding"))}'
    return f'<span class="plat-label" style="color:{meta["color"]}" data-tip="{ptip}">{esc(meta["name"])}</span>'


# ── Rendering helpers ────────────────────────────────────────────────────

def _render_finding(f: dict, finding_idx: int = 0) -> str:
    sev = f.get("Severity", "medium").lower()
    cat = f.get("Category", "unknown")
    subcat = f.get("Subcategory", "")
    title = f.get("Title", "Untitled")
    desc = f.get("Description", "")
    affected = f.get("AffectedCount", 0)
    affected_resources = f.get("AffectedResources", [])
    remediation = f.get("Remediation", {})
    platform = f.get("Platform", "cross-cutting")
    finding_id = f"finding-{esc(cat)}-{finding_idx}"

    rem_html = ""
    if remediation:
        rem_parts = []
        rem_desc = remediation.get("Description", "")
        if rem_desc:
            rem_parts.append(f'<div class="rem-desc">{esc(rem_desc)}</div>')
        for key, label in [("AzureCLI", "Azure CLI"), ("PowerShell", "PowerShell")]:
            cmd = remediation.get(key, "")
            if cmd:
                rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">{label}:</strong><pre>{esc(cmd)}</pre></div>')
        steps = remediation.get("PortalSteps", [])
        if steps:
            step_items = "".join(f"<li>{esc(s)}</li>" for s in steps)
            rem_parts.append(f'<div style="margin-top:6px"><strong style="font-size:11px;color:var(--text-muted)">Portal Steps:</strong><ol class="portal-steps">{step_items}</ol></div>')
        if rem_parts:
            rem_html = f'<div class="remediation-box"><h4>&#128736; Remediation</h4>{"".join(rem_parts)}</div>'

    affected_html = ""
    if affected_resources:
        sev_color_res = SEVERITY_COLORS.get(sev, "#A8A6A3")
        _STD = {"Name", "name", "Type", "type", "ResourceId", "resource_id", "id"}
        detail_keys = [k for k in (affected_resources[0] if affected_resources else {}) if k not in _STD] if affected_resources else []
        header = ('<th data-tip="Affected resource name.\n• As reported by the API (e.g., bot name, AI service, environment)\n• Use this to identify which resource needs remediation">Resource</th>'
                  '<th data-tip="Technology type of the resource.\n• Examples: CopilotStudioBot, AIService, ContentFilterPolicy\n• Helps route the fix to the right team">Type</th>'
                  '<th data-tip="Full resource identifier or GUID.\n• Azure resources starting with /subscriptions/ link to the portal\n• Copy the ID for CLI or PowerShell remediation commands">ID</th>')
        for dk in detail_keys:
            dk_label = dk.replace("_", " ").title()
            dk_tip = _DETAIL_KEY_TIPS.get(dk, _DETAIL_KEY_TIPS.get(dk_label, f'{esc(dk_label)} — additional detail from the API.\n• Shows the specific setting that triggered this finding\n• Collected during evidence gathering'))
            header += f'<th data-tip="{esc(dk_tip)}">{esc(dk_label)}</th>'
        header += '<th data-tip="Risk severity for this resource based on the finding.">Risk</th>'
        rows = ""
        for ar in affected_resources:
            if isinstance(ar, dict):
                rid_raw = str(ar.get("ResourceId", ar.get("resource_id", "—")))
                # Make Azure resource IDs clickable portal links
                if rid_raw.startswith("/subscriptions/"):
                    rid_cell = f'<a href="https://portal.azure.com/#@/resource{esc(rid_raw)}" target="_blank" rel="noopener" class="res-id" data-tip="Opens this resource in Azure Portal (new tab).\n• Use the portal to inspect settings and apply fixes\n• Requires at least Reader access to the resource">{esc(rid_raw)}</a>'
                else:
                    rid_cell = f'<span class="res-id">{esc(rid_raw)}</span>'
                rows += (f'<tr><td class="res-name">{esc(str(ar.get("Name", ar.get("name", "—"))))}</td>'
                         f'<td class="res-type">{esc(str(ar.get("Type", ar.get("type", "—"))))}</td>'
                         f'<td>{rid_cell}</td>')
                for dk in detail_keys:
                    rows += f'<td class="res-detail">{esc(str(ar.get(dk, "—")))}</td>'
                rows += f'<td class="res-risk" style="color:{sev_color_res};font-weight:700;font-size:11px;text-transform:uppercase">{esc(sev.upper())}</td>'
                rows += '</tr>'
        affected_html = (
            f'<details class="affected-details" open>'
            f'<summary><span class="aff-count">{affected}</span> affected resources</summary>'
            f'<table class="resource-table"><thead><tr>{header}</tr></thead><tbody>{rows}</tbody></table></details>'
        )

    return (
        f'<div class="finding-card {sev}" id="{finding_id}" data-severity="{sev}" data-category="{esc(cat)}" '
        f'data-subcategory="{esc(subcat)}" data-platform="{esc(platform)}" data-affected="{affected}" '
        f'tabindex="0" role="article">'
        f'<div class="finding-title">{esc(title)}</div>'
        f'<div class="finding-desc">{esc(desc)}</div>'
        f'{affected_html}{rem_html}</div>'
    )


def _bar_chart_svg(items: list[tuple[str, int, str]], width: int = 520) -> str:
    """Horizontal bar chart SVG for category scores."""
    if not items:
        return ""
    bar_h, gap, lbl_w = 24, 8, 160
    h = len(items) * (bar_h + gap) + gap
    max_val = max(v for _, v, _ in items) or 1
    chart_w = width - lbl_w - 60
    parts = [f'<svg width="{width}" height="{h}" viewBox="0 0 {width} {h}" role="img" style="display:block;margin:auto">']
    for i, (label, val, color) in enumerate(items):
        y = gap + i * (bar_h + gap)
        bw = (val / max_val) * chart_w if max_val else 0
        parts.append(f'<text x="{lbl_w - 8}" y="{y + bar_h // 2 + 4}" text-anchor="end" font-size="11" fill="var(--text-secondary)">{esc(label)}</text>')
        parts.append(f'<rect x="{lbl_w}" y="{y}" width="{bw:.1f}" height="{bar_h}" rx="4" fill="{color}" opacity="0.85"><animate attributeName="width" from="0" to="{bw:.1f}" dur="0.6s" fill="freeze"/></rect>')
        parts.append(f'<text x="{lbl_w + bw + 6}" y="{y + bar_h // 2 + 4}" font-size="11" fill="var(--text)" font-family="var(--font-mono)">{val}</text>')
    parts.append("</svg>")
    return "".join(parts)


# ── Report-specific CSS ─────────────────────────────────────────────────

def _as_css() -> str:
    return """
.top-nav{position:sticky;top:0;z-index:500;display:flex;align-items:center;gap:16px;padding:8px 24px;
  background:var(--bg-elevated);border-bottom:1px solid var(--border);font-size:13px;flex-wrap:wrap}
[id^="cat-"],section[id]{scroll-margin-top:56px}
.top-nav .brand{font-weight:700;color:var(--primary);font-size:14px;margin-right:12px}
.top-nav a{color:var(--text-secondary);text-decoration:none;padding:6px 10px;border-radius:6px;transition:all .2s}
.top-nav a:hover{color:var(--text);background:var(--bg-card)}
.nav-dropdown{position:relative}
.nav-dropdown>.nav-toggle{cursor:pointer;user-select:none;padding:6px 10px;border-radius:6px;color:var(--text-secondary);font-size:13px;display:inline-flex;align-items:center;gap:4px;transition:all .2s;min-height:36px;border:none;background:none;font-family:inherit}
.nav-dropdown>.nav-toggle:hover,.nav-dropdown:focus-within>.nav-toggle{color:var(--text);background:var(--bg-card)}
.nav-dropdown>.nav-toggle::after{content:'\\25BE';font-size:10px;margin-left:2px}
.nav-menu{display:none;position:absolute;top:100%;left:0;min-width:220px;background:var(--bg-elevated);border:1px solid var(--border);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.3);padding:6px 0;z-index:600;margin-top:4px;max-height:70vh;overflow-y:auto}
.nav-dropdown:hover>.nav-menu,.nav-dropdown:focus-within>.nav-menu{display:block}
.nav-menu a{display:flex;padding:8px 16px;color:var(--text-secondary);font-size:12px;border-radius:0;min-height:auto;white-space:nowrap}
.nav-menu a:hover{color:var(--text);background:var(--bg-card)}
.nav-menu .nav-sep{height:1px;background:var(--border);margin:4px 12px}
.full-width-content{padding:32px 40px;max-width:1200px;margin:0 auto}
.exec-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:20px;margin:24px 0}
.exec-panel{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:24px}
.exec-panel h3{font-size:14px;color:var(--text-secondary);margin-bottom:12px;border:none;padding:0}
.category-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:16px;margin:16px 0}
.category-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:20px;text-align:center;transition:all .3s;cursor:pointer}
.category-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md)}
.category-icon{font-size:28px;margin-bottom:6px}
.category-name{font-size:12px;color:var(--text-secondary);margin-bottom:4px}
.category-score{font-size:24px;font-weight:700;font-family:var(--font-mono)}
.category-level{font-size:10px;text-transform:uppercase;font-weight:600;letter-spacing:.5px;margin-top:2px}
.category-findings{font-size:11px;color:var(--text-muted);margin-top:4px}
.arch-tag{display:inline-block;font-size:9px;font-weight:700;letter-spacing:.3px;padding:1px 7px;border-radius:8px;margin-top:4px;text-transform:uppercase}
.arch-tag.arch-new{background:#e6f7ee;color:#107C10;border:1px solid #107C10}
.arch-tag.arch-classic{background:#e8f0fe;color:#0063B1;border:1px solid #0063B1}
.arch-tag.arch-both{background:#f3e8fd;color:#7B68EE;border:1px solid #7B68EE}
.score-display{display:flex;align-items:center;gap:40px;flex-wrap:wrap;margin:20px 0}
.level-badge{display:inline-block;padding:4px 12px;border-radius:6px;font-size:13px;font-weight:700;text-transform:uppercase}
.sev-bars{display:flex;flex-direction:column;gap:8px;margin:12px 0}
.sev-row{display:flex;align-items:center;gap:12px}
.sev-label{width:70px;font-size:12px;text-transform:uppercase;color:var(--text-secondary);font-weight:600}
.sev-track{flex:1;height:10px;background:var(--bar-bg);border-radius:5px;overflow:hidden}
.sev-fill{height:100%;border-radius:5px;transition:width .6s ease}
.sev-count{width:30px;text-align:right;font-family:var(--font-mono);font-size:13px}
.cat-group-header{background:color-mix(in srgb,var(--cat-color) 8%,transparent);border-left:3px solid var(--cat-color);padding:10px 16px;border-radius:6px;margin-top:24px;margin-bottom:0}
.finding-group .cat-toc-table,.finding-group .finding-card,.finding-group .affected-details{margin-left:28px}
.cat-toc-table{width:100%;border-collapse:collapse;margin:10px 0 20px;font-size:12px;background:var(--bg-card);border:1px solid var(--border);border-radius:8px;overflow:hidden}
.cat-toc-table thead th{padding:6px 12px;text-align:left;font-size:11px;font-weight:600;color:var(--text-muted);border-bottom:1px solid var(--border);background:var(--bg-elevated)}
.cat-toc-row td{padding:8px 12px;border-bottom:1px solid var(--border);vertical-align:top}
.cat-toc-row:last-child td{border-bottom:none}
.cat-toc-row:hover{background:var(--bg-card-hover)}
.cat-toc-info{min-width:200px}
.cat-toc-title{color:var(--text);text-decoration:none;font-weight:600;font-size:13px}
.cat-toc-title:hover{color:var(--primary);text-decoration:underline}
.cat-toc-check{white-space:nowrap}
.cat-toc-check code{font-size:10px;padding:2px 6px;border-radius:4px;background:var(--bg-elevated);color:var(--text-secondary)}
.finding-card{background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:14px 20px;margin-bottom:10px;transition:all .2s}
.finding-card:hover{background:var(--bg-card-hover)}
.finding-card:focus{outline:2px solid #0078D4;outline-offset:2px}
.finding-card.critical{border-left:3px solid #D13438}
.finding-card.high{border-left:3px solid #F7630C}
.finding-card.medium{border-left:3px solid #FFB900}
.finding-card.low{border-left:3px solid #107C10}
.finding-title{font-size:14px;font-weight:600;color:var(--text);line-height:1.4;margin-bottom:4px}
.finding-desc{font-size:12px;color:var(--text-secondary);line-height:1.5;margin-bottom:6px}
.aff-count{font-weight:700;font-family:var(--font-mono)}
.remediation-box{margin-top:10px;padding:14px;background:var(--remediation-bg);border-left:3px solid var(--remediation-border);border-radius:6px}
.remediation-box h4{font-size:12px;color:var(--remediation-border);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}
.remediation-box .rem-desc{font-size:13px;color:#A5D6A7;margin-bottom:8px;line-height:1.5}
.remediation-box pre{font-family:var(--font-mono);font-size:12px;background:var(--code-bg);border:1px solid var(--code-border);border-radius:4px;padding:10px;overflow-x:auto;color:var(--text);white-space:pre-wrap}
.remediation-box .portal-steps{margin:6px 0 0;padding-left:20px;font-size:12px;color:var(--text-secondary)}
.remediation-box .portal-steps li{margin-bottom:3px}
.affected-details summary{cursor:pointer;color:var(--primary);font-weight:500;font-size:12px;padding:6px 0}
.resource-table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;margin:8px 0;border:1px solid var(--border);border-radius:8px;overflow:hidden}
.resource-table thead{background:var(--bg-elevated)}
.resource-table th{padding:8px 12px;text-align:left;font-weight:600;color:var(--text-secondary);text-transform:uppercase;font-size:11px;border-bottom:2px solid var(--border)}
.resource-table td{padding:8px 12px;border-bottom:1px solid var(--border-light,var(--border));color:var(--text)}
.resource-table .res-name{font-weight:600}.resource-table .res-type{font-size:11px;color:var(--text-muted)}
.resource-table .res-id,.resource-table a.res-id{font-family:var(--font-mono);font-size:11px;color:var(--text-secondary);word-break:break-all;max-width:320px}
.resource-table a.res-id{color:var(--primary);text-decoration:none}
.resource-table a.res-id:hover{text-decoration:underline}
.resource-table .res-detail{font-size:11px;color:var(--text-secondary)}
.filter-bar{display:flex;align-items:center;gap:8px;margin-bottom:16px;flex-wrap:wrap;font-size:13px}
.filter-bar input[type="search"]{min-width:240px;padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.filter-bar select{padding:8px 12px;border:1px solid var(--border);border-radius:6px;background:var(--bg-elevated);color:var(--text);font-size:13px}
.platform-bar{display:flex;gap:16px;margin:24px 0;flex-wrap:wrap}
.platform-card{background:var(--bg-card);border:1px solid var(--border);border-radius:12px;padding:20px;flex:1;min-width:200px;text-align:center}
.platform-icon{font-size:36px;margin-bottom:8px}
.platform-name{font-size:14px;font-weight:600;margin-bottom:4px}
.platform-count{font-size:24px;font-weight:700;font-family:var(--font-mono)}
.tree-section{margin:24px 0}
.tree-platform{background:color-mix(in srgb,var(--plat-color,#888) 6%,var(--bg-card));border:1px solid color-mix(in srgb,var(--plat-color,#888) 18%,var(--border));border-radius:12px;margin-bottom:12px;overflow:hidden}
.tree-platform[open]>.tree-platform-header{border-bottom:1px solid var(--border)}
.tree-platform-header{display:flex;align-items:center;gap:12px;padding:16px 20px;cursor:pointer;list-style:none;user-select:none;transition:background .2s}
.tree-platform-header:hover{background:color-mix(in srgb,var(--bg-elevated) 80%,var(--primary))}
.tree-platform-header::-webkit-details-marker{display:none}
.tree-platform-header::before{content:'▶';font-size:10px;color:var(--text-muted);transition:transform .2s;flex-shrink:0}
.tree-platform[open]>.tree-platform-header::before{transform:rotate(90deg)}
.tree-plat-icon{font-size:28px;flex-shrink:0}
.tree-plat-info{flex:1;min-width:0}
.tree-plat-name{font-size:15px;font-weight:700;color:var(--text)}
.tree-plat-desc{font-size:11px;color:var(--text-secondary);margin-top:2px}
.tree-plat-stats{display:flex;gap:16px;flex-shrink:0;align-items:center}
.tree-plat-stat{text-align:center}
.tree-plat-stat .val{font-size:20px;font-weight:700;font-family:var(--font-mono)}
.tree-plat-stat .lbl{font-size:10px;color:var(--text-muted);text-transform:uppercase}
.tree-cats{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:12px;padding:16px 20px}
.resource-bar{display:flex;flex-wrap:wrap;gap:6px;padding:10px 20px;border-top:1px solid var(--border);background:var(--bg-elevated)}
.resource-pill{font-size:11px;padding:2px 8px;border-radius:10px;background:var(--bg-card);border:1px solid var(--border);color:var(--text-secondary);white-space:nowrap}
.resource-pill strong{color:var(--text);font-weight:600}
.resource-pill-warn{background:#fff3cd;border-color:#ffc107;color:#856404}
.variant-pill{font-weight:500;background:transparent}

.zoom-controls{display:flex;align-items:center;gap:4px;margin-left:auto}
.zoom-controls button{padding:4px 10px;border:1px solid var(--border);border-radius:4px;background:var(--bg-elevated);color:var(--text);cursor:pointer;font-size:14px;min-height:32px;transition:all .2s}
.zoom-controls button:hover{border-color:var(--primary);color:var(--primary)}
#zoom-label{font-family:var(--font-mono);font-size:12px;min-width:42px;text-align:center;color:var(--text-secondary)}
.sr-only{position:absolute;width:1px;height:1px;padding:0;margin:-1px;overflow:hidden;clip:rect(0,0,0,0);border:0}
#ciq-tooltip{position:fixed;z-index:99999;pointer-events:none;opacity:0;transition:opacity .18s ease;max-width:520px;min-width:220px;padding:14px 18px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  color:var(--text);
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  border-radius:10px;
  font-size:12.5px;line-height:1.6;font-weight:400;text-transform:none;letter-spacing:normal;white-space:normal;
  box-shadow:0 2px 6px rgba(0,0,0,.18),0 8px 24px rgba(0,0,0,.32),0 0 0 1px rgba(255,255,255,.06) inset;
}
#ciq-tooltip.visible{opacity:1}
#ciq-tooltip::before{content:'';position:absolute;width:12px;height:12px;
  background:linear-gradient(145deg,var(--bg-elevated),color-mix(in srgb,var(--bg-elevated) 90%,#000));
  border:1.5px solid color-mix(in srgb,var(--primary) 50%,var(--border));
  transform:rotate(45deg);z-index:-1}
#ciq-tooltip.arrow-bottom::before{bottom:-7px;left:var(--arrow-x,24px);border-top:none;border-left:none}
#ciq-tooltip.arrow-top::before{top:-7px;left:var(--arrow-x,24px);border-bottom:none;border-right:none}
#ciq-tooltip .t-sep{display:block;border-top:1px solid rgba(255,255,255,.15);margin:8px 0 4px;padding-top:6px;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--primary)}
#ciq-tooltip .tip-label{font-weight:600;color:var(--primary)}
@media(max-width:768px){.full-width-content{padding:16px}.exec-grid{grid-template-columns:1fr}.platform-bar{flex-direction:column}}
@media print{.top-nav,.filter-bar,.zoom-controls,.back-to-top{display:none!important}.full-width-content{padding:16px;max-width:100%}body{background:#fff;color:#000;font-size:12px}.finding-card,.category-card,.exec-panel,.stat-card{border:1px solid #ccc;background:#fff}.sev-text{print-color-adjust:exact;-webkit-print-color-adjust:exact}.remediation-box{background:#f1faf1;border-left-color:#107C10}.conf-notice{border:1px solid #999;background:#fff9e6}}
"""


def _as_js() -> str:
    return """
// Zoom
var zoomLevel=100;
function zoomIn(){zoomLevel=Math.min(zoomLevel+10,150);applyZoom()}
function zoomOut(){zoomLevel=Math.max(zoomLevel-10,70);applyZoom()}
function zoomReset(){zoomLevel=100;applyZoom()}
function applyZoom(){document.querySelector('.full-width-content').style.zoom=(zoomLevel/100);document.getElementById('zoom-label').textContent=zoomLevel+'%'}

function filterFindings(){
  var q=(document.getElementById('finding-filter').value||'').toLowerCase();
  var sev=(document.getElementById('filter-severity').value||'').toLowerCase();
  var cat=(document.getElementById('filter-category').value||'').toLowerCase();
  var plat=(document.getElementById('filter-platform').value||'').toLowerCase();
  var shown=0;
  document.querySelectorAll('.finding-card[data-category]').forEach(function(card){
    var match=true;
    if(q&&card.textContent.toLowerCase().indexOf(q)<0)match=false;
    if(sev&&card.getAttribute('data-severity')!==sev)match=false;
    if(cat&&card.getAttribute('data-category')!==cat)match=false;
    if(plat&&card.getAttribute('data-platform')!==plat)match=false;
    card.style.display=match?'':'none';
    if(match)shown++;
  });
  document.querySelectorAll('.finding-group').forEach(function(grp){
    var cards=grp.querySelectorAll('.finding-card[data-category]');
    var vis=0;
    cards.forEach(function(c){if(c.style.display!=='none')vis++;});
    grp.style.display=vis?'':'none';
    var cnt=grp.querySelector('.group-count');
    if(cnt)cnt.textContent=vis;
  });
  var hdr=document.getElementById('findings-heading');
  if(hdr)hdr.innerHTML='&#128270; All Findings ('+shown+')';
  var live=document.getElementById('findings-live');
  if(live)live.textContent='Showing '+shown+' findings';
}

// Keyboard navigation for finding cards
document.addEventListener('keydown',function(e){
  if(e.target.classList.contains('finding-card')){
    var cards=Array.from(document.querySelectorAll('.finding-card[tabindex]')).filter(function(c){return c.style.display!=='none';});
    var idx=cards.indexOf(e.target);
    if(e.key==='ArrowDown'&&idx<cards.length-1){e.preventDefault();cards[idx+1].focus();}
    else if(e.key==='ArrowUp'&&idx>0){e.preventDefault();cards[idx-1].focus();}
    else if(e.key==='Enter'||e.key===' '){e.preventDefault();e.target.classList.toggle('expanded');}
  }
});

// ── Tooltip engine (viewport-aware positioning) ──
(function(){
  var tip=document.getElementById('ciq-tooltip');
  if(!tip)return;
  var GAP=10,MARGIN=12;
  function show(ev){
    var tgt=ev.target.closest('[data-tip]');
    if(!tgt)return;
    var text=tgt.getAttribute('data-tip');
    if(!text)return;
    var d=document.createElement('span');d.textContent=text;var safe=d.innerHTML;
    safe=safe.replace(/\\n/g,'<br>');
    safe=safe.replace(/<br>\u2022 /g,'</p><p style="margin:2px 0 2px 12px;text-indent:-10px">\u2022 ');
    if(safe.indexOf('\u2022 ')===0){safe='<p style="margin:2px 0 2px 12px;text-indent:-10px">'+safe.substring(0);}
    safe=safe.replace(/(Checks:|Evaluates:|Zero findings:)/g,'<span class="tip-label">$1</span>');
    tip.innerHTML=safe;
    tip.classList.add('visible');
    tip.setAttribute('aria-hidden','false');
    requestAnimationFrame(function(){
      var r=tgt.getBoundingClientRect();
      var tw=tip.offsetWidth,th=tip.offsetHeight;
      var vw=window.innerWidth,vh=window.innerHeight;
      var above=r.top-GAP-th;
      var below=r.bottom+GAP;
      var top,arrow;
      if(above>=MARGIN){top=above;arrow='arrow-bottom';}
      else if(below+th<=vh-MARGIN){top=below;arrow='arrow-top';}
      else{top=Math.max(MARGIN,vh-th-MARGIN);arrow='';}
      var left=r.left+r.width/2-tw/2;
      left=Math.max(MARGIN,Math.min(left,vw-tw-MARGIN));
      var arrowX=r.left+r.width/2-left;
      arrowX=Math.max(16,Math.min(arrowX,tw-16));
      tip.style.top=top+'px';
      tip.style.left=left+'px';
      tip.style.setProperty('--arrow-x',arrowX+'px');
      tip.className='visible'+(arrow?' '+arrow:'');
    });
  }
  function hide(){
    tip.classList.remove('visible');
    tip.setAttribute('aria-hidden','true');
    tip.className='';
  }
  document.addEventListener('mouseenter',show,true);
  document.addEventListener('mouseleave',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
  document.addEventListener('focusin',show,true);
  document.addEventListener('focusout',function(ev){
    if(ev.target.closest('[data-tip]'))hide();
  },true);
})();

// ── Category card double-click → jump to findings section ──
document.addEventListener('dblclick',function(ev){
  var card=ev.target.closest('.category-card[data-cat]');
  if(!card)return;
  var cat=card.getAttribute('data-cat');
  var target=document.getElementById('cat-'+cat);
  if(target){target.scrollIntoView({behavior:'smooth',block:'start'});}
});

// SHA-256 report hash
(function(){
  var el=document.getElementById('report-hash');
  if(!el||typeof crypto==='undefined'||!crypto.subtle)return;
  var html=document.documentElement.outerHTML;
  var buf=new TextEncoder().encode(html);
  crypto.subtle.digest('SHA-256',buf).then(function(h){
    var a=Array.from(new Uint8Array(h));
    el.textContent=a.map(function(b){return b.toString(16).padStart(2,'0')}).join('');
  });
})();
"""


# ── Main generator ───────────────────────────────────────────────────────

def generate_ai_agent_security_report(results: dict, output_dir: str | pathlib.Path) -> pathlib.Path:
    """Generate the AI Agent Security Assessment HTML report."""
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "ai-agent-security-assessment.html"

    ts = format_date_short()
    report_id = str(uuid.uuid4())
    scores = results.get("AgentSecurityScores", {})
    findings = results.get("Findings", [])
    finding_count = results.get("FindingCount", len(findings))
    sub_count = results.get("SubscriptionCount", 0)
    assessed_at = results.get("AssessedAt", "")
    tenant_id = results.get("TenantId", "")
    tenant_display = results.get("TenantDisplay", "")
    evidence_summary = results.get("EvidenceSummary", {})

    overall_score = scores.get("OverallScore", 0)
    overall_level = scores.get("OverallLevel", "secure")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})
    plat_dist = scores.get("PlatformBreakdown", {})

    n_crit = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_med = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)

    lmeta = _LEVEL_META.get(overall_level, _LEVEL_META["medium"])
    level_color = lmeta["color"]

    score_ring = _ring_score_svg(overall_score, size=160)
    sev_donut = _donut_svg([
        ("Critical", n_crit, "#D13438"), ("High", n_high, "#F7630C"),
        ("Medium", n_med, "#FFB900"), ("Low", n_low, "#107C10"), ("Info", n_info, "#A8A6A3"),
    ], size=140)

    max_sev = max(n_crit, n_high, n_med, n_low, n_info, 1)
    sev_bars = ""
    for name, count, color in [("Critical", n_crit, "#D13438"), ("High", n_high, "#F7630C"),
                                ("Medium", n_med, "#FFB900"), ("Low", n_low, "#107C10"), ("Info", n_info, "#A8A6A3")]:
        pct = (count / max_sev) * 100 if max_sev > 0 else 0
        weight = 10 if name == "Critical" else 7.5 if name == "High" else 5 if name == "Medium" else 2.5 if name == "Low" else 1
        contrib = count * weight
        sev_tip = f'{name} severity findings.\n• {count} of {finding_count} total ({count*100//max(finding_count,1)}%)\n• Weight: {weight} points per finding\n• Risk contribution: {contrib:.0f} points total'
        sev_bars += (f'<div class="sev-row"><span class="sev-label" data-tip="{sev_tip}">{name}</span>'
                     f'<div class="sev-track"><div class="sev-fill" style="width:{pct:.0f}%;background:{color}"></div></div>'
                     f'<span class="sev-count">{count}</span></div>')

    # Hierarchical Platform → Category tree
    tree_html = ""
    total_plat_findings = sum(plat_dist.get(k, 0) for k in _PLATFORM_META)
    for plat_key, pmeta in _PLATFORM_META.items():
        plat_count = plat_dist.get(plat_key, 0)
        cat_keys = _PLATFORM_CATEGORIES.get(plat_key, [])
        # Aggregate platform-level score from child categories
        plat_cat_scores = [cat_scores.get(ck, {"Score": 0, "FindingCount": 0}) for ck in cat_keys]
        cats_with_findings = [s for s in plat_cat_scores if s.get("FindingCount", 0) > 0]
        plat_avg_score = (sum(s.get("Score", 0) for s in cats_with_findings) / len(cats_with_findings)) if cats_with_findings else 0
        plat_level = "critical" if plat_avg_score >= 75 else "high" if plat_avg_score >= 50 else "medium" if plat_avg_score >= 25 else "secure"
        plat_level_color = _LEVEL_META.get(plat_level, _LEVEL_META["secure"])["color"]
        plat_pct = (plat_count * 100 // max(total_plat_findings, 1))
        is_top = plat_count == max(plat_dist.values(), default=0) and plat_count > 0
        plat_tip = f'{esc(pmeta["name"])} — {esc(pmeta.get("desc", ""))}'
        plat_tip += f'\n• {plat_count} finding{"s" if plat_count != 1 else ""} ({plat_pct}% of {total_plat_findings} total)'
        plat_tip += f'\n• {len(cat_keys)} security categories under this platform'
        if is_top and plat_count > 0:
            plat_tip += f'\n• Highest-contributing platform — prioritize remediation here'
        elif plat_count == 0:
            plat_tip += f'\n• No issues detected in the current assessment'
        plat_tip += f'\n• Click to expand and see category details'

        # Build child category cards
        child_cards = ""
        for cat_key in cat_keys:
            meta = _CATEGORY_META.get(cat_key, {})
            if not meta:
                continue
            cs = cat_scores.get(cat_key, {"Score": 0, "Level": "secure", "FindingCount": 0})
            c_score = cs.get("Score", 0)
            c_level = cs.get("Level", "secure")
            c_count = cs.get("FindingCount", 0)
            c_color = _LEVEL_META.get(c_level, _LEVEL_META["secure"])["color"]
            cat_tip = f'{esc(meta["name"])} — {esc(meta["description"])}'
            cat_tip += f'\n• Score: {c_score:.0f}/100 ({c_level.upper()})'
            cat_tip += f'\n• {c_count} finding{"s" if c_count != 1 else ""} detected'
            cat_tip += f'\n• Platform: {esc(pmeta["name"])}'
            if c_count > 0:
                cat_tip += f'\n• Double-click this card to jump to its findings below'
            else:
                cat_tip += f'\n• No issues — this domain is well-configured'
            if meta.get("tooltip"):
                cat_tip += f'\n\n{esc(meta["tooltip"])}'
            # Architecture badge for Foundry categories
            arch_badge = ""
            arch_val = meta.get("arch", "")
            if arch_val == "new":
                arch_badge = '<div class="arch-tag arch-new">New</div>'
            elif arch_val == "classic":
                arch_badge = '<div class="arch-tag arch-classic">Classic</div>'
            elif arch_val == "classic & new":
                arch_badge = '<div class="arch-tag arch-both">Classic &amp; New</div>'
            child_cards += (
                f'<div class="category-card" data-tip="{cat_tip}" data-cat="{esc(cat_key)}" tabindex="0">'
                f'<div class="category-icon">{meta["icon"]}</div>'
                f'<div class="category-name">{esc(meta["name"])}</div>'
                f'{arch_badge}'
                f'<div class="category-score" style="color:{c_color}">{c_score:.0f}</div>'
                f'<div class="category-level" style="color:{c_color}">{esc(c_level.upper())}</div>'
                f'<div class="category-findings">{c_count} finding{"s" if c_count != 1 else ""}</div></div>'
            )

        open_attr = " open" if plat_count > 0 else ""
        # Build resource summary bar for this platform
        plat_evidence = evidence_summary.get(plat_key, {})
        resource_bar = ""
        if plat_evidence:
            # Variant breakdown pills (Classic vs New)
            variant_pills = ""
            _VARIANT_LABELS = {
                "ClassicHubs": ("Classic Hubs", "#6c757d"),
                "ClassicProjects": ("Classic Projects", "#6c757d"),
                "FoundryProjects": ("New Foundry Projects", "#0078D4"),
                "StandaloneML": ("Standalone ML", "#495057"),
            }
            for vk, (vlabel, vcolor) in _VARIANT_LABELS.items():
                vc = plat_evidence.get(vk, 0)
                if vc:
                    variant_pills += (
                        f'<span class="resource-pill variant-pill" '
                        f'style="border-color:{vcolor};color:{vcolor}">'
                        f'{esc(vlabel)}: <strong>{vc}</strong></span>'
                    )

            # Resource count pills (skip variant keys and access denied)
            _SKIP_KEYS = {"AccessDeniedErrors", "ClassicHubs", "ClassicProjects", "FoundryProjects", "StandaloneML"}
            resource_pills = "".join(
                f'<span class="resource-pill">{esc(k)}: <strong>{v}</strong></span>'
                for k, v in plat_evidence.items() if v and k not in _SKIP_KEYS
            )

            ad_count = plat_evidence.get("AccessDeniedErrors", 0)
            ad_pill = (
                f'<span class="resource-pill resource-pill-warn" data-tip="Some API calls were denied due to missing permissions. Results may be incomplete.">'
                f'&#9888; {ad_count} access denied</span>'
            ) if ad_count else ""

            all_pills = variant_pills + resource_pills + ad_pill
            if all_pills:
                resource_bar = f'<div class="resource-bar">{all_pills}</div>'

        tree_html += (
            f'<details class="tree-platform" style="--plat-color:{pmeta["color"]}" data-tip="{plat_tip}"{open_attr}>'
            f'<summary class="tree-platform-header">'
            f'<span class="tree-plat-icon">{pmeta["icon"]}</span>'
            f'<div class="tree-plat-info">'
            f'<div class="tree-plat-name" style="color:{pmeta["color"]}">{esc(pmeta["name"])}</div>'
            f'<div class="tree-plat-desc">{esc(pmeta.get("desc", ""))}</div>'
            f'</div>'
            f'<div class="tree-plat-stats">'
            f'<div class="tree-plat-stat"><div class="val" style="color:{plat_level_color}">{plat_avg_score:.0f}</div><div class="lbl">Score</div></div>'
            f'<div class="tree-plat-stat"><div class="val" style="color:{pmeta["color"]}">{plat_count}</div><div class="lbl">Findings</div></div>'
            f'<div class="tree-plat-stat"><div class="val">{len(cat_keys)}</div><div class="lbl">Categories</div></div>'
            f'</div>'
            f'</summary>'
            f'{resource_bar}'
            f'<div class="tree-cats">{child_cards}</div>'
            f'</details>'
        )

    # Findings grouped by category
    findings_by_cat: dict[str, list[dict]] = {}
    for f in findings:
        findings_by_cat.setdefault(f.get("Category", "unknown"), []).append(f)

    all_findings_html = ""
    cat_options = ""
    plat_options = ""
    findings_nav_items = ""
    seen_plats: set[str] = set()
    finding_counter = 0
    for cat_key in _CATEGORY_META:
        cat_findings = findings_by_cat.get(cat_key, [])
        if not cat_findings:
            continue
        meta = _CATEGORY_META[cat_key]
        cat_options += f'<option value="{esc(cat_key)}">{esc(meta["name"])} ({len(cat_findings)})</option>'
        findings_nav_items += f'      <a href="#cat-{esc(cat_key)}">{meta["icon"]} {esc(meta["name"])} ({len(cat_findings)})</a>\n'
        for f in cat_findings:
            p = f.get("Platform", "cross-cutting")
            if p not in seen_plats:
                seen_plats.add(p)
                pmeta = _PLATFORM_META.get(p, {"name": p})
                plat_options += f'<option value="{esc(p)}">{esc(pmeta["name"])}</option>'
        sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
        sorted_f = sorted(cat_findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))
        cat_color = meta.get("color", "#0078D4")
        all_findings_html += f'<div class="finding-group" data-group-cat="{esc(cat_key)}">'
        all_findings_html += (
            f'<h3 id="cat-{esc(cat_key)}" class="cat-group-header" '
            f'style="--cat-color:{cat_color}">{meta["icon"]} {esc(meta["name"])} '
            f'&mdash; <span class="group-count">{len(cat_findings)}</span> '
            f'finding{"s" if len(cat_findings) != 1 else ""}</h3>'
        )
        # Mini-index: compact table of distinct findings
        toc_rows = ""
        for fi, sf in enumerate(sorted_f):
            title_f = sf.get("Title", "Untitled")
            subcat_f = sf.get("Subcategory", "").replace("_", " ").title()
            anchor = f"finding-{esc(cat_key)}-{finding_counter + fi + 1}"
            toc_rows += (
                f'<tr class="cat-toc-row">'
                f'<td class="cat-toc-info">'
                f'<a href="#{anchor}" class="cat-toc-title">{esc(title_f)}</a>'
                f'</td>'
                f'<td class="cat-toc-check"><code>{esc(subcat_f)}</code></td>'
                f'</tr>'
            )
        all_findings_html += (
            f'<table class="cat-toc-table"><thead><tr>'
            f'<th>Finding</th><th>Check</th>'
            f'</tr></thead><tbody>{toc_rows}</tbody></table>'
        )
        for f in sorted_f:
            finding_counter += 1
            all_findings_html += _render_finding(f, finding_counter)
        all_findings_html += '</div>'

    # Evidence record count
    evidence_record_count = results.get("EvidenceCount", 0)

    # Assessment domain names (abbreviated)
    domain_names = ", ".join(m["name"] for m in list(_CATEGORY_META.values())[:10]) + f" (+{len(_CATEGORY_META) - 10} more)"

    # Executive summary
    if overall_level == "critical":
        exec_text = f"AI agent security posture is <strong>critically exposed</strong> with {n_crit} critical findings."
    elif overall_level == "high":
        exec_text = f"AI agent security is at <strong>elevated risk</strong> — {n_crit} critical and {n_high} high findings."
    elif overall_level == "medium":
        exec_text = f"AI agent security has <strong>moderate gaps</strong> requiring attention."
    else:
        exec_text = "AI agents are <strong>well-secured</strong>. No critical gaps found."

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Agent Security — EnterpriseSecurityIQ</title>
<style>{get_css()}{_as_css()}</style>
</head>
<body>
<a href="#main" class="skip-nav">Skip to content</a>

<nav class="top-nav" aria-label="Report sections">
  <span class="brand" data-tip="EnterpriseSecurityIQ AI Agent Security Assessment.\n• Automated security posture analysis for AI agents\n• Covers Copilot Studio, Microsoft Foundry, and custom agents\n• Read-only: no changes are made to your environment">&#129504; EnterpriseSecurityIQ AI Agent Security</span>
  <div class="nav-dropdown">
    <button class="nav-toggle" data-tip="Report metadata, confidentiality notice, and audit attestation.">Document</button>
    <div class="nav-menu">
      <a href="#doc-control" data-tip="Report ID, assessment scope, classification, and data integrity details.">Report Metadata</a>
      <a href="#doc-control" data-tip="Data handling and distribution restrictions for this report." onclick="setTimeout(function(){{document.querySelector('#doc-control .conf-notice').scrollIntoView({{behavior:'smooth'}})}},50)">Confidentiality Notice</a>
      <a href="#doc-control" data-tip="Attestation that this report was generated from read-only API data." onclick="setTimeout(function(){{document.querySelector('#doc-control h3').scrollIntoView({{behavior:'smooth'}})}},50)">Audit Attestation</a>
    </div>
  </div>
  <a href="#summary" data-tip="Executive summary: risk score, severity distribution, and stat cards.">Summary</a>
  <a href="#platforms" data-tip="Hierarchical view of platforms and their security categories.\n• Click a platform to expand its child categories\n• Double-click a category card to jump to its findings">Platforms &amp; Categories</a>
  <div class="nav-dropdown">
    <button class="nav-toggle" data-tip="Detailed finding cards grouped by security category.">All Findings</button>
    <div class="nav-menu">
      <a href="#findings" data-tip="View all findings with search, severity, platform, and category filters.">All Findings</a>
      <div class="nav-sep"></div>
{findings_nav_items}    </div>
  </div>
  <div class="zoom-controls" aria-label="Page zoom">
    <button onclick="zoomOut()" aria-label="Zoom out" data-tip="Decrease page zoom by 10%. Useful for seeing more of the report at once. Current zoom is shown between the +/- buttons.">&minus;</button>
    <span id="zoom-label">100%</span>
    <button onclick="zoomIn()" aria-label="Zoom in" data-tip="Increase page zoom by 10%. Useful for reading remediation commands and resource IDs in detail.">&plus;</button>
    <button onclick="zoomReset()" aria-label="Reset zoom" data-tip="Reset zoom back to 100% default." style="font-size:11px">Reset</button>
  </div>
  <button class="theme-btn" onclick="toggleTheme()" style="margin:0;padding:6px 14px"
          aria-label="Toggle dark and light theme" data-tip="Switch between dark and light viewing mode.\n• Dark mode reduces eye strain\n• Light mode is better for printing">Switch to Light</button>
</nav>

<main id="main" class="full-width-content">

<!-- ── Document Control ── -->
<section id="doc-control" class="section">
  <h1 class="page-title">EnterpriseSecurityIQ &mdash; AI Agent Security Assessment Report</h1>
  <table class="doc-control-table">
    <tr><th data-tip="Unique identifier for this report.
• Auto-generated UUID v4 assigned at creation time
• Use this ID in audit trails, remediation tickets, and ITSM workflows
• Ensures all stakeholders reference the same assessment">Report Identifier</th><td>{esc(report_id)}</td></tr>
    <tr><th data-tip="Assessment module that produced this report.
• AI Agent Security module of EnterpriseSecurityIQ
• Evaluates Copilot Studio, Microsoft Foundry, Azure AI, and custom agents
• Covers {len(_CATEGORY_META)} security domains">Assessment Name</th><td>EnterpriseSecurityIQ AI Agent Security Assessment</td></tr>
    <tr><th data-tip="When this report was generated.
• This is a point-in-time snapshot
• Your environment may have changed since then (especially after remediation)
• Compare with Assessment Period to understand the data collection window">Date Generated</th><td>{esc(ts)}</td></tr>
    <tr><th data-tip="Microsoft Entra ID (Azure AD) tenant that was assessed.
• Resolved from the credential used to run the assessment
• In multi-tenant orgs, verify this GUID matches your intended target">Tenant ID</th><td><code>{esc(tenant_id) if tenant_id else 'N/A'}</code></td></tr>
    <tr><th data-tip="Display name of the Entra ID tenant.
• Retrieved from Microsoft Graph API
• Confirms which organization this report covers">Tenant Name</th><td>{esc(tenant_display) if tenant_display else 'Unknown'}</td></tr>
    <tr><th data-tip="Security domains evaluated in this assessment.
• {len(_CATEGORY_META)} domains (e.g., authentication, network isolation, DLP, threat protection)
• Each domain maps to a set of analyzers
• Checks controls across Copilot Studio, Foundry, and Azure AI">Assessment Domains</th><td>{esc(domain_names)}</td></tr>
    <tr><th data-tip="Data classification level for this report.
• Contains sensitive details: resource IDs, tenant info, vulnerability descriptions
• Handle per your organization information protection policy
• Restrict distribution to authorized security personnel only">Classification</th><td>CONFIDENTIAL — Authorized Recipients Only</td></tr>
    <tr><th data-tip="Tool version that produced this report.
• EnterpriseSecurityIQ AI Agent v{VERSION}
• Version tracking ensures reproducibility
• Re-running with the same version guarantees consistent rules and scoring">Tool</th><td>EnterpriseSecurityIQ AI Agent v{VERSION}</td></tr>
    <tr><th data-tip="How evidence was collected from your environment.
• Power Platform Admin APIs (Copilot Studio)
• Azure AI Services REST APIs (Foundry/OpenAI)
• Azure Resource Manager (custom agents)
• All calls are read-only (GET only) — no tenant modifications">Collection Method</th><td>Power Platform Admin API + Azure AI Services REST API + Azure Resource Manager (Read-Only)</td></tr>
  </table>
  <div class="conf-notice">
    <strong>CONFIDENTIALITY NOTICE:</strong> This document contains sensitive security and compliance
    information about the assessed environment. Distribution is restricted to authorized personnel only.
  </div>
  <h3>Audit Attestation</h3>
  <table class="doc-control-table">
    <tr><th data-tip="Scope of this assessment.
• {sub_count} subscription(s) scanned
• {len(_CATEGORY_META)} security domains across {len(_PLATFORM_META)} platforms
• If you expected more subscriptions, verify credential has Reader access">Assessment Scope</th><td>AI agent security posture analysis across {sub_count} subscription(s) covering {len(_CATEGORY_META)} security domains and {len(_PLATFORM_META)} platforms</td></tr>
    <tr><th data-tip="Data integrity confirmation.
• All evidence collected via read-only (GET) API calls
• No write, delete, or update operations executed
• Azure Resource Manager, Power Platform Admin API, and AI Services REST endpoints queried without modification">Data Integrity</th><td>All evidence collected via read-only API calls; no tenant modifications were made</td></tr>
    <tr><th data-tip="Configuration records gathered during the assessment.
• {evidence_record_count:,} records collected and evaluated
• Each record = one API response (e.g., a bot, a deployment, a content filter)
• More records = broader inspection coverage">Evidence Records</th><td>{evidence_record_count:,} records collected and evaluated</td></tr>
    <tr><th data-tip="Tamper detection hash for this report.
• SHA-256 hash of the complete report HTML
• If the hash changes, the report has been modified
• Computed client-side at render time
• Save this hash for audit and legal defensibility">Report Hash (SHA-256)</th><td><code id="report-hash">Computed at render</code></td></tr>
    <tr><th data-tip="When evidence was collected from your environment.
• This is a point-in-time snapshot
• Changes after this date (including remediation) are NOT reflected
• Run a new assessment to capture current state">Assessment Period</th><td>{esc(assessed_at[:19]) if assessed_at else esc(ts)} (point-in-time snapshot)</td></tr>
  </table>
</section>

<section id="summary" class="section" aria-labelledby="summary-heading">
  <h2 id="summary-heading">&#128202; Executive Summary</h2>
  <p style="color:var(--text-secondary);font-size:14px;line-height:1.6;max-width:960px;margin:8px 0 20px">
    Security posture assessment of AI agents across Copilot Studio, Microsoft Foundry (Azure AI/OpenAI),
    Entra identity, AI infrastructure, agent orchestration, and custom agent deployments.
    Covers authentication, network isolation, content safety, identity management, threat protection,
    data connectors, diagnostics, model governance, and content leakage prevention.
  </p>

  <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin-bottom:24px;font-size:14px;line-height:1.65">
    {exec_text} Overall score: <strong>{overall_score:.0f}/100</strong> (higher = more risk).
  </div>

  <div class="stat-grid" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:12px;margin:24px 0">
    <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center" data-tip="Overall risk score for your AI agent environment.
• Score: {overall_score:.0f}/100 ({overall_level.upper()})
• Breakdown: {n_crit} critical, {n_high} high, {n_med} medium, {n_low} low
• Scoring: Critical=10, High=7.5, Medium=5, Low=2.5 points each
• Thresholds: 75+ Critical, 50-74 High, 25-49 Medium, <25 Secure
• Quick win: fixing {n_crit} critical findings would lower score by ~{n_crit * 10} points" tabindex="0">
      <div style="font-size:28px;font-weight:700;color:{level_color};font-family:var(--font-mono)">{overall_score:.0f}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Risk Score /100</div>
    </div>
    <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center" data-tip="Total security findings discovered in this assessment.
• {finding_count} findings across {len(cat_scores)} categories and {len(seen_plats)} platforms
• Each finding = a missing or misconfigured security control
• Filter by severity, platform, or category in All Findings below" tabindex="0">
      <div style="font-size:28px;font-weight:700;font-family:var(--font-mono)">{finding_count}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Findings</div>
    </div>
    <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center" data-tip="Critical severity — the most urgent security gaps.
• {n_crit} critical findings detected
• Examples: unauthenticated agents, public endpoints, missing encryption
• Each adds 10 points to the risk score ({n_crit * 10} points total)
• Remediate first using step-by-step instructions in each finding card" tabindex="0">
      <div style="font-size:28px;font-weight:700;color:#D13438;font-family:var(--font-mono)">{n_crit}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Critical</div>
    </div>
    <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center" data-tip="High severity — significant risks needing prompt attention.
• {n_high} high-severity findings detected
• Examples: missing DLP policies, disabled Defender, absent logging
• Each adds 7.5 points to the risk score ({n_high * 7.5:.0f} points total)
• Plan remediation within your next maintenance window" tabindex="0">
      <div style="font-size:28px;font-weight:700;color:#F7630C;font-family:var(--font-mono)">{n_high}</div>
      <div style="font-size:12px;color:var(--text-secondary)">High</div>
    </div>
    <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center" data-tip="Azure subscriptions scanned in this assessment.
• {sub_count} subscription(s) enumerated via Azure Resource Manager
• Checked for AI services, Copilot Studio environments, and agent configs
• If you expected more, verify the credential has Reader access to all targets" tabindex="0">
      <div style="font-size:28px;font-weight:700;font-family:var(--font-mono)">{sub_count}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Subscriptions</div>
    </div>
    <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:8px;padding:16px;text-align:center" data-tip="Security categories evaluated in this assessment.
• {len(cat_scores)} categories covering authentication, network isolation, DLP, threat protection, model governance, content safety, and more
• Each category groups related security checks
• See Category Breakdown section for per-category scores and finding counts" tabindex="0">
      <div style="font-size:28px;font-weight:700;font-family:var(--font-mono)">{len(cat_scores)}</div>
      <div style="font-size:12px;color:var(--text-secondary)">Categories</div>
    </div>
  </div>

  <div class="exec-grid">
    <div class="exec-panel" data-tip="Ring chart showing the overall risk score for your AI agent environment.
• Your score: {overall_score:.0f}/100 ({overall_level.upper()})
• Red arc (≥75) = Critical, Orange (50-74) = High, Yellow (25-49) = Medium, Green (<25) = Secure
• Arc length is proportional to the score
• Compare this ring across successive assessments to track improvement">
      <h3>Security Score</h3>
      <div class="score-display" style="justify-content:center">{score_ring}</div>
      <div style="display:flex;flex-wrap:wrap;justify-content:center;gap:8px;margin-top:10px;font-size:11px;color:var(--text-secondary)">
        <span data-tip="Score 75-100: Critical risk. Immediate action required — multiple high-impact gaps."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#D13438;vertical-align:middle"></span> Critical ≥75</span>
        <span data-tip="Score 50-74: High risk. Elevated exposure with significant security gaps."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#F7630C;vertical-align:middle"></span> High 50–74</span>
        <span data-tip="Score 25-49: Medium risk. Moderate gaps that should be addressed soon."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#FFB900;vertical-align:middle"></span> Medium 25–49</span>
        <span data-tip="Score 0-24: Secure. AI agents are well-configured with no critical gaps."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#107C10;vertical-align:middle"></span> Secure &lt;25</span>
      </div>
    </div>
    <div class="exec-panel" data-tip="Donut chart showing how {finding_count} findings are distributed by severity.
• Red = Critical ({n_crit}), Orange = High ({n_high}), Yellow = Medium ({n_med}), Green = Low ({n_low}), Gray = Info ({n_info})
• Each slice is proportional to its count
• Center number = {finding_count} total findings
• Mostly red/orange = urgent remediation needed">
      <h3>Severity Distribution</h3>
      <div style="text-align:center">{sev_donut}</div>
      <div style="display:flex;flex-wrap:wrap;justify-content:center;gap:8px;margin-top:10px;font-size:11px;color:var(--text-secondary)">
        <span data-tip="Critical findings: {n_crit}. Weight: 10 points each ({n_crit * 10} total). Fix these first."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#D13438;vertical-align:middle"></span> Critical ({n_crit})</span>
        <span data-tip="High findings: {n_high}. Weight: 7.5 points each ({n_high * 7.5:.0f} total). Plan prompt remediation."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#F7630C;vertical-align:middle"></span> High ({n_high})</span>
        <span data-tip="Medium findings: {n_med}. Weight: 5 points each ({n_med * 5} total). Schedule for backlog."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#FFB900;vertical-align:middle"></span> Medium ({n_med})</span>
        <span data-tip="Low findings: {n_low}. Weight: 2.5 points each ({n_low * 2.5:.0f} total). Best-practice improvements."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#107C10;vertical-align:middle"></span> Low ({n_low})</span>
        <span data-tip="Informational observations: {n_info}. Weight: 1 point each ({n_info} total). For awareness only."><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#A8A6A3;vertical-align:middle"></span> Info ({n_info})</span>
      </div>
    </div>
    <div class="exec-panel" data-tip="Horizontal bars showing exact finding counts per severity.
• Critical: {n_crit}, High: {n_high}, Medium: {n_med}, Low: {n_low}, Info: {n_info}
• Bar width is proportional to count — widest = most common severity
• Weight per finding: Critical=10, High=7.5, Medium=5, Low=2.5, Info=1">
      <h3>Severity Breakdown</h3>
      <div class="sev-bars">{sev_bars}</div>
      <div style="font-size:11px;color:var(--text-secondary);margin-top:8px;text-align:center">Bar width proportional to count. Weight: Critical=10, High=7.5, Medium=5, Low=2.5, Info=1 points per finding.</div>
    </div>
  </div>
</section>

<section id="platforms" class="section" aria-labelledby="platforms-heading">
  <h2 id="platforms-heading">&#128640; Platform &amp; Category Breakdown</h2>
  <p style="color:var(--text-secondary);font-size:12px;margin-bottom:12px">Click a platform to expand its security categories. Double-click a category card to jump to its findings.</p>
  <div class="tree-section">{tree_html}</div>
</section>

<section id="findings" class="section" aria-labelledby="findings-heading">
  <h2 id="findings-heading">&#128270; All Findings ({finding_count})</h2>
  <div class="filter-bar" role="search" aria-label="Filter findings">
    <label for="finding-filter" data-tip="Search all findings by keyword.
• Searches title, description, remediation steps, and resource names
• Results filter in real-time as you type
• Clear the field to show all {finding_count} findings">Search:</label>
    <input id="finding-filter" type="search" placeholder="Search…" oninput="filterFindings()">
    <label for="filter-severity" data-tip="Filter findings by severity level.
• {n_crit} Critical, {n_high} High, {n_med} Medium, {n_low} Low
• Select a level to focus on the most impactful findings first">Severity:</label>
    <select id="filter-severity" onchange="filterFindings()">
      <option value="">All</option>
      <option value="critical">Critical</option><option value="high">High</option>
      <option value="medium">Medium</option><option value="low">Low</option>
    </select>
    <label for="filter-platform" data-tip="Filter findings by technology platform.
• {len(seen_plats)} platform(s) have findings in this assessment
• Useful for assigning remediation to platform-specific teams">Platform:</label>
    <select id="filter-platform" onchange="filterFindings()">
      <option value="">All</option>{plat_options}
    </select>
    <label for="filter-category" data-tip="Filter findings by security category.
• {len(cat_scores)} categories were evaluated
• Select a category to view all related findings together">Category:</label>
    <select id="filter-category" onchange="filterFindings()">
      <option value="">All</option>{cat_options}
    </select>
  </div>
  <div id="findings-live" class="sr-only" aria-live="polite" aria-atomic="true"></div>
  {all_findings_html if all_findings_html else '<p>No findings — AI agents are secure!</p>'}
</section>

</main>

<div id="ciq-tooltip" role="tooltip" aria-hidden="true"></div>
<button class="back-to-top" aria-label="Back to top" data-tip="Scroll back to the top of the report.">&#8593;</button>

<script>{get_js()}</script>
<script>{_as_js()}</script>
</body>
</html>"""

    out_path.write_text(html, encoding="utf-8")
    log.info("[AIAgentSecurityReport] Written to %s (%d KB)", out_path, len(html) // 1024)
    return out_path


# ── Excel styling constants ──────────────────────────────────────────────

_XL_HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
_XL_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_XL_SECTION_FONT = Font(name="Segoe UI", bold=True, size=12, color="0078D4")
_XL_LABEL_FONT = Font(name="Segoe UI", bold=True, size=11)
_XL_VALUE_FONT = Font(name="Segoe UI", size=11)
_XL_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_XL_SEV_FILLS = {
    "critical": PatternFill(start_color="D13438", end_color="D13438", fill_type="solid"),
    "high": PatternFill(start_color="F7630C", end_color="F7630C", fill_type="solid"),
    "medium": PatternFill(start_color="FFB900", end_color="FFB900", fill_type="solid"),
    "low": PatternFill(start_color="107C10", end_color="107C10", fill_type="solid"),
    "informational": PatternFill(start_color="A8A6A3", end_color="A8A6A3", fill_type="solid"),
}


def _xl_apply_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _XL_HEADER_FONT
        cell.fill = _XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _XL_THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _xl_auto_width(ws, min_w: int = 10, max_w: int = 50):
    for col in ws.columns:
        length = min_w
        for cell in col:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


def generate_ai_agent_security_excel(results: dict, output_dir: str | pathlib.Path) -> pathlib.Path:
    """Generate a comprehensive multi-tab AI Agent Security Excel workbook.

    Sheets: Executive Summary, All Findings, Affected Resources,
    Category Scores, Platform Breakdown, Remediation.
    """
    out_dir = pathlib.Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "ai-agent-security.xlsx"

    scores = results.get("AgentSecurityScores", {})
    findings = results.get("Findings", [])
    overall = scores.get("OverallScore", 0)
    overall_level = scores.get("OverallLevel", "secure")
    sev_dist = scores.get("SeverityDistribution", {})
    cat_scores = scores.get("CategoryScores", {})
    plat_dist = scores.get("PlatformBreakdown", {})
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}

    level = (
        "Critical" if overall >= 75 else
        "High" if overall >= 50 else
        "Medium" if overall >= 25 else
        "Low"
    )

    n_critical = sev_dist.get("critical", 0)
    n_high = sev_dist.get("high", 0)
    n_medium = sev_dist.get("medium", 0)
    n_low = sev_dist.get("low", 0)
    n_info = sev_dist.get("informational", 0)

    assessed_at = results.get("AssessedAt", "")
    ts_now = assessed_at if assessed_at else datetime.now().strftime("%Y-%m-%d %H:%M")
    if "T" in ts_now:
        ts_now = ts_now[:16].replace("T", " ")

    wb = Workbook()

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"

    rows = [
        ("EnterpriseSecurityIQ — AI Agent Security Assessment", ""),
        ("", ""),
        ("Assessment Date", ts_now),
        ("Tenant ID", results.get("TenantId", "N/A")),
        ("Subscriptions", results.get("SubscriptionCount", 0)),
        ("Tool Version", f"EnterpriseSecurityIQ v{VERSION}"),
        ("", ""),
        ("RISK OVERVIEW", ""),
        ("Overall Risk Score", f"{overall:.0f}/100"),
        ("Risk Level", level),
        ("Total Findings", len(findings)),
        ("Critical", n_critical),
        ("High", n_high),
        ("Medium", n_medium),
        ("Low", n_low),
        ("Informational", n_info),
        ("Categories Evaluated", len(cat_scores)),
        ("Platforms Assessed", len(plat_dist)),
    ]
    for r, (label, value) in enumerate(rows, 1):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        if r == 1:
            c1.font = _XL_SECTION_FONT
        elif label in ("RISK OVERVIEW",):
            c1.font = _XL_SECTION_FONT
        else:
            c1.font = _XL_LABEL_FONT
            c2.font = _XL_VALUE_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # ── Sheet 2: All Findings ──
    ws2 = wb.create_sheet("All Findings")
    headers = ["Severity", "Platform", "Category", "Subcategory", "Title",
               "Description", "Affected Count", "Remediation"]
    _xl_apply_header(ws2, headers)
    sorted_findings = sorted(findings, key=lambda f: sev_order.get(f.get("Severity", "medium").lower(), 5))
    for r, f in enumerate(sorted_findings, 2):
        sev = f.get("Severity", "medium").lower()
        rem = f.get("Remediation", {})
        plat = f.get("Platform", "cross-cutting")
        p_meta = _PLATFORM_META.get(plat, {"name": plat})
        cat = f.get("Category", "")
        cat_m = _CATEGORY_META.get(cat, {"name": cat})
        vals = [
            sev.upper(),
            p_meta.get("name", plat),
            cat_m.get("name", cat),
            f.get("Subcategory", "").replace("_", " ").title(),
            f.get("Title", ""),
            f.get("Description", ""),
            f.get("AffectedCount", 0),
            rem.get("Description", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font = _XL_VALUE_FONT
            cell.border = _XL_THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sev_fill = _XL_SEV_FILLS.get(sev)
        if sev_fill:
            sev_cell = ws2.cell(row=r, column=1)
            sev_cell.fill = sev_fill
            sev_cell.font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
    _xl_auto_width(ws2)

    # ── Sheet 3: Affected Resources ──
    ws3 = wb.create_sheet("Affected Resources")
    res_headers = ["Finding Title", "Severity", "Platform", "Category",
                   "Resource Name", "Resource Type", "Resource ID"]
    _xl_apply_header(ws3, res_headers)
    row = 2
    for f in sorted_findings:
        sev = f.get("Severity", "medium").lower()
        plat = f.get("Platform", "cross-cutting")
        p_meta = _PLATFORM_META.get(plat, {"name": plat})
        cat = f.get("Category", "")
        cat_m = _CATEGORY_META.get(cat, {"name": cat})
        for ar in f.get("AffectedResources", []):
            if not isinstance(ar, dict):
                continue
            vals = [
                f.get("Title", ""),
                sev.upper(),
                p_meta.get("name", plat),
                cat_m.get("name", cat),
                str(ar.get("Name", ar.get("name", "—"))),
                str(ar.get("Type", ar.get("type", "—"))),
                str(ar.get("ResourceId", ar.get("resource_id", "—"))),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=c, value=v)
                cell.font = _XL_VALUE_FONT
                cell.border = _XL_THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            sev_fill = _XL_SEV_FILLS.get(sev)
            if sev_fill:
                ws3.cell(row=row, column=2).fill = sev_fill
                ws3.cell(row=row, column=2).font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
            row += 1
    _xl_auto_width(ws3)

    # ── Sheet 4: Category Scores ──
    ws4 = wb.create_sheet("Category Scores")
    cat_headers = ["Category", "Score", "Level", "Finding Count", "Description"]
    _xl_apply_header(ws4, cat_headers)
    for r, (cat_key, meta) in enumerate(_CATEGORY_META.items(), 2):
        cs = cat_scores.get(cat_key, {"Score": 0, "Level": "secure", "FindingCount": 0})
        vals = [
            meta["name"],
            cs.get("Score", 0),
            cs.get("Level", "secure").upper(),
            cs.get("FindingCount", 0),
            meta["description"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.font = _XL_VALUE_FONT
            cell.border = _XL_THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _xl_auto_width(ws4)

    # ── Sheet 5: Platform Breakdown ──
    ws5 = wb.create_sheet("Platform Breakdown")
    plat_headers = ["Platform", "Finding Count"]
    _xl_apply_header(ws5, plat_headers)
    for r, (plat_key, pmeta) in enumerate(_PLATFORM_META.items(), 2):
        count = plat_dist.get(plat_key, 0)
        ws5.cell(row=r, column=1, value=pmeta["name"]).font = _XL_VALUE_FONT
        ws5.cell(row=r, column=1).border = _XL_THIN_BORDER
        ws5.cell(row=r, column=2, value=count).font = _XL_VALUE_FONT
        ws5.cell(row=r, column=2).border = _XL_THIN_BORDER
    _xl_auto_width(ws5)

    # ── Sheet 6: Remediation ──
    ws6 = wb.create_sheet("Remediation")
    rem_headers = ["Finding Title", "Severity", "Remediation", "Azure CLI", "PowerShell", "Portal Steps"]
    _xl_apply_header(ws6, rem_headers)
    row = 2
    for f in sorted_findings:
        rem = f.get("Remediation", {})
        if not rem:
            continue
        sev = f.get("Severity", "medium").lower()
        steps = rem.get("PortalSteps", [])
        vals = [
            f.get("Title", ""),
            sev.upper(),
            rem.get("Description", ""),
            rem.get("AzureCLI", ""),
            rem.get("PowerShell", ""),
            "\n".join(f"{i}. {s}" for i, s in enumerate(steps, 1)) if steps else "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws6.cell(row=row, column=c, value=v)
            cell.font = _XL_VALUE_FONT
            cell.border = _XL_THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sev_fill = _XL_SEV_FILLS.get(sev)
        if sev_fill:
            ws6.cell(row=row, column=2).fill = sev_fill
            ws6.cell(row=row, column=2).font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
        row += 1
    _xl_auto_width(ws6)

    wb.save(out_path)
    log.info("[AIAgentSecurityExcel] Written to %s", out_path)
    return out_path
