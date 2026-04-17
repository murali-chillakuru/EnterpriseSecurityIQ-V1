"""
Custom Report Builder — Generate focused HTML / PDF / Excel reports
from session assessment data based on natural-language topics.

Reuses shared_theme CSS/JS, pdf_export, and openpyxl patterns from
existing report modules.  Called by the generate_custom_report tool.
"""

from __future__ import annotations

import pathlib
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.reports.shared_theme import (
    get_css, get_js, esc, severity_badge, score_color,
    format_date_short, sev_order, SEVERITY_COLORS, VERSION,
)
from app.reports.pdf_export import html_to_pdf
from app.logger import log

# ── Excel styling (matches existing reports) ──────────────────────────────

_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Segoe UI", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="EDEBE9"),
    right=Side(style="thin", color="EDEBE9"),
    top=Side(style="thin", color="EDEBE9"),
    bottom=Side(style="thin", color="EDEBE9"),
)
_SEVERITY_FILLS = {
    "critical": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "high": PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),
    "medium": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "low": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
}


# ── Data extraction from session state ────────────────────────────────────

_SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}


def _extract_findings(state: dict, topic: str, severity_filter: str | None = None,
                       category_filter: str | None = None) -> tuple[list[dict], dict]:
    """Pull relevant findings from session state based on topic keywords.

    Returns (findings_list, metadata_dict).
    """
    t = topic.lower()
    findings: list[dict] = []
    meta: dict[str, Any] = {"sources": [], "title": "Custom Report"}

    # Data Security
    ds = state.get("data_security_results")
    if ds:
        ds_findings = ds.get("Findings", [])
        ds_scores = ds.get("DataSecurityScores", {})
        # Include if topic mentions data-security keywords, or if no specific type detected (grab all)
        include_ds = any(kw in t for kw in [
            "data security", "storage", "database", "keyvault", "key vault",
            "encryption", "classification", "data lifecycle", "dlp", "blob",
            "sql", "cosmos", "data exposure",
        ])
        if include_ds or "all" in t:
            for f in ds_findings:
                f["_source"] = "Data Security"
            findings.extend(ds_findings)
            meta["sources"].append("Data Security")
            meta["data_security_score"] = ds_scores.get("OverallScore", 0)

    # Risk Analysis
    risk = state.get("risk_results")
    if risk:
        risk_findings = risk.get("Findings", [])
        risk_scores = risk.get("RiskScores", {})
        include_risk = any(kw in t for kw in [
            "risk", "threat", "vulnerability", "identity risk", "network risk",
            "defender", "insider", "config drift", "posture", "gap",
        ])
        if include_risk or "all" in t:
            for f in risk_findings:
                f["_source"] = "Risk Analysis"
            findings.extend(risk_findings)
            meta["sources"].append("Risk Analysis")
            meta["risk_score"] = risk_scores.get("OverallScore", 0)

    # Copilot Readiness
    cr = state.get("copilot_readiness_results")
    if cr:
        cr_findings = cr.get("Findings", [])
        cr_scores = cr.get("ReadinessScores", cr.get("CopilotReadinessScores", {}))
        include_cr = any(kw in t for kw in [
            "copilot", "m365", "oversharing", "sensitivity label",
            "readiness", "sharepoint search",
        ])
        if include_cr or "all" in t:
            for f in cr_findings:
                f["_source"] = "Copilot Readiness"
            findings.extend(cr_findings)
            meta["sources"].append("Copilot Readiness")
            meta["copilot_score"] = cr_scores.get("OverallScore", 0)

    # AI Agent Security
    ais = state.get("ai_agent_security_results")
    if ais:
        ais_findings = ais.get("Findings", [])
        ais_scores = ais.get("SecurityScores", ais.get("AIAgentSecurityScores", {}))
        include_ais = any(kw in t for kw in [
            "ai agent", "copilot studio", "foundry", "agent security",
            "content safety",
        ])
        if include_ais or "all" in t:
            for f in ais_findings:
                f["_source"] = "AI Agent Security"
            findings.extend(ais_findings)
            meta["sources"].append("AI Agent Security")
            meta["ai_agent_score"] = ais_scores.get("OverallScore", 0)

    # Compliance findings
    compliance_findings = state.get("findings", [])
    if compliance_findings:
        include_comp = any(kw in t for kw in [
            "compliance", "fedramp", "nist", "cis", "hipaa", "pci",
            "iso", "soc", "gdpr", "csf", "mcsb", "csa",
            "access control", "identity", "data protection",
            "logging", "network", "governance",
        ])
        if include_comp or "all" in t:
            for f in compliance_findings:
                f["_source"] = "Compliance"
            findings.extend(compliance_findings)
            meta["sources"].append("Compliance")
            summary = state.get("summary", {})
            meta["compliance_score"] = summary.get("ComplianceScore", 0)

    # If no specific match, grab everything available
    if not findings:
        for key, label in [
            ("data_security_results", "Data Security"),
            ("risk_results", "Risk Analysis"),
            ("copilot_readiness_results", "Copilot Readiness"),
            ("ai_agent_security_results", "AI Agent Security"),
        ]:
            res = state.get(key)
            if res:
                for f in res.get("Findings", []):
                    f["_source"] = label
                findings.extend(res.get("Findings", []))
                meta["sources"].append(label)
        comp = state.get("findings", [])
        if comp:
            for f in comp:
                f["_source"] = "Compliance"
            findings.extend(comp)
            meta["sources"].append("Compliance")

    # Apply severity filter
    if severity_filter:
        sev_list = [s.strip().lower() for s in severity_filter.split(",")]
        findings = [f for f in findings if f.get("Severity", "").lower() in sev_list]

    # Apply category filter
    if category_filter:
        cat_list = [c.strip().lower() for c in category_filter.split(",")]
        findings = [f for f in findings if f.get("Category", "").lower() in cat_list]

    # Sort by severity
    findings.sort(key=lambda x: _SEV_ORDER.get(x.get("Severity", "").lower(), 5))

    # Build title
    parts = []
    if severity_filter:
        parts.append(severity_filter.title())
    if category_filter:
        parts.append(category_filter.replace(",", ", ").title())
    parts.append(topic.title() if len(topic) < 60 else topic[:57] + "...")
    meta["title"] = " — ".join(parts) if len(parts) > 1 else parts[0]

    return findings, meta


# ── HTML report builder ───────────────────────────────────────────────────

def _build_html(findings: list[dict], meta: dict) -> str:
    """Build a full HTML report with shared theme styling."""
    title = esc(meta.get("title", "Custom Report"))
    now = format_date_short()
    sources = ", ".join(meta.get("sources", []))

    # Severity distribution
    sev_counts: dict[str, int] = {}
    for f in findings:
        s = f.get("Severity", "unknown").lower()
        sev_counts[s] = sev_counts.get(s, 0) + 1

    # Category distribution
    cat_counts: dict[str, int] = {}
    for f in findings:
        c = f.get("Category", f.get("Domain", "Other"))
        cat_counts[c] = cat_counts.get(c, 0) + 1

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} — EnterpriseSecurityIQ</title>
<style>{get_css()}</style>
</head>
<body>
<a href="#main" class="skip-nav">Skip to content</a>

<div class="layout">
<aside class="sidebar" role="navigation" aria-label="Report Navigation">
  <h2>EnterpriseSecurityIQ</h2>
  <p class="version">Custom Report &bull; v{VERSION}</p>
  <button class="theme-btn" onclick="toggleTheme()">Switch to Light</button>
  <nav>
    <div class="nav-group">Sections</div>
    <a href="#overview" class="active" aria-current="true">Overview</a>
    <a href="#severity">Severity Breakdown</a>
    <a href="#findings">Findings</a>
  </nav>
</aside>

<main class="content" id="main">
  <section id="overview" class="section">
    <h1 class="page-title">{title}</h1>
    <div class="meta-bar">
      <span>Generated: {esc(now)}</span>
      <span>Sources: {esc(sources)}</span>
      <span>Findings: {len(findings)}</span>
    </div>

    <div class="stat-grid">
      <div class="stat-card"><div class="stat-value">{len(findings)}</div><div class="stat-label">Total Findings</div></div>
      <div class="stat-card"><div class="stat-value" style="color:{SEVERITY_COLORS.get('critical', '#D13438')}">{sev_counts.get('critical', 0)}</div><div class="stat-label">Critical</div></div>
      <div class="stat-card"><div class="stat-value" style="color:{SEVERITY_COLORS.get('high', '#F7630C')}">{sev_counts.get('high', 0)}</div><div class="stat-label">High</div></div>
      <div class="stat-card"><div class="stat-value" style="color:{SEVERITY_COLORS.get('medium', '#FFB900')}">{sev_counts.get('medium', 0)}</div><div class="stat-label">Medium</div></div>
      <div class="stat-card"><div class="stat-value" style="color:{SEVERITY_COLORS.get('low', '#107C10')}">{sev_counts.get('low', 0)}</div><div class="stat-label">Low</div></div>
    </div>
"""

    # Score cards for available assessments
    score_cards = []
    for key, label in [("compliance_score", "Compliance"), ("data_security_score", "Data Security"),
                        ("risk_score", "Risk"), ("copilot_score", "Copilot Readiness"),
                        ("ai_agent_score", "AI Agent Security")]:
        val = meta.get(key)
        if val is not None:
            color = score_color(val)
            unit = "%" if key == "compliance_score" else "/100"
            score_cards.append(
                f'<div class="stat-card"><div class="stat-value" style="color:{color}">'
                f'{val:.0f}{unit}</div><div class="stat-label">{esc(label)}</div></div>'
            )
    if score_cards:
        html += f'<div class="stat-grid">{"".join(score_cards)}</div>\n'

    html += "</section>\n"

    # ── Severity breakdown ──
    html += '<section id="severity" class="section"><h2>Severity Breakdown</h2>\n'
    html += '<div class="severity-bars">\n'
    max_count = max(sev_counts.values()) if sev_counts else 1
    for sev in ("critical", "high", "medium", "low", "informational"):
        count = sev_counts.get(sev, 0)
        pct = (count / max_count * 100) if max_count else 0
        color = SEVERITY_COLORS.get(sev, "#A8A6A3")
        html += (
            f'<div class="sev-row"><span class="sev-label">{sev.title()}</span>'
            f'<div class="sev-track"><div class="sev-fill" style="width:{pct:.0f}%;background:{color}"></div></div>'
            f'<span class="sev-count">{count}</span></div>\n'
        )
    html += '</div>\n'

    # Category breakdown table
    if cat_counts:
        html += '<h3>By Category</h3>\n<table class="data-table"><thead><tr>'
        html += '<th>Category</th><th>Findings</th></tr></thead><tbody>\n'
        for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
            html += f'<tr><td>{esc(cat.replace("_", " ").title())}</td><td>{cnt}</td></tr>\n'
        html += '</tbody></table>\n'
    html += '</section>\n'

    # ── Detailed findings ──
    html += '<section id="findings" class="section"><h2>Findings</h2>\n'
    for f in findings:
        sev = f.get("Severity", "medium").lower()
        title_text = f.get("Title", f.get("Description", f.get("ControlTitle", "Untitled")))
        source = f.get("_source", "")
        cat = f.get("Category", f.get("Domain", ""))
        affected = f.get("AffectedCount", 0)
        desc = f.get("Description", "")
        ctrl_id = f.get("ControlId", "")
        remediation = ""
        rem_obj = f.get("Remediation")
        if isinstance(rem_obj, dict):
            remediation = rem_obj.get("Description", "")
        elif isinstance(rem_obj, str):
            remediation = rem_obj
        if not remediation:
            remediation = f.get("Recommendation", "")

        html += f'<div class="finding-card {sev}">\n'
        html += f'<div class="finding-header">{severity_badge(sev)}'
        if source:
            html += f' <span class="badge" style="background:var(--gray)">{esc(source)}</span>'
        if ctrl_id:
            html += f' <code>{esc(ctrl_id)}</code>'
        if cat:
            html += f' <span style="color:var(--text-muted);font-size:12px">{esc(cat.replace("_", " ").title())}</span>'
        html += f'</div>\n'
        html += f'<div class="finding-desc"><strong>{esc(title_text)}</strong></div>\n'
        if desc and desc != title_text:
            html += f'<div style="font-size:13px;color:var(--text-secondary);margin-top:4px">{esc(desc)}</div>\n'
        if affected:
            html += f'<div style="font-size:12px;color:var(--text-muted);margin-top:4px">Affected: {affected} resource(s)</div>\n'
        if remediation:
            html += f'<div class="remediation">{esc(remediation)}</div>\n'
        html += '</div>\n'

    if not findings:
        html += '<p class="empty">No findings match the selected criteria.</p>\n'
    html += '</section>\n'

    html += f"""
</main></div>
<button class="back-to-top" aria-label="Back to top">&uarr;</button>
<script>{get_js()}</script>
</body></html>"""

    return html


# ── Excel report builder ──────────────────────────────────────────────────

def _build_excel(findings: list[dict], meta: dict, out_path: pathlib.Path) -> pathlib.Path:
    """Build a multi-sheet Excel workbook from filtered findings."""
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    label_font = Font(name="Segoe UI", bold=True, size=11)
    value_font = Font(name="Segoe UI", size=11)

    summary_rows = [
        ("Report", meta.get("title", "Custom Report")),
        ("Generated", format_date_short()),
        ("Sources", ", ".join(meta.get("sources", []))),
        ("Total Findings", str(len(findings))),
    ]
    for key, label in [("compliance_score", "Compliance Score"),
                        ("data_security_score", "Data Security Score"),
                        ("risk_score", "Risk Score"),
                        ("copilot_score", "Copilot Readiness Score"),
                        ("ai_agent_score", "AI Agent Security Score")]:
        val = meta.get(key)
        if val is not None:
            unit = "%" if "compliance" in key else "/100"
            summary_rows.append((label, f"{val:.0f}{unit}"))

    # Severity counts
    sev_counts: dict[str, int] = {}
    for f in findings:
        s = f.get("Severity", "unknown").lower()
        sev_counts[s] = sev_counts.get(s, 0) + 1
    for sev in ("critical", "high", "medium", "low"):
        summary_rows.append((f"{sev.title()} Findings", str(sev_counts.get(sev, 0))))

    for i, (lbl, val) in enumerate(summary_rows, 1):
        ws.cell(row=i, column=1, value=lbl).font = label_font
        ws.cell(row=i, column=2, value=val).font = value_font
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # ── Sheet 2: All findings ──
    ws2 = wb.create_sheet("Findings")
    headers = ["Severity", "Source", "Category", "Title", "Description",
               "Affected", "Control ID", "Remediation"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws2.freeze_panes = "A2"

    for i, f in enumerate(findings, 2):
        sev = f.get("Severity", "").lower()
        title = f.get("Title", f.get("Description", f.get("ControlTitle", "")))
        desc = f.get("Description", "")
        remediation = ""
        rem_obj = f.get("Remediation")
        if isinstance(rem_obj, dict):
            remediation = rem_obj.get("Description", "")
        elif isinstance(rem_obj, str):
            remediation = rem_obj
        if not remediation:
            remediation = f.get("Recommendation", "")

        values = [
            sev.title(),
            f.get("_source", ""),
            f.get("Category", f.get("Domain", "")),
            title,
            desc if desc != title else "",
            f.get("AffectedCount", 0),
            f.get("ControlId", ""),
            remediation,
        ]
        col_fills: dict[int, PatternFill] = {}
        if sev in _SEVERITY_FILLS:
            col_fills[1] = _SEVERITY_FILLS[sev]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in col_fills:
                cell.fill = col_fills[col]

    # Auto-width
    for col, h in enumerate(headers, 1):
        best = min(len(h) + 4, 55)
        for row in range(2, min(ws2.max_row + 1, 52)):
            cell = ws2.cell(row=row, column=col)
            if cell.value:
                best = max(best, min(len(str(cell.value)) + 2, 55))
        ws2.column_dimensions[get_column_letter(col)].width = best

    wb.save(str(out_path))
    log.info("[CustomReport] Excel: %s (%d findings)", out_path.name, len(findings))
    return out_path


# ── Public API ────────────────────────────────────────────────────────────

async def build_custom_report(
    state: dict,
    topic: str,
    fmt: str,
    output_dir: pathlib.Path,
    severity_filter: str | None = None,
    category_filter: str | None = None,
) -> dict[str, pathlib.Path]:
    """Build a custom report and return {format: path} dict.

    Parameters
    ----------
    state : session state dict
    topic : natural-language description of what to include
    fmt : 'html', 'pdf', 'excel', or 'all'
    output_dir : directory to write files into
    severity_filter : optional comma-separated severities (e.g. 'critical,high')
    category_filter : optional comma-separated categories
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    findings, meta = _extract_findings(state, topic, severity_filter, category_filter)

    if not findings:
        raise ValueError("No findings match the requested topic/filters. Run an assessment first or broaden your criteria.")

    results: dict[str, pathlib.Path] = {}
    slug = topic.lower().replace(" ", "-")[:40]

    need_html = fmt in ("html", "pdf", "all")
    need_pdf = fmt in ("pdf", "all")
    need_excel = fmt in ("excel", "all")

    if need_html or need_pdf:
        html_content = _build_html(findings, meta)
        html_path = output_dir / f"custom-{slug}.html"
        html_path.write_text(html_content, encoding="utf-8")
        results["html"] = html_path
        log.info("[CustomReport] HTML: %s (%d findings)", html_path.name, len(findings))

        if need_pdf:
            pdf_path = await html_to_pdf(html_path)
            if pdf_path:
                results["pdf"] = pdf_path

    if need_excel:
        xlsx_path = output_dir / f"custom-{slug}.xlsx"
        _build_excel(findings, meta, xlsx_path)
        results["excel"] = xlsx_path

    return results
