"""
Excel Export — Per-framework .xlsx with Compliance Report, Gap Analysis, Executive Summary sheets.
No merged cells — every row/column independent and sortable/filterable.
"""

from __future__ import annotations
import pathlib
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.logger import log


# --- Colour palette ---------------------------------------------------------
_GREEN = "107C10"
_RED = "D13438"
_ORANGE = "F7630C"
_YELLOW = "FFB900"
_GREY = "A19F9D"
_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Segoe UI", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="EDEBE9"),
    right=Side(style="thin", color="EDEBE9"),
    top=Side(style="thin", color="EDEBE9"),
    bottom=Side(style="thin", color="EDEBE9"),
)

_STATUS_FILLS = {
    "compliant": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
    "non_compliant": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "partial": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "missing_evidence": PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid"),
    "not_assessed": PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid"),
}

_SEVERITY_FILLS = {
    "critical": PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid"),
    "high": PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),
    "medium": PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid"),
    "low": PatternFill(start_color="E6F5E6", end_color="E6F5E6", fill_type="solid"),
}


def _apply_header(ws, headers: list[str]):
    """Write header row with styling and auto-filter."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def _write_row(ws, row_num: int, values: list, col_fills: dict[int, PatternFill] | None = None):
    """Write a single data row."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_fills and col in col_fills:
            cell.fill = col_fills[col]


def _auto_width(ws, headers: list[str], max_width: int = 55):
    """Set column widths based on header length and data sample."""
    for col, h in enumerate(headers, 1):
        # Start with header length, cap at max_width
        best = min(len(h) + 4, max_width)
        # Sample first 50 data rows
        for row in range(2, min(ws.max_row + 1, 52)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = best


# ---------------------------------------------------------------------------
# Sheet 1: Compliance Report (all control results)
# ---------------------------------------------------------------------------
def _sheet_compliance(wb: Workbook, results: dict, framework: str):
    ws = wb.active
    ws.title = "Compliance Report"
    headers = [
        "Control ID", "Control Title", "Framework", "Domain", "Severity",
        "Status", "Finding Count", "Recommendation",
    ]
    _apply_header(ws, headers)

    controls = results.get("control_results", [])
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    controls_sorted = sorted(controls, key=lambda c: (sev_order.get(c.get("Severity", ""), 5), c.get("ControlId", "")))

    for i, c in enumerate(controls_sorted, 2):
        status = c.get("Status", "")
        severity = c.get("Severity", "")
        fills: dict[int, PatternFill] = {}
        if status in _STATUS_FILLS:
            fills[6] = _STATUS_FILLS[status]
        if severity in _SEVERITY_FILLS:
            fills[5] = _SEVERITY_FILLS[severity]

        _write_row(ws, i, [
            c.get("ControlId", ""),
            c.get("ControlTitle", ""),
            c.get("Framework", framework),
            c.get("Domain", ""),
            severity,
            status,
            c.get("FindingCount", 0),
            c.get("Recommendation", ""),
        ], fills)

    _auto_width(ws, headers)


# ---------------------------------------------------------------------------
# Sheet 2: Gap Analysis (non-compliant / partial / missing only)
# ---------------------------------------------------------------------------
def _sheet_gaps(wb: Workbook, results: dict, framework: str):
    ws = wb.create_sheet("Gap Analysis")
    headers = [
        "Control ID", "Control Title", "Domain", "Severity", "Status",
        "Finding Count", "Description", "Recommendation", "Priority",
    ]
    _apply_header(ws, headers)

    gap_statuses = {"non_compliant", "partial", "missing_evidence"}
    findings = [f for f in results.get("findings", []) if f.get("Status") in gap_statuses]
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    findings_sorted = sorted(findings, key=lambda f: (sev_order.get(f.get("Severity", ""), 5), f.get("ControlId", "")))

    # Deduplicate by ControlId — pick worst finding per control, add count
    seen: dict[str, dict] = {}
    seen_counts: dict[str, int] = {}
    for f in findings_sorted:
        cid = f.get("ControlId", "")
        seen_counts[cid] = seen_counts.get(cid, 0) + 1
        if cid not in seen:
            seen[cid] = f

    priority_map = {"critical": "P0 — Immediate", "high": "P1 — Short-term", "medium": "P2 — Medium-term", "low": "P3 — Long-term"}

    row = 2
    for cid, f in seen.items():
        severity = f.get("Severity", "")
        status = f.get("Status", "")
        fills: dict[int, PatternFill] = {}
        if severity in _SEVERITY_FILLS:
            fills[4] = _SEVERITY_FILLS[severity]
        if status in _STATUS_FILLS:
            fills[5] = _STATUS_FILLS[status]

        _write_row(ws, row, [
            cid,
            f.get("ControlTitle", ""),
            f.get("Domain", ""),
            severity,
            status,
            seen_counts.get(cid, 1),
            f.get("Description", ""),
            f.get("Recommendation", ""),
            priority_map.get(severity, "P3 — Long-term"),
        ], fills)
        row += 1

    # Also add missing evidence entries
    for m in results.get("missing_evidence", []):
        _write_row(ws, row, [
            m.get("ControlId", ""),
            "",
            "",
            m.get("Severity", ""),
            "missing_evidence",
            1,
            f"Missing evidence types: {', '.join(m.get('MissingTypes', []))}",
            "Ensure collectors have required permissions and data sources are available.",
            priority_map.get(m.get("Severity", ""), "P3 — Long-term"),
        ])
        row += 1

    _auto_width(ws, headers)


# ---------------------------------------------------------------------------
# Sheet 3: Executive Summary
# ---------------------------------------------------------------------------
def _sheet_summary(wb: Workbook, results: dict, framework: str):
    ws = wb.create_sheet("Executive Summary")
    headers = ["Metric", "Value"]
    _apply_header(ws, headers)

    summary = results.get("summary", {})

    kv_pairs = [
        ("Framework", framework),
        ("Overall Compliance Score (%)", summary.get("ComplianceScore", "N/A")),
        ("Total Controls Evaluated", summary.get("TotalControls", 0)),
        ("Compliant", summary.get("Compliant", 0)),
        ("Non-Compliant", summary.get("NonCompliant", 0)),
        ("Partial", summary.get("Partial", 0)),
        ("Missing Evidence", summary.get("MissingEvidence", 0)),
        ("Not Assessed", summary.get("NotAssessed", 0)),
        ("Total Findings", summary.get("TotalFindings", 0)),
        ("Critical Findings", summary.get("CriticalFindings", 0)),
        ("High Findings", summary.get("HighFindings", 0)),
        ("Medium Findings", summary.get("MediumFindings", 0)),
        ("Total Evidence Records", summary.get("TotalEvidence", 0)),
    ]

    for i, (metric, value) in enumerate(kv_pairs, 2):
        _write_row(ws, i, [metric, value])

    # Domain scores section
    domain_scores = summary.get("DomainScores", {})
    if domain_scores:
        row = len(kv_pairs) + 3
        _write_row(ws, row, ["", ""])
        row += 1
        domain_headers = ["Domain", "Score (%)", "Compliant", "Total"]
        for col, h in enumerate(domain_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.border = _THIN_BORDER
        row += 1
        for domain, scores in sorted(domain_scores.items()):
            _write_row(ws, row, [
                domain,
                scores.get("Score", 0),
                scores.get("Compliant", 0),
                scores.get("Total", 0),
            ])
            row += 1

    _auto_width(ws, headers)
    # Widen for domain section too
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def generate_excel_report(
    results: dict[str, Any],
    output_dir: str,
    framework: str = "",
) -> str:
    """Generate a single .xlsx with 3 sheets for one framework. Returns path."""
    if not results:
        raise ValueError("Cannot generate Excel report: results dict is empty")

    out = pathlib.Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    fw_label = framework or "compliance"
    filename = f"{fw_label.lower()}-report.xlsx"
    filepath = out / filename

    wb = Workbook()
    _sheet_compliance(wb, results, framework)
    _sheet_gaps(wb, results, framework)
    _sheet_summary(wb, results, framework)
    wb.save(str(filepath))
    log.info("Excel report: %s", filepath)
    return str(filepath)
